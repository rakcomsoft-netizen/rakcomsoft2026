"""
Rakcomsoft - ระบบขายหน้าร้าน v2.1
UI: Modern Dark — ฟอนต์ใหญ่, การ์ด, เส้นแยกสี
"""

from config import DB_FILE, PERMISSIONS, ROLE_DEFAULT_PERMS, RECEIPT_LANG
RL = RECEIPT_LANG["th"]
from theme import (C_SURFACE2, C_BORDER,
                   C_ACCENT, C_ACCENT2, C_YELLOW, C_BLUE,
                   C_GREEN, C_MUTED, C_TEXT, C_WHITE, ORANGE,
                   F_LOGO, F_H1, F_H2, F_BODY, F_SMALL,
                  F_MONO, F_NUM, F_NUM_SM)

from helpers import (
    get_db, get_setting, set_setting,
    to_thai_dt, to_thai_date,
    get_vat_config, calc_vat,
    load_staff_permissions,
    _promptpay_payload,
    _do_export_csv, _open_path,
)
from barcode_utils import (
    validate_barcode,
    generate_barcode,
    check_barcode_duplicate,
    _parse_rakcom_excel,
)
from db_product import (
    get_product,
    get_product_name_stock,
    get_product_by_barcode,
    search_products,
    get_all_products_active,
    get_products_for_sale_list,
    search_products_filtered,
    get_low_stock_products,
    get_negative_stock_products,
    count_low_stock,
    count_negative_stock,
    get_distinct_categories,
    get_distinct_units,
)
from db_sale import (
    get_sale,
    get_sale_with_customer,
    get_sale_items,
    get_sale_item_product_id,
    get_sales_by_customer,
    get_customer_debt,
    get_today_sales_kpi,
    get_daily_sales_range,
    get_top_products,
    get_eod_payment_summary,
    get_eod_summary,
    get_eod_cost,
    get_eod_debt_summary,
    get_report_data,
    # Session 9 — ReportTab filter layer
    get_customer_names,
    get_report_bills,
    get_report_bill_ids_by_product,
    get_report_top_products,
    get_report_daily_summary,
    get_product_sales_report,
    get_product_sale_detail,
)
from db_promotion import (
    # Session 12 — promotion READ layer swap
    get_promotion,
    get_all_promotions,
    get_product_promotions,
    calc_promo_discount,
    get_promo_discount_for_product,
)
# ── Accounts Receivable (Session AR) ──
from db_receivable import (
    init_receivable_schema,
    get_customer_outstanding_balance,   # ★ SaleTab debt label (single source of truth)
    get_total_debt_summary,             # ★ EODTab / Dashboard debt rows (single source of truth)
)
from receivable_tab import ReceivableTab
import stock_service
import promotion_engine
import unit_service
from ui_widgets import (
    card, section_label, divider, stat_card,
    pill_btn, accent_btn, danger_btn, ghost_btn,
    field, make_tree,
)
import tkinter as tk
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import sqlite3
from datetime import datetime, date
import os, sys, csv
APP_VERSION = "2.2.0"   # เวอร์ชันโปรแกรม (ต้องตรงกับ updater.APP_VERSION)
try:
    import updater as _updater   # ระบบอัปเดต (อาจไม่มีไฟล์ในบางเครื่อง)
except Exception:
    _updater = None
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

try:
    from escpos.printer import Usb
    HAS_ESCPOS = True
except ImportError:
    HAS_ESCPOS = False

try:
    import qrcode
    from PIL import Image
    from PIL import ImageTk as ITk
    HAS_QR = True
except ImportError:
    HAS_QR = False
    Image = None
    ITk = None



# ══ BARCODE PRINTER ══════════════════════════════════════
try:
    from barcode_printer import open_barcode_printer, BarcodePrinterWindow
    HAS_BARCODE = True
except ImportError:
    HAS_BARCODE = False

# ══ FONT MANAGER ════════════════════════════════════════
try:
    from font_manager import FM, apply_fonts
    HAS_FONT_MGR = True
except ImportError:
    HAS_FONT_MGR = False
    class _FM:
        primary=FM.primary
        logo=(FM.primary,19,"bold"); h1=(FM.primary,16,"bold")
        h2=(FM.primary,13,"bold"); body=(FM.primary,12)
        small=(FM.primary,11); num=(FM.primary,26,"bold")
        num_sm=("Segoe UI",18,"bold"); mono=("Consolas",12)
        btn=(FM.primary,12,"bold"); table=(FM.primary,11)
    FM=_FM()
    def apply_fonts(root): pass

# ══ SERVICE LAYER (Cloud Ready Architecture) ═════════════
try:
    from db_services import svc, DB_CONFIG, DB_MODES, set_db_mode, Services
    HAS_SERVICES = True
except ImportError:
    HAS_SERVICES = False
    svc = None

# ══ FEATURE FLAGS & EDITION SYSTEM ═══════════════════════
try:
    from feature_flags import has_feature, get_edition_name, locked_screen, EDITION
    HAS_FEATURE_FLAGS = True
except ImportError:
    # Fallback: ทุก feature เปิดใช้งานได้ (ระหว่างพัฒนา)
    HAS_FEATURE_FLAGS = False
    EDITION = "plus"
    def has_feature(k): return True
    def get_edition_name(): return "Plus Edition (Dev)"
    def locked_screen(parent, key): pass

# ══ BACKUP SYSTEM ════════════════════════════════════════
try:
    from backup_system import (
        auto_backup_if_needed, create_local_backup,
        create_multi_location_backup, get_multi_status_summary,
        get_backup_status, list_local_backups, list_gdrive_backups,
        connect_gdrive, disconnect_gdrive, restore_from_zip,
        restore_from_gdrive, upload_to_gdrive
    )
    HAS_BACKUP = True
except ImportError:
    HAS_BACKUP = False

# ══ LICENSE SYSTEM ═══════════════════════════════════════
try:
    from license_system import check_license, can_sell, is_expired, TrialManager, HardwareFingerprint
    HAS_LICENSE = True
except ImportError:
    HAS_LICENSE = False
    def can_sell(): return True
    def is_expired(): return False
    def check_license(): return {"status":"trial","edition":"trial","days_left":7,"message":"","hardware_id":"N/A"}

current_staff = {"id":0,"name":"เจ้าของร้าน","role":"owner"}

# ── Permission module integration ──────────────────────────
import permissions as _perm_mod

def has_permission(perm_key):
    """
    Unified permission check.
    - owner/admin → always True
    - can_xxx keys  → check permissions.CURRENT_EMPLOYEE (new system)
    - old sale_xxx keys → check current_staff["permissions"] set (legacy)
    """
    role = current_staff.get("role", "")
    if role in ("owner", "admin"):
        return True
    # New column-based permissions (can_xxx)
    if perm_key.startswith("can_"):
        emp = _perm_mod.get_current_employee()
        if emp:
            return _perm_mod.has_permission(perm_key)
        # Fallback: if new session not set, allow non-cashier
        return role != "cashier"
    # Legacy set-based permissions (sale_xxx, product_xxx, etc.)
    perms = current_staff.get("permissions", set())
    return perm_key in perms

def require_permission(perm_key, parent=None):
    """Show warning popup if permission denied. Returns True if allowed."""
    if has_permission(perm_key):
        return True
    import tkinter.messagebox as _mb
    _mb.showwarning(
        "ไม่มีสิทธิ์",
        f"คุณไม่มีสิทธิ์: {PERMISSIONS.get(perm_key, perm_key)}\n\nกรุณาติดต่อเจ้าของร้าน",
        parent=parent,
    )
    return False

# ════════ VAT SERVICE ════════════════════════════════════════════


def record_movement(product_id, qty_change, movement_type,
                    ref_id=0, ref_type="", note="", cost=0.0):
    """
    บันทึก Stock Movement ทุกครั้งที่สต็อกเปลี่ยน
    movement_type: sale | restock | return | adjust | import | manual
    เรียกหลัง conn.commit() เสมอ เพื่อหลีกเลี่ยง DB lock
    """
    try:
        conn = get_db()
        # ตรวจว่า table มีอยู่
        tbl = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='stock_movements'"
        ).fetchone()
        if not tbl:
            print(f"[StockCard] ⚠️ stock_movements table ไม่มี — สร้างใหม่")
            conn.executescript("""
                CREATE TABLE IF NOT EXISTS stock_movements (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    product_id INTEGER NOT NULL,
                    product_name TEXT NOT NULL,
                    movement_type TEXT NOT NULL,
                    qty_change INTEGER NOT NULL,
                    qty_before INTEGER NOT NULL,
                    qty_after INTEGER NOT NULL,
                    cost REAL DEFAULT 0,
                    ref_id INTEGER DEFAULT 0,
                    ref_type TEXT DEFAULT '',
                    note TEXT DEFAULT '',
                    staff_id INTEGER DEFAULT 0,
                    staff_name TEXT DEFAULT '',
                    movement_date TEXT NOT NULL
                );
            """)
            conn.commit()
        p = get_product_name_stock(product_id)  # Session 6: swap → db_product
        if not p:
            print(f"[StockCard] ⚠️ ไม่พบสินค้า id={product_id}")
            conn.close(); return
        qty_before = p["stock"]
        qty_after  = qty_before + qty_change
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sname = current_staff.get("name", "")
        sid   = current_staff.get("id", 0)
        conn.execute(
            "INSERT INTO stock_movements "
            "(product_id,product_name,movement_type,qty_change,"
            "qty_before,qty_after,cost,ref_id,ref_type,"
            "note,staff_id,staff_name,movement_date) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (product_id, p["name"], movement_type,
             qty_change, qty_before, qty_after,
             cost, ref_id, ref_type,
             note, sid, sname, now)
        )
        conn.commit(); conn.close()
        print(f"[StockCard] ✅ {p['name']} | {movement_type} | {qty_change:+} | {qty_before}→{qty_after}")
    except Exception as e:
        print(f"[StockCard] ❌ record_movement error: {e}")




# ══ RESPONSIVE MANAGER ═══════════════════════════════════════════════════════
class ResponsiveManager:
    """
    ตรวจสอบขนาดหน้าจอและ DPI แล้วกำหนด UI Mode อัตโนมัติ
    Mode: compact | standard | large | touch
    """
    def __init__(self, root):
        self.root = root
        self.mode = "standard"
        self.scale = 1.0
        self._detect()

    def _detect(self):
        # Get screen size
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        # Get DPI scaling
        try:
            dpi = self.root.winfo_fpixels('1i')
            self.scale = dpi / 96.0
        except:
            self.scale = 1.0

        # Detect mode by resolution
        if sw >= 2560:
            self.mode = "large"
        elif sw >= 1920:
            self.mode = "standard"
        elif sw <= 1366:
            self.mode = "compact"
        else:
            self.mode = "standard"

        # Override to touch if very high DPI (tablet/touch POS)
        if self.scale >= 1.75 and sh <= 900:
            self.mode = "touch"

        print(f"[Responsive] Screen: {sw}x{sh}, DPI Scale: {self.scale:.2f}, Mode: {self.mode}")
        self._apply()

    def _apply(self):
        """ปรับ font constants ตาม mode"""
        global F_H2, F_BODY, F_SMALL, F_MONO, F_NUM, F_NUM_SM

        cfg = {
            "compact":  dict(logo=14,h1=12,h2=10,body=10,small=9, mono=10,num=20,num_sm=13),
            "standard": dict(logo=19,h1=16,h2=13,body=12,small=11,mono=12,num=26,num_sm=18),
            "large":    dict(logo=22,h1=18,h2=15,body=14,small=13,mono=14,num=32,num_sm=22),
            "touch":    dict(logo=22,h1=18,h2=15,body=15,small=13,mono=14,num=34,num_sm=24),
        }
        s = cfg.get(self.mode, cfg["standard"])

        _FNR = FM.primary
        F_LOGO   = (_FNR, s["logo"],   "bold")
        F_H1     = (_FNR, s["h1"],     "bold")
        F_H2     = (_FNR, s["h2"],     "bold")
        F_BODY   = (_FNR, s["body"])
        F_SMALL  = (_FNR, s["small"])
        F_MONO   = ("Consolas", s["mono"])
        F_NUM    = ("Segoe UI", s["num"],    "bold")
        F_NUM_SM = ("Segoe UI", s["num_sm"], "bold")

    def right_panel_width(self):
        """ความกว้าง right panel ตาม mode"""
        return {"compact":260,"standard":360,"large":420,"touch":400}.get(self.mode,360)

    def row_height(self):
        """ความสูง row ใน treeview ตาม mode"""
        return {"compact":26,"standard":36,"large":42,"touch":52}.get(self.mode,36)

    def btn_pady(self):
        """padding ปุ่มตาม mode"""
        return {"compact":4,"standard":8,"large":12,"touch":16}.get(self.mode,8)

    def checkout_font_size(self):
        """ขนาด font ปุ่มชำระเงินตาม mode"""
        return {"compact":11,"standard":16,"large":18,"touch":20}.get(self.mode,16)

    def maximize(self):
        """เปิดหน้าต่างแบบ Maximized"""
        try:
            self.root.state("zoomed")  # Windows
        except:
            try:
                self.root.attributes("-zoomed", True)  # Linux
            except:
                sw = self.root.winfo_screenwidth()
                sh = self.root.winfo_screenheight()
                self.root.geometry(f"{sw}x{sh}+0+0")


# Global responsive manager instance
_responsive = None

def get_responsive():
    return _responsive


# ══ DATABASE ══════════════════════════════════════════════


def init_db():
    conn = get_db(); c = conn.cursor()
    c.executescript("""
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            barcode TEXT UNIQUE, name TEXT NOT NULL,
            price REAL NOT NULL, cost REAL DEFAULT 0,
            stock INTEGER DEFAULT 0, min_stock INTEGER DEFAULT 5,
            unit TEXT DEFAULT 'ชิ้น', active INTEGER DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT UNIQUE, name TEXT NOT NULL,
            phone TEXT, points INTEGER DEFAULT 0,
            total_spent REAL DEFAULT 0, created_at TEXT,
            price_level TEXT DEFAULT 'price'
        );
        CREATE TABLE IF NOT EXISTS promotions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL, type TEXT NOT NULL,
            value REAL NOT NULL, min_amount REAL DEFAULT 0,
            product_id INTEGER DEFAULT 0,
            start_date TEXT, end_date TEXT, active INTEGER DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS sales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sale_date TEXT NOT NULL, total REAL NOT NULL,
            discount REAL DEFAULT 0, promo_disc REAL DEFAULT 0,
            net REAL NOT NULL, paid REAL NOT NULL,
            change_amt REAL NOT NULL, customer_id INTEGER DEFAULT 0, note TEXT, payment_method TEXT DEFAULT "cash"
        );
        CREATE TABLE IF NOT EXISTS sale_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sale_id INTEGER NOT NULL, product_id INTEGER NOT NULL,
            name TEXT NOT NULL, price REAL NOT NULL,
            cost REAL DEFAULT 0, qty INTEGER NOT NULL, subtotal REAL NOT NULL
        );
        CREATE TABLE IF NOT EXISTS store_settings (
            key TEXT PRIMARY KEY, value TEXT
        );
        CREATE TABLE IF NOT EXISTS staff (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL, pin TEXT NOT NULL,
            role TEXT DEFAULT "cashier", active INTEGER DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS returns (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            return_date TEXT NOT NULL,
            sale_id INTEGER NOT NULL,
            product_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            price REAL NOT NULL,
            qty INTEGER NOT NULL,
            refund REAL NOT NULL,
            reason TEXT,
            restock INTEGER DEFAULT 1
        );
    """)
    for k,v in [("store_name","Rakcomsoft"),("store_address",""),("store_phone",""),
                ("receipt_footer","ขอบคุณที่ใช้บริการ"),("point_rate","10"),("low_stock_alert","1"),("printer_vid",""),("printer_pid",""),("promptpay_id",""),("quick_sale_mode","0"),("quick_sale_payment","cash"),("quick_sale_print","1"),("quick_sale_new_bill","1"),("neg_stock_policy","warn"),("shift_required","0"),
                ("paper_size","80mm"),("auto_cut","1"),("open_drawer","0"),("preview_before_print","1"),("auto_print_receipt","1"),("print_copy","0"),
                ("printer_name",""),("receipt_header",""),("left_margin","0"),("right_margin","0"),
                ("show_vat","1"),("show_staff","0"),("show_points","1"),("logo_path",""),("receipt_logo_size","medium"),
                ("logo_enabled","1"),("receipt_font","auto"),("receipt_font_size","medium"),("receipt_line_spacing","normal"),("receipt_sharpen","1"),
                ("printer_mode","usb_raw"),("com_port",""),("baudrate","9600"),("escpos_direct","1"),("encoding","tis-620")]:
        c.execute("INSERT OR IGNORE INTO store_settings VALUES (?,?)",(k,v))
    # Add min_stock column if not exists (backward compat)
    try:
        conn.execute("ALTER TABLE products ADD COLUMN category TEXT DEFAULT ''")
    except: pass
    try:
        conn.execute("ALTER TABLE products ADD COLUMN price_a REAL DEFAULT 0")
    except: pass
    try:
        conn.execute("ALTER TABLE products ADD COLUMN price_b REAL DEFAULT 0")
    except: pass
    try:
        conn.execute("ALTER TABLE products ADD COLUMN price_c REAL DEFAULT 0")
    except: pass
    # price_level migration must be independent — NOT nested inside price_c block.
    # If price_c already existed the outer except skipped this silently.
    try:
        conn.execute("ALTER TABLE customers ADD COLUMN price_level TEXT DEFAULT 'price'")
    except: pass
    try:
        c.execute("ALTER TABLE products ADD COLUMN min_stock INTEGER NOT NULL DEFAULT 5")
    except: pass
    try:
        c.execute("ALTER TABLE products ADD COLUMN cost REAL NOT NULL DEFAULT 0")
    except: pass
    try:
        c.execute("ALTER TABLE sales ADD COLUMN promo_disc REAL DEFAULT 0")
    except: pass
    try:
        c.execute("ALTER TABLE sales ADD COLUMN net REAL DEFAULT 0")
    except: pass
    try:
        c.execute("ALTER TABLE sales ADD COLUMN customer_id INTEGER DEFAULT 0")
    except: pass
    try:
        c.execute("ALTER TABLE sales ADD COLUMN payment_method TEXT DEFAULT 'cash'")
    except: pass
    try:
        c.execute("ALTER TABLE sales ADD COLUMN note TEXT DEFAULT ''")
    except: pass
    # ── สถานะบิล + audit การยกเลิก (Phase 2) ──
    try:
        c.execute("ALTER TABLE sales ADD COLUMN status TEXT DEFAULT 'normal'")
    except: pass
    try:
        c.execute("ALTER TABLE sales ADD COLUMN void_at TEXT DEFAULT ''")
    except: pass
    try:
        c.execute("ALTER TABLE sales ADD COLUMN void_by TEXT DEFAULT ''")
    except: pass
    try:
        c.execute("ALTER TABLE sales ADD COLUMN void_reason TEXT DEFAULT ''")
    except: pass
    c.executescript("""
        CREATE TABLE IF NOT EXISTS sale_audit (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sale_id INTEGER NOT NULL,
            action TEXT NOT NULL,
            by_name TEXT DEFAULT '',
            at TEXT NOT NULL,
            reason TEXT DEFAULT ''
        );
    """)
    # suppliers table
    c.executescript("""
        CREATE TABLE IF NOT EXISTS suppliers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT UNIQUE NOT NULL,
            name TEXT NOT NULL,
            contact TEXT DEFAULT '',
            phone TEXT DEFAULT '',
            email TEXT DEFAULT '',
            address TEXT DEFAULT '',
            tax_id TEXT DEFAULT '',
            credit_days INTEGER DEFAULT 0,
            note TEXT DEFAULT '',
            active INTEGER DEFAULT 1,
            created_at TEXT
        );
    """)
    # product_barcodes (Multi-Barcode future)
    c.execute('''
        CREATE TABLE IF NOT EXISTS product_barcodes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL,
            barcode TEXT NOT NULL UNIQUE,
            barcode_type TEXT DEFAULT "main",
            unit TEXT DEFAULT "",
            qty_per_unit REAL DEFAULT 1,
            note TEXT DEFAULT ""
        );
    ''')
    c.execute("UPDATE sales SET net=total-discount WHERE net=0 AND total>0")
    # ── shifts (Open/Close Shift — Feature Phase 1) ─────────
    c.executescript("""
        CREATE TABLE IF NOT EXISTS shifts (
            id                    INTEGER PRIMARY KEY AUTOINCREMENT,
            staff_id              INTEGER NOT NULL,
            staff_name            TEXT,
            open_time             TEXT NOT NULL,
            close_time            TEXT,
            opening_cash          REAL NOT NULL DEFAULT 0,
            closing_cash_counted  REAL,
            expected_cash         REAL,
            cash_variance         REAL,
            total_sales           REAL,
            total_bills           INTEGER,
            status                TEXT NOT NULL DEFAULT 'open',
            note                  TEXT
        );
    """)
    try: c.execute("ALTER TABLE sales ADD COLUMN shift_id INTEGER")
    except: pass
    c.executescript("""
        CREATE TABLE IF NOT EXISTS staff_permissions (
            staff_id INTEGER NOT NULL,
            perm_key TEXT NOT NULL,
            granted INTEGER DEFAULT 1,
            PRIMARY KEY (staff_id, perm_key)
        );
    """)
    try: c.execute("ALTER TABLE staff ADD COLUMN permissions TEXT DEFAULT ''")
    except: pass
    # ── Permission columns migration (can_xxx) ──────────────
    # Ensure all can_xxx columns exist on the staff table
    _perm_cols = [
        "can_settings","can_reports","can_dashboard","can_view_profit",
        "can_delete_bill","can_edit_bill","can_edit_price",
        "can_manage_employees","can_manage_products","can_import_products",
        "can_manage_promotions","can_manage_debt","can_stock_adjust",
        "can_manage_suppliers","can_barcode_print","can_system_tools",
    ]
    _existing = {row[1] for row in conn.execute("PRAGMA table_info(staff)").fetchall()}
    for _col in _perm_cols:
        if _col not in _existing:
            try:
                conn.execute(f"ALTER TABLE staff ADD COLUMN {_col} INTEGER NOT NULL DEFAULT 0")
            except Exception as _ce:
                pass  # already exists or concurrent add
    conn.commit()
    # ── Sync permissions module DB_PATH ─────────────────────
    import permissions as _pm
    _pm.DB_PATH = DB_FILE
    c.execute("SELECT COUNT(*) FROM staff")
    if c.fetchone()[0]==0:
        c.execute("INSERT INTO staff (name,pin,role) VALUES (?,?,?)",("เจ้าของร้าน","1234","owner"))
    c.execute("SELECT COUNT(*) FROM products")
    if c.fetchone()[0]==0:
        for row in [("001","น้ำดื่ม 600ml",7,4,100,3,"ขวด"),("002","ขนมปัง แผ่น",25,18,50,5,"ถุง"),
                    ("003","บะหมี่กึ่งสำเร็จรูป",6,4,4,5,"ซอง"),("004","น้ำอัดลม 325ml",15,10,80,10,"กระป๋อง"),
                    ("005","ไข่ไก่ (แผง 10 ฟอง)",45,38,2,5,"แผง")]:
            c.execute("INSERT INTO products (barcode,name,price,cost,stock,min_stock,unit) VALUES (?,?,?,?,?,?,?)",row)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_products_barcode ON products(barcode)")
    # ── index เร่งรายงานรายสินค้า + drill-down (สำคัญตอนสินค้า 30k-50k) ──
    conn.execute("CREATE INDEX IF NOT EXISTS idx_sale_items_product ON sale_items(product_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_sale_items_sale ON sale_items(sale_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_sales_date ON sales(sale_date)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_products_category ON products(category)")
    conn.commit(); conn.close()





# ══ DESIGN TOKENS ══════════════════════════════════════════
# Palette — Deep Navy + Electric Teal + Warm Accent
C_BG       = "#0d1117"   # page bg
C_SURFACE  = "#161b22"   # card/panel
C_SURFACE2 = "#21262d"   # elevated
C_BORDER   = "#30363d"   # subtle border
C_ACCENT   = "#00d4aa"   # teal primary
C_ACCENT2  = "#ff6b6b"   # coral danger/badge
C_YELLOW   = "#ffd166"   # warning/money
C_BLUE     = "#58a6ff"   # info/link
C_GREEN    = "#3fb950"   # success/profit
C_MUTED    = "#8b949e"   # secondary text
C_TEXT     = "#e6edf3"   # primary text
C_WHITE    = "#ffffff"
ORANGE     = "#e17055"

# Fonts
# Font constants — ใช้ FM.primary (Noto Sans Thai ถ้ามี, ไม่งั้น Segoe UI)
# nametofont ใน apply_fonts จะ apply font ทั้งระบบอัตโนมัติ
_FN = FM.primary  # "Noto Sans Thai" ถ้ามี .ttf ไม่งั้น FM.primary
F_LOGO  = (_FN, 19, "bold")
F_H1    = (_FN, 16, "bold")
F_H2    = (_FN, 13, "bold")
F_BODY  = (_FN, 12)
F_SMALL = (_FN, 11)
F_MONO  = ("Consolas", 12)
F_NUM   = ("Segoe UI", 26, "bold")
F_NUM_SM= ("Segoe UI", 18, "bold")


def _style():
    s = ttk.Style(); s.theme_use("default")
    # Treeview
    s.configure("R.Treeview", background=C_SURFACE2, foreground=C_TEXT,
                 fieldbackground=C_SURFACE2, rowheight=get_responsive().row_height() if get_responsive() else 36, font=F_BODY,
                 borderwidth=0, relief="flat")
    s.configure("R.Treeview.Heading", background=C_SURFACE, foreground=C_MUTED,
                 font=F_H2, relief="flat", borderwidth=0)
    s.map("R.Treeview", background=[("selected","#1f3a5f")], foreground=[("selected",C_ACCENT)])
    # Notebook
    s.configure("R.TNotebook", background=C_BG, borderwidth=0,
                 tabmargins=[0,0,0,0], tabposition="nw")
    s.configure("R.TNotebook.Tab", background=C_SURFACE, foreground=C_MUTED,
                 font=F_H2, padding=[18,8,18,8], borderwidth=0,
                 focuscolor=C_BG)
    s.map("R.TNotebook.Tab",
          background=[("selected",C_BG),("active",C_SURFACE)],
          foreground=[("selected",C_ACCENT),("active",C_TEXT)],
          focuscolor=[("selected",C_BG),("active",C_BG)],
          padding=[("selected",[18,8,18,8])])
    # Separator
    s.configure("TSeparator", background=C_BORDER)
    # Scrollbar
    s.configure("R.Vertical.TScrollbar", background=C_SURFACE2, troughcolor=C_SURFACE,
                 arrowcolor=C_MUTED, borderwidth=0, relief="flat", width=8)

# ══ WIDGET HELPERS ══════════════════════════════════════════
# def card(parent, **kw):
#     f = tk.Frame(parent, bg=C_SURFACE, highlightbackground=C_BORDER,
#                  highlightthickness=1, **kw)
#     return f

# def section_label(parent, text, color=C_MUTED):
#     tk.Label(parent, text=text.upper(), font=F_SMALL, bg=C_SURFACE,
#              fg=color, anchor="w").pack(fill=tk.X, padx=16, pady=(12,4))

# def divider(parent, bg=C_SURFACE):
#     tk.Frame(parent, bg=C_BORDER, height=1).pack(fill=tk.X, padx=0, pady=4)

# def stat_card(parent, label, value_var, icon="", color=C_TEXT, wide=False):
#     w = 2 if wide else 1
#     f = tk.Frame(parent, bg=C_SURFACE2, highlightbackground=C_BORDER,
#                  highlightthickness=1, padx=16, pady=12)
#     f.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=4, pady=4)
#     top = tk.Frame(f, bg=C_SURFACE2); top.pack(fill=tk.X)
#     tk.Label(top, text=icon, font=(FM.primary,17), bg=C_SURFACE2, fg=color).pack(side=tk.LEFT)
#     tk.Label(top, text=label, font=F_SMALL, bg=C_SURFACE2, fg=C_MUTED).pack(side=tk.LEFT, padx=6)
#     tk.Label(f, textvariable=value_var, font=F_NUM_SM, bg=C_SURFACE2, fg=color).pack(anchor="w", pady=(4,0))
#     return f

# def pill_btn(parent, text, cmd, bg=C_SURFACE2, fg=C_TEXT, pad_x=14, pad_y=5):
#     return tk.Button(parent, text=text, command=cmd, font=F_H2,
#                      bg=bg, fg=fg, relief=tk.FLAT, cursor="hand2",
#                      padx=pad_x, pady=pad_y, activebackground=bg,
#                      activeforeground=fg, bd=0)

# def accent_btn(parent, text, cmd, pad_x=16, pad_y=8):
#     return pill_btn(parent, text, cmd, bg=C_ACCENT, fg=C_BG, pad_x=pad_x, pad_y=pad_y)

# def danger_btn(parent, text, cmd):
#     return pill_btn(parent, text, cmd, bg=C_ACCENT2, fg=C_WHITE)

# def ghost_btn(parent, text, cmd):
#     return pill_btn(parent, text, cmd, bg=C_SURFACE2, fg=C_MUTED)

# def field(parent, width=24, justify="left", font=F_BODY, fg=C_TEXT, bg=C_SURFACE2, show=""):
#     kw = dict(font=font, bg=bg, fg=fg, insertbackground=C_ACCENT,
#               relief=tk.FLAT, width=width, justify=justify,
#               highlightbackground=C_BORDER, highlightthickness=1,
#               highlightcolor=C_ACCENT)
#     if show: kw["show"]=show
#     return tk.Entry(parent, **kw)

# def make_tree(parent, cols, widths, height=14, style="R.Treeview"):



# ══ RECEIPT LANGUAGE RESOURCE ════════════════════════════════════════════════
# ตั้งค่าเป็นภาษาไทย (ค่าเริ่มต้น)
# รองรับการเพิ่มภาษาอังกฤษในอนาคตโดยเปลี่ยน R = RECEIPT_LANG["en"]



# ══ RECEIPT ══════════════════════════════════════════════
def _build_lines(sale, items, cust, W=42, staff_name=None):
    """สร้างบรรทัดใบเสร็จ — delegate ไป receipt_service (text building)"""
    import receipt_service
    return receipt_service.build_lines(sale, items, cust, W, staff_name)

# ── โค้ดเพจไทยของเครื่องพิมพ์ (ESC t n) ─────────────────────────────
# เครื่องส่วนใหญ่ (Xprinter/Gprinter/EPSON-compat) ใช้ cp874 ที่ page 21
# ถ้าไทยออกมาเป็นขยะ/กล่อง ลองเปลี่ยนเลขนี้: 30, 255, 20, 0
THAI_CODEPAGE = 21
RECEIPT_WIDTH = 42      # 80mm ฟอนต์ในตัวรับได้ ~48 ; 42 เผื่อขอบ (เท่าของเดิม)


def _win_print_raw(printer_name, data):
    """ส่ง RAW bytes (ESC/POS) ไปเครื่องพิมพ์ Windows ตามชื่อ ผ่าน winspool (ไม่ต้องมี pywin32)
       เขียนข้อมูลลงไฟล์ชั่วคราวก่อน แล้วให้ PowerShell อ่านจากไฟล์
       (กันปัญหา command line ยาวเกินลิมิตเมื่อรูปใหญ่)
       คืน (ok: bool, err: str)"""
    import subprocess, tempfile, os
    safe = (printer_name or "").replace('"', '')
    fd = tempfile.NamedTemporaryFile(suffix=".prn", delete=False)
    fd.write(data); fd.close()
    path = fd.name.replace("\\", "/")          # .NET รับ forward slash ได้
    ps = r'''
$ErrorActionPreference = "Stop"
$src = @"
using System;
using System.Runtime.InteropServices;
public class RawPrint {
  [StructLayout(LayoutKind.Sequential, CharSet=CharSet.Unicode)]
  public struct DOCINFO { public string pDocName; public string pOutputFile; public string pDataType; }
  [DllImport("winspool.drv", CharSet=CharSet.Unicode, SetLastError=true)]
  public static extern bool OpenPrinter(string src, out IntPtr h, IntPtr p);
  [DllImport("winspool.drv", SetLastError=true)] public static extern bool ClosePrinter(IntPtr h);
  [DllImport("winspool.drv", CharSet=CharSet.Unicode, SetLastError=true)]
  public static extern bool StartDocPrinter(IntPtr h, int level, ref DOCINFO di);
  [DllImport("winspool.drv", SetLastError=true)] public static extern bool EndDocPrinter(IntPtr h);
  [DllImport("winspool.drv", SetLastError=true)] public static extern bool StartPagePrinter(IntPtr h);
  [DllImport("winspool.drv", SetLastError=true)] public static extern bool EndPagePrinter(IntPtr h);
  [DllImport("winspool.drv", SetLastError=true)]
  public static extern bool WritePrinter(IntPtr h, byte[] buf, int n, out int written);
  public static bool Send(string printer, byte[] bytes) {
    IntPtr h;
    if (!OpenPrinter(printer, out h, IntPtr.Zero)) return false;
    DOCINFO di = new DOCINFO(); di.pDocName="RakComSoft"; di.pDataType="RAW";
    bool ok=false;
    if (StartDocPrinter(h, 1, ref di)) {
      if (StartPagePrinter(h)) { int w; ok = WritePrinter(h, bytes, bytes.Length, out w); EndPagePrinter(h); }
      EndDocPrinter(h);
    }
    ClosePrinter(h);
    return ok;
  }
}
"@
Add-Type -TypeDefinition $src -Language CSharp
$bytes = [System.IO.File]::ReadAllBytes("%s")
if (-not [RawPrint]::Send("%s", $bytes)) { Write-Error "WritePrinter failed"; exit 1 }
''' % (path, safe)
    flags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
    try:
        r = subprocess.run(["powershell", "-NoProfile", "-Command", ps],
                           capture_output=True, text=True, timeout=25, creationflags=flags)
        return (r.returncode == 0), (r.stderr or "").strip()
    except Exception as ex:
        return False, str(ex)
    finally:
        try: os.unlink(fd.name)
        except Exception: pass



def _escpos_bytes(lines):
    """ห่อ build_lines เป็น ESC/POS bytes: init + ชิดซ้าย + โค้ดเพจไทย + ข้อความ + ตัดกระดาษ"""
    ESC = b"\x1b"; GS = b"\x1d"
    out = bytearray()
    out += ESC + b"@"                          # init
    out += b"\x1c\x2e"                         # FS .  ยกเลิกโหมดอักษรจีน (กันไทยกลายเป็นจีน) ★ ตัวแก้หลัก
    out += ESC + b"a" + b"\x00"                # ชิดซ้าย (ให้ build_lines คุม alignment ทางเดียว)
    out += ESC + b"t" + bytes([THAI_CODEPAGE]) # เลือกโค้ดเพจไทย
    body = "\n".join(lines) + "\n\n\n"
    out += body.encode("cp874", "replace")     # ไทย = cp874/TIS-620
    out += GS + b"V" + b"\x00"                 # ตัดกระดาษ (full cut)
    return bytes(out)


def _thermal_send(sale, items, cust):
    """เส้นทาง thermal USB (escpos) — converge ไป build_lines ตัวเดียว"""
    if not HAS_ESCPOS: return False
    vid = get_setting("printer_vid").strip()
    pid = get_setting("printer_pid").strip()
    if not vid or not pid: return False
    try:
        p = Usb(int(vid, 16), int(pid, 16), timeout=0, in_ep=0x82, out_ep=0x01)
        lines = _build_lines(sale, items, cust, W=RECEIPT_WIDTH,
                             staff_name=current_staff.get("name") if get_setting("show_staff") == "1" else None)
        import receipt_raster
        p._raw(receipt_raster.lines_to_escpos_raster(lines))
        p.close()
        return True
    except Exception as e:
        print("Thermal err:", e); return False


def _win_print_escpos(printer_name, sale, items, cust):
    """เส้นทาง Windows printer — ส่ง RAW ESC/POS (ฟอนต์ในตัวเครื่อง) ผ่าน build_lines ตัวเดียว
       คืน (ok, err)"""
    lines = _build_lines(sale, items, cust, W=RECEIPT_WIDTH,
                        staff_name=current_staff.get("name") if get_setting("show_staff") == "1" else None)
    import receipt_raster
    return _win_print_raw(printer_name, receipt_raster.lines_to_escpos_raster(lines))

def _win_print_text(printer_name, text):
    """พิมพ์ข้อความไปเครื่องพิมพ์ Windows ตามชื่อ (PowerShell Out-Printer — ไม่ต้องมี pywin32)
       คืน (ok: bool, err: str)"""
    import subprocess, tempfile
    fd = tempfile.NamedTemporaryFile("w", suffix=".txt", delete=False, encoding="utf-8-sig")
    fd.write(text); fd.close(); path = fd.name
    safe = (printer_name or "").replace('"', '')
    ps = f'Get-Content -LiteralPath "{path}" -Encoding UTF8 | Out-Printer -Name "{safe}"'
    flags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
    try:
        r = subprocess.run(["powershell", "-NoProfile", "-Command", ps],
                           capture_output=True, text=True, timeout=25, creationflags=flags)
        return (r.returncode == 0), (r.stderr or "").strip()
    except Exception as ex:
        return False, str(ex)
    finally:
        try: os.unlink(path)
        except Exception: pass


def _receipt_promo_label(prod):
    """ฉลากโปรสั้นๆ ต่อบรรทัดใบเสร็จ จาก promo ปัจจุบันของสินค้า ('' = ไม่มี/หมดอายุ)"""
    try:
        if not prod or not promotion_engine.is_active(prod): return ""
        t=(prod.get("promotion_type","") or "")
        v=prod.get("promotion_value",0) or 0
        sp=prod.get("promotion_price",0) or 0
        X=int(prod.get("promotion_buy_qty",0) or 0)
        Y=int(prod.get("promotion_free_qty",0) or 0)
        if t=="percent": return f"ลด {v:g}%"
        if t=="fixed":   return f"ลด {v:g} บ./ชิ้น"
        if t=="special": return f"ราคาพิเศษ {sp:g} บ."
        if t=="bulk":    return f"ซื้อครบ {X} เหลือ {sp:g} บ./ชิ้น"
        if t=="buyx":    return f"ซื้อ {X} แถม {Y}"
        if t=="buyy":    return f"ซื้อ {X} แถมสินค้าอื่น {Y}"
    except Exception: pass
    return ""


# ══ INPUT VALIDATION (กันตัวเลขผิดปกติ) ══════════════════════
def _num_ok(proposed, max_val, decimals):
    """ใช้กับ validatecommand %P — รับเฉพาะตัวเลขที่อยู่ในขอบเขต
    บล็อก: ตัวอักษร, e/E, +/-, ช่องว่าง, inf, nan, scientific notation,
           ทศนิยมเกินกำหนด, ค่าเกิน max_val"""
    s = (proposed or "").replace(",", "")        # อนุญาต comma (จาก auto-format)
    if s in ("", "."): return True               # ระหว่างพิมพ์
    if any(ch in s for ch in "eE+-  \t") : return False
    if s.count(".") > 1: return False
    body = s.replace(".", "")
    if not body.isdigit(): return False          # มีตัวอักษร/อักขระแปลก → บล็อก
    if decimals == 0 and "." in s: return False
    if "." in s and len(s.split(".", 1)[1]) > decimals: return False
    try:
        if float(s) > max_val: return False       # เกินเพดาน → บล็อก
    except ValueError:
        return False
    return True

def _maxlen_ok(proposed, n):
    return len(proposed or "") <= n

def _attach_num_validator(entry, win, max_val, decimals=2):
    """ผูก validator ตัวเลข (เรียก *หลัง* ใส่ค่าเริ่มต้นแล้ว ไม่งั้น validation ปิดเอง)"""
    try:
        vc = (win.register(lambda P: _num_ok(P, max_val, decimals)), "%P")
        entry.config(validate="key", validatecommand=vc)
    except Exception as e:
        print(f"[validator num] {e}")

def _attach_len_validator(entry, win, n):
    try:
        vc = (win.register(lambda P: _maxlen_ok(P, n)), "%P")
        entry.config(validate="key", validatecommand=vc)
    except Exception as e:
        print(f"[validator len] {e}")

def _attach_money_format(entry):
    """comma + 2 ตำแหน่ง ตอนออกจากช่อง / ลบ comma ตอนเข้าช่อง (แก้ง่าย)"""
    def _out(_=None):
        s = entry.get().replace(",", "").strip()
        if s in ("", "."): return
        try: entry.delete(0, tk.END); entry.insert(0, f"{float(s):,.2f}")
        except Exception: pass
    def _in(_=None):
        s = entry.get().replace(",", "")
        entry.delete(0, tk.END); entry.insert(0, s)
    entry.bind("<FocusOut>", _out); entry.bind("<FocusIn>", _in)

def _attach_int_format(entry):
    def _out(_=None):
        s = entry.get().replace(",", "").strip()
        if s == "": return
        try: entry.delete(0, tk.END); entry.insert(0, f"{int(float(s)):,}")
        except Exception: pass
    def _in(_=None):
        s = entry.get().replace(",", "")
        entry.delete(0, tk.END); entry.insert(0, s)
    entry.bind("<FocusOut>", _out); entry.bind("<FocusIn>", _in)

def _safe_float(v, default=0.0):
    try: return float(str(v).replace(",", "").strip() or default)
    except Exception: return default

def _safe_int(v, default=0):
    try: return int(float(str(v).replace(",", "").strip() or default))
    except Exception: return default


def print_receipt(sale_id):
    sale =get_sale(sale_id)           # db_sale — Session 8
    items=get_sale_items(sale_id)     # db_sale — Session 8
    # enrich ฉลากโปรต่อบรรทัด (ข้ามบรรทัดของแถม — ชื่อบอกอยู่แล้ว)
    _eitems=[]
    for it in items:
        d=dict(it); nm=str(d.get("name","") or "")
        d["promo_label"]="" if "(แถม" in nm else _receipt_promo_label(get_product(d.get("product_id"))) if d.get("product_id") else ""
        _eitems.append(d)
    items=_eitems
    conn=get_db()
    cust =conn.execute("SELECT * FROM customers WHERE id=?",(sale["customer_id"],)).fetchone() if sale["customer_id"] else None
    conn.close()
    lines2=_build_lines(sale,items,cust,
                        staff_name=current_staff.get("name") if get_setting("show_staff")=="1" else None)
    text="\n".join(lines2)
    if get_setting("print_copy")=="1":
        # สำเนาเพิ่ม (text-level — ลูกค้า/ร้าน) ; thermal direct เป็น Phase B
        text = text + "\n\n" + "- - - - - -  สำเนา (COPY)  - - - - - -".center(42) + "\n\n" + text
    os.makedirs("receipts",exist_ok=True)
    fname="receipts/receipt_%d_%s.txt"%(sale_id,datetime.now().strftime("%Y%m%d_%H%M%S"))
    with open(fname,"w",encoding="utf-8") as f: f.write(text)

    pname = (get_setting("printer_name") or "").strip()
    preview_on = get_setting("preview_before_print") != "0"

    # auto-print ไปเครื่องพิมพ์ Windows ที่เลือก เมื่อปิด "แสดงตัวอย่างก่อนพิมพ์"
    auto_ok = None
    if pname and not preview_on:
        auto_ok, _werr = _win_print_escpos(pname, sale, items, cust)
    # thermal USB เดิม เฉพาะกรณีไม่ได้เลือกเครื่องพิมพ์ Windows
    thermal_ok = _thermal_send(sale,items,cust) if not pname else False
    ok = bool(auto_ok) or thermal_ok

    win=tk.Toplevel(); win.title(f'{RL["win_title"]} #{sale_id}')
    win.configure(bg=C_BG); win.geometry("500x580"); win.lift()
    hbg=C_ACCENT if ok else C_SURFACE; hfg=C_BG if ok else C_TEXT
    hf=tk.Frame(win,bg=hbg,pady=10); hf.pack(fill=tk.X)
    title_txt = (f'✅ พิมพ์แล้ว #{sale_id}') if ok else (f'{RL["win_title"]} #{sale_id}')
    tk.Label(hf,text=title_txt,font=F_H1,bg=hbg,fg=hfg).pack()
    if ok: tk.Label(hf,text=RL["printed_ok"],font=F_SMALL,bg=C_ACCENT,fg=C_BG).pack()
    t=tk.Text(win,font=(FM.primary,11),bg="#1a1a2e",fg="#e8e8e8",relief=tk.FLAT,padx=14,pady=12)
    t.pack(fill=tk.BOTH,expand=True,padx=10,pady=8)
    t.insert("1.0",text); t.config(state=tk.DISABLED)
    bf=tk.Frame(win,bg=C_BG,pady=8); bf.pack(fill=tk.X,padx=10)
    if pname:
        # เลือกเครื่องพิมพ์ Windows แล้ว → ปุ่มพิมพ์ออกเครื่องนั้น (พิมพ์ซ้ำได้)
        def win_print():
            wok,werr=_win_print_escpos(pname,sale,items,cust)
            if wok: messagebox.showinfo("✅",f'ส่งงานพิมพ์ไปที่ "{pname}" แล้ว',parent=win)
            else: messagebox.showerror("พิมพ์ไม่สำเร็จ",
                    f'พิมพ์ไปที่ "{pname}" ไม่ได้\n{werr}\n\nตรวจสอบว่าเครื่องพิมพ์พร้อมใช้งาน',parent=win)
        accent_btn(bf,f'🖨️  พิมพ์ ({pname})',win_print,pad_x=12,pad_y=6).pack(side=tk.LEFT,padx=3)
    elif not ok:
        # ไม่ได้เลือกเครื่องพิมพ์ Windows → ทาง ESC/POS USB เดิม
        def retry():
            r=_thermal_send(sale,items,cust)
            if r: messagebox.showinfo("✅","พิมพ์เรียบร้อยแล้ว",parent=win)
            elif not HAS_ESCPOS: messagebox.showinfo("เลือกเครื่องพิมพ์",
                    "เลือกเครื่องพิมพ์ที่ ตั้งค่า → เครื่องพิมพ์\n\nหรือต่อ thermal USB: pip install python-escpos + กรอก VID/PID",parent=win)
            else: messagebox.showerror("ข้อผิดพลาด","USB/VID-PID ผิดพลาด ตรวจสอบในตั้งค่า",parent=win)
        accent_btn(bf,RL["btn_print"],retry,pad_x=12,pad_y=6).pack(side=tk.LEFT,padx=3)
    if sys.platform=="win32":
        pill_btn(bf,RL["btn_open"],lambda:os.startfile(fname),bg=C_SURFACE2,fg=C_TEXT).pack(side=tk.LEFT,padx=3)
    pill_btn(bf,RL["btn_copy"],lambda:(win.clipboard_clear(),win.clipboard_append(text)),bg=C_SURFACE2,fg=C_TEXT).pack(side=tk.LEFT,padx=3)
    pill_btn(bf,RL["btn_close"],win.destroy,bg=C_SURFACE,fg=C_MUTED).pack(side=tk.RIGHT,padx=3)





def make_promptpay_qr(phone, amount=None):
    """Return PhotoImage of PromptPay QR or None"""
    if not HAS_QR:
        return None
    try:
        payload = _promptpay_payload(phone, amount)
        qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_M,
                           box_size=6, border=2)
        qr.add_data(payload)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        return img
    except Exception as e:
        print("QR error:", e)
        return None

def check_low_stock(parent_win=None):
    if get_setting("low_stock_alert")!="1": return
    rows = get_low_stock_products()  # Session 6: swap → db_product
    if not rows: return
    def _show():
        root=parent_win or tk._default_root
        win=tk.Toplevel(root)
        win.title("สต็อกใกล้หมด"); win.configure(bg=C_BG); win.geometry("440x360")
        if root:
            rx=root.winfo_x(); ry=root.winfo_y()
            rw=root.winfo_width(); rh=root.winfo_height()
            win.geometry("440x360+{}+{}".format(rx+rw//2-220, ry+rh//2-180))
        win.lift(); win.focus_force()
        hf=tk.Frame(win,bg=C_YELLOW,pady=10); hf.pack(fill=tk.X)
        tk.Label(hf,text="⚠️  สต็อกใกล้หมด / หมด",font=F_H1,bg=C_YELLOW,fg=C_BG).pack()
        frm,tv=make_tree(win,("ชื่อสินค้า","คงเหลือ","ขั้นต่ำ","สถานะ"),(200,80,80,80),height=8)
        frm.pack(fill=tk.BOTH,expand=True,padx=12,pady=8)
        for r in rows:
            tag="out" if r["stock"]==0 else "low"
            min_s = r["min_stock"] if "min_stock" in r.keys() else 5
            tv.insert("","end",values=(r["name"],r["stock"],min_s,"หมด!" if r["stock"]==0 else "ต่ำ"),tags=(tag,))
        accent_btn(win,"ปิด",win.destroy,pad_x=24,pad_y=8).pack(pady=(0,12))
    root=parent_win or tk._default_root
    if root:
        root.after(800, _show)
    else:
        _show()




def _do_export_excel(path, sheets):
    import os
    os.makedirs(os.path.dirname(path) if os.path.dirname(path) else ".", exist_ok=True)
    name = path  # use path directly from file dialog
    if not HAS_EXCEL:
        return None
    os.makedirs("exports", exist_ok=True)
    path = os.path.join("exports", name + "_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".xlsx")
    wb = openpyxl.Workbook(); first = True
    for sname, headers, rows in sheets:
        ws = wb.active if first else wb.create_sheet(sname)
        if first: ws.title = sname; first = False
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F3460")
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append(list(row))
        for col in ws.columns:
            ml = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(ml + 4, 40)
    wb.save(path); return path




# ══ LOGIN ════════════════════════════════════════════════
def show_login(root, on_success):
    global current_staff
    win=tk.Toplevel(root)
    win.title("เข้าสู่ระบบ")
    win.configure(bg=C_BG)
    win.resizable(False,False)
    win.grab_set()
    win.protocol("WM_DELETE_WINDOW", root.destroy)
    win.update_idletasks()
    w = 360
    h = 680  # พอดีจอ 1080p มี taskbar
    x = (win.winfo_screenwidth() - w) // 2
    y = (win.winfo_screenheight() - h) // 2
    win.geometry(f"{w}x{h}+{x}+{y}")

    tk.Frame(win,bg=C_ACCENT,height=4).pack(fill=tk.X)
    tk.Label(win,text="◆ Rakcomsoft",font=(FM.primary,22,"bold"),bg=C_BG,fg=C_ACCENT,pady=16).pack()
    tk.Label(win,text="เลือกพนักงานและกรอก PIN",font=F_SMALL,bg=C_BG,fg=C_MUTED).pack()

    conn=get_db()
    staff_list=conn.execute("SELECT * FROM staff WHERE active=1 ORDER BY role DESC,name").fetchall()
    conn.close()
    names=[]
    for s in staff_list:
        suffix = "(เจ้าของ)" if s["role"]=="owner" else "(แคชเชียร์)"
        names.append("{} {}".format(s["name"], suffix))

    tk.Label(win,text="พนักงาน:",font=F_BODY,bg=C_BG,fg=C_TEXT).pack(anchor="w",padx=30,pady=(10,2))
    staff_var=tk.StringVar()
    cb=tk.OptionMenu(win,staff_var,*names)
    cb.config(font=F_BODY,bg=C_SURFACE2,fg=C_TEXT,
              activebackground=C_ACCENT,activeforeground=C_BG,
              relief=tk.FLAT,highlightthickness=0,
              indicatoron=1,bd=0)
    cb["menu"].config(bg=C_SURFACE2,fg=C_TEXT,
                      activebackground=C_ACCENT,activeforeground=C_BG,
                      font=F_BODY)
    if names: staff_var.set(names[0])
    cb.pack(padx=30,pady=(0,10),ipady=4)

    tk.Label(win,text="PIN:",font=F_BODY,bg=C_BG,fg=C_TEXT).pack(anchor="w",padx=30)
    pin_dots=tk.Label(win,text="",font=(FM.primary,26,"bold"),bg=C_SURFACE,fg=C_ACCENT,width=12,pady=8)
    pin_dots.pack(padx=30,pady=(2,4),fill=tk.X)
    err_lbl=tk.Label(win,text="",font=F_SMALL,bg=C_BG,fg=C_ACCENT2)
    err_lbl.pack()

    pin=[""]

    def press(v):
        if len(pin[0])>=6: return
        pin[0]+=str(v)
        pin_dots.config(text="●"*len(pin[0]))
        err_lbl.config(text="")

    def back():
        pin[0]=pin[0][:-1]
        pin_dots.config(text="●"*len(pin[0]))

    def login():
        global current_staff
        if not staff_list: return
        idx=names.index(staff_var.get()) if staff_var.get() in names else 0
        if idx<0: return
        s=dict(staff_list[idx])
        if s["pin"]==pin[0]:
            # ── Load legacy set-based permissions ──
            s["permissions"] = load_staff_permissions(s["id"])
            current_staff = s
            # ── Load new can_xxx column permissions into perm module ──
            _full = _perm_mod.load_employee_permissions(s["id"])
            if _full:
                _perm_mod.set_current_employee(_full)
            else:
                # Fallback: build a minimal employee dict from current_staff
                _fallback = dict(s)
                _fallback.setdefault("role", s.get("role","cashier"))
                _perm_mod.set_current_employee(_fallback)
            win.destroy()
            on_success()
        else:
            err_lbl.config(text="PIN ไม่ถูกต้อง กรุณาลองใหม่")
            pin[0]=""
            pin_dots.config(text="")

    nf=tk.Frame(win,bg=C_BG); nf.pack(pady=4)
    # rows: 1-2-3 / 4-5-6 / 7-8-9 / ←-0-C
    for row in [[1,2,3],[4,5,6],[7,8,9],["←",0,"C"]]:
        rf=tk.Frame(nf,bg=C_BG); rf.pack()
        for val in row:
            if val=="C":    bg,fg,cmd=C_ACCENT2,C_WHITE,lambda:(pin.__setitem__(0,"") or pin_dots.config(text=""))
            elif val=="←": bg,fg,cmd=C_SURFACE2,C_MUTED,back
            else:           bg,fg,cmd=C_SURFACE,C_TEXT,lambda v=val:press(v)
            tk.Button(rf,text=str(val),command=cmd,
                      font=(FM.primary,15,"bold"),bg=bg,fg=fg,
                      width=4,height=1,relief=tk.FLAT,cursor="hand2",
                      activebackground=bg,bd=0).pack(side=tk.LEFT,padx=3,pady=2)
    # separator + login button
    tk.Frame(win,bg=C_BORDER,height=1).pack(fill=tk.X,padx=20,pady=(4,0))
    tk.Button(win,text="→  เข้าสู่ระบบ",command=login,
              font=(FM.primary,13,"bold"),bg=C_ACCENT,fg=C_BG,
              relief=tk.FLAT,cursor="hand2",pady=8,
              activebackground=C_ACCENT,bd=0).pack(fill=tk.X,padx=20,pady=(6,12))
    win.bind("<Return>",lambda e:login())
    win.bind("<KP_Enter>",lambda e:login())

    # ── พิมพ์ PIN จากแป้นพิมพ์ได้ (เลข/Backspace/Esc) ──
    def _clear_pin():
        pin[0]=""; pin_dots.config(text=""); err_lbl.config(text="")
    def _on_key(e):
        k=e.keysym; ch=e.char
        if k=="BackSpace":            back();        return "break"
        if k in ("Escape","Delete"):  _clear_pin();  return "break"
        if ch and ch.isdigit():       press(int(ch)); return "break"
        return None
    win.bind("<Key>", _on_key)
    cb.bind("<Key>", _on_key)   # กันโฟกัสค้างที่ช่องเลือกพนักงาน
    try: win.after(120, win.focus_force)
    except Exception: pass


# ══ TAB: STAFF ════════════════════════════════════════════════
class StaffTab(tk.Frame):
    def __init__(self,parent):
        super().__init__(parent,bg=C_BG)
        self._build(); self.load()

    def _build(self):
        top=tk.Frame(self,bg=C_BG,pady=10,padx=14); top.pack(fill=tk.X)
        tk.Label(top,text="พนักงานและสิทธิ์",font=F_H1,bg=C_BG,fg=C_TEXT).pack(side=tk.LEFT)
        bf=tk.Frame(top,bg=C_BG); bf.pack(side=tk.RIGHT)
        accent_btn(bf,"+ เพิ่มพนักงาน",self._add,pad_x=14,pad_y=6).pack(side=tk.LEFT,padx=3)
        pill_btn(bf,"แก้ไข",self._edit,bg=C_SURFACE2,fg=C_YELLOW).pack(side=tk.LEFT,padx=3)
        pill_btn(bf,"ลบ",self._delete,bg=C_SURFACE2,fg=C_ACCENT2).pack(side=tk.LEFT,padx=3)

        info=tk.Frame(self,bg=C_SURFACE2,padx=16,pady=8); info.pack(fill=tk.X,padx=14,pady=(0,8))
        tk.Label(info,text="พนักงานที่ Login อยู่:",font=F_SMALL,bg=C_SURFACE2,fg=C_MUTED).pack(side=tk.LEFT)
        self.cur_lbl=tk.Label(info,text="",font=F_BODY,bg=C_SURFACE2,fg=C_ACCENT)
        self.cur_lbl.pack(side=tk.LEFT,padx=8)
        pill_btn(info,"ออกจากระบบ",self._logout,bg=C_SURFACE,fg=C_MUTED).pack(side=tk.RIGHT)

        cols=("ID","ชื่อ","Role","สิทธิ์","สถานะ")
        self.tree=ttk.Treeview(self,columns=cols,show="headings",style="R.Treeview")
        for col,w in zip(cols,[40,180,90,300,70]):
            self.tree.heading(col,text=col)
            self.tree.column(col,width=w,anchor="w" if col in ("ชื่อ","สิทธิ์") else "center")
        self.tree.pack(fill=tk.BOTH,expand=True,padx=14,pady=(0,8))
        self.tree.tag_configure("owner",  foreground=C_ACCENT)
        self.tree.tag_configure("manager",foreground="#58a6ff")
        self.tree.tag_configure("cashier",foreground=C_TEXT)
        tk.Label(self,text="💡 เจ้าของ=ทุกสิทธิ์  |  ผู้จัดการ=ขาย+สินค้า+รายงาน  |  แคชเชียร์=ขาย+เติมสต็อก  |  กำหนดสิทธิ์เองได้ในหน้าแก้ไข",
                 font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(anchor="w",padx=14,pady=(0,8))

    def load(self,*_):
        role_th = "เจ้าของ" if current_staff.get("role")=="owner" else "แคชเชียร์"
        self.cur_lbl.config(text="{} ({})".format(current_staff.get("name",""), role_th))
        conn=get_db()
        rows=conn.execute("SELECT * FROM staff ORDER BY role DESC,name").fetchall()
        conn.close()
        self.tree.delete(*self.tree.get_children())
        # Show/hide A/B/C columns (index 6,7,8)
        try:
            show_abc=getattr(self,"show_abc",None) and self.show_abc.get()
            abc_w=55 if show_abc else 0
            for ci in (6,7,8): self.tree.column("#{}".format(ci+1),width=abc_w,minwidth=0,stretch=False)
        except: pass
        role_map={"owner":"เจ้าของ","manager":"ผู้จัดการ","cashier":"แคชเชียร์"}
        for r in rows:
            rth=role_map.get(r["role"],"แคชเชียร์")
            sth="ใช้งาน" if r["active"] else "ปิด"
            if r["role"]=="owner":
                psumm="ทุกสิทธิ์"
            else:
                pp=load_staff_permissions(r["id"])
                names_=[PERMISSIONS[k] for k in list(pp)[:3] if k in PERMISSIONS]
                psumm=", ".join(names_)+(f" +{len(pp)-3}" if len(pp)>3 else "")
            self.tree.insert("","end",iid=r["id"],tags=(r["role"],),
                values=(r["id"],r["name"],rth,psumm,sth))

    def _sel(self):
        s=self.tree.selection()
        if not s: messagebox.showwarning("เลือก","กรุณาเลือกพนักงานก่อน"); return None
        return int(s[0])

    def _form(self,sid=None):
        conn=get_db()
        s=conn.execute("SELECT * FROM staff WHERE id=?",(sid,)).fetchone() if sid else None
        conn.close()
        win=tk.Toplevel(self); win.title("พนักงาน"); win.configure(bg=C_BG)
        win.grab_set(); win.resizable(False,False)
        tk.Frame(win,bg=C_ACCENT,height=3).pack(fill=tk.X)
        tk.Label(win,text="แก้ไขพนักงาน" if sid else "เพิ่มพนักงาน",
                 font=F_H1,bg=C_BG,fg=C_TEXT,pady=10).pack(anchor="w",padx=16)

        tk.Label(win,text="ชื่อพนักงาน:",font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(anchor="w",padx=16,pady=(8,2))
        name_e=field(win,width=28); name_e.pack(fill=tk.X,padx=16,ipady=5,pady=(0,8))
        if s: name_e.insert(0,s["name"])

        tk.Label(win,text="PIN (ตัวเลข 4-6 หลัก):",font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(anchor="w",padx=16,pady=(6,2))
        pin_row=tk.Frame(win,bg=C_BG); pin_row.pack(fill=tk.X,padx=16,pady=(0,2))
        # show_pin toggle
        show_pin=tk.BooleanVar(value=False)
        pin_e=field(pin_row,width=22,show="●")
        pin_e.pack(side=tk.LEFT,fill=tk.X,expand=True,ipady=5)
        if s: pin_e.insert(0,s["pin"])

        def _toggle_show():
            pin_e.config(show="" if show_pin.get() else "●")
        tk.Checkbutton(pin_row,text="แสดง",variable=show_pin,command=_toggle_show,
                       bg=C_BG,fg=C_MUTED,selectcolor=C_SURFACE2,
                       activebackground=C_BG,font=F_SMALL).pack(side=tk.LEFT,padx=(6,0))

        # ถ้าแก้ไข: แสดง PIN hint และปุ่ม reset
        if s:
            pin_hint=tk.Label(win,
                text=f"💡 PIN ปัจจุบัน: {'●'*len(s['pin'])} ({len(s['pin'])} หลัก)   กดติ๊ก 'แสดง' เพื่อดู หรือพิมพ์ใหม่เพื่อเปลี่ยน",
                font=F_SMALL,bg=C_BG,fg=C_MUTED)
            pin_hint.pack(anchor="w",padx=16,pady=(0,8))
        else:
            tk.Frame(win,bg=C_BG,height=8).pack()

        tk.Label(win,text="สิทธิ์:",font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(anchor="w",padx=16,pady=(6,2))
        rv=tk.StringVar(value=s["role"] if s else "cashier")
        rf=tk.Frame(win,bg=C_BG); rf.pack(anchor="w",padx=16,pady=(0,12))
        for v,l,c_ in [("owner","เจ้าของร้าน — ทุกสิทธิ์",C_ACCENT),
                     ("manager","ผู้จัดการ — ขาย+สินค้า+รายงาน","#58a6ff"),
                     ("cashier","แคชเชียร์ — ขายสินค้า+เติมสต็อก",C_TEXT)]:
            tk.Radiobutton(rf,text=l,variable=rv,value=v,bg=C_BG,fg=c_,
                           selectcolor=C_SURFACE2,activebackground=C_BG,
                           activeforeground=c_,font=F_BODY,
                           command=lambda:_upd_perms()).pack(anchor="w")
        tk.Frame(win,bg=C_BORDER,height=1).pack(fill=tk.X,padx=16,pady=(10,4))
        tk.Label(win,text="สิทธิ์การใช้งาน (กำหนดเอง):",
                 font=(FM.primary,11,"bold"),bg=C_BG,fg=C_TEXT).pack(anchor="w",padx=16,pady=(0,4))
        cur_perms=load_staff_permissions(sid) if sid else set(ROLE_DEFAULT_PERMS["cashier"])
        pf=tk.Frame(win,bg=C_SURFACE2,padx=14,pady=10); pf.pack(fill=tk.X,padx=16,pady=(0,8))
        perm_vars={}
        pkeys=list(PERMISSIONS.items()); mid=(len(pkeys)+1)//2
        lc=tk.Frame(pf,bg=C_SURFACE2); lc.pack(side=tk.LEFT,fill=tk.BOTH,expand=True)
        rc=tk.Frame(pf,bg=C_SURFACE2); rc.pack(side=tk.LEFT,fill=tk.BOTH,expand=True)
        for i,(k,lbl) in enumerate(pkeys):
            pc=lc if i<mid else rc
            var=tk.BooleanVar(value=k in cur_perms); perm_vars[k]=var
            tk.Checkbutton(pc,text=lbl,variable=var,bg=C_SURFACE2,fg=C_TEXT,
                           selectcolor=C_BG,activebackground=C_SURFACE2,
                           font=(FM.primary,10)).pack(anchor="w",pady=1)
        def _upd_perms():
            rp=set(ROLE_DEFAULT_PERMS.get(rv.get(),[]))
            for k,v in perm_vars.items(): v.set(k in rp)

        def save():
            nm=name_e.get().strip(); pn=pin_e.get().strip()
            if not nm:
                messagebox.showerror("ผิดพลาด","กรุณากรอกชื่อ",parent=win); return
            if not pn.isdigit() or not (4<=len(pn)<=6):
                messagebox.showerror("ผิดพลาด","PIN ต้องเป็นตัวเลข 4-6 หลัก",parent=win); return
            conn2=get_db()
            if sid: conn2.execute("UPDATE staff SET name=?,pin=?,role=? WHERE id=?",(nm,pn,rv.get(),sid))
            else:   conn2.execute("INSERT INTO staff (name,pin,role) VALUES (?,?,?)",(nm,pn,rv.get()))
            conn2.commit()
            sid2=sid or conn2.execute("SELECT last_insert_rowid()").fetchone()[0]
            conn2.execute("DELETE FROM staff_permissions WHERE staff_id=?",(sid2,))
            rdef=set(ROLE_DEFAULT_PERMS.get(rv.get(),[]))
            for k,v in perm_vars.items():
                if v.get()!=(k in rdef):
                    conn2.execute("INSERT OR REPLACE INTO staff_permissions (staff_id,perm_key,granted) VALUES (?,?,?)",
                                  (sid2,k,1 if v.get() else 0))
            conn2.commit(); conn2.close()
            win.destroy(); self.load()

        accent_btn(win,"💾 บันทึก",save,pad_x=24,pad_y=8).pack(pady=12)

    def _add(self): self._form()
    def _edit(self):
        sid=self._sel()
        if sid: self._form(sid)

    def _delete(self):
        sid=self._sel()
        if not sid: return
        if current_staff.get("id")==sid:
            messagebox.showwarning("ไม่ได้","ไม่สามารถลบบัญชีที่ใช้งานอยู่"); return
        if messagebox.askyesno("ยืนยัน","ลบพนักงานนี้?"):
            conn=get_db(); conn.execute("DELETE FROM staff WHERE id=?",(sid,))
            conn.commit(); conn.close(); self.load()

    def _logout(self):
        if messagebox.askyesno("ออกจากระบบ","ต้องการออกจากระบบ?"):
            self.winfo_toplevel().withdraw()
            show_login(self.winfo_toplevel(), lambda: self.winfo_toplevel().deiconify())


# ══ TAB: SALE ════════════════════════════════════════════

def check_neg_stock_policy(parent, product_name, stock_now, qty_want):
    """ตรวจสอบ Negative Stock Policy ก่อนขาย
    Returns True=ขายได้, False=ยกเลิก
    Policies: allow | warn (default) | strict
    """
    import cart_manager
    policy = get_setting("neg_stock_policy") or "warn"
    will_be = stock_now - qty_want
    decision = cart_manager.check_stock(stock_now, qty_want, policy)
    if decision == "ok":
        return True
    if decision == "block":
        messagebox.showwarning(
            "สินค้าไม่เพียงพอ",
            f"สินค้า: {product_name}\nคงเหลือ: {stock_now}\n"
            f"ต้องการขาย: {qty_want}\n\nไม่สามารถขายได้ (โหมดเข้มงวด)",
            parent=parent)
        return False
    # warn
    return messagebox.askokcancel(
        "⚠️  สต็อกไม่เพียงพอ",
        f"สินค้า: {product_name}\nคงเหลือ: {stock_now}\n"
        f"ต้องการขาย: {qty_want}\nสต็อกจะติดลบ: {will_be}\n\n"
        "ต้องการดำเนินการต่อหรือไม่?",
        parent=parent)

class SaleTab(tk.Frame):
    def __init__(self,parent):
        super().__init__(parent,bg=C_BG)
        self.cart=[]; self.customer=None; self.last_sale_id=None; self._build(); self.parked_bills=[]

    def _build(self):
        # ── Left ──
        left=tk.Frame(self,bg=C_BG)
        left.pack(side=tk.LEFT,fill=tk.BOTH,expand=True,padx=(12,6),pady=12)

        # Quick Sale Mode status bar
        self.qs_bar=tk.Frame(left,bg="#e67e22")
        self.qs_bar_lbl=tk.Label(self.qs_bar,text="",
            font=(FM.primary,12,"bold"),bg="#e67e22",fg="#ffffff",pady=6,padx=14)
        self.qs_bar_lbl.pack(fill=tk.X)
        # will pack/unpack dynamically via _refresh_qs_bar

        # search bar
        self.sb_card=card(left); sb=self.sb_card; sb.pack(fill=tk.X,pady=(0,8))
        sf=tk.Frame(sb,bg=C_SURFACE,pady=10,padx=12); sf.pack(fill=tk.X)
        tk.Label(sf,text="🔍",font=(FM.primary,17),bg=C_SURFACE,fg=C_ACCENT).pack(side=tk.LEFT)
        tk.Label(sf,text="Double-click เปิดรายการ",font=(FM.primary,10),bg=C_SURFACE,fg=C_MUTED).pack(side=tk.RIGHT,padx=6)
        tk.Label(sf,text="F1 = เริ่มขายใหม่  |  F5 = บันทึก/ชำระเงิน  |  Double Click = แก้ไขรายการ  |  Delete = ลบรายการ",font=(FM.primary,10),bg=C_SURFACE,fg=C_MUTED).pack(side=tk.RIGHT,padx=6)
        self.sv=tk.StringVar()
        self.se=field(sf,width=32,font=(FM.primary,13)); self.se.configure(textvariable=self.sv)
        self.se.pack(side=tk.LEFT,padx=8,ipady=5,fill=tk.X,expand=True)
        self.se.bind("<Return>",self._search)
        self.se.bind("<Double-Button-1>",lambda e:self._browse_products())
        self.se.bind("<Button-1>",lambda e:self._on_search_click(e))
        accent_btn(sf,"เพิ่มสินค้า",self._open_picker,pad_x=14,pad_y=6).pack(side=tk.LEFT)
        pill_btn(sf,"🖨️ พิมพ์ล่าสุด",self._print_last,bg=C_SURFACE2,fg=C_TEXT).pack(side=tk.LEFT,padx=(6,0))
        # customer
        cf2=tk.Frame(sb,bg=C_SURFACE,padx=12,pady=6); cf2.pack(fill=tk.X)
        tk.Label(cf2,text="👤",font=(FM.primary,13),bg=C_SURFACE,fg=C_MUTED).pack(side=tk.LEFT)
        self.clbl=tk.Label(cf2,text="ไม่ระบุลูกค้า",font=F_BODY,bg=C_SURFACE,fg=C_MUTED)
        self.clbl.pack(side=tk.LEFT,padx=8)
        # ยอดค้างชำระ — แสดงเมื่อเลือกลูกค้า
        self.debt_lbl=tk.Label(cf2,text="",font=(FM.primary,11,"bold"),
                               bg=C_SURFACE,fg="#ff6b6b")
        self.debt_lbl.pack(side=tk.LEFT,padx=(0,8))
        pill_btn(cf2,"👤 ลูกค้า",self._pick_cust,bg=C_SURFACE2,fg=C_BLUE).pack(side=tk.LEFT)
        pill_btn(cf2,"⏸ พักบิล",self._park_recall,bg=C_SURFACE2,fg=C_YELLOW).pack(side=tk.LEFT,padx=(4,0))
        pill_btn(cf2,"✕",self._clear_cust,bg=C_SURFACE,fg=C_MUTED,pad_x=6).pack(side=tk.LEFT,padx=(2,0))
        # cart table
        ct=card(left); ct.pack(fill=tk.BOTH,expand=True)
        # header row
        hdr=tk.Frame(ct,bg=C_SURFACE,pady=8); hdr.pack(fill=tk.X)
        tk.Label(hdr,text="🛒  ตะกร้าสินค้า",font=F_H1,bg=C_SURFACE,fg=C_TEXT,padx=16).pack(side=tk.LEFT)

        tk.Label(hdr,text="Double Click = แก้ไขรายการ  ·  Delete = ลบรายการ",font=F_SMALL,bg=C_SURFACE,fg=C_MUTED,padx=16).pack(side=tk.RIGHT)
        tk.Frame(ct,bg=C_ACCENT,height=2).pack(fill=tk.X)  # accent line top

        cols=("ชื่อสินค้า","ราคา/ชิ้น","จำนวน","โปรโมชั่น","ส่วนลด/หน่วย","รวม (฿)")
        frm,self.tree=make_tree(ct,cols,(200,80,60,80,80,90),height=13)
        frm.pack(fill=tk.BOTH,expand=True,padx=0,pady=0)
        self.tree.configure(style="R.Treeview")
        self.tree.bind("<Delete>",self._del); self.tree.bind("<Double-1>",self._edit_qty)

        # ── Right ──
        _rpw = get_responsive().right_panel_width() if get_responsive() else 360
        right=tk.Frame(self,bg=C_BG,width=_rpw)
        right.pack(side=tk.RIGHT,fill=tk.BOTH,padx=(6,8),pady=6)
        right.pack_propagate(False)

        # ── bottom_action_frame — pack BOTTOM first to reserve space ──
        bot=tk.Frame(right,bg=C_BG)
        bot.pack(side=tk.BOTTOM,fill=tk.X,pady=(4,4))

        # ── scrollable content area for right panel ──
        right_canvas=tk.Canvas(right,bg=C_BG,highlightthickness=0)
        right_vsb=ttk.Scrollbar(right,orient="vertical",command=right_canvas.yview,
                                 style="R.Vertical.TScrollbar")
        right_canvas.configure(yscrollcommand=right_vsb.set)
        right_vsb.pack(side=tk.RIGHT,fill=tk.Y)
        right_canvas.pack(side=tk.LEFT,fill=tk.BOTH,expand=True)
        right_inner=tk.Frame(right_canvas,bg=C_BG)
        _rc_win=right_canvas.create_window((0,0),window=right_inner,anchor="nw")
        right_canvas.bind("<Configure>",
            lambda e:right_canvas.itemconfig(_rc_win,width=e.width))
        right_inner.bind("<Configure>",
            lambda e:right_canvas.configure(scrollregion=right_canvas.bbox("all")))
        right_canvas.bind_all("<MouseWheel>",
            lambda e:right_canvas.yview_scroll(int(-1*(e.delta/120)),"units"))
        # alias so all subsequent widget creation targets right_inner
        right=right_inner

        # ── content frame — fill พื้นที่ที่เหลือ ──
        # summary card
        sc=card(right); sc.pack(fill=tk.X,pady=(0,4))
        tk.Frame(sc,bg=C_ACCENT,height=3).pack(fill=tk.X)
        tk.Label(sc,text="สรุปการขาย",font=F_H1,bg=C_SURFACE,fg=C_TEXT,pady=4).pack(anchor="w",padx=12)

        self.tv=tk.StringVar(value="฿ 0.00")
        self.pv=tk.StringVar(value="฿ 0.00")
        self.dv=tk.StringVar(value="฿ 0.00")
        self.nv=tk.StringVar(value="฿ 0.00")
        self.cv=tk.StringVar(value="฿ 0.00")
        self.ptv=tk.StringVar(value="—")

        def srow(lbl,var,color=C_MUTED,big=False):
            f=tk.Frame(sc,bg=C_SURFACE,padx=10,pady=1); f.pack(fill=tk.X)
            f.columnconfigure(0,weight=1); f.columnconfigure(1,weight=1)
            tk.Label(f,text=lbl,font=F_SMALL,bg=C_SURFACE,fg=C_MUTED,anchor="w").grid(row=0,column=0,sticky="w")
            tk.Label(f,textvariable=var,font=(FM.primary,18,"bold") if big else F_BODY,
                     bg=C_SURFACE,fg=color,anchor="e").grid(row=0,column=1,sticky="e")

        srow("ยอดรวม",self.tv,C_TEXT)
        srow("โปรโมชั่น",self.pv,C_YELLOW)
        srow("ส่วนลดเพิ่ม",self.dv,C_YELLOW)
        # 🎁 แถว "ของแถม" — โชว์เฉพาะมีของแถม
        self.fv=tk.StringVar(value="—")
        self.free_row=tk.Frame(sc,bg=C_SURFACE,padx=10,pady=1)
        self.free_row.columnconfigure(0,weight=1); self.free_row.columnconfigure(1,weight=1)
        tk.Label(self.free_row,text="🎁 ของแถม",font=F_SMALL,bg=C_SURFACE,fg=C_GREEN,anchor="w").grid(row=0,column=0,sticky="w")
        tk.Label(self.free_row,textvariable=self.fv,font=F_BODY,bg=C_SURFACE,fg=C_GREEN,anchor="e").grid(row=0,column=1,sticky="e")
        self._sum_div=tk.Frame(sc,bg=C_BORDER,height=1); self._sum_div.pack(fill=tk.X,padx=12,pady=2)
        srow("ยอดสุทธิ",self.nv,"#19E5C5",big=True)
        # VAT rows (ซ่อน/แสดงตาม setting)
        self.vat_row=tk.Frame(sc,bg=C_SURFACE,padx=10,pady=1); self.vat_row.pack(fill=tk.X)
        self.vat_row.columnconfigure(0,weight=1); self.vat_row.columnconfigure(1,weight=1)
        self.vat_lbl_hdr=tk.Label(self.vat_row,text="ภาษี VAT 7%",font=F_SMALL,bg=C_SURFACE,fg=C_MUTED,anchor="w")
        self.vat_lbl_hdr.grid(row=0,column=0,sticky="w")
        self.vat_lbl=tk.Label(self.vat_row,text="฿ 0.00",font=(FM.primary,12),bg=C_SURFACE,fg="#ffd166",anchor="e")
        self.vat_lbl.grid(row=0,column=1,sticky="e")
        self.vat_row.pack_forget()
        self.grand_row=tk.Frame(sc,bg=C_SURFACE,padx=10,pady=2); self.grand_row.pack(fill=tk.X)
        self.grand_row.columnconfigure(0,weight=1); self.grand_row.columnconfigure(1,weight=1)
        tk.Label(self.grand_row,text="รวมทั้งสิ้น (รวม VAT)",font=(FM.primary,11,"bold"),bg=C_SURFACE,fg=C_TEXT,anchor="w").grid(row=0,column=0,sticky="w")
        self.grand_lbl=tk.Label(self.grand_row,text="฿ 0.00",font=(FM.primary,14,"bold"),bg=C_SURFACE,fg=C_ACCENT,anchor="e")
        self.grand_lbl.grid(row=0,column=1,sticky="e")
        self.grand_row.pack_forget()
        srow("แต้มที่ได้",self.ptv,C_BLUE)

        # discount input — รองรับ ฿ และ %
        dc=card(right); dc.pack(fill=tk.X,pady=(0,4))
        # header + toggle
        dh=tk.Frame(dc,bg=C_SURFACE,padx=10,pady=3); dh.pack(fill=tk.X)
        self.disc_lbl=tk.Label(dh,text="ส่วนลดเพิ่มเติม",font=F_SMALL,
                               bg=C_SURFACE,fg=C_MUTED)
        self.disc_lbl.pack(side=tk.LEFT)
        # toggle ฿ / %
        self.disc_mode=tk.StringVar(value="thb")   # "thb" | "pct"
        tf=tk.Frame(dh,bg=C_SURFACE); tf.pack(side=tk.RIGHT)
        self._disc_toggle_btns={}
        for code,lbl in [("thb","฿"),("pct","%")]:
            b=tk.Button(tf,text=lbl,font=(FM.primary,11,"bold"),
                        bg=C_SURFACE2,fg=C_MUTED,relief=tk.FLAT,
                        cursor="hand2",padx=8,pady=2,bd=0,
                        activebackground=C_SURFACE2,
                        command=lambda c=code:self._set_disc_mode(c))
            b.pack(side=tk.LEFT,padx=1)
            self._disc_toggle_btns[code]=b
        self._set_disc_mode("thb", _init=True)   # init — no permission check
        self.disc=field(dc,font=(FM.primary,13),justify="right")
        self.disc.insert(0,"0")
        self.disc.pack(fill=tk.X,padx=10,pady=(0,6),ipady=4)
        self.disc.bind("<KeyRelease>",self._upd)
        try: _attach_num_validator(self.disc, self, 999999.99, 2)
        except Exception as _ve: print(f"[disc validation] {_ve}")

        # payment method
        pm=card(right); pm.pack(fill=tk.X,pady=(0,3))
        tk.Label(pm,text="ช่องทางชำระ",font=F_SMALL,bg=C_SURFACE,fg=C_MUTED,pady=3).pack(anchor="w",padx=10)
        self.pay_method=tk.StringVar(value="cash")
        self._pay_btns={}
        # แถว 1: เงินสด QR บัตร
        btnrow1=tk.Frame(pm,bg=C_SURFACE); btnrow1.pack(fill=tk.X,padx=6,pady=(0,1))
        for code,label,color in [("cash","💵 เงินสด",C_GREEN),("qr","📱 QR",C_BLUE),
                                   ("credit","💳 บัตร",C_YELLOW)]:
            b=tk.Button(btnrow1,text=label,font=(FM.primary,13,"bold"),
                        bg=C_SURFACE2,fg=C_MUTED,relief=tk.FLAT,cursor="hand2",
                        padx=4,pady=4,activebackground=C_SURFACE2,bd=0,
                        command=lambda c=code,cl=color:self._set_pay(c,cl))
            b.pack(side=tk.LEFT,fill=tk.X,expand=True,padx=2)
            self._pay_btns[code]=b
        # แถว 2: เงินเชื่อ + ผสม
        btnrow2=tk.Frame(pm,bg=C_SURFACE); btnrow2.pack(fill=tk.X,padx=6,pady=(1,4))
        for code,label,color in [("debt","📋 เงินเชื่อ","#e67e22"),("mixed","🔀 ผสม","#9b59b6")]:
            b=tk.Button(btnrow2,text=label,font=(FM.primary,13,"bold"),
                        bg=C_SURFACE2,fg=C_MUTED,relief=tk.FLAT,cursor="hand2",
                        padx=6,pady=6,activebackground=C_SURFACE2,bd=0,
                        command=lambda c=code,cl=color:self._set_pay(c,cl))
            b.pack(side=tk.LEFT,fill=tk.X,expand=True,padx=2)
            self._pay_btns[code]=b

        # paid input
        pc2=card(right); pc2.pack(fill=tk.X,pady=(0,3))
        self.paid_lbl=tk.Label(pc2,text="รับเงินสด (฿)",font=F_SMALL,bg=C_SURFACE,fg="#4FC3F7",pady=3)
        self.paid_lbl.pack(anchor="w",padx=16)
        self._set_pay("cash",C_GREEN)  # init default
        self.paid_e=field(pc2,font=(FM.primary,20,"bold"),justify="right",fg="#4FC3F7")
        self.paid_e.pack(fill=tk.X,padx=10,pady=(0,2),ipady=3)
        self.paid_e.bind("<KeyRelease>",self._chg)
        try: _attach_num_validator(self.paid_e, self, 9999999.99, 2)
        except Exception as _ve: print(f"[paid validation] {_ve}")

        # ปุ่มแบงค์ + พอดี
        bill_f=tk.Frame(pc2,bg=C_SURFACE,padx=6,pady=2); bill_f.pack(fill=tk.X)
        for denom in [20,50,100,500,1000]:
            tk.Button(bill_f,text="฿{}".format(denom),
                      font=(FM.primary,11,"bold"),
                      bg=C_SURFACE2,fg=C_YELLOW,
                      relief=tk.FLAT,cursor="hand2",
                      padx=3,pady=3,
                      activebackground=C_SURFACE2,
                      command=lambda d=denom:self._add_bill(d)
                      ).pack(side=tk.LEFT,fill=tk.X,expand=True,padx=2)
        # ปุ่มพอดี
        tk.Button(bill_f,text="พอดี",
                  font=(FM.primary,11,"bold"),
                  bg=C_ACCENT,fg=C_BG,
                  relief=tk.FLAT,cursor="hand2",
                  padx=3,pady=3,
                  activebackground=C_ACCENT,
                  command=self._exact_amount
                  ).pack(side=tk.LEFT,fill=tk.X,expand=True,padx=2)

        chf=tk.Frame(pc2,bg=C_SURFACE,padx=10,pady=4); chf.pack(fill=tk.X)
        chf.columnconfigure(0,weight=1)
        chf.columnconfigure(1,weight=2)
        self.chg_lbl_var=tk.StringVar(value="เงินทอน")
        self.chg_lbl_w=tk.Label(chf,textvariable=self.chg_lbl_var,font=F_SMALL,bg=C_SURFACE,fg="#FFD700",anchor="w")
        self.chg_lbl_w.grid(row=0,column=0,sticky="w")
        self.cv_lbl=tk.Label(chf,textvariable=self.cv,font=(FM.primary,24,"bold"),bg=C_SURFACE,fg="#FFD700",anchor="e")
        self.cv_lbl.grid(row=0,column=1,sticky="ew")

        # ── ใส่ปุ่มใน bot frame ที่ pack ไว้แล้วด้านบน ──
        tk.Button(bot,text="✅  ชำระเงิน  (F5)",command=self._checkout,
                  font=(FM.primary,get_responsive().checkout_font_size() if get_responsive() else 15,"bold"),
                  bg=C_ACCENT,fg=C_BG,relief=tk.FLAT,cursor="hand2",
                  pady=get_responsive().btn_pady() if get_responsive() else 6,
                  activebackground=C_ACCENT,bd=0
                  ).pack(fill=tk.X,pady=(0,4))
        tk.Button(bot,text="🗑️   ล้างตะกร้า",command=self._clear,
                  font=F_H2,bg=C_SURFACE2,fg=C_MUTED,
                  relief=tk.FLAT,cursor="hand2",pady=4,
                  activebackground=C_SURFACE2,bd=0
                  ).pack(fill=tk.X)

        self.after(100,self._bind_hotkeys)
        self._refresh_qs_bar()

    def _refresh_qs_bar(self):
        try:
            if get_setting("quick_sale_mode")=="1":
                pay_map={"cash":"💵 เงินสด","qr":"📱 QR","transfer":"🏦 เงินโอน"}
                pay_th=pay_map.get(get_setting("quick_sale_payment"),"💵 เงินสด")
                txt="⚡ QUICK SALE MODE  |ชำระ: {}  |F1=เริ่ม  F5=บันทึก+พิมพ์ทันที".format(pay_th)
                self.qs_bar_lbl.config(text=txt)
                self.qs_bar.pack(fill=tk.X,pady=(0,4),before=self.sb_card)
            else:
                self.qs_bar.pack_forget()
        except: pass

    def _bind_hotkeys(self):
        root=self.winfo_toplevel()
        root.bind("<F1>",lambda e:self._qs_start())
        root.bind("<F5>",lambda e:self._qs_save() if get_setting("quick_sale_mode")=="1" else self._checkout())
        root.bind("<Escape>",lambda e:self._close_popups())

    def _close_popups(self):
        for w in self.winfo_toplevel().winfo_children():
            if isinstance(w,tk.Toplevel):
                try: w.destroy()
                except: pass

    # ── logic (same as before) ──
    def _add_bill(self,denom):
        if self.pay_method.get() not in ("cash","mixed"): return
        try: cur=float(self.paid_e.get() or 0)
        except: cur=0
        self.paid_e.delete(0,tk.END)
        self.paid_e.insert(0,"%.0f"%(cur+denom))
        self._chg()

    def _exact_amount(self):
        """กรอกยอดพอดีเท่ากับยอดที่ต้องชำระ (รวม VAT)"""
        try:
            net=float(self.nv.get().replace("฿","").replace(",","").strip())
            payable=self._get_payable(net)
            self.paid_e.delete(0,tk.END)
            self.paid_e.insert(0,"%.0f"%payable)
            self._chg()
        except: pass

    def _set_pay(self,code,color):
        self.pay_method.set(code)
        labels={"cash":"รับเงินสด (฿)","qr":"QR / พร้อมเพย์ (฿)","credit":"บัตรเครดิต (฿)","mixed":"รับเงิน (รวมทุกช่องทาง) (฿)","debt":"เงินเชื่อ — ไม่รับเงิน (฿)"}
        self.paid_lbl.config(text=labels.get(code,"รับเงิน (฿)"))
        for c,b in self._pay_btns.items():
            if c==code: b.config(bg=color,fg=C_BG)
            else:        b.config(bg=C_SURFACE2,fg=C_MUTED)
        if code=="qr":
            self._show_qr_hint()

    def _show_qr_hint(self):
        phone=get_setting("store_phone").strip()
        if not phone:
            messagebox.showwarning("ยังไม่ได้ตั้งค่า",
                "กรุณาตั้งค่าเบอร์โทรศัพท์/พร้อมเพย์ก่อน\nไปที่แถบ ⚙️ ตั้งค่า",parent=self.winfo_toplevel())
            return
        try: net=float(self.nv.get().replace("฿","").replace(",","").strip())
        except: net=0
        try: total=float(self.tv.get().replace("฿","").replace(",","").strip())
        except: total=net
        try:
            disc_val=float(self.disc.get() or 0)
            if hasattr(self,"disc_mode") and self.disc_mode.get()=="pct":
                md=round((total-sum(i["pd"]*i["qty"] for i in self.cart))*disc_val/100,2)
            else:
                md=disc_val
        except: md=0
        pd=sum(i["pd"]*i["qty"] for i in self.cart)
        # ยอดที่ลูกค้าต้องชำระ = net + VAT
        payable = self._get_payable(net)
        vat_info = calc_vat(net)
        win=tk.Toplevel(self); win.title("QR พร้อมเพย์"); win.configure(bg=C_BG)
        win.lift(); win.resizable(False,False)
        tk.Frame(win,bg=C_BLUE,height=4).pack(fill=tk.X)
        tk.Label(win,text="📱  QR พร้อมเพย์",font=F_H1,bg=C_BG,fg=C_TEXT,pady=10).pack()
        # Amount display
        if payable>0:
            af=tk.Frame(win,bg=C_SURFACE,pady=6); af.pack(fill=tk.X,padx=16,pady=(0,6))
            tk.Label(af,text="ยอดที่ต้องชำระ",font=F_SMALL,bg=C_SURFACE,fg=C_MUTED).pack()
            tk.Label(af,text=f"฿ {payable:,.2f}",font=(FM.primary,26,"bold"),bg=C_SURFACE,fg=C_ACCENT).pack()
            if vat_info["enabled"] and vat_info["vat_amount"]>0:
                tk.Label(af,text=f"(รวม VAT {vat_info['percent']:.0f}%  ฿{vat_info['vat_amount']:,.2f})",
                         font=F_SMALL,bg=C_SURFACE,fg=C_MUTED).pack()
        # QR Code
        qr_frame=tk.Frame(win,bg=C_BG); qr_frame.pack(pady=4)
        if HAS_QR:
            qr_img = make_promptpay_qr(phone, payable if payable>0 else None)
            if qr_img:
                try:
                    # convert to PIL RGBA then resize
                    pil = qr_img.convert("RGB") if hasattr(qr_img,"convert") else qr_img.get_image().convert("RGB")
                    pil = pil.resize((260,260), Image.LANCZOS)
                    photo = ITk.PhotoImage(pil)
                    lbl = tk.Label(qr_frame,image=photo,bg="white",relief=tk.FLAT)
                    lbl.image = photo
                    lbl.pack(padx=12,pady=4)
                except Exception as e:
                    tk.Label(qr_frame,text="QR Error: %s"%str(e),
                             font=F_SMALL,bg=C_BG,fg=C_ACCENT2,wraplength=260).pack(pady=10)
            else:
                tk.Label(qr_frame,text="ไม่สามารถสร้าง QR ได้\nตรวจสอบเบอร์พร้อมเพย์",
                         font=F_SMALL,bg=C_BG,fg=C_ACCENT2,wraplength=260).pack(pady=10)
        else:
            install_f=tk.Frame(win,bg=C_SURFACE2,padx=16,pady=10); install_f.pack(fill=tk.X,padx=16,pady=4)
            tk.Label(install_f,text="ติดตั้ง qrcode เพื่อแสดง QR",font=F_SMALL,bg=C_SURFACE2,fg=C_YELLOW).pack()
            tk.Label(install_f,text="pip install qrcode pillow",
                     font=("Consolas",12),bg=C_SURFACE2,fg=C_TEXT).pack(pady=4)
            tk.Label(win,text="เบอร์พร้อมเพย์:",font=F_SMALL,bg=C_BG,fg=C_MUTED).pack()
            tk.Label(win,text=phone,font=(FM.primary,26,"bold"),bg=C_BG,fg=C_BLUE,pady=6).pack()
        phone_type = "นิติบุคคล" if len(phone.replace("-","").replace(" ",""))==13 else "บุคคลธรรมดา"
        tk.Label(win,text=f"พร้อมเพย์ {phone_type}: {phone}",font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(pady=(0,4))
        tk.Label(win,text="สแกน QR แล้วโอนเงิน จากนั้นกดปุ่มด้านล่าง",
                 font=F_SMALL,bg=C_BG,fg=C_MUTED,wraplength=300).pack(pady=4)
        def _confirm_qr():
            win.destroy()
            self._save_sale(payable,total,md,pd,payable,0.0,"QR/พร้อมเพย์")

        accent_btn(win,"✅  รับเงินแล้ว",_confirm_qr,pad_x=24,pad_y=10).pack(pady=8)
        win.update_idletasks()
        win.minsize(win.winfo_reqwidth()+20, win.winfo_reqheight()+20)

    def _on_search_click(self,event):
        pass  # single click — do nothing extra

    def _browse_products(self):
        win=tk.Toplevel(self); win.title("เลือกสินค้า")
        win.configure(bg=C_BG); win.geometry("720x520"); win.lift()
        tk.Frame(win,bg=C_ACCENT,height=3).pack(fill=tk.X)
        hf=tk.Frame(win,bg=C_BG,padx=14,pady=10); hf.pack(fill=tk.X)
        tk.Label(hf,text="เลือกสินค้า",font=F_H1,bg=C_BG,fg=C_TEXT).pack(side=tk.LEFT)
        tk.Label(hf,text="Double-click หรือ Enter = เพิ่มเข้าตะกร้า",font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(side=tk.RIGHT)
        sf2=tk.Frame(win,bg=C_BG,padx=14); sf2.pack(fill=tk.X,pady=(0,6))
        bsv=tk.StringVar()
        bse=field(sf2,width=36,font=(FM.primary,13)); bse.configure(textvariable=bsv)
        bse.pack(side=tk.LEFT,padx=(0,6),ipady=5,fill=tk.X,expand=True)
        tf=tk.Frame(win,bg=C_BG); tf.pack(fill=tk.BOTH,expand=True,padx=14,pady=(0,6))
        cols=("บาร์โค้ด","ชื่อสินค้า","ราคา","สต็อก","หน่วย")
        tv=ttk.Treeview(tf,columns=cols,show="headings",height=16,style="R.Treeview")
        for col,w in zip(cols,[90,300,80,70,60]):
            tv.heading(col,text=col)
            tv.column(col,width=w,anchor="center" if col!="ชื่อสินค้า" else "w")
        sb2=ttk.Scrollbar(tf,orient="vertical",command=tv.yview,style="R.Vertical.TScrollbar")
        tv.configure(yscrollcommand=sb2.set)
        tv.pack(side=tk.LEFT,fill=tk.BOTH,expand=True); sb2.pack(side=tk.RIGHT,fill=tk.Y)
        tv.tag_configure("out",foreground=C_MUTED); tv.tag_configure("low",foreground=C_YELLOW)
        all_prods=[]
        def load_prods(*_):
            q=bsv.get().strip()
            rows=search_products(q) if q else get_all_products_active()  # Session 7: swap → db_product
            all_prods.clear(); all_prods.extend(rows)
            tv.delete(*tv.get_children())
            _mids=unit_service.get_multi_ids(DB_FILE)
            for r in rows:
                tag="out" if r["stock"]==0 else ("low" if r["stock"]<=r["min_stock"] else "")
                _m=r["id"] in _mids
                tv.insert("","end",iid=r["id"],tags=(tag,),
                    values=(r["barcode"] or "-",r["name"],"฿{:.2f}".format(r["price"]),r["stock"],
                            (f"📦 {r['unit']}" if _m else r["unit"])))
        def add_sel(event=None):
            # ใช้ identify_row สำหรับ double-click เพราะ selection อาจยังไม่ update
            if event and hasattr(event,"y"):
                iid=tv.identify_row(event.y)
                if iid: tv.selection_set(iid)
            sel=tv.selection()
            if not sel: return
            p=next((x for x in all_prods if x["id"]==int(sel[0])),None)
            if p:
                self._add(p)
                tv.selection_remove(sel)
        bsv.trace_add("write",load_prods)
        tv.bind("<Double-1>",add_sel); tv.bind("<Return>",add_sel)
        load_prods(); bse.focus_set()
        bf2=tk.Frame(win,bg=C_BG,padx=14,pady=8); bf2.pack(fill=tk.X)
        accent_btn(bf2,"+ เพิ่มเข้าตะกร้า",add_sel,pad_x=16,pad_y=8).pack(side=tk.RIGHT)
        pill_btn(bf2,"ปิด",win.destroy,bg=C_SURFACE,fg=C_MUTED,pad_x=16,pad_y=8).pack(side=tk.RIGHT,padx=(0,6))

    def _search(self,event=None):
        q=self.se.get().strip()
        if not q: return
        rows=search_products(q, limit=10)  # Session 7: swap → db_product
        if not rows:
            self.winfo_toplevel().lift()
            messagebox.showwarning("ไม่พบสินค้า",f"ไม่พบสินค้า '{q}' ในระบบ\nกรุณาตรวจสอบบาร์โค้ดหรือชื่อสินค้า",parent=self.winfo_toplevel())
            self.sv.set("")
        elif len(rows)==1:
            self._add(rows[0]); self.sv.set("")  # rows[0] เป็น dict แล้ว ไม่ต้อง dict()
        else:
            self._pick_prod(rows)
        self.se.focus_set()

    def _open_picker(self):
        """เปิดหน้าเลือกสินค้า — แสดงทุกรายการ ค้นหาได้"""
        all_prods=get_all_products_active()  # Session 7: swap → db_product

        win=tk.Toplevel(self); win.title("เลือกสินค้า")
        win.configure(bg=C_BG); win.grab_set()
        win.geometry("680x520")
        tk.Frame(win,bg=C_ACCENT,height=3).pack(fill=tk.X)

        # header
        hf=tk.Frame(win,bg=C_BG,padx=14,pady=8); hf.pack(fill=tk.X)
        tk.Label(hf,text="เลือกสินค้า",font=F_H1,bg=C_BG,fg=C_TEXT).pack(side=tk.LEFT)
        tk.Label(hf,text="Double-click หรือ Enter = เพิ่มเข้าตะกร้า",
                 font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(side=tk.RIGHT)

        # search
        sf2=tk.Frame(win,bg=C_BG,padx=12,pady=4); sf2.pack(fill=tk.X)
        sv2=tk.StringVar()
        se2=field(sf2,width=30); se2.configure(textvariable=sv2)
        se2.pack(side=tk.LEFT,ipady=5,fill=tk.X,expand=True)
        tk.Label(sf2,text=" 🔍",font=(FM.primary,14),bg=C_BG,fg=C_MUTED).pack(side=tk.LEFT)
        bmv2=tk.BooleanVar(value=False)
        tk.Checkbutton(sf2,text="📦 หลายหน่วย",variable=bmv2,command=lambda:_load(sv2.get().strip()),
                       bg=C_BG,fg="#4aa3ff",selectcolor="#0a1f33",activebackground=C_BG,
                       activeforeground="#4aa3ff",font=(FM.primary,11,"bold")).pack(side=tk.LEFT,padx=(8,0))

        # tree
        frm,tv=make_tree(win,
            ("บาร์โค้ด","ชื่อสินค้า","ราคา","สต็อก","หน่วย"),
            (100,240,80,70,60),height=14)
        frm.pack(fill=tk.BOTH,expand=True,padx=12,pady=(0,4))
        tv.tag_configure("multi",background="#13283b")

        def _load(q=""):
            tv.delete(*tv.get_children())
            mids=unit_service.get_multi_ids(DB_FILE); only_multi=bmv2.get()
            for i,r in enumerate(all_prods):
                if q and q.lower() not in r["name"].lower() \
                       and q not in (r["barcode"] or ""):
                    continue
                _ismulti=r["id"] in mids
                if only_multi and not _ismulti: continue
                tags=("multi",) if _ismulti else ("alt" if i%2==0 else "",)
                tv.insert("","end",iid=r["id"],tags=tags,
                    values=(r["barcode"] or "—",r["name"],
                            f"฿{r['price']:.2f}",r["stock"],
                            (f"📦 {r['unit']}" if _ismulti else (r["unit"] or ""))))
        _load()

        def _on_search(*_):
            _load(sv2.get().strip())
        sv2.trace_add("write",_on_search)

        def add_sel(event=None):
            sel=tv.selection()
            if not sel: return
            if event and hasattr(event,"y"):
                iid=tv.identify_row(event.y)
                if iid: tv.selection_set(iid); sel=(iid,)
            p=next((dict(r) for r in all_prods if r["id"]==int(sel[0])),None)
            if p:
                self._add(p)
                tv.selection_remove(sel)

        tv.bind("<Double-1>",add_sel)
        tv.bind("<Return>",add_sel)

        # ปุ่ม
        bf2=tk.Frame(win,bg=C_BG,padx=12,pady=8); bf2.pack(fill=tk.X)
        accent_btn(bf2,"+ เพิ่มเข้าตะกร้า",add_sel,pad_x=16,pad_y=8).pack(side=tk.LEFT)
        pill_btn(bf2,"ปิด",win.destroy,bg=C_SURFACE2,fg=C_MUTED,pad_x=16,pad_y=8).pack(side=tk.RIGHT)

        win.update_idletasks()
        win.minsize(win.winfo_reqwidth(), win.winfo_reqheight())
        se2.focus_set()

    def _pick_prod(self,rows):
        win=tk.Toplevel(self); win.title("เลือกสินค้า"); win.configure(bg=C_BG); win.grab_set()
        tk.Frame(win,bg=C_ACCENT,height=3).pack(fill=tk.X)
        tk.Label(win,text="เลือกสินค้า",font=F_H1,bg=C_BG,fg=C_TEXT,pady=10).pack(anchor="w",padx=16)
        frm,tv=make_tree(win,("บาร์โค้ด","ชื่อสินค้า","ราคา","คงเหลือ"),(100,200,80,80),height=8)
        frm.pack(padx=12,pady=(0,8))
        for r in rows: tv.insert("","end",iid=r["id"],values=(r["barcode"],r["name"],f"{r['price']:.2f}",r["stock"]))
        def sel():
            s=tv.selection()
            if s: self._add(dict(next(r for r in rows if r["id"]==int(s[0])))); self.sv.set(""); win.destroy()
        accent_btn(win,"✔  เลือก",sel,pad_x=20,pad_y=8).pack(pady=8)
        tv.bind("<Double-1>",lambda e:sel())

    def _promo_disc(self,pid,price,prod=None):
        # โปรกลางเดิม (ตาราง promotions)
        central = get_promo_discount_for_product(pid,price)
        # โปรระดับสินค้า (เฟส A: percent/fixed/special)
        pdisc = 0.0
        try:
            if prod is None: prod = get_product(pid)
            if prod: pdisc,_lbl = promotion_engine.unit_discount(prod, price)
        except Exception as _e:
            pdisc = 0.0
        return round(max(central, pdisc), 2)

    def _recompute_promo(self, item):
        """คิดส่วนลดโปรทั้งบรรทัด แล้วเก็บเป็น pd ต่อหน่วย (pd × qty = ส่วนลดบรรทัด)
           ยกเว้น buyx/buyy → ของแถมจัดเป็น 'บรรทัดแถม' (ราคา 0) ไม่ลดราคาบรรทัดนี้"""
        try:
            qty = int(item.get("qty",1) or 1)
            if int(item.get("factor",1) or 1) != 1:   # หน่วยใหญ่ (ลัง/แพ็ค) ไม่คิดโปรต่อหน่วย
                item["pd"]=0.0; item["promo_label"]=""; return
            prod = get_product(item["product_id"])
            ptype = (prod.get("promotion_type","") if prod else "") or ""
            if prod and ptype in ("buyx","buyy"):
                line_disc, lbl = 0.0, ""
            else:
                line_disc, lbl, _free = promotion_engine.line_discount(prod, item["price"], qty) if prod else (0.0,"",0)
            central_line = get_promo_discount_for_product(item["product_id"], item["price"]) * qty
            if central_line >= line_disc:
                line_disc, lbl = central_line, ""
            item["pd"] = round(line_disc/qty, 4) if qty else 0.0
            item["promo_label"] = lbl
        except Exception:
            pass

    def _strip_free(self):
        """ลบบรรทัดของแถมอัตโนมัติทั้งหมด (จะคำนวณใหม่)"""
        self.cart[:] = [it for it in self.cart if not it.get("_free")]

    def _apply_free_items(self):
        """เพิ่มบรรทัด 'ของแถม' อัตโนมัติ (ซื้อ X แถม X/Y) ราคา 0 → โชว์ในตะกร้า + ตัดสต๊อก
           เคารพนโยบายสต๊อกติดลบ (strict=ตัดไม่ให้เกิน, warn/allow=ให้เต็มแต่ทำเครื่องหมาย ⚠️)"""
        self._strip_free()
        strict = (get_setting("neg_stock_policy") or "warn") == "strict"
        # สต๊อกที่ paid lines ใช้ไปแล้ว ต่อสินค้า
        paid_used={}
        for it in self.cart:
            paid_used[it["product_id"]]=paid_used.get(it["product_id"],0)+int(it.get("qty",0) or 0)
        remaining={}   # สต๊อกคงเหลือหลังหัก paid + ของแถมที่จัดไปแล้ว
        extras=[]
        for it in list(self.cart):
            try:
                if int(it.get("factor",1) or 1) != 1: continue   # หน่วยใหญ่ ไม่แถมต่อหน่วย
                prod=get_product(it["product_id"])
                if not prod: continue
                t=(prod.get("promotion_type","") or "")
                if t not in ("buyx","buyy"): continue
                if not promotion_engine.is_active(prod): continue
                X=int(prod.get("promotion_buy_qty",0) or 0)
                Y=int(prod.get("promotion_free_qty",0) or 0)
                if X<=0 or Y<=0: continue
                paid_qty=int(it.get("qty",0) or 0)
                free_qty=(paid_qty//X)*Y
                if free_qty<=0: continue
                if t=="buyx":
                    fp=prod
                else:
                    fpid=int(prod.get("promotion_free_product_id",0) or 0)
                    if not fpid: continue
                    fp=get_product(fpid)
                    if not fp: continue
                fpid=fp["id"]; stock=int(fp.get("stock",0) or 0)
                if fpid not in remaining:
                    remaining[fpid]=stock - paid_used.get(fpid,0)
                give=free_qty
                if strict:
                    give=min(free_qty, max(0, remaining[fpid]))
                    if give<=0: continue
                neg = (remaining[fpid]-free_qty) < 0
                remaining[fpid]-=give
                nm=f"{fp.get('name','')} (แถม)"
                if neg and not strict: nm=f"{fp.get('name','')} (แถม⚠️สต๊อกไม่พอ)"
                extras.append({"product_id":fpid,"name":nm,
                               "price":0.0,"cost":float(fp.get("cost",0) or 0),
                               "qty":give,"subtotal":0.0,"pd":0.0,
                               "item_disc_amt":0.0,"item_disc_mode":"thb","item_disc_val":0,
                               "stock":stock,"_free":True})
            except Exception:
                continue
        self.cart.extend(extras)

    def _add(self,p):
        import cart_manager
        self._strip_free()
        root = self.winfo_toplevel()
        base_unit = (p["unit"] or "ชิ้น")
        base_price = self._get_price_for_customer(p)
        # หน่วยขาย (รวมหน่วยฐาน) — ถ้ามีหลายหน่วยให้เลือกก่อน
        try: units = unit_service.get_units(DB_FILE, p["id"], base_unit=base_unit, base_price=base_price)
        except Exception: units = [{"unit":base_unit,"factor":1,"price":base_price,"is_base":True}]
        if len(units) > 1:
            chosen = self._pick_unit(p, units)
            if not chosen: return
        else:
            chosen = units[0]
        unit_name = chosen["unit"]; factor = int(chosen["factor"] or 1); use_price = float(chosen["price"] or 0)
        pd = self._promo_disc(p["id"], use_price, p) if factor==1 else 0.0
        idx = self._find_cart_line(p["id"], unit_name)
        if idx >= 0:
            need = (self.cart[idx]["qty"]+1)*factor
            if not check_neg_stock_policy(root,p["name"],p["stock"],need): return
            cart_manager.increment(self.cart[idx])
            if factor==1: self._recompute_promo(self.cart[idx])
            self._redraw(); return
        if not check_neg_stock_policy(root,p["name"],p["stock"],factor): return
        it = cart_manager.new_item(p, use_price, pd)
        it["unit"]=unit_name; it["factor"]=factor
        if factor!=1: it["name"]=f'{p["name"]} ({unit_name})'
        self.cart.append(it)
        if factor==1: self._recompute_promo(it)
        self._redraw()

    def _find_cart_line(self, pid, unit):
        """หาบรรทัดในตะกร้าที่ตรงทั้ง product_id และหน่วย (ข้ามบรรทัดของแถม)"""
        for i,it in enumerate(self.cart):
            if it.get("_free"): continue
            if it["product_id"]==pid and (it.get("unit") or "")==unit:
                return i
        return -1

    def _pick_unit(self, p, units):
        """popup เลือกหน่วยขาย → คืน dict หน่วย หรือ None ถ้ายกเลิก"""
        root=self.winfo_toplevel()
        win=tk.Toplevel(root); win.title("เลือกหน่วยขาย"); win.configure(bg=C_BG)
        win.grab_set(); win.resizable(False,False)
        tk.Frame(win,bg=C_ACCENT,height=3).pack(fill=tk.X)
        tk.Label(win,text=f'เลือกหน่วยขาย\n{p["name"]}',font=F_H2,bg=C_BG,fg=C_TEXT,pady=8,
                 justify="left",wraplength=320).pack(anchor="w",padx=16)
        result={"v":None}
        def choose(u): result["v"]=u; win.destroy()
        for u in units:
            txt=f'{u["unit"]}   ฿{u["price"]:,.2f}'
            if not u["is_base"]: txt+=f'   ({u["factor"]} {p["unit"] or "ชิ้น"})'
            tk.Button(win,text=txt,font=(FM.primary,13,"bold"),bg=C_SURFACE2,fg=C_TEXT,
                      relief=tk.FLAT,cursor="hand2",anchor="w",padx=16,pady=10,bd=0,
                      command=lambda uu=u: choose(uu)).pack(fill=tk.X,padx=16,pady=3)
        tk.Button(win,text="ยกเลิก",font=F_SMALL,bg=C_BG,fg=C_MUTED,relief=tk.FLAT,bd=0,
                  cursor="hand2",command=win.destroy).pack(pady=(4,10))
        win.update_idletasks()
        win.geometry(f"+{root.winfo_rootx()+max(root.winfo_width()//2-160,0)}+{root.winfo_rooty()+160}")
        win.wait_window()
        return result["v"]

    def _refresh(self):
        self.tree.delete(*self.tree.get_children())
        for i,item in enumerate(self.cart):
            tag="alt" if i%2==0 else ""
            pd_line=item["pd"]*item["qty"]                 # โปรรวมทั้งบรรทัด
            ds=f"-฿{pd_line:,.2f}" if pd_line>0 else "—"
            ida=item.get("item_disc_amt",0)
            if ida>0:
                idm=item.get("item_disc_mode","thb")
                idv=item.get("item_disc_val",0)
                ids=f"-฿{ida:.2f}" if idm=="thb" else f"-{idv:.1f}% (฿{ida:.2f})"
            else:
                ids="—"
            line_total=item["subtotal"]-pd_line            # ยอดบรรทัดหลังหักโปร
            self.tree.insert("","end",iid=i,tags=(tag,),
                values=(item["name"],f"฿{item['price']:.2f}",item["qty"],ds,ids,f"฿{line_total:,.2f}"))

    def _redraw(self):
        """วาดตะกร้าใหม่ + sync ยอด — รวม _refresh() (Treeview) กับ _upd() (StringVar)"""
        self._apply_free_items()
        self._refresh()
        self._upd()

    def _del(self,event=None):
        import cart_manager
        if not has_permission("sale_delete_item"):
            require_permission("sale_delete_item",self.winfo_toplevel()); return
        s=self.tree.selection()
        if s: cart_manager.remove(self.cart,int(s[0])); self._redraw()

    def _edit_qty(self,event=None):
        s=self.tree.selection()
        if not s: return
        idx=int(s[0]); item=self.cart[idx]
        root=self.winfo_toplevel()

        win=tk.Toplevel(root)
        win.title("แก้ไขรายการ")
        win.configure(bg=C_BG)
        win.grab_set()
        win.resizable(False,False)

        # ── accent top ──────────────────────────────────
        tk.Frame(win,bg=C_ACCENT,height=3).pack(fill=tk.X)

        # ── ชื่อสินค้า ───────────────────────────────────
        tk.Label(win,text=item["name"],font=F_H2,bg=C_BG,fg=C_TEXT,
                 pady=8,padx=20,wraplength=460,justify="left").pack(anchor="w",fill=tk.X)

        # ── label row ────────────────────────────────────
        lf=tk.Frame(win,bg=C_BG,padx=20); lf.pack(fill=tk.X,pady=(4,2))
        tk.Label(lf,text="จำนวน",font=F_SMALL,bg=C_BG,fg=C_MUTED,
                 width=12,anchor="w").pack(side=tk.LEFT)
        tk.Label(lf,text="ส่วนลดรายการ",font=F_SMALL,bg=C_BG,
                 fg=C_MUTED).pack(side=tk.LEFT)

        # ── input row ────────────────────────────────────
        ef=tk.Frame(win,bg=C_BG,padx=20,pady=4); ef.pack(fill=tk.X)

        qty_e=tk.Entry(ef,width=8,font=(FM.primary,14),justify="right",
                       bg=C_SURFACE2,fg=C_TEXT,insertbackground=C_ACCENT,
                       relief=tk.FLAT,highlightthickness=1,
                       highlightbackground=C_BORDER,highlightcolor=C_ACCENT)
        qty_e.insert(0,str(item["qty"]))
        qty_e.pack(side=tk.LEFT,ipady=6)

        tk.Frame(ef,bg=C_BG,width=16).pack(side=tk.LEFT)

        disc_e=tk.Entry(ef,width=10,font=(FM.primary,14),justify="right",
                        bg=C_SURFACE2,fg=C_TEXT,insertbackground=C_ACCENT,
                        relief=tk.FLAT,highlightthickness=1,
                        highlightbackground=C_BORDER,highlightcolor=C_ACCENT)
        disc_e.insert(0,f"{item.get('item_disc_val',0):.2f}")
        disc_e.pack(side=tk.LEFT,ipady=6)

        # toggle ฿ %
        disc_mode=tk.StringVar(value=item.get("item_disc_mode","thb"))
        _tbns={}
        tbf=tk.Frame(ef,bg=C_BG); tbf.pack(side=tk.LEFT,padx=(6,0))

        # preview label
        prev_lbl=tk.Label(ef,text="",font=(FM.primary,13,"bold"),
                          bg=C_BG,fg=C_ACCENT,padx=10)
        prev_lbl.pack(side=tk.LEFT)

        # นิยาม calc_disc และ upd_prev ก่อน set_mode (closure ต้องการ)
        def calc_disc(price,mode,val):
            try: v=float(val or 0)
            except: v=0
            if mode=="pct": return min(round(price*v/100,2),price)
            return min(v,price)

        def upd_prev(*_):
            try: q=int(qty_e.get() or 1)
            except: q=1
            a=calc_disc(item["price"],disc_mode.get(),disc_e.get())
            try:
                prod=get_product(item["product_id"])
                ld,_lb,_f=promotion_engine.line_discount(prod,item["price"],q) if prod else (0.0,"",0)
                cen=get_promo_discount_for_product(item["product_id"],item["price"])*q
                promo=max(ld,cen)
            except Exception:
                promo=item.get("pd",0)*q
            total=max(item["price"]*q - promo - a*q, 0)
            prev_lbl.config(text=f"รวม ฿{total:,.2f}")

        def set_mode(m):
            disc_mode.set(m)
            for k,b in _tbns.items():
                b.config(bg=C_YELLOW if k==m else C_SURFACE2,
                         fg=C_BG    if k==m else C_MUTED)
            upd_prev()

        for code,lbl in [("thb","฿"),("pct","%")]:
            b=tk.Button(tbf,text=lbl,font=(FM.primary,12,"bold"),
                        bg=C_SURFACE2,fg=C_MUTED,relief=tk.FLAT,
                        cursor="hand2",padx=12,pady=6,bd=0,
                        activebackground=C_SURFACE2,
                        command=lambda c=code:set_mode(c))
            b.pack(side=tk.LEFT,padx=2)
            _tbns[code]=b
        set_mode(item.get("item_disc_mode","thb"))

        qty_e.bind("<KeyRelease>",upd_prev)
        disc_e.bind("<KeyRelease>",upd_prev)
        # validation: จำนวน (int ≤99,999), ส่วนลด (≤999,999.99)
        try:
            _attach_num_validator(qty_e, win, 99999, 0)
            _attach_num_validator(disc_e, win, 999999.99, 2)
        except Exception as _ve: print(f"[edit_qty validation] {_ve}")
        upd_prev()

        # ── ปุ่ม + separator pack(BOTTOM) ก่อน content ──
        # (tkinter pack จาก BOTTOM ขึ้น ต้องสร้างก่อนจึงไม่ถูก clip)
        bf=tk.Frame(win,bg=C_BG,padx=20); bf.pack(fill=tk.X,side=tk.BOTTOM,pady=10)
        sep=tk.Frame(win,bg=C_BORDER,height=1); sep.pack(fill=tk.X,side=tk.BOTTOM,pady=(10,5))

        def do_save():
            import cart_manager
            try: nq=max(int(qty_e.get() or 1),1)
            except: nq=1
            _fac=int(item.get("factor",1) or 1)
            if not check_neg_stock_policy(root,item["name"],item["stock"],nq*_fac):
                win.destroy(); return
            cart_manager.update_item(item, nq, disc_mode.get(), disc_e.get())
            self._recompute_promo(item)
            self._redraw(); win.destroy()

        win.bind("<Return>",lambda e:do_save())

        tk.Button(bf,text="✔  บันทึก",command=do_save,
                  font=(FM.primary,13,"bold"),
                  bg=C_ACCENT,fg=C_BG,relief=tk.FLAT,cursor="hand2",
                  padx=24,pady=10,activebackground=C_ACCENT,bd=0
                  ).pack(side=tk.LEFT,padx=(0,10))
        tk.Button(bf,text="ยกเลิก",command=win.destroy,
                  font=(FM.primary,13),
                  bg=C_SURFACE2,fg=C_MUTED,relief=tk.FLAT,cursor="hand2",
                  padx=16,pady=10,activebackground=C_SURFACE2,bd=0
                  ).pack(side=tk.LEFT)

        qty_e.focus_set(); qty_e.select_range(0,tk.END)

        # คำนวณ geometry จาก widget จริงหลังสร้างทั้งหมด
        win.update_idletasks()
        req_w=win.winfo_reqwidth()
        req_h=win.winfo_reqheight()
        win.geometry(f"{req_w+20}x{req_h+20}")
        win.minsize(req_w+20, req_h+20)
        """
        คืนราคาที่เหมาะสมตามระดับราคาของลูกค้า
        price_level: 'price' (ปกติ) | 'price_a' | 'price_b' | 'price_c'
        ถ้าราคา level นั้น = 0 หรือไม่มี → ใช้ราคาปกติ
        """
        if not self.customer:
            return product["price"]
        level = self.customer.get("price_level") or "price"
        if level == "price":
            return product["price"]
        val = product.get(level, 0)
        try:
            val = float(val or 0)
        except (TypeError, ValueError):
            val = 0.0
        return val if val > 0 else product["price"]

    def _get_price_for_customer(self, product: dict) -> float:
        """คืนราคาตามระดับลูกค้า price/price_a/price_b/price_c ถ้า 0 ใช้ราคาปกติ"""
        if not self.customer:
            return product["price"]
        level = self.customer.get("price_level") or "price"
        if level == "price":
            return product["price"]
        val = product.get(level, 0)
        try: val = float(val or 0)
        except: val = 0.0
        return val if val > 0 else product["price"]

    def _load_debt(self, customer_id: int) -> float:
        """ยอดค้างชำระของลูกค้า — ★ single source of truth (db_receivable)"""
        return get_customer_outstanding_balance(customer_id)

    def _update_debt_label(self):
        """อัปเดต debt_lbl หลังเลือก/ล้างลูกค้า — แสดงยอดค้าง + ระดับราคา"""
        if not self.customer:
            self.debt_lbl.config(text="")
            return
        debt  = self._load_debt(self.customer["id"])
        level = self.customer.get("price_level") or "price"
        level_map = {
            "price":   ("ราคาปกติ", "#8b949e"),
            "price_a": ("ราคา A",   "#4a9eff"),
            "price_b": ("ราคา B",   "#3fb950"),
            "price_c": ("ราคา C",   "#ffd166"),
        }
        level_txt, level_color = level_map.get(level, ("ราคาปกติ", "#8b949e"))
        if debt > 0:
            self.debt_lbl.config(
                text=f"⚠️ ค้างชำระ ฿{debt:,.2f}  [{level_txt}]",
                fg="#ff4444")
        else:
            self.debt_lbl.config(
                text=f"✅ ไม่มียอดค้าง  [{level_txt}]",
                fg="#3fb950")

    def _pick_cust(self):
        conn=get_db(); custs=conn.execute("SELECT * FROM customers ORDER BY name").fetchall(); conn.close()
        win=tk.Toplevel(self); win.title("เลือกลูกค้า"); win.configure(bg=C_BG); win.grab_set()
        tk.Frame(win,bg=C_ACCENT,height=3).pack(fill=tk.X)
        tk.Label(win,text="เลือกลูกค้า",font=F_H1,bg=C_BG,fg=C_TEXT,pady=10).pack(anchor="w",padx=16)
        sf2=tk.Frame(win,bg=C_BG); sf2.pack(fill=tk.X,padx=12,pady=(0,8))
        tk.Label(sf2,text="ค้นหา:",font=F_BODY,bg=C_BG,fg=C_MUTED).pack(side=tk.LEFT)
        sv2=tk.StringVar(); fe=field(sf2,width=22); fe.pack(side=tk.LEFT,padx=6,ipady=4); fe.configure(textvariable=sv2)
        frm,tv=make_tree(win,("รหัส","ชื่อ","โทร","แต้ม","ยอดสะสม"),(80,160,100,70,100),height=10)
        frm.pack(padx=12)
        def load(q=""):
            tv.delete(*tv.get_children())
            for i,r in enumerate(custs):
                if q.lower() in r["name"].lower() or q in (r["code"] or "") or q in (r["phone"] or ""):
                    tv.insert("","end",iid=r["id"],tags=("alt" if i%2==0 else "",),
                        values=(r["code"],r["name"],r["phone"] or "",r["points"],f"฿{r['total_spent']:,.2f}"))
        sv2.trace_add("write",lambda *a:load(sv2.get())); load()
        def sel():
            s=tv.selection()
            if s:
                self.customer=dict(next(r for r in custs if r["id"]==int(s[0])))
                name  = self.customer["name"]
                pts   = self.customer["points"]
                self.clbl.config(text=f"👤  {name}   แต้ม: {pts}", fg=C_BLUE)
                self._update_debt_label()
                win.destroy()
        accent_btn(win,"✔  เลือก",sel,pad_x=20,pad_y=8).pack(pady=8)
        tv.bind("<Double-1>",lambda e:sel())

    def _clear_cust(self):
        self.customer=None
        self.clbl.config(text="ไม่ระบุลูกค้า", fg=C_MUTED)
        self.debt_lbl.config(text="")
        self._upd()

    def _set_disc_mode(self, mode: str, _init: bool = False):
        """สลับโหมดส่วนลด ฿ หรือ %
        _init=True  → called at widget build time, skip permission check
        """
        if not _init and not has_permission("sale_discount"):
            require_permission("sale_discount", self.winfo_toplevel())
            return
        self.disc_mode.set(mode)
        for code, btn in self._disc_toggle_btns.items():
            if code == mode:
                btn.config(bg=C_YELLOW, fg=C_BG)
            else:
                btn.config(bg=C_SURFACE2, fg=C_MUTED)
        label = "ส่วนลดเพิ่มเติม (฿)" if mode == "thb" else "ส่วนลดเพิ่มเติม (%)"
        self.disc_lbl.config(text=label)
        self._upd()

    def _upd(self,event=None):
        import sale_service
        total=sale_service.calc_totals(self.cart)["total"]   # gross total ผ่าน service (SSoT)
        pd=sum(i["pd"]*i["qty"] for i in self.cart)
        # ส่วนลดรายการรวม (item-level discount)
        item_disc_total=sum(i.get("item_disc_amt",0)*i["qty"] for i in self.cart)
        try: disc_val=float(self.disc.get() or 0)
        except: disc_val=0
        # คำนวณ md ตามโหมด ฿ หรือ %
        if hasattr(self,"disc_mode") and self.disc_mode.get()=="pct":
            md=round((total-pd-item_disc_total)*disc_val/100, 2)
            md=min(md, total-pd-item_disc_total)
        else:
            md=disc_val
        net=max(total-pd-item_disc_total-md,0)
        try: rate=int(get_setting("point_rate") or 10)
        except: rate=10
        pts=int(net//rate)
        total_disc=pd+item_disc_total+md
        self.tv.set(f"฿ {total:,.2f}"); self.pv.set(f"-฿ {pd:,.2f}" if pd else "—")
        # 🎁 ของแถม (บรรทัด _free) — โชว์เฉพาะมี
        free_qty=sum(int(i.get("qty",0) or 0) for i in self.cart if i.get("_free"))
        if free_qty>0:
            self.fv.set(f"{free_qty} ชิ้น")
            self.free_row.pack(fill=tk.X,before=self._sum_div)
        else:
            self.fv.set("—"); self.free_row.pack_forget()
        self.dv.set(f"-฿ {md+item_disc_total:,.2f}" if (md+item_disc_total)>0 else "—")
        self.nv.set(f"฿ {net:,.2f}")
        # ── VAT display ──────────────────────────────────
        _vat = calc_vat(net)
        if _vat["enabled"] and hasattr(self,"vat_row"):
            self.vat_lbl_hdr.config(text=f"ภาษี VAT {_vat['percent']:.0f}%")
            self.vat_lbl.config(text=f"฿ {_vat['vat_amount']:,.2f}")
            self.grand_lbl.config(text=f"฿ {_vat['total']:,.2f}")
            self.vat_row.pack(fill=tk.X,padx=0,pady=1)
            self.grand_row.pack(fill=tk.X,padx=0,pady=(1,4))
        elif hasattr(self,"vat_row"):
            self.vat_row.pack_forget()
            self.grand_row.pack_forget()
        self.ptv.set(f"{pts} แต้ม" if self.customer else "—"); self._chg()

    def _chg(self,event=None):
        import payment_service
        try:
            net=float(self.nv.get().replace("฿","").replace(",","").strip())
            # Use payable (VAT-inclusive) as the amount customer must pay
            net=self._get_payable(net)
            paid_str=self.paid_e.get().strip()
            paid=float(paid_str) if paid_str else 0.0
            st=payment_service.classify_payment(paid, net)
            if st["status"]=="empty":
                # ยังไม่ได้กรอกเงิน
                self.cv.set("—")
                try:
                    self.chg_lbl_var.set("เงินทอน")
                    self.chg_lbl_w.config(fg=C_MUTED)
                    self.cv_lbl.config(fg=C_MUTED)
                except: pass
            elif st["status"]=="short":
                # ขาดเงิน
                self.cv.set(f"฿ {st['shortfall']:,.2f}")
                try:
                    self.chg_lbl_var.set("ขาดเงิน")
                    self.chg_lbl_w.config(fg="#FF5252")
                    self.cv_lbl.config(fg="#FF5252")
                except: pass
            else:
                # เงินทอน
                self.cv.set(f"฿ {st['change']:,.2f}")
                try:
                    self.chg_lbl_var.set("เงินทอน")
                    self.chg_lbl_w.config(fg="#FFD700")
                    self.cv_lbl.config(fg="#FFD700")
                except: pass
        except:
            self.cv.set("—")
            try: self.chg_lbl_var.set("เงินทอน")
            except: pass

    def _get_payable(self, net: float) -> float:
        """ยอดที่ลูกค้าต้องชำระจริง = net บวก VAT (ถ้าเปิด VAT)"""
        vat = calc_vat(net)
        return round(vat["total"] if vat["enabled"] else net, 2)

    def _shift_guard(self):
        """ตรวจกะก่อนขาย (opt-in ผ่าน setting 'shift_required')
        Returns:
            True  — ขายต่อได้ (มีกะเปิด หรือ ไม่บังคับกะ)
            False — ถูกบล็อก (บังคับเปิดกะ แต่ยังไม่เปิด) → caller ต้อง return
        """
        import shift_service
        if shift_service.get_open_shift():
            return True
        if (get_setting("shift_required") or "0") == "1":
            messagebox.showwarning(
                "ยังไม่เปิดกะ",
                "ระบบกำหนดให้เปิดกะก่อนขายสินค้า\n\nกรุณาเปิดกะก่อนเริ่มขาย",
                parent=self.winfo_toplevel())
            return False
        return True   # ไม่บังคับ → ขายได้ (บิลจะไม่ผูกกะ)

    def _active_shift_id(self):
        """shift_id ของกะที่เปิดอยู่ หรือ None — ใช้แปะตอนบันทึกบิล"""
        import shift_service
        sh = shift_service.get_open_shift()
        return sh["id"] if sh else None

    def _checkout(self):
        """F5 = ชำระเงิน — route ตาม payment method ที่เลือก"""
        if not self.cart:
            messagebox.showwarning("ว่างเปล่า","กรุณาเพิ่มสินค้าก่อน",parent=self.winfo_toplevel()); return
        if not self._shift_guard(): return
        try:
            net=float(self.nv.get().replace("฿","").replace(",","").strip())
            total=float(self.tv.get().replace("฿","").replace(",","").strip())
            pd=sum(it["pd"]*it["qty"] for it in self.cart)
            try:
                disc_val=float(self.disc.get() or 0)
                if hasattr(self,"disc_mode") and self.disc_mode.get()=="pct":
                    md=round((total-pd)*disc_val/100, 2)
                    md=min(md, total-pd)
                else:
                    md=disc_val
            except: md=0
        except: return
        # ── ยอดที่ลูกค้าต้องชำระจริง (รวม VAT ถ้าเปิด) ──
        payable = self._get_payable(net)

        method=self.pay_method.get()
        # ตรวจสอบเงินที่รับมาก่อนเปิด popup (เฉพาะเงินสด)
        if method=="cash":
            import payment_service
            try:
                paid_now=float(self.paid_e.get() or 0)
                if payment_service.classify_payment(paid_now, payable)["status"]=="short":
                    messagebox.showwarning(
                        "เงินไม่พอ",
                        "รับเงินมา ฿{:,.2f}\nยอดที่ต้องชำระ ฿{:,.2f}\nขาดอีก ฿{:,.2f}\n\nกรุณากรอกเงินให้ครบก่อนชำระ".format(
                            paid_now, payable, payable-paid_now),
                        parent=self.winfo_toplevel())
                    self.paid_e.focus_set()
                    self.paid_e.select_range(0,tk.END)
                    return
            except: pass
        if method=="qr":
            self._show_qr_hint()
            return
        elif method=="mixed":
            result=self._mixed_payment_dialog(payable)
            if result is None: return
            paid,change,mix_detail=result
            method_th="ผสม("+mix_detail+")"
            self._save_sale(payable,total,md,pd,paid,change,method_th)
            return
        elif method=="credit":
            # บัตร — paid=payable, no change
            self._save_sale(payable,total,md,pd,payable,0.0,"บัตรเครดิต")
            return
        elif method=="debt":
            if not self.customer:
                messagebox.showwarning(
                    "ต้องระบุลูกค้า",
                    "การบันทึกเงินเชื่อต้องเลือกลูกค้าก่อน\n\n"
                    "กรุณากดปุ่ม 👤 ลูกค้า เพื่อเลือกลูกค้า",
                    parent=self.winfo_toplevel())
                return
            ans = messagebox.askokcancel(
                "📋  บันทึกเงินเชื่อ",
                f"ลูกค้า: {self.customer['name']}\n"
                f"ยอดเชื่อ: ฿{payable:,.2f}\n\n"
                "ยืนยันบันทึกเป็นเงินเชื่อ?",
                parent=self.winfo_toplevel())
            if not ans: return
            self._save_sale(payable,total,md,pd,0.0,0.0,"เงินเชื่อ")
            return
        # เงินสด — เปิด fast popup พร้อมยอดที่กรอกไว้
        try: preset=float(self.paid_e.get() or 0)
        except: preset=0.0
        self._fast_payment_popup(payable, total, md, pd, preset)

    def _save_sale(self,net,total,md,pd,paid,change,method_th):
        """บันทึกบิล — delegate DB write ไป sale_service, Stock Card หลัง commit"""
        import sale_service, payment_service
        cid=self.customer["id"] if self.customer else 0
        _pm=payment_service.normalize_method(method_th)
        try: rate=int(get_setting("point_rate") or 10)
        except: rate=10

        # ── Step 1-3: DB write ผ่าน service (atomic: sales+items+stock+points) ──
        cart_snapshot = list(self.cart)  # snapshot ก่อน clear
        items=[{"product_id":it["product_id"],"name":it["name"],"price":it["price"],
                "cost":it["cost"]*int(it.get("factor",1) or 1),"qty":it["qty"],"subtotal":it["subtotal"]}
               for it in cart_snapshot]
        sid=sale_service.save_sale(
            items, total=total, net=net, discount=md, promo_disc=pd,
            paid=paid, change=change, payment_method=_pm,
            note=method_th, customer_id=cid, point_rate=rate,
            shift_id=self._active_shift_id())

        # ── Step 5: Stock Card บันทึกหลัง commit ───────
        # แยก connection ใหม่ ไม่ล็อค DB หลัก
        try:
            for item in cart_snapshot:
                _fac=int(item.get("factor",1) or 1)
                record_movement(item["product_id"], -item["qty"]*_fac, "sale",
                    ref_id=sid, ref_type="sale",
                    note=f"ขายบิล#{sid}", cost=item.get("cost",0))
        except Exception as e:
            print(f"[StockCard] {e}")  # ไม่ block flow

        # ── Step 6: แจ้งผล ─────────────────────────────
        if method_th == "เงินเชื่อ":
            msg = f"📋 บันทึกเงินเชื่อสำเร็จ\nลูกค้า: {self.customer['name'] if self.customer else '-'}\nยอดเชื่อ: ฿{net:,.2f}"
        else:
            icon={"เงินสด":"💵","QR/พร้อมเพย์":"📱","บัตรเครดิต":"💳"}.get(method_th,"✅")
            msg="{} ชำระเงินสำเร็จ\nยอดชำระ ฿{:,.2f} | {}\nเงินทอน ฿{:,.2f}".format(icon,net,method_th,change)
        messagebox.showinfo("✅ สำเร็จ",msg,parent=self.winfo_toplevel())

        # ── Step 6b: Refresh debt badge + AR tab after save ────
        # ต้องเรียกก่อน _clear() ล้าง self.customer
        if self.customer:
            self._update_debt_label()
        # Refresh ReceivableTab ถ้ามีอยู่
        try:
            root = self.winfo_toplevel()
            if hasattr(root, "_ar_tab"):
                root._ar_tab.refresh_all()
        except Exception as _e:
            print(f"[SaleTab] AR refresh after save: {_e}")

        # ── Step 7: พิมพ์ใบเสร็จ → clear → focus ──────
        self.last_sale_id = sid
        if get_setting("auto_print_receipt") != "0":
            print_receipt(sid)
        self._clear()
        self.se.focus_set()
        check_low_stock(self.winfo_toplevel())

    def _print_last(self):
        """พิมพ์ใบเสร็จของบิลล่าสุด (ใช้กับปุ่ม 🖨️ พิมพ์ล่าสุด)"""
        sid = getattr(self, "last_sale_id", None)
        if not sid:
            messagebox.showinfo("ยังไม่มีบิล", "ยังไม่มีบิลล่าสุดให้พิมพ์",
                                parent=self.winfo_toplevel())
            return
        print_receipt(sid)

    def _fast_payment_popup(self, net, total, md, pd, preset=0.0):
        """Popup ชำระเงิน — Modern POS UI พร้อมปุ่มยืนยันใหญ่ + Quick Denom"""
        # ── Setup ──────────────────────────────────────────────────────────────
        root = self.winfo_toplevel()
        root.unbind("<F5>")

        win = tk.Toplevel(self)
        win.title("ชำระเงิน")
        win.configure(bg=C_BG)
        # resizable จะถูก lock หลัง geometry — ดูด้านล่าง

        # ── VAT breakdown ──────────────────────────────────────────────────────
        try:
            _nv_str = self.nv.get().replace("฿","").replace(",","").strip()
            _net_before_vat = float(_nv_str) if _nv_str else net
        except:
            _net_before_vat = net
        _vat = calc_vat(_net_before_vat)

        # ── Header bar ─────────────────────────────────────────────────────────
        tk.Frame(win, bg=C_ACCENT, height=4).pack(fill=tk.X)
        hf = tk.Frame(win, bg=C_BG, pady=8); hf.pack(fill=tk.X)
        tk.Label(hf, text="ชำระเงิน", font=F_H1, bg=C_BG, fg=C_TEXT).pack(side=tk.LEFT, padx=16)
        tk.Label(hf, text="[เงินสด]", font=F_SMALL, bg=C_BG, fg=C_MUTED).pack(side=tk.RIGHT, padx=16)

        # ── Section: ยอดรวม ────────────────────────────────────────────────────
        sf = tk.Frame(win, bg=C_SURFACE, padx=18, pady=14,
                      highlightbackground=C_BORDER, highlightthickness=1)
        sf.pack(fill=tk.X, padx=14, pady=(0, 8))

        # แถว 1: ยอดรวม
        r1 = tk.Frame(sf, bg=C_SURFACE); r1.pack(fill=tk.X)
        lbl_total = "รวมทั้งสิ้น (รวม VAT):" if _vat["enabled"] and _vat["vat_amount"] > 0 else "ยอดรวมสุทธิ:"
        tk.Label(r1, text=lbl_total, font=(FM.primary,13), bg=C_SURFACE, fg=C_MUTED).pack(side=tk.LEFT)
        tk.Label(r1, text="฿{:,.2f}".format(net),
                 font=(FM.primary,28,"bold"), bg=C_SURFACE, fg=C_ACCENT).pack(side=tk.RIGHT)

        # แถว VAT breakdown
        if _vat["enabled"] and _vat["vat_amount"] > 0:
            r1b = tk.Frame(sf, bg=C_SURFACE); r1b.pack(fill=tk.X, pady=(2,0))
            tk.Label(r1b,
                     text="  ก่อน VAT: ฿{:,.2f}   VAT {:.0f}%: ฿{:,.2f}".format(
                         _vat["before_vat"], _vat["percent"], _vat["vat_amount"]),
                     font=(FM.primary,10), bg=C_SURFACE, fg=C_MUTED).pack(side=tk.LEFT)

        tk.Frame(sf, bg=C_BORDER, height=1).pack(fill=tk.X, pady=(10,8))

        # แถว 2: รับเงินมา
        r2 = tk.Frame(sf, bg=C_SURFACE); r2.pack(fill=tk.X, pady=(0,6))
        tk.Label(r2, text="รับเงินมา:", font=(FM.primary,13), bg=C_SURFACE, fg=C_TEXT).pack(side=tk.LEFT)
        init_val = preset if preset >= net else net
        paid_var = tk.StringVar(value="{:.2f}".format(init_val))
        paid_e = tk.Entry(r2, textvariable=paid_var,
                          font=(FM.primary,22,"bold"),
                          bg=C_SURFACE2, fg=C_YELLOW,
                          insertbackground=C_YELLOW,
                          justify="right", relief=tk.FLAT, width=12,
                          highlightbackground=C_ACCENT,
                          highlightthickness=2, highlightcolor=C_ACCENT)
        paid_e.pack(side=tk.RIGHT, ipady=4)

        # แถว 3: เงินทอน
        r3 = tk.Frame(sf, bg=C_SURFACE); r3.pack(fill=tk.X)
        change_lbl_text = tk.StringVar(value="เงินทอน:")
        tk.Label(r3, textvariable=change_lbl_text,
                 font=(FM.primary,13), bg=C_SURFACE, fg=C_MUTED).pack(side=tk.LEFT)
        change_var = tk.StringVar(value="฿ 0.00")
        change_lbl = tk.Label(r3, textvariable=change_var,
                              font=(FM.primary,28,"bold"), bg=C_SURFACE, fg=C_GREEN)
        change_lbl.pack(side=tk.RIGHT)

        # ── Section: Quick Denomination Buttons ────────────────────────────────
        qf = tk.Frame(win, bg=C_BG, padx=14); qf.pack(fill=tk.X, pady=4)
        # ── แถว 1: +20 +50 +100 +500 +1000 (pack แทน grid — ไม่ block reqheight)
        btn_row1 = tk.Frame(qf, bg=C_BG); btn_row1.pack(fill=tk.X)

        DENOMS = [20, 50, 100, 500, 1000]
        def _make_denom_cmd(d):
            def _cmd():
                if paid_e["state"] == "disabled": return
                try:
                    cur = float(paid_var.get() or 0)
                except:
                    cur = 0
                paid_var.set("{:.0f}".format(cur + d))
                paid_e.icursor(tk.END)
                paid_e.focus_set()
            return _cmd

        for d in DENOMS:
            tk.Button(btn_row1,
                      text="+{:,}".format(d),
                      command=_make_denom_cmd(d),
                      font=(FM.primary,11,"bold"),
                      bg=C_SURFACE2, fg=C_TEXT,
                      relief=tk.FLAT, cursor="hand2",
                      pady=7, bd=0,
                      highlightbackground=C_BORDER,
                      highlightthickness=1,
                      activebackground=C_SURFACE,
                      activeforeground=C_ACCENT
                      ).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)

        # ── แถว 2: พอดี / ล้าง / ⌫
        def _set_exact():
            if paid_e["state"] == "disabled": return
            paid_var.set("{:.0f}".format(net))
            paid_e.focus_set()

        def _clear_paid():
            if paid_e["state"] == "disabled": return
            paid_var.set("")
            paid_e.focus_set()

        def _backspace():
            if paid_e["state"] == "disabled": return
            cur = paid_var.get()
            paid_var.set(cur[:-1] if cur else "")
            paid_e.focus_set()

        btn_row2 = tk.Frame(qf, bg=C_BG); btn_row2.pack(fill=tk.X, pady=(4,0))
        for lbl, cmd, color in [
            ("พอดี", _set_exact, C_BLUE),
            ("ล้าง",  _clear_paid, C_MUTED),
            ("⌫",    _backspace,  C_MUTED),
        ]:
            tk.Button(btn_row2, text=lbl, command=cmd,
                      font=(FM.primary,11,"bold"),
                      bg=C_SURFACE2, fg=color,
                      relief=tk.FLAT, cursor="hand2",
                      pady=7, bd=0,
                      highlightbackground=C_BORDER,
                      highlightthickness=1,
                      activebackground=C_SURFACE,
                      activeforeground=color
                      ).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)

        # ── State ─────────────────────────────────────────────────────────────
        state = {"confirmed": False}

        def _rebind_f5():
            try:
                root.bind("<F5>", lambda e: self._qs_save() if get_setting("quick_sale_mode")=="1" else self._checkout())
            except: pass

        def _close_popup():
            win.destroy()
            _rebind_f5()

        def calc_change(*_):
            import payment_service
            try:
                paid = float(paid_var.get().strip() or 0)
            except:
                change_var.set("---"); return
            st = payment_service.classify_payment(paid, net)
            if st["status"] == "empty":
                change_lbl_text.set("เงินทอน:")
                change_var.set("—")
                change_lbl.config(fg=C_MUTED)
            elif st["status"] == "short":
                change_lbl_text.set("ขาดเงิน:")
                change_var.set("฿ {:,.2f}".format(st["shortfall"]))
                change_lbl.config(fg=C_ACCENT2)
            else:
                change_lbl_text.set("เงินทอน:")
                change_var.set("฿ {:,.2f}".format(st["change"]))
                change_lbl.config(fg=C_YELLOW)

        def on_paid_change(*_):
            calc_change()
            if state["confirmed"]:
                state["confirmed"] = False
                confirm_btn.config(
                    text="✅   ยืนยันชำระเงิน   (Enter)",
                    bg=C_ACCENT, pady=20)
                paid_e.config(state="normal")

        paid_var.trace_add("write", on_paid_change)
        calc_change()

        def do_confirm():
            import payment_service
            try:
                paid = float(paid_var.get() or 0)
            except:
                paid_e.focus_set(); return

            if not payment_service.is_sufficient(paid, net):
                change_lbl_text.set("ขาดเงิน:")
                change_var.set("฿ {:,.2f}".format(net - paid))
                change_lbl.config(fg=C_ACCENT2)
                confirm_btn.config(
                    text="⛔   เงินไม่พอ — กรอกใหม่",
                    bg="#c0392b", pady=20)
                paid_e.focus_set(); paid_e.select_range(0, tk.END)
                return

            if not state["confirmed"]:
                # Enter ครั้งแรก: lock input แสดงเงินทอน รอยืนยัน
                state["confirmed"] = True
                confirm_btn.config(
                    text="✅   ยืนยันบันทึก   (Enter อีกครั้ง)",
                    bg="#27ae60", pady=20)
                paid_e.config(state="disabled")
                return

            # Enter ครั้งที่สอง: บันทึก + จบ
            change = payment_service.change_due(paid, net)
            win.destroy()
            _rebind_f5()
            self._save_sale(net, total, md, pd, paid, change, "เงินสด")

        # ── ปุ่มยืนยัน — ใหญ่ เด่น กดง่าย ──────────────────────────────────
        btn_zone = tk.Frame(win, bg=C_BG, padx=14)
        btn_zone.pack(fill=tk.X, pady=6)

        confirm_btn = tk.Button(
            btn_zone,
            text="✅   ยืนยันชำระเงิน   (Enter)",
            command=do_confirm,
            font=(FM.primary, 16, "bold"),
            bg=C_ACCENT,
            fg=C_BG,
            relief=tk.FLAT,
            cursor="hand2",
            pady=16,
            activebackground="#00b899",
            activeforeground=C_BG,
            bd=0,
        )
        confirm_btn.pack(fill=tk.X, pady=(0, 4))

        # ── ปุ่มยกเลิก ────────────────────────────────────────────────────────
        tk.Button(
            btn_zone,
            text="✕   ยกเลิก   (Esc)",
            command=_close_popup,
            font=(FM.primary, 12),
            bg=C_SURFACE2,
            fg=C_MUTED,
            relief=tk.FLAT,
            cursor="hand2",
            pady=8,
            activebackground=C_BORDER,
            activeforeground=C_TEXT,
            bd=0,
        ).pack(fill=tk.X)

        # ── Hint bar ───────────────────────────────────────────────────────────
        tk.Label(win,
                 text="F5 / Enter = ยืนยัน   ·   Esc = ยกเลิก",
                 font=(FM.primary, 10), bg=C_BG, fg=C_BORDER).pack(pady=(4, 8))

        # ── Keybindings ────────────────────────────────────────────────────────
        win.bind("<Return>", lambda e: do_confirm())
        win.bind("<Escape>", lambda e: _close_popup())
        win.bind("<F5>",     lambda e: do_confirm())
        win.protocol("WM_DELETE_WINDOW", _close_popup)

        # ── Geometry: วัดจาก reqwidth/reqheight หลัง update_idletasks ──────────
        win.update_idletasks()
        rw = win.winfo_reqwidth()
        rh = win.winfo_reqheight()
        sw = win.winfo_screenwidth()
        sh = win.winfo_screenheight()
        # ให้กว้างอย่างน้อย 420px และสูงเท่าที่ content ต้องการจริง
        w = max(rw, 420)
        h = rh
        x = (sw - w) // 2
        y = max(40, (sh - h) // 2)
        win.geometry("{}x{}+{}+{}".format(w, h, x, y))
        win.resizable(False, False)
        win.grab_set()
        win.lift()
        win.focus_force()

        paid_e.focus_set()
        paid_e.select_range(0, tk.END)

    def _mixed_payment_dialog(self,net):
        win=tk.Toplevel(self)
        win.title("ชำระแบบผสม")
        win.configure(bg=C_BG)
        win.grab_set()
        win.resizable(False,False)
        win.lift()

        tk.Frame(win,bg=ORANGE,height=4).pack(fill=tk.X)
        tk.Label(win,text="ชำระแบบผสม",font=F_H1,bg=C_BG,fg=C_TEXT,pady=10).pack()

        # net parameter here is already the payable amount (VAT-inclusive)
        # re-derive vat breakdown for display
        try:
            _nv_str2 = self.nv.get().replace("฿","").replace(",","").strip()
            _net_before_vat2 = float(_nv_str2) if _nv_str2 else net
        except:
            _net_before_vat2 = net
        _vat_mix = calc_vat(_net_before_vat2)

        hf=tk.Frame(win,bg=C_SURFACE,padx=16,pady=10)
        hf.pack(fill=tk.X,padx=16,pady=(0,8))
        tk.Label(hf,text="ยอดรวมทั้งสิ้น (รวม VAT)" if _vat_mix["enabled"] and _vat_mix["vat_amount"]>0 else "ยอดที่ต้องชำระ",
                 font=F_SMALL,bg=C_SURFACE,fg=C_MUTED).pack(anchor="w")
        tk.Label(hf,text="฿ %.2f"%net,font=(FM.primary,24,"bold"),bg=C_SURFACE,fg=C_ACCENT).pack(anchor="w")
        if _vat_mix["enabled"] and _vat_mix["vat_amount"]>0:
            tk.Label(hf,text=f"(ก่อน VAT: ฿{_vat_mix['before_vat']:,.2f}  +  VAT {_vat_mix['percent']:.0f}%: ฿{_vat_mix['vat_amount']:,.2f})",
                     font=(FM.primary,10),bg=C_SURFACE,fg=C_MUTED).pack(anchor="w")

        # เงินสด
        fc=tk.Frame(win,bg=C_SURFACE,bd=1,relief=tk.FLAT)
        fc.pack(fill=tk.X,padx=16,pady=3)
        rc=tk.Frame(fc,bg=C_SURFACE,padx=10,pady=6); rc.pack(fill=tk.X)
        tk.Frame(rc,bg=C_GREEN,width=4).pack(side=tk.LEFT,fill=tk.Y,padx=(0,8))
        tk.Label(rc,text="💵  เงินสด (฿)",font=F_BODY,bg=C_SURFACE,fg=C_TEXT).pack(side=tk.LEFT)
        b_cash=tk.Button(rc,text="เติมยอดที่เหลือ",font=(FM.primary,10),bg=C_SURFACE2,fg=C_MUTED,relief=tk.FLAT,cursor="hand2")
        b_cash.pack(side=tk.RIGHT)
        e_cash=field(fc,font=(FM.primary,16,"bold"),justify="right",fg=C_GREEN)
        e_cash.insert(0,"0"); e_cash.pack(fill=tk.X,padx=10,pady=(0,8),ipady=5)

        # QR
        fq=tk.Frame(win,bg=C_SURFACE,bd=1,relief=tk.FLAT)
        fq.pack(fill=tk.X,padx=16,pady=3)
        rq=tk.Frame(fq,bg=C_SURFACE,padx=10,pady=6); rq.pack(fill=tk.X)
        tk.Frame(rq,bg=C_BLUE,width=4).pack(side=tk.LEFT,fill=tk.Y,padx=(0,8))
        tk.Label(rq,text="📱  QR / พร้อมเพย์ (฿)",font=F_BODY,bg=C_SURFACE,fg=C_TEXT).pack(side=tk.LEFT)
        b_qr=tk.Button(rq,text="เติมยอดที่เหลือ",font=(FM.primary,10),bg=C_SURFACE2,fg=C_MUTED,relief=tk.FLAT,cursor="hand2")
        b_qr.pack(side=tk.RIGHT)
        e_qr=field(fq,font=(FM.primary,16,"bold"),justify="right",fg=C_BLUE)
        e_qr.insert(0,"0"); e_qr.pack(fill=tk.X,padx=10,pady=(0,8),ipady=5)

        # บัตร
        fk=tk.Frame(win,bg=C_SURFACE,bd=1,relief=tk.FLAT)
        fk.pack(fill=tk.X,padx=16,pady=3)
        rk=tk.Frame(fk,bg=C_SURFACE,padx=10,pady=6); rk.pack(fill=tk.X)
        tk.Frame(rk,bg=C_YELLOW,width=4).pack(side=tk.LEFT,fill=tk.Y,padx=(0,8))
        tk.Label(rk,text="💳  บัตรเครดิต (฿)",font=F_BODY,bg=C_SURFACE,fg=C_TEXT).pack(side=tk.LEFT)
        b_card=tk.Button(rk,text="เติมยอดที่เหลือ",font=(FM.primary,10),bg=C_SURFACE2,fg=C_MUTED,relief=tk.FLAT,cursor="hand2")
        b_card.pack(side=tk.RIGHT)
        e_card=field(fk,font=(FM.primary,16,"bold"),justify="right",fg=C_YELLOW)
        e_card.insert(0,"0"); e_card.pack(fill=tk.X,padx=10,pady=(0,8),ipady=5)

        # สรุป
        sf=tk.Frame(win,bg=C_SURFACE2,padx=16,pady=8)
        sf.pack(fill=tk.X,padx=16,pady=6)
        row_s=tk.Frame(sf,bg=C_SURFACE2); row_s.pack(fill=tk.X)
        tk.Label(row_s,text="รวมที่กรอก:",font=F_BODY,bg=C_SURFACE2,fg=C_MUTED).pack(side=tk.LEFT)
        sum_var=tk.StringVar(value="฿ 0.00")
        tk.Label(row_s,textvariable=sum_var,font=F_H2,bg=C_SURFACE2,fg=C_TEXT).pack(side=tk.RIGHT)
        remain_var=tk.StringVar(value="ขาดอีก ฿ %.2f"%net)
        remain_lbl=tk.Label(sf,textvariable=remain_var,font=F_SMALL,bg=C_SURFACE2,fg=C_ACCENT2)
        remain_lbl.pack(anchor="e")

        result=[None]

        def upd(*a):
            import payment_service
            try: s=payment_service.split_total([e_cash.get(),e_qr.get(),e_card.get()])
            except: s=0
            sum_var.set("฿ %.2f"%s)
            if s>=net:
                remain_var.set("เงินทอน ฿ %.2f"%(s-net)); remain_lbl.config(fg=C_GREEN)
            else:
                remain_var.set("ขาดอีก ฿ %.2f"%(net-s)); remain_lbl.config(fg=C_ACCENT2)

        def fill(tgt):
            import payment_service
            try:
                others={"cash":[e_qr.get(),e_card.get()],
                        "qr":[e_cash.get(),e_card.get()],
                        "card":[e_cash.get(),e_qr.get()]}[tgt]
                rem=payment_service.split_remaining(others, net)
                e={"cash":e_cash,"qr":e_qr,"card":e_card}[tgt]
                e.delete(0,tk.END); e.insert(0,"%.2f"%rem); upd()
            except: pass

        b_cash.config(command=lambda:fill("cash"))
        b_qr.config(command=lambda:fill("qr"))
        b_card.config(command=lambda:fill("card"))
        e_cash.bind("<KeyRelease>",upd)
        e_qr.bind("<KeyRelease>",upd)
        e_card.bind("<KeyRelease>",upd)
        # validation: แต่ละช่อง ≤ 9,999,999.99 (กันเลขยาวผิดปกติ)
        try:
            for _pe in (e_cash, e_qr, e_card):
                _attach_num_validator(_pe, win, 9999999.99, 2)
        except Exception as _ve: print(f"[mixed validation] {_ve}")

        def confirm():
            import payment_service
            try: s=payment_service.split_total([e_cash.get(),e_qr.get(),e_card.get()])
            except: s=0
            if s<net:
                messagebox.showwarning("ยอดไม่พอ","รวม ฿%.2f\nต้องชำระ ฿%.2f\nขาดอีก ฿%.2f"%(s,net,net-s),parent=win)
                return
            vc=float(e_cash.get() or 0); vq=float(e_qr.get() or 0); vk=float(e_card.get() or 0)
            _parts,detail=payment_service.build_split_detail([("เงินสด",vc),("QR",vq),("บัตร",vk)])
            result[0]=(s,s-net,detail)
            win.destroy()

        tk.Button(win,text="✅  ยืนยันชำระ",command=confirm,
                  font=(FM.primary,14,"bold"),bg=C_ACCENT,fg=C_BG,
                  relief=tk.FLAT,cursor="hand2",pady=10,
                  activebackground=C_ACCENT,bd=0).pack(fill=tk.X,padx=16,pady=(0,12))
        win.wait_window()
        return result[0]

    def _clear(self):
        """Reset หน้าขายทุก state — เรียกหลัง checkout สำเร็จ"""
        import cart_manager
        # 1. ล้างตะกร้า
        cart_manager.clear(self.cart)
        self._refresh()
        # 2. reset ส่วนลด
        self.paid_e.delete(0,tk.END)
        self.disc.delete(0,tk.END); self.disc.insert(0,"0")
        if hasattr(self,"disc_mode"): self._set_disc_mode("thb")
        # 3. reset ลูกค้า
        self.customer=None
        self.clbl.config(text="ไม่ระบุลูกค้า",fg=C_MUTED)
        if hasattr(self,"debt_lbl"): self.debt_lbl.config(text="")
        # 4. reset ช่องทางชำระ → เงินสด
        self.pay_method.set("cash")
        self._set_pay("cash","#27ae60")
        # 5. อัปเดตยอด
        self._upd()


    def _park_recall(self):
        if self.cart: self._park_bill()
        else: self._recall_bills()

    def _park_bill(self):
        import copy
        self.parked_bills.append({
            "cart": copy.deepcopy(self.cart),
            "customer": self.customer,
            "disc": self.disc.get(),
            "label": "บิล #{} ({} รายการ)".format(len(self.parked_bills)+1, len(self.cart))
        })
        self._clear()
        self._update_park_btn()
        messagebox.showinfo("พักบิลแล้ว",
            "พักบิลสำเร็จ มี {} บิลที่พักไว้\nเพิ่มสินค้าเพื่อเริ่มบิลใหม่ หรือกด F6 เรียกคืน".format(len(self.parked_bills)),
            parent=self.winfo_toplevel())

    def _recall_bills(self):
        if not self.parked_bills:
            messagebox.showinfo("ไม่มีบิลที่พัก","ยังไม่มีบิลที่พักไว้",parent=self.winfo_toplevel()); return
        import copy
        win=tk.Toplevel(self); win.title("บิลที่พักไว้")
        win.configure(bg=C_BG); win.resizable(False,False); win.lift(); win.grab_set()
        tk.Frame(win,bg=C_YELLOW,height=3).pack(fill=tk.X)
        tk.Label(win,text="⏸  บิลที่พักไว้ ({} บิล)".format(len(self.parked_bills)),
                 font=F_H1,bg=C_BG,fg=C_TEXT,pady=10).pack(anchor="w",padx=16)
        for idx in range(len(self.parked_bills)):
            bill=self.parked_bills[idx]
            bf=tk.Frame(win,bg=C_SURFACE,highlightbackground=C_BORDER,highlightthickness=1)
            bf.pack(fill=tk.X,padx=16,pady=4)
            info=tk.Frame(bf,bg=C_SURFACE,padx=12,pady=8); info.pack(fill=tk.X)
            total=sum(it["subtotal"] for it in bill["cart"])
            cname=bill["customer"]["name"] if bill["customer"] else "ไม่ระบุลูกค้า"
            tk.Label(info,text=bill["label"],font=F_H2,bg=C_SURFACE,fg=C_TEXT).pack(anchor="w")
            tk.Label(info,text="ลูกค้า: {}  |  ยอด: ฿{:,.2f}  |  {} รายการ".format(cname,total,len(bill["cart"])),
                     font=F_SMALL,bg=C_SURFACE,fg=C_MUTED).pack(anchor="w")
            btnf=tk.Frame(bf,bg=C_SURFACE,padx=12); btnf.pack(fill=tk.X,pady=(0,8))
            def recall(i=idx,w=win):
                b=self.parked_bills.pop(i)
                if self.cart:
                    if not messagebox.askyesno("แทนที่บิล","ตะกร้ามีสินค้าอยู่ ต้องการแทนที่ด้วยบิลที่พัก?",parent=w):
                        self.parked_bills.insert(i,b); return
                self.cart=b["cart"]; self.customer=b["customer"]
                self.disc.delete(0,tk.END); self.disc.insert(0,b["disc"])
                if self.customer:
                    self.clbl.config(text="👤 {}  (แต้ม:{})".format(self.customer["name"],self.customer["points"]),fg=C_BLUE)
                    self._update_debt_label()
                else:
                    self.debt_lbl.config(text="")
                self._redraw(); self._update_park_btn(); w.destroy()
            def del_park(i=idx,w=win):
                if messagebox.askyesno("ลบบิล","ลบบิลที่พักนี้ทิ้ง?",parent=w):
                    self.parked_bills.pop(i); self._update_park_btn(); w.destroy()
                    if self.parked_bills: self._recall_bills()
            accent_btn(btnf,"↩ เรียกคืนบิลนี้",recall,pad_x=12,pad_y=4).pack(side=tk.LEFT,padx=(0,6))
            pill_btn(btnf,"ลบทิ้ง",del_park,bg=C_SURFACE2,fg=C_ACCENT2,pad_x=10,pad_y=4).pack(side=tk.LEFT)
        pill_btn(win,"ปิด",win.destroy,bg=C_SURFACE,fg=C_MUTED,pad_x=20,pad_y=8).pack(pady=12)

    def _update_park_btn(self):
        try:
            n=len(self.parked_bills)
            self.park_btn.config(
                text="⏸ พักบิล ({})".format(n) if n>0 else "⏸ พักบิล",
                fg=C_YELLOW if n>0 else C_MUTED)
        except: pass

    def _qs_start(self):
        """F1 = เริ่มขาย — ล้างบิลและโฟกัสช่องสแกน"""
        if self.cart:
            if not messagebox.askyesno("เริ่มบิลใหม่",
                "มีสินค้าในตะกร้าอยู่\nต้องการล้างเพื่อเริ่มบิลใหม่?",
                parent=self.winfo_toplevel()): 
                self.se.focus_set(); return
        self._clear()
        self.se.focus_set()
        self.se.select_range(0,tk.END)

    def _qs_save(self):
        """F5 = บันทึกและพิมพ์ (Quick Sale Mode) หรือยอดพอดี (โหมดปกติ)"""
        if get_setting("quick_sale_mode")!="1":
            self._exact_amount(); return
        if not self.cart:
            messagebox.showwarning("ว่างเปล่า","กรุณาเพิ่มสินค้าก่อน",parent=self.winfo_toplevel()); return
        if not self._shift_guard(): return
        # ชำระเงินแบบ Quick Sale ไม่ต้องเปิดหน้าต่าง
        try:
            net=float(self.nv.get().replace("฿","").replace(",","").strip())
            total=float(self.tv.get().replace("฿","").replace(",","").strip())
            pd=sum(it["pd"]*it["qty"] for it in self.cart)
            try: md=float(self.disc.get() or 0)
            except: md=0
        except: return
        method=get_setting("quick_sale_payment") or "cash"
        import payment_service
        method_th=payment_service.method_label(method)
        # Quick sale — ยอดชำระจริง รวม VAT
        payable_qs = self._get_payable(net)
        # Quick sale — ไม่ต้องเปิดหน้าต่าง
        import sale_service
        cid=self.customer["id"] if self.customer else 0
        try: rate=int(get_setting("point_rate") or 10)
        except: rate=10
        cart_snapshot=list(self.cart)
        items=[{"product_id":it["product_id"],"name":it["name"],"price":it["price"],
                "cost":it["cost"]*int(it.get("factor",1) or 1),"qty":it["qty"],"subtotal":it["subtotal"]}
               for it in cart_snapshot]
        sid=sale_service.save_sale(
            items, total=total, net=payable_qs, discount=md, promo_disc=pd,
            paid=payable_qs, change=0.0, payment_method=method,
            note=method_th, customer_id=cid, point_rate=rate,
            shift_id=self._active_shift_id())
        # Stock Card — บันทึกหลัง commit ไม่ lock DB
        try:
            for item in cart_snapshot:
                _fac=int(item.get("factor",1) or 1)
                record_movement(item["product_id"],-item["qty"]*_fac,"sale",
                    ref_id=sid,ref_type="sale",note=f"ขายบิล#{sid}",cost=item.get("cost",0))
        except Exception as e:
            print(f"[StockCard/QS] {e}")
        self.last_sale_id = sid
        if get_setting("quick_sale_print")=="1":
            print_receipt(sid)
        else:
            messagebox.showinfo("✅ Quick Sale",
                "บิล #{} | ฿{:,.2f} | {}".format(sid,net,method_th),
                parent=self.winfo_toplevel())
        if get_setting("quick_sale_new_bill")=="1":
            self._clear(); self.se.focus_set()
        check_low_stock(self.winfo_toplevel())

# ══ TAB: PRODUCTS (Enhanced) ══════════════════════════════
class CategoryManager:
    @staticmethod
    def ensure():
        conn=get_db()
        conn.execute("CREATE TABLE IF NOT EXISTS categories (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE NOT NULL, created_at TEXT)")
        now=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for r in conn.execute("SELECT DISTINCT category FROM products WHERE category!=''").fetchall():
            conn.execute("INSERT OR IGNORE INTO categories(name,created_at) VALUES(?,?)",(r[0],now))
        conn.commit(); conn.close()
    @staticmethod
    def get_all():
        conn=get_db()
        try:
            names=[r[0] for r in conn.execute("SELECT name FROM categories").fetchall()]
        except Exception:
            names=[]
        for r in conn.execute("SELECT DISTINCT category FROM products WHERE category!=''").fetchall():
            if r[0] not in names: names.append(r[0])
        conn.close()
        return sorted(set(names))
    @staticmethod
    def count_usage(name):
        conn=get_db()
        n=conn.execute("SELECT COUNT(*) FROM products WHERE category=?",(name,)).fetchone()[0]
        conn.close(); return n
    @staticmethod
    def add(name):
        name=(name or "").strip()
        if not name: return (False,"กรุณากรอกชื่อหมวด")
        CategoryManager.ensure()
        conn=get_db()
        try:
            now=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            cur=conn.execute("INSERT OR IGNORE INTO categories(name,created_at) VALUES(?,?)",(name,now))
            conn.commit()
            return (True,"") if cur.rowcount else (False,"มีหมวดนี้อยู่แล้ว")
        finally: conn.close()
    @staticmethod
    def rename(old,new):
        new=(new or "").strip()
        if not new: return (False,"ชื่อใหม่ว่าง")
        if new==old: return (True,"")
        CategoryManager.ensure()
        conn=get_db()
        try:
            exists=conn.execute("SELECT 1 FROM categories WHERE name=?",(new,)).fetchone()
            conn.execute("UPDATE products SET category=? WHERE category=?",(new,old))
            if exists: conn.execute("DELETE FROM categories WHERE name=?",(old,))
            else:      conn.execute("UPDATE categories SET name=? WHERE name=?",(new,old))
            conn.commit(); return (True,"")
        finally: conn.close()
    @staticmethod
    def delete(name):
        used=CategoryManager.count_usage(name)
        if used>0: return (False,"มีสินค้าใช้หมวดนี้อยู่ %d รายการ"%used)
        CategoryManager.ensure()
        conn=get_db()
        try:
            conn.execute("DELETE FROM categories WHERE name=?",(name,)); conn.commit(); return (True,"")
        finally: conn.close()

class UnitManager:
    DEFAULT=["\u0e0a\u0e34\u0e49\u0e19","\u0e02\u0e27\u0e14","\u0e01\u0e25\u0e48\u0e2d\u0e07","\u0e41\u0e1e\u0e47\u0e04","\u0e25\u0e31\u0e07","\u0e42\u0e2b\u0e25","\u0e16\u0e38\u0e07","\u0e0b\u0e2d\u0e07","\u0e2d\u0e31\u0e19","\u0e01\u0e01.","\u0e40\u0e21\u0e15\u0e23","\u0e04\u0e39\u0e48"]
    @staticmethod
    def ensure():
        conn=get_db()
        existed=conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='units'").fetchone()
        conn.execute("CREATE TABLE IF NOT EXISTS units (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE NOT NULL, created_at TEXT)")
        now=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        seed=set(UnitManager.DEFAULT) if not existed else set()
        for r in conn.execute("SELECT DISTINCT unit FROM products WHERE unit!=''").fetchall():
            seed.add(r[0])
        for u in seed:
            conn.execute("INSERT OR IGNORE INTO units(name,created_at) VALUES(?,?)",(u,now))
        conn.commit(); conn.close()
    @staticmethod
    def get_all():
        conn=get_db()
        try:
            names=[r[0] for r in conn.execute("SELECT name FROM units").fetchall()]
        except Exception:
            names=[]
        for r in conn.execute("SELECT DISTINCT unit FROM products WHERE unit!=''").fetchall():
            if r[0] not in names: names.append(r[0])
        conn.close()
        return sorted(set(names))
    @staticmethod
    def count_usage(name):
        conn=get_db()
        n=conn.execute("SELECT COUNT(*) FROM products WHERE unit=?",(name,)).fetchone()[0]
        conn.close(); return n
    @staticmethod
    def add(name):
        name=(name or "").strip()
        if not name: return (False,"กรุณากรอกชื่อหน่วย")
        UnitManager.ensure()
        conn=get_db()
        try:
            now=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            cur=conn.execute("INSERT OR IGNORE INTO units(name,created_at) VALUES(?,?)",(name,now))
            conn.commit()
            return (True,"") if cur.rowcount else (False,"มีหน่วยนี้อยู่แล้ว")
        finally: conn.close()
    @staticmethod
    def rename(old,new):
        new=(new or "").strip()
        if not new: return (False,"ชื่อใหม่ว่าง")
        if new==old: return (True,"")
        UnitManager.ensure()
        conn=get_db()
        try:
            exists=conn.execute("SELECT 1 FROM units WHERE name=?",(new,)).fetchone()
            conn.execute("UPDATE products SET unit=? WHERE unit=?",(new,old))
            if exists: conn.execute("DELETE FROM units WHERE name=?",(old,))
            else:      conn.execute("UPDATE units SET name=? WHERE name=?",(new,old))
            conn.commit(); return (True,"")
        finally: conn.close()
    @staticmethod
    def delete(name):
        used=UnitManager.count_usage(name)
        if used>0: return (False,"มีสินค้าใช้หน่วยนี้อยู่ %d รายการ"%used)
        UnitManager.ensure()
        conn=get_db()
        try:
            conn.execute("DELETE FROM units WHERE name=?",(name,)); conn.commit(); return (True,"")
        finally: conn.close()

# ══════════════════════════════════════════════════════════════════════════════
# SHARED THAI DATE HELPERS — ใช้ร่วมกันทุก Tab
# รูปแบบแสดงผล : DD/MM/BBBB (พ.ศ.)  รูปแบบ DB : YYYY-MM-DD (ค.ศ.)
# ══════════════════════════════════════════════════════════════════════════════

def _th_to_iso(s):
    """แปลง DD/MM/BBBB → YYYY-MM-DD  (คืน s เดิมถ้า parse ไม่ได้)"""
    try:
        p = s.strip().split("/")
        if len(p) == 3:
            d, m, y = int(p[0]), int(p[1]), int(p[2])
            if y > 2500: y -= 543          # พ.ศ. → ค.ศ.
            return f"{y:04d}-{m:02d}-{d:02d}"
    except Exception:
        pass
    return s

def _iso_to_th(ymd):
    """แปลง YYYY-MM-DD → DD/MM/BBBB"""
    try:
        p = ymd.split("-")
        return f"{int(p[2]):02d}/{int(p[1]):02d}/{int(p[0])+543}"
    except Exception:
        return ymd

def _today_th():
    """วันนี้ในรูปแบบ DD/MM/BBBB"""
    return _iso_to_th(date.today().strftime("%Y-%m-%d"))

def show_thai_calendar(entry_widget, root_window=None):
    """
    เปิด popup ปฏิทินภาษาไทย (พ.ศ.) ใต้ entry_widget
    เขียนค่าลงใน entry_widget เป็น DD/MM/BBBB
    """
    import calendar as _cal

    # ── parse ค่าปัจจุบัน ──────────────────────────────────
    try:
        iso = _th_to_iso(entry_widget.get().strip())
        from datetime import datetime as _dt2
        cur_date = _dt2.strptime(iso, "%Y-%m-%d").date()
    except Exception:
        cur_date = date.today()

    # ── สร้าง Toplevel ────────────────────────────────────
    parent_win = root_window or entry_widget.winfo_toplevel()
    win = tk.Toplevel(parent_win)
    win.title("เลือกวันที่")
    win.configure(bg=C_BG)
    win.resizable(False, False)
    win.overrideredirect(True)   # ไม่มี title bar → ดูสะอาด
    win.grab_set()
    win.lift()

    # ── วางตำแหน่ง (จะเรียกหลัง render เสร็จ) ────────────
    def _place():
        win.update_idletasks()
        ex = entry_widget.winfo_rootx()
        ey = entry_widget.winfo_rooty() + entry_widget.winfo_height() + 2
        pw = win.winfo_reqwidth()
        ph = win.winfo_reqheight()
        sw = win.winfo_screenwidth()
        sh = win.winfo_screenheight()
        if ex + pw > sw: ex = sw - pw - 4
        if ey + ph > sh: ey = entry_widget.winfo_rooty() - ph - 2
        ex = max(ex, 0); ey = max(ey, 0)
        win.geometry(f"+{ex}+{ey}")

    # ── ตกแต่ง ─────────────────────────────────────────────
    # border frame
    outer = tk.Frame(win, bg=C_ACCENT, padx=1, pady=1)
    outer.pack()
    frm = tk.Frame(outer, bg=C_BG, padx=10, pady=10)
    frm.pack()

    THAI_MONTHS = ["","ม.ค.","ก.พ.","มี.ค.","เม.ย.","พ.ค.","มิ.ย.",
                   "ก.ค.","ส.ค.","ก.ย.","ต.ค.","พ.ย.","ธ.ค."]
    DAYS = ["จ.", "อ.", "พ.", "พฤ.", "ศ.", "ส.", "อา."]

    state = {"year": cur_date.year, "month": cur_date.month}

    # ── Header ─────────────────────────────────────────────
    hdr = tk.Frame(frm, bg=C_BG); hdr.pack(fill=tk.X, pady=(0, 6))
    prev_btn = tk.Button(hdr, text="◀", font=F_BODY, bg=C_BG, fg=C_MUTED,
                         relief=tk.FLAT, cursor="hand2", bd=0, padx=6,
                         activebackground=C_SURFACE2, activeforeground=C_TEXT)
    prev_btn.pack(side=tk.LEFT)
    title_lbl = tk.Label(hdr, text="", font=F_H2, bg=C_BG, fg=C_TEXT, width=14, anchor="center")
    title_lbl.pack(side=tk.LEFT, expand=True)
    next_btn = tk.Button(hdr, text="▶", font=F_BODY, bg=C_BG, fg=C_MUTED,
                         relief=tk.FLAT, cursor="hand2", bd=0, padx=6,
                         activebackground=C_SURFACE2, activeforeground=C_TEXT)
    next_btn.pack(side=tk.RIGHT)

    # ── Day-of-week headers ─────────────────────────────────
    dh = tk.Frame(frm, bg=C_BG); dh.pack(fill=tk.X, pady=(0, 2))
    for i, d in enumerate(DAYS):
        fg = C_ACCENT2 if i == 6 else (C_MUTED if i < 5 else C_YELLOW)
        tk.Label(dh, text=d, font=(_FN, 9, "bold"), bg=C_BG,
                 fg=fg, width=4, anchor="center").pack(side=tk.LEFT)

    # ── Calendar grid ───────────────────────────────────────
    cal_f = tk.Frame(frm, bg=C_BG); cal_f.pack()

    def _select(dt):
        entry_widget.delete(0, tk.END)
        entry_widget.insert(0, _iso_to_th(dt.strftime("%Y-%m-%d")))
        win.destroy()

    def render():
        for w in cal_f.winfo_children(): w.destroy()
        y, m = state["year"], state["month"]
        thai_year = y + 543
        title_lbl.config(text=f"{THAI_MONTHS[m]}  {thai_year}")
        fd = date(y, m, 1).weekday()           # 0=Mon
        days_in = _cal.monthrange(y, m)[1]
        row = tk.Frame(cal_f, bg=C_BG); row.pack(anchor="w")
        col = 0
        for _ in range(fd):
            tk.Label(row, text="", width=4, bg=C_BG).pack(side=tk.LEFT)
            col += 1
        for day in range(1, days_in + 1):
            d = date(y, m, day)
            is_today = d == date.today()
            is_sel   = d == cur_date
            is_sun   = d.weekday() == 6
            if is_sel:
                bg, fg = C_ACCENT, C_BG
            elif is_today:
                bg, fg = C_SURFACE2, C_ACCENT
            else:
                bg, fg = C_BG, (C_ACCENT2 if is_sun else C_TEXT)
            weight = "bold" if (is_today or is_sel) else "normal"
            btn = tk.Button(row, text=str(day), width=3, padx=0, pady=4,
                            font=(_FN, 10, weight), bg=bg, fg=fg,
                            relief=tk.FLAT, cursor="hand2", bd=0,
                            activebackground=C_SURFACE2, activeforeground=C_TEXT,
                            command=lambda dt=d: _select(dt))
            btn.pack(side=tk.LEFT)
            col += 1
            if col % 7 == 0 and day < days_in:
                row = tk.Frame(cal_f, bg=C_BG); row.pack(anchor="w")

    def prev_month():
        if state["month"] == 1: state["year"] -= 1; state["month"] = 12
        else: state["month"] -= 1
        render()

    def next_month():
        if state["month"] == 12: state["year"] += 1; state["month"] = 1
        else: state["month"] += 1
        render()

    prev_btn.config(command=prev_month)
    next_btn.config(command=next_month)

    # ── Footer: วันนี้ + ปิด ────────────────────────────────
    foot = tk.Frame(frm, bg=C_BG); foot.pack(fill=tk.X, pady=(8, 0))
    tk.Button(foot, text="📅 วันนี้", font=F_SMALL, bg=C_SURFACE2, fg=C_ACCENT,
              relief=tk.FLAT, cursor="hand2", bd=0, padx=10, pady=4,
              activebackground=C_BORDER,
              command=lambda: _select(date.today())).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(0, 4))
    tk.Button(foot, text="✖", font=F_SMALL, bg=C_SURFACE2, fg=C_MUTED,
              relief=tk.FLAT, cursor="hand2", bd=0, padx=10, pady=4,
              activebackground=C_BORDER,
              command=win.destroy).pack(side=tk.RIGHT)

    render()
    win.after(1, _place)
    # ── ปิดเมื่อคลิกนอก popup ─────────────────────────────
    win.bind("<FocusOut>", lambda e: win.destroy() if win.winfo_exists() else None)


class ThaiDateEntry(tk.Frame):
    """
    Entry + ▼ button สำหรับเลือกวันที่ภาษาไทย (พ.ศ.)
    ใช้แทน field() + Button ▼ แบบเดิม
    get()  → DD/MM/BBBB
    get_iso() → YYYY-MM-DD
    set(iso_or_th) → ตั้งค่า
    """
    def __init__(self, parent, width=11, **kw):
        super().__init__(parent, bg=C_BG)
        self._ent = tk.Entry(self, width=width, font=F_BODY,
                             bg=C_SURFACE2, fg=C_TEXT,
                             insertbackground=C_ACCENT,
                             relief=tk.FLAT,
                             highlightthickness=1,
                             highlightbackground=C_BORDER,
                             highlightcolor=C_ACCENT)
        self._ent.pack(side=tk.LEFT, ipady=4)
        tk.Button(self, text="▼", font=(_FN, 9), bg=C_SURFACE2, fg=C_TEXT,
                  relief=tk.FLAT, cursor="hand2", bd=0, padx=5,
                  activebackground=C_BORDER,
                  command=self._open_cal).pack(side=tk.LEFT, padx=(1, 0))
        # คลิกที่ entry ก็เปิด calendar ได้เลย
        self._ent.bind("<Button-1>", lambda e: self._open_cal())

    def _open_cal(self):
        show_thai_calendar(self._ent)

    def get(self):
        return self._ent.get()

    def get_iso(self):
        return _th_to_iso(self._ent.get().strip())

    def set(self, value):
        """รับ YYYY-MM-DD หรือ DD/MM/BBBB"""
        self._ent.delete(0, tk.END)
        if value and "-" in value and len(value) == 10:
            self._ent.insert(0, _iso_to_th(value))
        else:
            self._ent.insert(0, value)

    def delete(self, *a):  self._ent.delete(*a)
    def insert(self, *a):  self._ent.insert(*a)
    def configure(self, **kw): self._ent.configure(**kw)


# ══ STOCK CARD TAB ════════════════════════════════════════════════════════════
class StockCardTab(tk.Frame):
    _TYPE_COLORS = {"sale":"#ff6b6b","restock":"#3fb950","return":"#58a6ff",
                    "adjust":"#ffd166","import":"#a8dadc","manual":"#8b949e"}
    _TYPE_TH     = {"sale":"ขาย","restock":"รับเข้า","return":"คืนสินค้า",
                    "adjust":"ปรับสต็อก","import":"Import","manual":"Manual"}

    def __init__(self,parent):
        super().__init__(parent,bg=C_BG); self._build()

    def _build(self):
        tk.Frame(self,bg=C_BLUE,height=3).pack(fill=tk.X)
        hf=tk.Frame(self,bg=C_BG,padx=14,pady=10); hf.pack(fill=tk.X)
        tk.Label(hf,text="📋  Stock Card — ประวัติเคลื่อนไหวสินค้า",
                 font=F_H1,bg=C_BG,fg=C_TEXT).pack(side=tk.LEFT)

        # filter bar
        fb=tk.Frame(self,bg=C_BG,padx=12); fb.pack(fill=tk.X,pady=(0,4))
        tk.Label(fb,text="สินค้า:",font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(side=tk.LEFT)
        self.prod_sv=tk.StringVar()
        pe=field(fb,width=22); pe.configure(textvariable=self.prod_sv)
        pe.pack(side=tk.LEFT,ipady=4,padx=(4,10))
        tk.Label(fb,text="ประเภท:",font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(side=tk.LEFT)
        self.type_var=tk.StringVar(value="ทั้งหมด")
        ttk.Combobox(fb,textvariable=self.type_var,
                     values=["ทั้งหมด","ขาย","รับเข้า","คืนสินค้า","ปรับสต็อก","Import"],
                     font=F_BODY,width=12,state="readonly").pack(side=tk.LEFT,padx=(4,10))
        tk.Label(fb,text="จาก:",font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(side=tk.LEFT)
        self._df=ThaiDateEntry(fb,width=10); self._df.pack(side=tk.LEFT,padx=(4,4))
        tk.Label(fb,text="ถึง:",font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(side=tk.LEFT)
        self._dt=ThaiDateEntry(fb,width=10); self._dt.pack(side=tk.LEFT,padx=(4,8))
        accent_btn(fb,"🔍 ดู",self.load,pad_x=12,pad_y=4).pack(side=tk.LEFT)
        pill_btn(fb,"✖ ล้าง",self._clear,bg=C_SURFACE2,fg=C_MUTED,pad_x=8,pad_y=4).pack(side=tk.LEFT,padx=4)

        # summary cards
        sf=tk.Frame(self,bg=C_BG,padx=12); sf.pack(fill=tk.X,pady=(0,4))
        self._sv={}
        for k,lbl,col in [("in","รับเข้ารวม","#3fb950"),("out","จ่ายออกรวม","#ff6b6b"),
                           ("net","สุทธิ",C_ACCENT),("cnt","รายการ",C_MUTED)]:
            v=tk.StringVar(value="—"); self._sv[k]=v
            c2=tk.Frame(sf,bg=C_SURFACE2,highlightbackground=C_BORDER,highlightthickness=1)
            c2.pack(side=tk.LEFT,fill=tk.X,expand=True,padx=3)
            tk.Label(c2,text=lbl,font=F_SMALL,bg=C_SURFACE2,fg=C_MUTED,pady=4).pack(anchor="w",padx=10)
            tk.Label(c2,textvariable=v,font=F_NUM_SM,bg=C_SURFACE2,fg=col).pack(anchor="w",padx=10,pady=(0,6))

        # tree
        cols=("วันที่เวลา","สินค้า","ประเภท","เปลี่ยน","ก่อน","หลัง","ต้นทุน","พนักงาน","หมายเหตุ")
        frm,self.tree=make_tree(self,cols,(130,180,70,60,60,60,80,90,140))
        frm.pack(fill=tk.BOTH,expand=True,padx=12,pady=(0,4))
        for t,c3 in self._TYPE_COLORS.items():
            self.tree.tag_configure(t,foreground=c3)
        self.count_lbl=tk.Label(self,text="",font=F_SMALL,bg=C_BG,fg=C_MUTED)
        self.count_lbl.pack(anchor="e",padx=14,pady=(0,6))

        # init date (30 วันล่าสุด)
        from datetime import timedelta
        today=date.today(); d30=today-timedelta(days=30)
        self._df.set(d30.strftime("%Y-%m-%d")); self._dt.set(today.strftime("%Y-%m-%d"))

    def _from_thai(self,s):
        try:
            p=s.strip().split("/")
            if len(p)==3 and int(p[2])>2500:
                return f"{int(p[2])-543:04d}-{int(p[1]):02d}-{int(p[0]):02d}"
        except: pass
        return s

    def _clear(self):
        from datetime import timedelta
        self.prod_sv.set(""); self.type_var.set("ทั้งหมด")
        today=date.today(); d30=today-timedelta(days=30)
        self._df.set(d30.strftime("%Y-%m-%d")); self._dt.set(today.strftime("%Y-%m-%d"))
        self.load()

    def load(self,*_):
        d0=self._df.get_iso()
        d1=self._dt.get_iso()
        q=self.prod_sv.get().strip()
        typ=self.type_var.get()
        type_map={"ขาย":"sale","รับเข้า":"restock","คืนสินค้า":"return",
                  "ปรับสต็อก":"adjust","Import":"import"}
        where="date(m.movement_date) BETWEEN ? AND ?"; params=[d0,d1]
        if q:
            where+=(" AND (m.product_name LIKE ?"
                    " OR EXISTS(SELECT 1 FROM products p"
                    "  WHERE p.id=m.product_id AND p.barcode LIKE ?))")
            params.extend([f"%{q}%", f"%{q}%"])
        if typ!="ทั้งหมด" and typ in type_map:
            where+=" AND m.movement_type=?"; params.append(type_map[typ])
        conn=get_db()
        rows=conn.execute(
            f"SELECT m.* FROM stock_movements m WHERE {where} ORDER BY m.movement_date DESC LIMIT 500",
            params).fetchall()
        conn.close()
        ti=sum(r["qty_change"] for r in rows if r["qty_change"]>0)
        to=sum(r["qty_change"] for r in rows if r["qty_change"]<0)
        self._sv["in"].set(f"+{ti:,}"); self._sv["out"].set(f"{to:,}")
        self._sv["net"].set(f"{ti+to:+,}"); self._sv["cnt"].set(f"{len(rows):,}")
        self.tree.delete(*self.tree.get_children())
        for i,r in enumerate(rows):
            mt=r["movement_type"]; sign="+" if r["qty_change"]>0 else ""
            self.tree.insert("","end",iid=r["id"],tags=(mt,"alt" if i%2==0 else ""),
                values=(r["movement_date"],r["product_name"],
                        self._TYPE_TH.get(mt,mt),
                        f"{sign}{r['qty_change']}",
                        r["qty_before"],r["qty_after"],
                        f"฿{r['cost']:,.2f}" if r["cost"] else "—",
                        r["staff_name"] or "—",r["note"] or "—"))
        self.count_lbl.config(text=f"แสดง {len(rows):,} รายการ (สูงสุด 500)")




    

class ProductTab(tk.Frame):
    def __init__(self,parent):
        super().__init__(parent,bg=C_BG); self._build(); self.load()

    def _build(self):
        # Dashboard summary cards
        dc=tk.Frame(self,bg=C_BG,padx=12); dc.pack(fill=tk.X,pady=(8,4))
        self.stat_cards={}
        # 3 color groups
        INFO  = "#4a9eff"   # ฟ้าอมเขียว — ข้อมูลทั่วไป
        WARN  = "#e67e22"   # ส้ม — แจ้งเตือน
        COST  = "#5dade2"   # ฟ้า — ต้นทุน
        SELL  = "#f4c542"   # เหลืองทอง — ราคาขาย
        PROF  = "#58d68d"   # เขียว — กำไร
        for key,lbl,col,dflt in [
            ("total","\U0001f4e6 \u0e2a\u0e34\u0e19\u0e04\u0e49\u0e32\u0e17\u0e31\u0e49\u0e07\u0e2b\u0e21\u0e14",INFO,"0 \u0e23\u0e32\u0e22\u0e01\u0e32\u0e23"),
            ("cats","\U0001f3f7\ufe0f \u0e2b\u0e21\u0e27\u0e14\u0e2b\u0e21\u0e39\u0e48",INFO,"0 \u0e2b\u0e21\u0e27\u0e14"),
            ("low","\u26a0\ufe0f \u0e43\u0e01\u0e25\u0e49\u0e2b\u0e21\u0e14",WARN,"0 \u0e23\u0e32\u0e22\u0e01\u0e32\u0e23"),
            ("vcost","\U0001f4b0 \u0e21\u0e39\u0e25\u0e04\u0e48\u0e32\u0e17\u0e38\u0e19",COST,"\u0e3f0"),
            ("vsell","\U0001f4b5 \u0e21\u0e39\u0e25\u0e04\u0e48\u0e32\u0e02\u0e32\u0e22",SELL,"\u0e3f0"),
            ("profit","\U0001f4c8 \u0e01\u0e33\u0e44\u0e23",PROF,"\u0e3f0"),
        ]:
            cf=tk.Frame(dc,bg=C_SURFACE,highlightbackground=col,highlightthickness=2)
            cf.pack(side=tk.LEFT,fill=tk.X,expand=True,padx=3)
            tk.Frame(cf,bg=col,height=3).pack(fill=tk.X)
            tk.Label(cf,text=lbl,font=(FM.primary,10),bg=C_SURFACE,fg=C_MUTED,pady=3).pack(anchor="w",padx=8)
            vl=tk.Label(cf,text=dflt,font=(FM.primary,16,"bold"),bg=C_SURFACE,fg=col,pady=3)
            vl.pack(anchor="w",padx=8); self.stat_cards[key]=vl

        # Toolbar
        top=tk.Frame(self,bg=C_BG,pady=6,padx=12); top.pack(fill=tk.X)
        tk.Label(top,text="\U0001f4e6  \u0e2a\u0e34\u0e19\u0e04\u0e49\u0e32\u0e41\u0e25\u0e30\u0e2a\u0e15\u0e47\u0e2d\u0e01",font=F_H1,bg=C_BG,fg=C_TEXT).pack(side=tk.LEFT)
        bf=tk.Frame(top,bg=C_BG); bf.pack(side=tk.RIGHT)
        accent_btn(bf,"\u2795 \u0e40\u0e1e\u0e34\u0e48\u0e21",self._add,pad_x=12,pad_y=5).pack(side=tk.LEFT,padx=2)
        pill_btn(bf,"\u270f\ufe0f \u0e41\u0e01\u0e49\u0e44\u0e02",self._edit,bg=C_SURFACE2,fg=C_YELLOW,pad_x=10,pad_y=5).pack(side=tk.LEFT,padx=2)
        pill_btn(bf,"\U0001f4e6 \u0e40\u0e15\u0e34\u0e21\u0e2a\u0e15\u0e47\u0e2d\u0e01",self._restock,bg=C_SURFACE2,fg=C_BLUE,pad_x=10,pad_y=5).pack(side=tk.LEFT,padx=2)
        pill_btn(bf,"📥 รับสินค้า",self._receive_advanced,bg=C_SURFACE2,fg=C_GREEN,pad_x=10,pad_y=5).pack(side=tk.LEFT,padx=2)
        pill_btn(bf,"📜 ประวัติซื้อ",self._purchase_history,bg=C_SURFACE2,fg=C_MUTED,pad_x=10,pad_y=5).pack(side=tk.LEFT,padx=2)
        pill_btn(bf,"\U0001f3f7\ufe0f \u0e2b\u0e21\u0e27\u0e14\u0e2b\u0e21\u0e39\u0e48",self._manage_cats,bg=C_SURFACE2,fg=C_GREEN,pad_x=10,pad_y=5).pack(side=tk.LEFT,padx=2)
        pill_btn(bf,"\U0001f4cf \u0e2b\u0e19\u0e48\u0e27\u0e22\u0e19\u0e31\u0e1a",self._manage_units,bg=C_SURFACE2,fg=C_BLUE,pad_x=10,pad_y=5).pack(side=tk.LEFT,padx=2)
        pill_btn(bf,"\U0001f5d1\ufe0f \u0e25\u0e1a",self._delete,bg=C_ACCENT2,fg=C_WHITE,pad_x=10,pad_y=5).pack(side=tk.LEFT,padx=2)
        pill_btn(bf,"\u21ba",self.load,bg=C_SURFACE,fg=C_MUTED,pad_x=8,pad_y=5).pack(side=tk.LEFT,padx=2)
        self.show_abc=tk.BooleanVar(value=False)
        tk.Checkbutton(bf,text="\u0e41\u0e2a\u0e14\u0e07\u0e23\u0e32\u0e04\u0e32 A/B/C",
            variable=self.show_abc,command=self.load,
            bg=C_BG,fg=C_MUTED,selectcolor=C_SURFACE2,
            activebackground=C_BG,activeforeground=C_TEXT,
            font=(FM.primary,11)).pack(side=tk.LEFT,padx=8)

        # ── Search Bar ────────────────────────────────────────
        sf=tk.Frame(self,bg=C_BG,padx=12); sf.pack(fill=tk.X,pady=(0,2))
        search_wrap=tk.Frame(sf,bg=C_SURFACE2,highlightbackground=C_BORDER,highlightthickness=1)
        search_wrap.pack(side=tk.LEFT,fill=tk.X,expand=True,ipady=2)
        tk.Label(search_wrap,text="🔍",font=(FM.primary,13),bg=C_SURFACE2,fg=C_MUTED).pack(side=tk.LEFT,padx=(8,2))
        self.sv=tk.StringVar()
        fe=tk.Entry(search_wrap,textvariable=self.sv,font=(FM.primary,12),bg=C_SURFACE2,
                    fg=C_TEXT,insertbackground=C_ACCENT,relief=tk.FLAT,bd=0,highlightthickness=0)
        fe.pack(side=tk.LEFT,fill=tk.X,expand=True,ipady=6,padx=(0,8))
        fe.bind("<FocusIn>", lambda e:search_wrap.config(highlightbackground=C_ACCENT,highlightthickness=2))
        fe.bind("<FocusOut>",lambda e:search_wrap.config(highlightbackground=C_BORDER,highlightthickness=1))
        fe.bind("<Double-Button-1>",lambda e:self._clear_filters())
        self.low_only=tk.BooleanVar(value=False)
        self.neg_only=tk.BooleanVar(value=False)
        self.multi_only=tk.BooleanVar(value=False)
        def _on_search(*_):
            if self.sv.get().strip():
                if self.low_only.get() or self.neg_only.get():
                    self.low_only.set(False); self.neg_only.set(False)
                    self._update_filter_bar()
            self.load()
        self.sv.trace_add("write",_on_search)
        fb=tk.Frame(sf,bg=C_BG); fb.pack(side=tk.LEFT,padx=(8,0))
        tk.Checkbutton(fb,text="⚠️ สต็อกต่ำ",variable=self.low_only,command=self._on_filter_change,
                       bg=C_BG,fg="#e67e22",selectcolor="#2d1a00",activebackground=C_BG,
                       activeforeground="#e67e22",font=(FM.primary,11,"bold")).pack(side=tk.LEFT,padx=4)
        tk.Checkbutton(fb,text="📉 ติดลบ",variable=self.neg_only,command=self._on_filter_change,
                       bg=C_BG,fg="#ff4444",selectcolor="#2d0000",activebackground=C_BG,
                       activeforeground="#ff4444",font=(FM.primary,11,"bold")).pack(side=tk.LEFT,padx=(0,4))
        tk.Checkbutton(fb,text="📦 หลายหน่วย",variable=self.multi_only,command=self.load,
                       bg=C_BG,fg="#4aa3ff",selectcolor="#0a1f33",activebackground=C_BG,
                       activeforeground="#4aa3ff",font=(FM.primary,11,"bold")).pack(side=tk.LEFT,padx=(0,4))
        self.count_lbl=tk.Label(sf,text="",font=F_SMALL,bg=C_BG,fg=C_MUTED); self.count_lbl.pack(side=tk.RIGHT,padx=8)
        self._filter_bar=tk.Frame(self,bg=C_BG)
        self._filter_status=tk.Label(self._filter_bar,text="",font=(FM.primary,11,"bold"),bg=C_BG,fg="#e67e22",anchor="w")
        self._filter_status.pack(side=tk.LEFT,padx=(12,0),pady=4)
        tk.Button(self._filter_bar,text="✖ ล้างตัวกรอง",command=self._clear_filters,
                  font=(FM.primary,10),bg=C_SURFACE2,fg=C_MUTED,relief=tk.FLAT,cursor="hand2",
                  padx=8,pady=2,bd=0).pack(side=tk.LEFT,padx=8)

        # Table
        cols=("ID","\u0e1a\u0e32\u0e23\u0e4c\u0e42\u0e04\u0e49\u0e14","\u0e0a\u0e37\u0e48\u0e2d\u0e2a\u0e34\u0e19\u0e04\u0e49\u0e32","\u0e2b\u0e21\u0e27\u0e14\u0e2b\u0e21\u0e39\u0e48","\u0e23\u0e32\u0e04\u0e32\u0e02\u0e32\u0e22","\u0e15\u0e49\u0e19\u0e17\u0e38\u0e19","A","B","C","\u0e2a\u0e15\u0e47\u0e2d\u0e01","\u0e02\u0e31\u0e49\u0e19\u0e15\u0e48\u0e33","\u0e2b\u0e19\u0e48\u0e27\u0e22","\u0e2a\u0e16\u0e32\u0e19\u0e30")
        frm,self.tree=make_tree(self,cols,(40,90,190,90,75,75,55,55,55,60,60,55,65))
        frm.pack(fill=tk.BOTH,expand=True,padx=12,pady=(0,4))
        self._build_ctx_menu()
        self.tree.bind("<Button-3>", self._show_ctx_menu)
        self.tree.bind("<Double-1>", lambda e: self._edit())

        # Footer
        ff=tk.Frame(self,bg=C_SURFACE2,padx=16,pady=8); ff.pack(fill=tk.X,padx=12,pady=(0,8))
        self.foot={}
        foot_colors={"cnt":"#d0d6e0","qty":"#58d68d","vcost":"#5dade2","vsell":"#f4c542","profit":"#58d68d"}
        for key,lbl in [("cnt","\u0e2a\u0e34\u0e19\u0e04\u0e49\u0e32"),("qty","\u0e04\u0e07\u0e40\u0e2b\u0e25\u0e37\u0e2d"),("vcost","\u0e21\u0e39\u0e25\u0e04\u0e48\u0e32\u0e17\u0e38\u0e19"),("vsell","\u0e21\u0e39\u0e25\u0e04\u0e48\u0e32\u0e02\u0e32\u0e22"),("profit","\u0e01\u0e33\u0e44\u0e23")]:
            col=tk.Frame(ff,bg=C_SURFACE2); col.pack(side=tk.LEFT,fill=tk.X,expand=True)
            tk.Label(col,text=lbl,font=(FM.primary,10),bg=C_SURFACE2,fg=C_MUTED).pack(anchor="w")
            vl=tk.Label(col,text="\u2014",font=(FM.primary,13,"bold"),bg=C_SURFACE2,fg=foot_colors.get(key,C_TEXT)); vl.pack(anchor="w"); self.foot[key]=vl

    def _on_filter_change(self):
        self._update_filter_bar(); self.load()

    def _update_filter_bar(self):
        low=self.low_only.get(); neg=self.neg_only.get()
        if low or neg:
            msg,color,bg2 = ("📉  กำลังแสดง: สต็อกติดลบ","#ff4444","#1a0000") if neg else ("⚠️  กำลังแสดง: สินค้าสต็อกต่ำ","#e67e22","#1a0e00")
            self._filter_bar.config(bg=bg2)
            self._filter_status.config(text=msg,fg=color,bg=bg2)
            # insert filter_bar ก่อน tree (index 3 ของ self children)
            try: self._filter_bar.pack(fill=tk.X,before=self.winfo_children()[3])
            except: self._filter_bar.pack(fill=tk.X)
        else:
            self._filter_bar.pack_forget()

    def _clear_filters(self):
        self.low_only.set(False); self.neg_only.set(False)
        self.sv.set(""); self._update_filter_bar(); self.load()

    def reset_filters(self):
        """เรียกจาก _on_tab เมื่อออกจากหน้า"""
        self.low_only.set(False); self.neg_only.set(False)
        self._update_filter_bar()

    def load(self,*_):
        rows=search_products_filtered(self.sv.get().strip(),self.low_only.get(),self.neg_only.get())
        all_rows=get_all_products_active()
        total=len(all_rows); cats=len(set(r["category"] for r in all_rows if r["category"]))
        low=sum(1 for r in all_rows if 0<r["stock"]<=r["min_stock"])
        vc=sum((r["cost"] or 0)*r["stock"] for r in all_rows)
        vs=sum((r["price"] or 0)*r["stock"] for r in all_rows)
        self.stat_cards["total"].config(text=f"{total:,} \u0e23\u0e32\u0e22\u0e01\u0e32\u0e23")
        self.stat_cards["cats"].config(text=f"{cats} \u0e2b\u0e21\u0e27\u0e14")
        self.stat_cards["low"].config(text=f"{low} \u0e23\u0e32\u0e22\u0e01\u0e32\u0e23",
            fg="#e67e22" if low>0 else "#e74c3c" if low==0 else "#e67e22")
        self.stat_cards["vcost"].config(text=f"\u0e3f{vc:,.0f}")
        self.stat_cards["vsell"].config(text=f"\u0e3f{vs:,.0f}")
        profit=vs-vc
        self.stat_cards["profit"].config(text=f"\u0e3f{profit:,.0f}",
            fg="#58d68d" if profit>=0 else "#e74c3c")
        self.tree.delete(*self.tree.get_children())
        # Row alternating colors
        self.tree.tag_configure("row0", background="#161B22", foreground="#d0d6e0")
        self.tree.tag_configure("row1", background="#1B212B", foreground="#d0d6e0")
        # Status tags override row color for status column only (applied via foreground)
        self.tree.tag_configure("ok",  foreground="#58d68d")  # เขียว
        self.tree.tag_configure("low", foreground="#e67e22")  # ส้ม
        self.tree.tag_configure("out", foreground="#e74c3c")  # แดง
        self.tree.tag_configure("neg", foreground="#ff2222")  # แดงสด: ติดลบ
        # 📦 แถวสินค้าหลายหน่วย: พื้นหลังฟ้าเข้ม + คงสีสถานะ
        self.tree.tag_configure("multi_ok",  background="#13283b", foreground="#58d68d")
        self.tree.tag_configure("multi_low", background="#13283b", foreground="#e67e22")
        self.tree.tag_configure("multi_out", background="#13283b", foreground="#e74c3c")
        self.tree.tag_configure("multi_neg", background="#13283b", foreground="#ff2222")

        # ── ซ่อน/แสดง column A B C ตาม show_abc ──────────
        show_abc = self.show_abc.get()
        for col,w in [("A",60),("B",60),("C",60)]:
            self.tree.column(col, width=w if show_abc else 0,
                             minwidth=0, stretch=False)
            self.tree.heading(col, text=col if show_abc else "")
        cnt=qty=vcost=vsell=0
        multi_ids=unit_service.get_multi_ids(DB_FILE)   # 📦 สินค้าที่มีหน่วยหลายขนาด
        if self.multi_only.get():
            rows=[r for r in rows if r["id"] in multi_ids]
        for i,r in enumerate(rows):
            min_s=r["min_stock"] if "min_stock" in r.keys() else 5
            if r["stock"]<0: tag,stat="neg",f"\u0e15\u0e34\u0e14\u0e25\u0e1a {r['stock']}"
            elif r["stock"]==0: tag,stat="out","\u0e2b\u0e21\u0e14!"
            elif r["stock"]<=min_s: tag,stat="low","\u0e15\u0e48\u0e33"
            else: tag,stat="ok","\u0e1b\u0e01\u0e15\u0e34"
            pa=r["price_a"] if "price_a" in r.keys() and r["price_a"] else None
            pb=r["price_b"] if "price_b" in r.keys() and r["price_b"] else None
            pc=r["price_c"] if "price_c" in r.keys() and r["price_c"] else None
            _ismulti = r["id"] in multi_ids
            _rowtags = ("multi_"+tag,) if _ismulti else ("row{}".format(i%2),tag)
            self.tree.insert("","end",iid=r["id"],tags=_rowtags,
                values=(r["id"],r["barcode"] or "\u2014",r["name"],r["category"] or "\u2014",
                        f"{r['price']:.2f}",f"{r['cost']:.2f}",
                        f"{pa:.2f}" if pa else "\u2014",f"{pb:.2f}" if pb else "\u2014",f"{pc:.2f}" if pc else "\u2014",
                        r["stock"],min_s,(f"📦 {r['unit']}" if _ismulti else r["unit"]),stat))
            cnt+=1; qty+=r["stock"]; vcost+=r["cost"]*r["stock"]; vsell+=r["price"]*r["stock"]
        self.count_lbl.config(text=f"\u0e41\u0e2a\u0e14\u0e07 {cnt:,} / {total:,} \u0e23\u0e32\u0e22\u0e01\u0e32\u0e23")
        self.foot["cnt"].config(text=f"{cnt:,} \u0e23\u0e32\u0e22\u0e01\u0e32\u0e23")
        self.foot["qty"].config(text=f"{qty:,} \u0e2b\u0e19\u0e48\u0e27\u0e22")
        self.foot["vcost"].config(text=f"\u0e3f{vcost:,.2f}")
        self.foot["vsell"].config(text=f"\u0e3f{vsell:,.2f}")
        self.foot["profit"].config(text=f"\u0e3f{vsell-vcost:,.2f}")

    def _sel(self):
        s=self.tree.selection()
        if not s: messagebox.showwarning("\u0e40\u0e25\u0e37\u0e2d\u0e01","\u0e01\u0e23\u0e38\u0e13\u0e32\u0e40\u0e25\u0e37\u0e2d\u0e01\u0e2a\u0e34\u0e19\u0e04\u0e49\u0e32\u0e01\u0e48\u0e2d\u0e19"); return None
        return int(s[0])

    def _build_ctx_menu(self):
        """เมนูคลิกขวา — reuse command เดิม; รายการที่ยังไม่มี = disabled"""
        m=tk.Menu(self,tearoff=0,bg=C_SURFACE,fg=C_TEXT,
                  activebackground=C_ACCENT,activeforeground=C_WHITE,bd=0,font=(FM.primary,11))
        m.add_command(label="✏️  แก้ไขสินค้า",command=self._edit)
        m.add_command(label="➕  เพิ่มสินค้า",command=self._add)
        m.add_separator()
        m.add_command(label="📦  เติมสต็อก",command=self._restock)
        m.add_command(label="📥  รับสินค้า",command=self._receive_advanced)
        m.add_command(label="🎁  โปรโมชั่นสินค้า",command=self._edit)        # โปรอยู่ในฟอร์มแก้ไข
        m.add_command(label="📐  หน่วยขายหลายขนาด",command=self._edit)       # หน่วยอยู่ในฟอร์มแก้ไข
        m.add_separator()
        m.add_command(label="📋  ประวัติสินค้า (เร็วๆ นี้)",state="disabled")
        m.add_command(label="🧾  ประวัติขาย (เร็วๆ นี้)",state="disabled")
        m.add_command(label="📜  ประวัติซื้อ",command=self._purchase_history)
        m.add_separator()
        m.add_command(label="🏷️  พิมพ์บาร์โค้ด (เร็วๆ นี้)",state="disabled")
        m.add_command(label="⧉  คัดลอกบาร์โค้ด",command=self._copy_barcode)
        m.add_separator()
        m.add_command(label="🗑️  ลบสินค้า",command=self._delete,foreground="#ff6b6b")
        self._ctx_menu=m

    def _show_ctx_menu(self,event):
        iid=self.tree.identify_row(event.y)
        if iid: self.tree.selection_set(iid)
        try: self._ctx_menu.tk_popup(event.x_root,event.y_root)
        finally: self._ctx_menu.grab_release()

    def _copy_barcode(self):
        sel=self.tree.selection()
        if not sel: return
        vals=self.tree.item(sel[0],"values")
        bc=(vals[1] if vals and len(vals)>1 else "") or ""
        if bc in ("\u2014","-",""):
            messagebox.showinfo("บาร์โค้ด","สินค้านี้ไม่มีบาร์โค้ด",parent=self.winfo_toplevel()); return
        try:
            self.clipboard_clear(); self.clipboard_append(str(bc)); self.update_idletasks()
        except Exception: pass
        self.count_lbl.config(text=f"✓ คัดลอกบาร์โค้ด {bc}")

    def _manage_cats(self):
        CategoryManager.ensure()
        win=tk.Toplevel(self); win.title("หมวดหมู่สินค้า"); win.configure(bg=C_BG)
        win.geometry("400x520"); win.grab_set(); win.lift()
        tk.Frame(win,bg=C_GREEN,height=3).pack(fill=tk.X)
        tk.Label(win,text="🏷️  หมวดหมู่สินค้า",font=F_H1,bg=C_BG,fg=C_TEXT,pady=8).pack(anchor="w",padx=16)
        lb=tk.Listbox(win,font=F_BODY,bg=C_SURFACE2,fg=C_TEXT,selectbackground=C_ACCENT,selectforeground=C_BG,relief=tk.FLAT,highlightthickness=0,height=12)
        lb.pack(fill=tk.BOTH,expand=True,padx=16,pady=(0,4))
        names=[]
        def ref():
            lb.delete(0,tk.END); names[:]=CategoryManager.get_all()
            for c in names:
                u=CategoryManager.count_usage(c)
                lb.insert(tk.END, f"{c}      • ใช้ {u} รายการ" if u else f"{c}      • ยังไม่ใช้")
        ref()
        tk.Label(win,text="พิมพ์ชื่อ → กด เพิ่ม  |  เลือกในลิสต์ → แก้ไขชื่อ/ลบ",font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(anchor="w",padx=16,pady=(4,2))
        e=field(win,width=28); e.pack(fill=tk.X,padx=16,ipady=5,pady=(0,6))
        def cur():
            sel=lb.curselection(); return names[sel[0]] if sel else None
        lb.bind("<<ListboxSelect>>",lambda _:(e.delete(0,tk.END),e.insert(0,cur())) if cur() else None)
        bf2=tk.Frame(win,bg=C_BG); bf2.pack(fill=tk.X,padx=16,pady=(0,12))
        def add():
            ok,msg=CategoryManager.add(e.get())
            if not ok: messagebox.showwarning("เพิ่มไม่ได้",msg,parent=win); return
            ref(); e.delete(0,tk.END); self.load()
        def ren():
            old=cur()
            if not old: messagebox.showinfo("แก้ไขชื่อ","เลือกหมวดในลิสต์ก่อน",parent=win); return
            ok,msg=CategoryManager.rename(old,e.get())
            if not ok: messagebox.showwarning("แก้ไขไม่ได้",msg,parent=win); return
            ref(); e.delete(0,tk.END); self.load()
        def dlt():
            old=cur()
            if not old: messagebox.showinfo("ลบ","เลือกหมวดในลิสต์ก่อน",parent=win); return
            u=CategoryManager.count_usage(old)
            if u>0:
                messagebox.showwarning("ลบไม่ได้",f"มีสินค้าใช้หมวด '{old}' อยู่ {u} รายการ\nเปลี่ยนหมวดสินค้าก่อน จึงจะลบได้",parent=win); return
            if messagebox.askyesno("ยืนยัน",f"ลบหมวด '{old}' ?",parent=win):
                ok,msg=CategoryManager.delete(old)
                if not ok: messagebox.showwarning("ลบไม่ได้",msg,parent=win); return
                ref(); e.delete(0,tk.END); self.load()
        accent_btn(bf2,"➕ เพิ่ม",add,pad_x=10,pad_y=5).pack(side=tk.LEFT,padx=(0,4))
        pill_btn(bf2,"✏️ แก้ไขชื่อ",ren,bg=C_SURFACE2,fg=C_YELLOW,pad_x=10,pad_y=5).pack(side=tk.LEFT,padx=(0,4))
        pill_btn(bf2,"🗑️ ลบ",dlt,bg=C_ACCENT2,fg=C_WHITE,pad_x=10,pad_y=5).pack(side=tk.LEFT)
        pill_btn(bf2,"ปิด",win.destroy,bg=C_SURFACE,fg=C_MUTED,pad_x=10,pad_y=5).pack(side=tk.RIGHT)

    def _manage_units(self):
        UnitManager.ensure()
        win=tk.Toplevel(self); win.title("หน่วยนับ"); win.configure(bg=C_BG)
        win.geometry("400x520"); win.grab_set(); win.lift()
        tk.Frame(win,bg=C_BLUE,height=3).pack(fill=tk.X)
        tk.Label(win,text="📏  หน่วยนับ",font=F_H1,bg=C_BG,fg=C_TEXT,pady=8).pack(anchor="w",padx=16)
        lb=tk.Listbox(win,font=F_BODY,bg=C_SURFACE2,fg=C_TEXT,selectbackground=C_ACCENT,selectforeground=C_BG,relief=tk.FLAT,highlightthickness=0,height=12)
        lb.pack(fill=tk.BOTH,expand=True,padx=16,pady=(0,4))
        names=[]
        def ref():
            lb.delete(0,tk.END); names[:]=UnitManager.get_all()
            for u in names:
                n=UnitManager.count_usage(u)
                lb.insert(tk.END, f"{u}      • ใช้ {n} รายการ" if n else f"{u}      • ยังไม่ใช้")
        ref()
        tk.Label(win,text="พิมพ์ชื่อ → กด เพิ่ม  |  เลือกในลิสต์ → แก้ไขชื่อ/ลบ",font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(anchor="w",padx=16,pady=(4,2))
        e=field(win,width=28); e.pack(fill=tk.X,padx=16,ipady=5,pady=(0,6))
        def cur():
            sel=lb.curselection(); return names[sel[0]] if sel else None
        lb.bind("<<ListboxSelect>>",lambda _:(e.delete(0,tk.END),e.insert(0,cur())) if cur() else None)
        bf2=tk.Frame(win,bg=C_BG); bf2.pack(fill=tk.X,padx=16,pady=(0,12))
        def add():
            ok,msg=UnitManager.add(e.get())
            if not ok: messagebox.showwarning("เพิ่มไม่ได้",msg,parent=win); return
            ref(); e.delete(0,tk.END); self.load()
        def ren():
            old=cur()
            if not old: messagebox.showinfo("แก้ไขชื่อ","เลือกหน่วยในลิสต์ก่อน",parent=win); return
            ok,msg=UnitManager.rename(old,e.get())
            if not ok: messagebox.showwarning("แก้ไขไม่ได้",msg,parent=win); return
            ref(); e.delete(0,tk.END); self.load()
        def dlt():
            old=cur()
            if not old: messagebox.showinfo("ลบ","เลือกหน่วยในลิสต์ก่อน",parent=win); return
            n=UnitManager.count_usage(old)
            if n>0:
                messagebox.showwarning("ลบไม่ได้",f"มีสินค้าใช้หน่วย '{old}' อยู่ {n} รายการ\nเปลี่ยนหน่วยสินค้าก่อน จึงจะลบได้",parent=win); return
            if messagebox.askyesno("ยืนยัน",f"ลบหน่วย '{old}' ?",parent=win):
                ok,msg=UnitManager.delete(old)
                if not ok: messagebox.showwarning("ลบไม่ได้",msg,parent=win); return
                ref(); e.delete(0,tk.END); self.load()
        accent_btn(bf2,"➕ เพิ่ม",add,pad_x=10,pad_y=5).pack(side=tk.LEFT,padx=(0,4))
        pill_btn(bf2,"✏️ แก้ไขชื่อ",ren,bg=C_SURFACE2,fg=C_YELLOW,pad_x=10,pad_y=5).pack(side=tk.LEFT,padx=(0,4))
        pill_btn(bf2,"🗑️ ลบ",dlt,bg=C_ACCENT2,fg=C_WHITE,pad_x=10,pad_y=5).pack(side=tk.LEFT)
        pill_btn(bf2,"ปิด",win.destroy,bg=C_SURFACE,fg=C_MUTED,pad_x=10,pad_y=5).pack(side=tk.RIGHT)

    def _form(self, pid=None):
        p = get_product(pid) if pid else None
        cats=CategoryManager.get_all(); units=UnitManager.get_all()

        win=tk.Toplevel(self)
        win.title("แก้ไขสินค้า" if pid else "เพิ่มสินค้าใหม่")
        win.configure(bg=C_BG); win.grab_set(); win.resizable(False,False)
        tk.Frame(win,bg=C_ACCENT,height=3).pack(fill=tk.X)
        tk.Label(win,text="แก้ไขสินค้า" if pid else "เพิ่มสินค้าใหม่",
                 font=F_H1,bg=C_BG,fg=C_TEXT,pady=10).pack(anchor="w",padx=16)

        body=tk.Frame(win,bg=C_BG); body.pack(fill=tk.BOTH,expand=True,padx=6)
        left=tk.Frame(body,bg=C_BG); left.pack(side=tk.LEFT,fill=tk.BOTH,expand=True,padx=(0,6))
        tk.Frame(body,bg=C_BORDER,width=1).pack(side=tk.LEFT,fill=tk.Y,pady=8)
        right=tk.Frame(body,bg=C_BG); right.pack(side=tk.LEFT,fill=tk.BOTH,expand=True,padx=(6,0))

        entries={}

        def row(lbl,key,default=""):
            tk.Label(left,text=lbl,font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(anchor="w",padx=16,pady=(6,2))
            e=field(left,width=22); e.pack(fill=tk.X,padx=16,pady=(0,2),ipady=5)
            val=str(p[key] if p and p[key] is not None else default)
            if val: e.insert(0,val)
            entries[key]=e; return e

        def combo_row(lbl,key,values,default=""):
            tk.Label(left,text=lbl,font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(anchor="w",padx=16,pady=(6,2))
            var=tk.StringVar(value=str(p[key] if p and p[key] is not None else default))
            cb=ttk.Combobox(left,textvariable=var,values=values,font=F_BODY,width=20)
            cb.pack(fill=tk.X,padx=16,pady=(0,2),ipady=4); entries[key]=var

        # ── Barcode Field + Validation ─────────────────────
        tk.Label(left,text="บาร์โค้ด",font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(anchor="w",padx=16,pady=(6,2))
        bc_row=tk.Frame(left,bg=C_BG); bc_row.pack(fill=tk.X,padx=16,pady=(0,2))

        bc_e=field(bc_row,width=22); bc_e.pack(side=tk.LEFT,fill=tk.X,expand=True,ipady=5)
        if p and p["barcode"]: bc_e.insert(0,p["barcode"])
        entries["barcode"]=bc_e

        # Validation label
        bc_err=tk.Label(left,text="",font=F_SMALL,bg=C_BG,fg="#ff4444",padx=16)
        bc_err.pack(anchor="w")

        def _validate_bc_live(*_):
            bc=bc_e.get()
            ok,msg=validate_barcode(bc) if bc else (True,"")
            if not ok:
                bc_err.config(text="⚠️ "+msg)
                bc_e.config(highlightbackground="#ff4444",highlightthickness=2,highlightcolor="#ff4444")
            else:
                bc_err.config(text="")
                bc_e.config(highlightbackground=C_BORDER,highlightthickness=1,highlightcolor=C_ACCENT)

        bc_e.bind("<KeyRelease>",_validate_bc_live)

        # ปุ่ม Generate + Auto checkbox
        auto_var=tk.BooleanVar(value=False)
        def _gen_barcode():
            new_bc=generate_barcode("RK")
            bc_e.config(state=tk.NORMAL)
            bc_e.delete(0,tk.END); bc_e.insert(0,new_bc)
            bc_err.config(text="✅ Barcode สร้างอัตโนมัติ",fg="#3fb950")

        def _toggle_auto():
            if auto_var.get():
                _gen_barcode()
                bc_e.config(state="readonly",fg=C_MUTED)
            else:
                bc_e.config(state=tk.NORMAL,fg=C_TEXT)

        tk.Button(bc_row,text="⚡ Auto",font=(FM.primary,10,"bold"),
                  bg=C_YELLOW,fg=C_BG,relief=tk.FLAT,cursor="hand2",
                  padx=8,pady=4,bd=0,command=_gen_barcode
                  ).pack(side=tk.LEFT,padx=(6,2))
        tk.Checkbutton(bc_row,text="อัตโนมัติ",variable=auto_var,
                       command=_toggle_auto,
                       bg=C_BG,fg=C_MUTED,selectcolor=C_SURFACE2,
                       activebackground=C_BG,font=(FM.primary,10)
                       ).pack(side=tk.LEFT)

        row("ชื่อสินค้า","name")
        combo_row("หมวดหมู่","category",cats)
        combo_row("หน่วยนับ","unit",units,"ชิ้น")

        # 📦 หน่วยขายหลายขนาด (Multi-Unit) — เก็บเป็น list, persist ตอน save
        try: unit_rows = unit_service.get_alt_units(DB_FILE, pid) if pid else []
        except Exception: unit_rows = []
        munit_btn = tk.Button(left,text="",font=(FM.primary,10,"bold"),bg=C_SURFACE2,fg=C_TEXT,
                              relief=tk.FLAT,cursor="hand2",anchor="w",padx=10,pady=6,bd=0)
        munit_btn.pack(fill=tk.X,padx=16,pady=(6,2))
        def _munit_lbl(): munit_btn.config(text=f"📦 หน่วยขายหลายขนาด ({len(unit_rows)})  ▸")
        def _manage_punits():
            pop=tk.Toplevel(win); pop.title("หน่วยขายหลายขนาด"); pop.configure(bg=C_BG)
            pop.grab_set(); pop.geometry("460x470")
            tk.Frame(pop,bg=C_ACCENT,height=3).pack(fill=tk.X)
            tk.Label(pop,text="📦 หน่วยขายหลายขนาด",font=F_H2,bg=C_BG,fg=C_TEXT,pady=8).pack(anchor="w",padx=14)
            base_u=entries["unit"].get() or "ชิ้น"
            tk.Label(pop,text=f"หน่วยฐาน: นับสต๊อกเป็น \"{base_u}\"  (factor = จำนวน {base_u}/หน่วย)",
                     font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(anchor="w",padx=14)
            lb=tk.Listbox(pop,font=F_BODY,bg=C_SURFACE2,fg=C_TEXT,selectbackground=C_ACCENT,
                          selectforeground=C_BG,relief=tk.FLAT,highlightthickness=0,height=8)
            lb.pack(fill=tk.BOTH,expand=True,padx=14,pady=(6,4))
            def _refresh_lb():
                lb.delete(0,tk.END)
                for u in unit_rows:
                    nm=u.get("unit_name") or u.get("unit"); f=int(u.get("factor",1) or 1)
                    lb.insert(tk.END,f"{nm}  =  {f} {base_u}    @ {float(u.get('price',0) or 0):,.2f} บ.")
            fr=tk.Frame(pop,bg=C_BG); fr.pack(fill=tk.X,padx=14,pady=(2,4))
            uc=tk.Frame(fr,bg=C_BG); uc.pack(side=tk.LEFT,padx=2)
            tk.Label(uc,text="หน่วย",font=(FM.primary,9),bg=C_BG,fg=C_MUTED).pack(anchor="w")
            uvar=tk.StringVar(); ttk.Combobox(uc,textvariable=uvar,values=units,font=F_BODY,width=8).pack()
            fc=tk.Frame(fr,bg=C_BG); fc.pack(side=tk.LEFT,padx=2)
            tk.Label(fc,text=f"{base_u}/หน่วย",font=(FM.primary,9),bg=C_BG,fg=C_MUTED).pack(anchor="w")
            fe=field(fc,width=7,justify="right"); fe.pack(ipady=3)
            pc=tk.Frame(fr,bg=C_BG); pc.pack(side=tk.LEFT,padx=2)
            tk.Label(pc,text="ราคา/หน่วย",font=(FM.primary,9),bg=C_BG,fg=C_MUTED).pack(anchor="w")
            pe=field(pc,width=8,justify="right"); pe.pack(ipady=3)
            def _add_row():
                nm=uvar.get().strip()
                try: f=int(float(fe.get() or 0))
                except Exception: f=0
                try: pr=float(pe.get() or 0)
                except Exception: pr=0.0
                if not nm or f<=1:
                    messagebox.showwarning("ไม่ถูกต้อง",f"กรอกหน่วย และจำนวน {base_u}/หน่วย (มากกว่า 1)",parent=pop); return
                unit_rows[:]=[u for u in unit_rows if (u.get("unit_name") or u.get("unit"))!=nm]
                unit_rows.append({"unit_name":nm,"factor":f,"price":pr})
                unit_rows.sort(key=lambda u:-int(u.get("factor",1) or 1))
                _refresh_lb(); _munit_lbl(); uvar.set(""); fe.delete(0,tk.END); pe.delete(0,tk.END)
            ac=tk.Frame(fr,bg=C_BG); ac.pack(side=tk.LEFT,padx=2,anchor="s")
            tk.Label(ac,text="",font=(FM.primary,9),bg=C_BG).pack()
            accent_btn(ac,"➕ เพิ่ม",_add_row,pad_x=10,pad_y=3).pack()
            def _del_row():
                s=lb.curselection()
                if s: del unit_rows[s[0]]; _refresh_lb(); _munit_lbl()
            bf=tk.Frame(pop,bg=C_BG); bf.pack(fill=tk.X,padx=14,pady=(2,10))
            tk.Button(bf,text="🗑 ลบที่เลือก",command=_del_row,font=(FM.primary,10),bg=C_SURFACE2,
                      fg="#ff6b6b",relief=tk.FLAT,bd=0,padx=10,pady=4,cursor="hand2").pack(side=tk.LEFT)
            accent_btn(bf,"ปิด",pop.destroy,pad_x=20,pad_y=4).pack(side=tk.RIGHT)
            _refresh_lb()
        munit_btn.config(command=_manage_punits); _munit_lbl()

        tk.Frame(left,bg=C_BORDER,height=1).pack(fill=tk.X,padx=16,pady=(10,4))
        tk.Label(left,text="ราคา",font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(anchor="w",padx=16)
        pg=tk.Frame(left,bg=C_BG); pg.pack(fill=tk.X,padx=16,pady=(2,4))
        for lbl,key in [("ต้นทุน","cost"),("ราคาขาย","price"),("A","price_a"),("B","price_b"),("C","price_c")]:
            col=tk.Frame(pg,bg=C_BG); col.pack(side=tk.LEFT,fill=tk.X,expand=True,padx=2)
            tk.Label(col,text=lbl,font=(FM.primary,10),bg=C_BG,fg=C_MUTED).pack(anchor="w")
            e=field(col,width=8,justify="right",font=(FM.primary,12,"bold"),fg=C_YELLOW)
            e.pack(fill=tk.X,ipady=5)
            val=str(p[key] if p and p[key] is not None else "0"); e.insert(0,val); entries[key]=e

        tk.Frame(left,bg=C_BORDER,height=1).pack(fill=tk.X,padx=16,pady=(8,4))
        sg=tk.Frame(left,bg=C_BG); sg.pack(fill=tk.X,padx=16,pady=(0,4))
        for lbl,key,def_ in [("สต็อก","stock","0"),("ขั้นต่ำ","min_stock","5")]:
            col=tk.Frame(sg,bg=C_BG); col.pack(side=tk.LEFT,fill=tk.X,expand=True,padx=2)
            tk.Label(col,text=lbl,font=(FM.primary,10),bg=C_BG,fg=C_MUTED).pack(anchor="w")
            e=field(col,width=12,justify="right"); e.pack(fill=tk.X,ipady=5)
            val=str(p[key] if p and p[key] is not None else def_); e.insert(0,val); entries[key]=e

        # ── 🎁 โปรโมชั่นสินค้า (เฟส A) ──
        def _pg(key,default=0):
            return (p[key] if (p and key in p.keys() and p[key] is not None) else default)
        tk.Label(right,text="🎁 โปรโมชั่นสินค้า",font=F_SMALL,bg=C_BG,fg=C_GREEN).pack(anchor="w",padx=16,pady=(6,2))
        promo_on=tk.BooleanVar(value=bool(_pg("promotion_enabled",0)))
        tk.Checkbutton(right,text="เปิดใช้งานโปรโมชั่น",variable=promo_on,bg=C_BG,fg=C_TEXT,
                       selectcolor=C_SURFACE2,activebackground=C_BG,font=F_BODY).pack(anchor="w",padx=16)
        promo_type=tk.StringVar(value=str(_pg("promotion_type","") or "percent"))
        ptf=tk.Frame(right,bg=C_BG); ptf.pack(anchor="w",padx=16,pady=(2,0))
        ptf2=tk.Frame(right,bg=C_BG); ptf2.pack(anchor="w",padx=16,pady=(0,2))
        for fr,opts in [(ptf,[("percent","ลด %"),("fixed","ลด ฿"),("special","ราคาพิเศษ")]),
                        (ptf2,[("buyx","ซื้อ X แถม X"),("buyy","ซื้อ X แถม Y"),("bulk","ซื้อครบลด")])]:
            for v,l in opts:
                tk.Radiobutton(fr,text=l,variable=promo_type,value=v,bg=C_BG,fg=C_TEXT,
                               selectcolor=C_SURFACE2,activebackground=C_BG,font=F_SMALL).pack(side=tk.LEFT,padx=4)
        pvf=tk.Frame(right,bg=C_BG); pvf.pack(fill=tk.X,padx=16,pady=(2,2))
        def _pcol(parent,lbl,initv,w=10):
            col=tk.Frame(parent,bg=C_BG); col.pack(side=tk.LEFT,fill=tk.X,expand=True,padx=2)
            lblw=tk.Label(col,text=lbl,font=(FM.primary,10),bg=C_BG,fg=C_MUTED); lblw.pack(anchor="w")
            ee=field(col,width=w,justify="right"); ee.insert(0,str(initv)); ee.pack(fill=tk.X,ipady=5)
            ee._col=col; ee._lbl=lblw
            return ee
        promo_value_e=_pcol(pvf,"ค่าโปร (%/฿)",_pg("promotion_value",0) or 0)
        promo_price_e=_pcol(pvf,"ราคาพิเศษ/ชิ้น",_pg("promotion_price",0) or 0)
        promo_prio_e =_pcol(pvf,"ลำดับ",_pg("promotion_priority",1) or 1,6)
        pvf2=tk.Frame(right,bg=C_BG); pvf2.pack(fill=tk.X,padx=16,pady=(0,2))
        promo_buy_e =_pcol(pvf2,"ซื้อ X (ชิ้น)",_pg("promotion_buy_qty",0) or 0,8)
        promo_free_e=_pcol(pvf2,"แถม Y (ชิ้น)",_pg("promotion_free_qty",0) or 0,8)
        pvf2_sp=tk.Frame(pvf2,bg=C_BG); pvf2_sp.pack(side=tk.LEFT,fill=tk.X,expand=True)

        # ── สินค้าแถม Y (เฉพาะ "ซื้อ X แถม Y") ──
        free_pid_var=tk.IntVar(value=int(_pg("promotion_free_product_id",0) or 0))
        pyf=tk.Frame(right,bg=C_BG)   # โชว์เฉพาะ buyy (toggle)
        tk.Label(pyf,text="สินค้าแถม (Y):",font=(FM.primary,10),bg=C_BG,fg=C_GREEN).pack(anchor="w")
        pyf_row=tk.Frame(pyf,bg=C_BG); pyf_row.pack(fill=tk.X)
        free_disp=field(pyf_row,width=24); free_disp.pack(side=tk.LEFT,fill=tk.X,expand=True,ipady=4)
        free_disp.config(state="readonly")
        def _set_free_disp(pidv):
            txt=""
            if pidv:
                try:
                    fp=get_product(pidv)
                    if fp: txt=f"{fp.get('barcode','') or '-'}  {fp.get('name','')}"
                except Exception: pass
            free_disp.config(state=tk.NORMAL); free_disp.delete(0,tk.END)
            free_disp.insert(0,txt); free_disp.config(state="readonly")
        def _open_free_picker():
            pk=tk.Toplevel(win); pk.title("เลือกสินค้าแถม"); pk.configure(bg=C_BG); pk.grab_set(); pk.geometry("480x480")
            tk.Label(pk,text="เลือกสินค้าแถม (Y)",font=F_H2,bg=C_BG,fg=C_TEXT,pady=8).pack(anchor="w",padx=12)
            se=field(pk,width=44); se.pack(fill=tk.X,padx=12,ipady=5); se.focus_set()
            lb=tk.Listbox(pk,font=F_BODY,bg=C_SURFACE2,fg=C_TEXT,selectbackground=C_ACCENT,
                          selectforeground=C_BG,relief=tk.FLAT,highlightthickness=0,height=15)
            lb.pack(fill=tk.BOTH,expand=True,padx=12,pady=8)
            rows=[]
            def refresh(*_):
                kw=se.get().strip()
                try: res=search_products(kw) if kw else get_products_for_sale_list()
                except Exception: res=[]
                rows[:]=res[:300]; lb.delete(0,tk.END)
                for r in rows: lb.insert(tk.END,f"{r.get('barcode','') or '-'}   {r.get('name','')}")
            def choose(*_):
                sel=lb.curselection()
                if not sel: return
                r=rows[sel[0]]; free_pid_var.set(r["id"]); _set_free_disp(r["id"])
                _update_promo_preview(); pk.destroy()
            se.bind("<KeyRelease>",refresh); lb.bind("<Double-1>",choose)
            accent_btn(pk,"✔ เลือก",choose,pad_x=20,pad_y=6).pack(pady=(0,10))
            refresh()
        tk.Button(pyf_row,text="...",font=(FM.primary,11,"bold"),bg=C_BLUE,fg=C_WHITE,
                  relief=tk.FLAT,cursor="hand2",padx=12,bd=0,command=_open_free_picker).pack(side=tk.LEFT,padx=(4,0))
        _set_free_disp(free_pid_var.get())

        phelp=tk.Label(right,text="",font=(FM.primary,9),bg=C_BG,fg=C_MUTED,justify=tk.LEFT,wraplength=300)
        phelp.pack(anchor="w",padx=16)
        def _toggle_promo_fields(*_):
            t=promo_type.get()
            # rule ต่อชนิด: field ที่ต้องแสดง + label (priority แสดงทุกชนิด)
            RULES={
                "percent":{"value":"ส่วนลด (%)"},
                "fixed":  {"value":"ส่วนลด (฿/ชิ้น)"},
                "special":{"price":"ราคาพิเศษ/ชิ้น"},
                "buyx":   {"buy":"ซื้อ X (ชิ้น)","free":"แถมเพิ่ม (ชิ้น)"},
                "buyy":   {"buy":"ซื้อ X (ชิ้น)","free":"แถม Y (ชิ้น)","picker":1},
                "bulk":   {"buy":"ซื้อครบ X (ชิ้น)","price":"ราคาพิเศษ/ชิ้น"},
            }
            rule=RULES.get(t,{})
            def _clear(e):
                try: e.delete(0,tk.END); e.insert(0,"0")
                except Exception: pass
            # 1) ซ่อนทุก field ก่อน
            for e in (promo_value_e,promo_price_e,promo_prio_e,promo_buy_e,promo_free_e):
                e._col.pack_forget()
            pvf2_sp.pack_forget(); pyf.pack_forget()
            # 2) relabel + เคลียร์ค่าที่ไม่ใช้
            if "value" in rule: promo_value_e._lbl.config(text=rule["value"])
            else: _clear(promo_value_e)
            if "price" in rule: promo_price_e._lbl.config(text=rule["price"])
            else: _clear(promo_price_e)
            if "buy" in rule: promo_buy_e._lbl.config(text=rule["buy"])
            else: _clear(promo_buy_e)
            if "free" in rule: promo_free_e._lbl.config(text=rule["free"])
            else: _clear(promo_free_e)
            if "picker" not in rule:
                free_pid_var.set(0); _set_free_disp(0)
            def _seed(e):
                v=(e.get() or "").strip()
                if v in ("","0","0.0"):
                    e.delete(0,tk.END); e.insert(0,"1")
            if "buy" in rule: _seed(promo_buy_e)
            if "free" in rule: _seed(promo_free_e)
            # 3) re-pack เฉพาะที่เกี่ยว (ลำดับคงที่: value→price→ลำดับ | ซื้อX→แถมY)
            if "value" in rule: promo_value_e._col.pack(side=tk.LEFT,fill=tk.X,expand=True,padx=2)
            if "price" in rule: promo_price_e._col.pack(side=tk.LEFT,fill=tk.X,expand=True,padx=2)
            promo_prio_e._col.pack(side=tk.LEFT,fill=tk.X,expand=True,padx=2)
            if "buy" in rule: promo_buy_e._col.pack(side=tk.LEFT,fill=tk.X,expand=True,padx=2)
            if "free" in rule: promo_free_e._col.pack(side=tk.LEFT,fill=tk.X,expand=True,padx=2)
            pvf2_sp.pack(side=tk.LEFT,fill=tk.X,expand=True)
            if "picker" in rule: pyf.pack(fill=tk.X,padx=16,pady=(2,2),before=phelp)
            # 4) helper text
            phelp.config(text="• "+{
                "percent":"ลดเปอร์เซ็นต์ทุกชิ้น",
                "fixed":"ลดเป็นบาท/ชิ้น ทุกชิ้น",
                "special":"ขายราคาพิเศษต่อชิ้น (กรอกช่อง ราคาพิเศษ/ชิ้น)",
                "buyx":"ซื้อสินค้าชนิดเดียวกัน แล้วแถมตัวเดิม\nระบบตัดสต๊อกรวมอัตโนมัติ",
                "buyy":"ซื้อสินค้า X แล้วแถมสินค้าอีกชนิด (เลือกสินค้าแถมด้านบน)",
                "bulk":"ซื้อครบ X ชิ้น เหลือชิ้นละ (ราคาพิเศษ/ชิ้น)",
            }.get(t,""))
            # 5) refresh preview ให้ตรง type ปัจจุบัน
            _update_promo_preview()
        pdf=tk.Frame(right,bg=C_BG); pdf.pack(fill=tk.X,padx=16,pady=(4,2))
        tk.Label(pdf,text="เริ่ม:",font=(FM.primary,10),bg=C_BG,fg=C_MUTED).pack(side=tk.LEFT)
        promo_start_tde=ThaiDateEntry(pdf,width=12)
        if _pg("promotion_start",""): promo_start_tde.set(_pg("promotion_start",""))
        promo_start_tde.pack(side=tk.LEFT,padx=(2,8))
        tk.Label(pdf,text="สิ้นสุด:",font=(FM.primary,10),bg=C_BG,fg=C_MUTED).pack(side=tk.LEFT)
        promo_end_tde=ThaiDateEntry(pdf,width=12)
        if _pg("promotion_end",""): promo_end_tde.set(_pg("promotion_end",""))
        promo_end_tde.pack(side=tk.LEFT,padx=2)

        # ── 🔎 พรีวิวโปรโมชั่น realtime ──
        prev_box=tk.Frame(right,bg=C_SURFACE2,highlightbackground=C_GREEN,highlightthickness=1)
        prev_box.pack(fill=tk.X,padx=16,pady=(8,2))
        tk.Label(prev_box,text="🔎 ผลโปรโมชั่น",font=(FM.primary,10,"bold"),
                 bg=C_SURFACE2,fg=C_GREEN).pack(anchor="w",padx=10,pady=(6,0))
        promo_prev=tk.Label(prev_box,text="",font=F_SMALL,bg=C_SURFACE2,fg=C_TEXT,
                            justify=tk.LEFT,anchor="w")
        promo_prev.pack(fill=tk.X,padx=10,pady=(2,8))

        def _update_promo_preview(*_):
            try: price=float((entries["price"].get() or "0").replace(",",""))
            except Exception: price=0.0
            def _f(e):
                try: return float((e.get() or "0").replace(",",""))
                except Exception: return 0.0
            def _i(e):
                try: return int(_f(e))
                except Exception: return 0
            t=promo_type.get()
            fake={"promotion_enabled":1,"promotion_type":t,
                  "promotion_value":_f(promo_value_e),"promotion_price":_f(promo_price_e),
                  "promotion_buy_qty":_i(promo_buy_e),"promotion_free_qty":_i(promo_free_e),
                  "promotion_start":"","promotion_end":"","promotion_days":"",
                  "promotion_time_start":"","promotion_time_end":"","promotion_member_only":0}
            if not promo_on.get():
                lines=["(ปิดโปรโมชั่น — ขายราคาปกติ)"]
            elif price<=0:
                lines=["กรอกราคาขายก่อน เพื่อดูผลโปรโมชั่น"]
            elif t in ("percent","fixed","special"):
                ud,lbl=promotion_engine.unit_discount(fake,price)
                X=_i(promo_buy_e); exq=X if X>0 else 1
                head=f"ซื้อ {X} ชิ้นขึ้นไป {lbl}" if X>0 else (lbl or "—")
                lines=[head,
                       f"ราคาปกติ: {price:,.2f} บาท",
                       f"ราคาหลังลด: {price-ud:,.2f} บาท",
                       f"ส่วนลดต่อชิ้น: {ud:,.2f} บาท",
                       f"ส่วนลดรวม{(' ('+str(X)+' ชิ้น)') if X>0 else ''}: {ud*exq:,.2f} บาท"]
            elif t=="buyx":
                X=_i(promo_buy_e); Y=_i(promo_free_e)
                if X>0 and Y>0 and price>0:
                    avg=(X*price)/(X+Y)
                    lines=[f"ซื้อ {X} แถม {Y}",
                           f"ลูกค้าได้รับรวม {X+Y} ชิ้น",
                           f"เฉลี่ยราคาต่อชิ้น: {avg:,.2f} บาท",
                           f"ระบบตัดสต๊อกรวม {X+Y} ชิ้น"]
                else:
                    lines=["กรอกจำนวน ซื้อ X และ แถมเพิ่ม"]
            elif t=="buyy":
                X=_i(promo_buy_e); Y=_i(promo_free_e); fpid=free_pid_var.get()
                fname=""
                if fpid:
                    try:
                        _fp=get_product(fpid); fname=_fp.get("name","") if _fp else ""
                    except Exception: fname=""
                if X>0 and Y>0 and fpid:
                    lines=[f"ซื้อสินค้านี้ {X} ชิ้น",
                           f"แถม: {fname or '(สินค้า Y)'} {Y} ชิ้น",
                           "ระบบจะตัดสต๊อก:",
                           f"   • สินค้านี้ {X}",
                           f"   • {fname or 'Y'} {Y}",
                           "(แถมจริงตอนขาย = Part 2)"]
                else:
                    lines=["กรอก ซื้อ X, แถม Y และเลือกสินค้าแถม (Y)"]
            elif t=="bulk":
                X=_i(promo_buy_e); sp=_f(promo_price_e)
                if X>0 and 0<sp<price:
                    d=price-sp
                    lines=[f"ซื้อครบ {X} ชิ้น เหลือชิ้นละ {sp:,.2f}",
                           f"ราคาปกติ: {price:,.2f} บาท",
                           f"ราคาหลังลด: {sp:,.2f} บาท",
                           f"ส่วนลดต่อชิ้น: {d:,.2f} บาท",
                           f"ส่วนลดรวม ({X} ชิ้น): {d*X:,.2f} บาท"]
                else:
                    lines=["กรอก ซื้อครบ X (ชิ้น) + ราคาพิเศษ/ชิ้น (น้อยกว่าราคาปกติ)"]
            else:
                lines=["—"]
            promo_prev.config(text="\n".join(lines))

        promo_on.trace_add("write",_update_promo_preview)
        promo_type.trace_add("write",_toggle_promo_fields)
        for _e in (promo_value_e,promo_price_e,promo_buy_e,promo_free_e,entries["price"]):
            _e.bind("<KeyRelease>",_update_promo_preview)
        _toggle_promo_fields()

        # ── ผูก validation + auto-format ทุกช่อง (หลัง prefill) ──
        try:
            _attach_len_validator(entries["barcode"], win, 50)
            if "name" in entries: _attach_len_validator(entries["name"], win, 255)
            for _k in ("cost","price","price_a","price_b","price_c"):
                if _k in entries:
                    _attach_num_validator(entries[_k], win, 999999.99, 2)
                    _attach_money_format(entries[_k])
            if "stock" in entries:
                _attach_num_validator(entries["stock"], win, 999999, 0)
                _attach_int_format(entries["stock"])
            if "min_stock" in entries:
                _attach_num_validator(entries["min_stock"], win, 99999, 0)
                _attach_int_format(entries["min_stock"])
            for _pe in (promo_value_e, promo_price_e):
                _attach_num_validator(_pe, win, 999999.99, 2)
            for _pe in (promo_buy_e, promo_free_e, promo_prio_e):
                _attach_num_validator(_pe, win, 99999, 0)
        except Exception as _ve:
            print(f"[form validation] {_ve}")

        # ── Save with full validation ──────────────────────
        def save():
            try:
                d={k:v.get().strip() if hasattr(v,"get") else str(v)
                   for k,v in entries.items()}
                # Validate name
                if not d["name"]: entries["name"].focus_set(); raise ValueError("กรุณากรอกชื่อสินค้า")
                if len(d["name"]) > 255: entries["name"].focus_set(); raise ValueError("ชื่อสินค้ายาวเกิน 255 ตัวอักษร")
                # Validate barcode
                bc=d["barcode"]
                if len(bc) > 50: entries["barcode"].focus_set(); raise ValueError("บาร์โค้ดยาวเกิน 50 ตัวอักษร")
                if bc:  # barcode optional แต่ถ้ากรอกต้องถูกต้อง
                    ok,msg=validate_barcode(bc)
                    if not ok: entries["barcode"].focus_set(); raise ValueError(msg)
                    # Check duplicate
                    if check_barcode_duplicate(bc, exclude_id=pid):
                        entries["barcode"].focus_set(); raise ValueError(f"Barcode '{bc}' มีอยู่ในระบบแล้ว")
                # แปลงตัวเลข (strip comma) + ตรวจช่วงค่า
                MAXP = 999999.99
                for _k,_lbl in [("price","ราคาขาย"),("cost","ต้นทุน"),
                                ("price_a","ราคา A"),("price_b","ราคา B"),("price_c","ราคา C")]:
                    _v = _safe_float(d.get(_k,0))
                    if _v < 0 or _v > MAXP:
                        if _k in entries: entries[_k].focus_set()
                        raise ValueError(f"{_lbl} ต้องอยู่ระหว่าง 0 ถึง 999,999.99")
                    d[_k]=_v
                d["stock"]=_safe_int(d.get("stock",0))
                if not (0 <= d["stock"] <= 999999):
                    entries["stock"].focus_set(); raise ValueError("สต็อกต้องอยู่ระหว่าง 0 ถึง 999,999")
                d["min_stock"]=_safe_int(d.get("min_stock",5))
                if not (0 <= d["min_stock"] <= 99999):
                    entries["min_stock"].focus_set(); raise ValueError("ขั้นต่ำต้องอยู่ระหว่าง 0 ถึง 99,999")
                d["unit"]=d.get("unit","") or "ชิ้น"
            except ValueError as e:
                messagebox.showerror("ผิดพลาด",str(e),parent=win); return
            conn2=get_db()
            if pid:
                conn2.execute(
                    "UPDATE products SET barcode=?,name=?,category=?,unit=?,"
                    "price=?,cost=?,price_a=?,price_b=?,price_c=?,stock=?,min_stock=? WHERE id=?",
                    (d["barcode"],d["name"],d["category"],d["unit"],
                     d["price"],d["cost"],d["price_a"],d["price_b"],d["price_c"],
                     d["stock"],d["min_stock"],pid))
            else:
                conn2.execute(
                    "INSERT INTO products (barcode,name,category,unit,price,cost,"
                    "price_a,price_b,price_c,stock,min_stock) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                    (d["barcode"],d["name"],d["category"],d["unit"],
                     d["price"],d["cost"],d["price_a"],d["price_b"],d["price_c"],
                     d["stock"],d["min_stock"]))
            _id = pid or conn2.execute("SELECT last_insert_rowid()").fetchone()[0]
            # ── บันทึกโปรโมชั่นสินค้า (เฟส A) ──
            try:
                _ps = promo_start_tde.get_iso() if promo_start_tde.get().strip() else ""
                _pe = promo_end_tde.get_iso() if promo_end_tde.get().strip() else ""
                _pval = _safe_float(promo_value_e.get())
                if promo_type.get() == "percent" and _pval > 100: _pval = 100.0   # cap ลด % ที่ 100
                conn2.execute(
                    "UPDATE products SET promotion_enabled=?,promotion_type=?,promotion_value=?,"
                    "promotion_price=?,promotion_start=?,promotion_end=?,promotion_priority=?,"
                    "promotion_buy_qty=?,promotion_free_qty=?,promotion_free_product_id=? WHERE id=?",
                    (1 if promo_on.get() else 0, promo_type.get(),
                     _pval, _safe_float(promo_price_e.get()),
                     _ps, _pe, _safe_int(promo_prio_e.get(),1),
                     _safe_int(promo_buy_e.get()), _safe_int(promo_free_e.get()),
                     _safe_int(free_pid_var.get()), _id))
            except Exception as _pe2:
                print("[promo save]", _pe2)
            conn2.commit(); conn2.close()
            # ── บันทึกหน่วยขายหลายขนาด (เฟส4) — หลัง commit กัน DB lock ──
            try:
                unit_service.set_units(DB_FILE, _id, unit_rows)
            except Exception as _ue:
                print("[unit save]", _ue)
            win.destroy(); self.load()

        accent_btn(win,"💾  บันทึก",save,pad_x=24,pad_y=10).pack(fill=tk.X,padx=16,pady=12)
        win.bind("<Return>",lambda e:save())
        win.update_idletasks()
        win.minsize(win.winfo_reqwidth(), win.winfo_reqheight())

    def _add(self):
        if not has_permission("product_add"):
            require_permission("product_add",self.winfo_toplevel()); return
        self._form()
    def _edit(self):
        if not has_permission("product_edit"):
            require_permission("product_edit",self.winfo_toplevel()); return
        pid=self._sel()
        if pid: self._form(pid)
    def _unit_choices(self, pid):
        """คืน (display_names, map) สำหรับเลือกหน่วยรับเข้า — หน่วยฐานก่อน + หน่วยใหญ่
           map: display -> (unit_name, factor)"""
        try:
            prod=get_product(pid); base_unit=((prod["unit"] if prod else "ชิ้น") or "ชิ้น")
        except Exception: base_unit="ชิ้น"
        try: us=unit_service.get_units(DB_FILE, pid, base_unit=base_unit, base_price=0)
        except Exception: us=[{"unit":base_unit,"factor":1,"is_base":True}]
        ordered=[u for u in us if u.get("is_base")]+[u for u in us if not u.get("is_base")]
        names=[]; mp={}
        for u in ordered:
            disp=u["unit"] if u.get("is_base") else f'{u["unit"]} (×{u["factor"]})'
            names.append(disp); mp[disp]=(u["unit"], int(u["factor"] or 1))
        return names, mp

    def _restock(self):
        pid=self._sel()
        if not pid: return
        p=get_product_name_stock(pid)  # Session 6: swap → db_product
        ci=stock_service.get_cost_info(pid, DB_FILE) or {}
        win=tk.Toplevel(self); win.title("\u0e40\u0e15\u0e34\u0e21\u0e2a\u0e15\u0e47\u0e2d\u0e01"); win.configure(bg=C_BG); win.grab_set(); win.resizable(False,False)
        tk.Frame(win,bg=C_BLUE,height=3).pack(fill=tk.X)
        tk.Label(win,text=f"\u0e40\u0e15\u0e34\u0e21\u0e2a\u0e15\u0e47\u0e2d\u0e01: {p['name']}",font=F_H2,bg=C_BG,fg=C_TEXT,pady=10).pack(anchor="w",padx=16)
        tk.Label(win,text=f"\u0e2a\u0e15\u0e47\u0e2d\u0e01\u0e1b\u0e31\u0e08\u0e08\u0e38\u0e1a\u0e31\u0e19: {p['stock']}  |  \u0e15\u0e49\u0e19\u0e17\u0e38\u0e19\u0e40\u0e09\u0e25\u0e35\u0e48\u0e22: \u0e3f{ci.get('avg_cost',0):,.2f}",font=F_BODY,bg=C_BG,fg=C_MUTED).pack(anchor="w",padx=16)
        tk.Label(win,text="\u0e08\u0e33\u0e19\u0e27\u0e19\u0e17\u0e35\u0e48\u0e40\u0e15\u0e34\u0e21:",font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(anchor="w",padx=16,pady=(8,2))
        e=field(win,width=20,justify="right",font=(FM.primary,19,"bold"),fg=C_BLUE)
        e.insert(0,"0"); e.pack(fill=tk.X,padx=16,ipady=8)
        tk.Label(win,text="\u0e15\u0e49\u0e19\u0e17\u0e38\u0e19/\u0e2b\u0e19\u0e48\u0e27\u0e22 (\u0e40\u0e27\u0e49\u0e19\u0e27\u0e48\u0e32\u0e07 = \u0e43\u0e0a\u0e49\u0e15\u0e49\u0e19\u0e17\u0e38\u0e19\u0e40\u0e14\u0e34\u0e21):",font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(anchor="w",padx=16,pady=(8,2))
        ec=field(win,width=20,justify="right",font=(FM.primary,15),fg=C_TEXT)
        ec.pack(fill=tk.X,padx=16,ipady=6)
        _unames,_umap=self._unit_choices(pid); _uvar=tk.StringVar(value=_unames[0])
        if len(_unames)>1:
            tk.Label(win,text="หน่วยที่รับ:",font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(anchor="w",padx=16,pady=(8,2))
            ttk.Combobox(win,textvariable=_uvar,values=_unames,state="readonly",font=F_BODY).pack(fill=tk.X,padx=16,ipady=2)
        # validation: จำนวน (int ≤999,999), ต้นทุน (money ≤999,999.99)
        try:
            _attach_num_validator(e, win, 999999, 0)
            _attach_num_validator(ec, win, 999999.99, 2); _attach_money_format(ec)
        except Exception as _ve: print(f"[restock validation] {_ve}")
        def do():
            try: qty=_safe_int(e.get())
            except: messagebox.showerror("\u0e1c\u0e34\u0e14\u0e1e\u0e25\u0e32\u0e14","\u0e01\u0e23\u0e38\u0e13\u0e32\u0e01\u0e23\u0e2d\u0e01\u0e15\u0e31\u0e27\u0e40\u0e25\u0e02",parent=win); return
            if qty<=0: messagebox.showwarning("แจ้งเตือน","กรุณากรอกจำนวนมากกว่า 0",parent=win); return
            _un,_fac=_umap.get(_uvar.get(),("",1)); base_qty=qty*_fac
            cost_str=ec.get().strip().replace(",","")
            if cost_str:
                try: ucost=float(cost_str)
                except: messagebox.showerror("ผิดพลาด","ต้นทุนไม่ถูกต้อง",parent=win); return
                res=stock_service.receive_stock(pid, base_qty, (ucost/_fac if _fac else ucost), DB_FILE,
                        staff_id=current_staff.get("id",0), staff_name=current_staff.get("name",""),
                        note=(f"รับ {qty} {_un}" if _fac>1 else ""))
            else:
                res=stock_service.quick_add_stock(pid, base_qty, DB_FILE,
                        staff_id=current_staff.get("id",0), staff_name=current_staff.get("name",""),
                        note=(f"เติม {qty} {_un}" if _fac>1 else ""))
            if not res.get("success"):
                messagebox.showerror("ผิดพลาด",res.get("error","รับสินค้าไม่สำเร็จ"),parent=win); return
            if cost_str:
                messagebox.showinfo("สำเร็จ",
                    f"รับเข้า +{qty} {_un}" + (f" (= {base_qty} ชิ้น)" if _fac>1 else "") +
                    f"\nต้นทุนเฉลี่ย: ฿{res['old_avg']:,.2f} → ฿{res['new_avg']:,.2f}",parent=win)
            win.destroy(); self.load()
        accent_btn(win,"\U0001f4be  \u0e1a\u0e31\u0e19\u0e17\u0e36\u0e01",do,pad_x=24,pad_y=10).pack(pady=12)
        win.bind("<Return>",lambda _:do()); e.focus_set(); e.select_range(0,tk.END)

    # ── เฟส 2: รับสินค้าแบบละเอียด (supplier/invoice/VAT/วันที่) ──
    def _receive_advanced(self):
        pid=self._sel()
        if not pid: return
        ci=stock_service.get_cost_info(pid, DB_FILE) or {}
        # ดึงรายชื่อผู้จำหน่าย
        sup_map={"— ไม่ระบุ —":(0,"")}
        try:
            conn=get_db()
            for s in conn.execute("SELECT id,name FROM suppliers WHERE active=1 ORDER BY name"):
                sup_map[s["name"]]=(s["id"],s["name"])
            conn.close()
        except Exception: pass

        win=tk.Toplevel(self); win.title("รับสินค้าเข้า (ละเอียด)"); win.configure(bg=C_BG)
        win.grab_set(); win.resizable(False,False)
        tk.Frame(win,bg=C_BLUE,height=3).pack(fill=tk.X)
        tk.Label(win,text=f"📥 รับสินค้า: {ci.get('name','')}",font=F_H2,bg=C_BG,fg=C_TEXT,pady=8).pack(anchor="w",padx=16)
        tk.Label(win,text=f"สต็อกปัจจุบัน: {ci.get('stock',0)}  |  ต้นทุนเฉลี่ย: ฿{ci.get('avg_cost',0):,.2f}",
                 font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(anchor="w",padx=16,pady=(0,6))
        body=tk.Frame(win,bg=C_BG,padx=16); body.pack(fill=tk.X)
        def row(lbl):
            tk.Label(body,text=lbl,font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(anchor="w",pady=(6,2))
        row("จำนวนที่รับ:")
        e_qty=field(body,width=24,justify="right",font=(FM.primary,17,"bold"),fg=C_BLUE); e_qty.insert(0,"0"); e_qty.pack(fill=tk.X,ipady=6)
        _unames,_umap=self._unit_choices(pid); _uvar=tk.StringVar(value=_unames[0])
        if len(_unames)>1:
            row("หน่วยที่รับ:")
            ttk.Combobox(body,textvariable=_uvar,values=_unames,state="readonly",font=F_BODY).pack(fill=tk.X,ipady=2)
        row("ต้นทุน/หน่วย (ก่อน VAT):")
        e_cost=field(body,width=24,justify="right",font=(FM.primary,15),fg=C_TEXT)
        e_cost.insert(0,f"{ci.get('last_cost',0):.2f}"); e_cost.pack(fill=tk.X,ipady=6)
        row("ผู้จำหน่าย:")
        cb_sup=ttk.Combobox(body,values=list(sup_map.keys()),state="readonly",font=F_BODY)
        cb_sup.current(0); cb_sup.pack(fill=tk.X,ipady=2)
        row("เลขเอกสาร/บิลซื้อ:")
        e_inv=field(body,width=24,justify="left",font=F_BODY,fg=C_TEXT); e_inv.pack(fill=tk.X,ipady=6)
        row("VAT % (เว้น 0 = ไม่มี):")
        e_vat=field(body,width=24,justify="right",font=F_BODY,fg=C_TEXT); e_vat.insert(0,"0"); e_vat.pack(fill=tk.X,ipady=6)
        row("วันที่ซื้อ:")
        e_date=field(body,width=24,justify="left",font=F_BODY,fg=C_TEXT)
        e_date.insert(0,datetime.now().strftime("%Y-%m-%d")); e_date.pack(fill=tk.X,ipady=6)
        row("หมายเหตุ:")
        e_note=field(body,width=24,justify="left",font=F_BODY,fg=C_TEXT); e_note.pack(fill=tk.X,ipady=6)

        # validation: จำนวน/ต้นทุน/VAT/เลขเอกสาร/หมายเหตุ
        try:
            _attach_num_validator(e_qty, win, 999999, 0)
            _attach_num_validator(e_cost, win, 999999.99, 2); _attach_money_format(e_cost)
            _attach_num_validator(e_vat, win, 100, 2)
            _attach_len_validator(e_inv, win, 50)
            _attach_len_validator(e_note, win, 255)
        except Exception as _ve: print(f"[receive validation] {_ve}")

        def do():
            try: qty=_safe_int(e_qty.get())
            except: messagebox.showerror("ผิดพลาด","จำนวนไม่ถูกต้อง",parent=win); return
            if qty<=0: messagebox.showwarning("แจ้งเตือน","จำนวนต้องมากกว่า 0",parent=win); return
            try: ucost=_safe_float(e_cost.get())
            except: messagebox.showerror("ผิดพลาด","ต้นทุนไม่ถูกต้อง",parent=win); return
            try: vat=_safe_float(e_vat.get())
            except: messagebox.showerror("ผิดพลาด","VAT ไม่ถูกต้อง",parent=win); return
            sid,sname=sup_map.get(cb_sup.get(),(0,""))
            d=e_date.get().strip()
            pdate=f"{d} {datetime.now().strftime('%H:%M:%S')}" if d else None
            _un,_fac=_umap.get(_uvar.get(),("",1)); base_qty=qty*_fac
            _note=e_note.get().strip()
            if _fac>1: _note=(f"รับ {qty} {_un}" + (f" | {_note}" if _note else ""))
            res=stock_service.receive_advanced(pid, base_qty, (ucost/_fac if _fac else ucost), DB_FILE,
                    supplier_id=sid, supplier_name=sname, invoice_no=e_inv.get().strip(),
                    vat_rate=vat, purchase_date=pdate, note=_note,
                    staff_id=current_staff.get("id",0), staff_name=current_staff.get("name",""))
            if not res.get("success"):
                messagebox.showerror("ผิดพลาด",res.get("error","รับสินค้าไม่สำเร็จ"),parent=win); return
            messagebox.showinfo("สำเร็จ",
                f"รับเข้า +{qty} {_un}" + (f" (= {base_qty} ชิ้น)" if _fac>1 else "") +
                f"\nต้นทุนเฉลี่ย: ฿{res['old_avg']:,.2f} → ฿{res['new_avg']:,.2f}\n"
                f"ยอดเอกสาร (รวม VAT): ฿{res.get('total',0):,.2f}",parent=win)
            win.destroy(); self.load()
        accent_btn(win,"💾  บันทึกรับสินค้า",do,pad_x=24,pad_y=10).pack(pady=12)
        e_qty.focus_set(); e_qty.select_range(0,tk.END)

    # ── เฟส 2: ดูประวัติการซื้อของสินค้า ──
    def _purchase_history(self):
        pid=self._sel()
        if not pid: return
        ci=stock_service.get_cost_info(pid, DB_FILE) or {}
        rows=stock_service.get_purchase_history(pid, DB_FILE, limit=200)
        win=tk.Toplevel(self); win.title("ประวัติการซื้อ"); win.configure(bg=C_BG); win.geometry("720x460")
        tk.Frame(win,bg=C_BLUE,height=3).pack(fill=tk.X)
        tk.Label(win,text=f"📜 ประวัติซื้อ: {ci.get('name','')}",font=F_H2,bg=C_BG,fg=C_TEXT,pady=8).pack(anchor="w",padx=16)
        if not rows:
            tk.Label(win,text="ยังไม่มีประวัติการซื้อสินค้านี้",font=F_BODY,bg=C_BG,fg=C_MUTED,pady=30).pack()
            return
        cols=("date","sup","inv","qty","cost","vat","total")
        heads={"date":"วันที่","sup":"ผู้จำหน่าย","inv":"เลขเอกสาร","qty":"จำนวน",
               "cost":"ต้นทุน/หน่วย","vat":"VAT","total":"ยอดรวม"}
        wid={"date":120,"sup":160,"inv":110,"qty":60,"cost":90,"vat":70,"total":100}
        tvf=tk.Frame(win,bg=C_BG); tvf.pack(fill=tk.BOTH,expand=True,padx=16,pady=8)
        tv=ttk.Treeview(tvf,columns=cols,show="headings",height=14)
        for c in cols:
            tv.heading(c,text=heads[c])
            tv.column(c,width=wid[c],anchor=("e" if c in("qty","cost","vat","total") else "w"))
        for r in rows:
            tv.insert("","end",values=(
                (r["purchase_date"] or "")[:16], r["supplier_name"] or "—",
                r["invoice_no"] or "—", r["qty"], f"฿{r['unit_cost']:,.2f}",
                (f"{r['vat_rate']:.0f}%" if r["vat_rate"] else "—"), f"฿{r['total']:,.2f}"))
        tv.pack(fill=tk.BOTH,expand=True)

    def _delete(self):
        pid = self._sel()
        if not pid:
            return

        conn = get_db()
        used = conn.execute(
            "SELECT 1 FROM sale_items WHERE product_id=? LIMIT 1", (pid,)
        ).fetchone()

        if used:
            if messagebox.askyesno("ยืนยัน", "สินค้านี้มีประวัติการขาย\nจะเปลี่ยนสถานะเป็น 'ยกเลิก' แทนการลบ?"):
                conn.execute(
                    "UPDATE products SET status='discontinued' WHERE id=?", (pid,)
                )
                conn.commit()
        else:
            if messagebox.askyesno("ยืนยัน", "ลบสินค้านี้?"):
                conn.execute("DELETE FROM products WHERE id=?", (pid,))
                conn.commit()

        conn.close()
        self.load()



# ══ TAB: SUPPLIER MASTER ══════════════════════════════════════════════════════
class SupplierTab(tk.Frame):
    def __init__(self,parent):
        super().__init__(parent,bg=C_BG); self._build()
    def _build(self):
        tk.Frame(self,bg=C_BLUE,height=3).pack(fill=tk.X)
        hf=tk.Frame(self,bg=C_BG,padx=16,pady=10); hf.pack(fill=tk.X)
        tk.Label(hf,text="🏭  ผู้จำหน่าย",font=F_H1,bg=C_BG,fg=C_TEXT).pack(side=tk.LEFT)
        sf=tk.Frame(self,bg=C_BG,padx=12); sf.pack(fill=tk.X,pady=(0,4))
        self.sv=tk.StringVar(); self.sv.trace_add("write",lambda *a:self.load())
        fe=field(sf,width=30); fe.configure(textvariable=self.sv); fe.pack(side=tk.LEFT,ipady=5)
        tk.Label(sf,text="  🔍",font=(FM.primary,14),bg=C_BG,fg=C_MUTED).pack(side=tk.LEFT)
        accent_btn(sf,"+ เพิ่มผู้จำหน่าย",self._add,pad_x=14,pad_y=6).pack(side=tk.RIGHT)
        cols=("รหัส","ชื่อบริษัท/ร้าน","ผู้ติดต่อ","โทรศัพท์","เครดิต(วัน)","เลขผู้เสียภาษี")
        frm,self.tree=make_tree(self,cols,(80,220,130,110,80,130))
        frm.pack(fill=tk.BOTH,expand=True,padx=12,pady=(0,4))
        self.tree.bind("<Double-1>",lambda e:self._edit())
        self.count_lbl=tk.Label(self,text="",font=F_SMALL,bg=C_BG,fg=C_MUTED)
        self.count_lbl.pack(anchor="e",padx=14,pady=(0,6))
    def load(self,*_):
        q=self.sv.get().strip()
        conn=get_db()
        rows=conn.execute(
            "SELECT * FROM suppliers WHERE active=1 AND "
            "(name LIKE ? OR code LIKE ? OR phone LIKE ? OR contact LIKE ?) ORDER BY name",
            (f"%{q}%",)*4).fetchall()
        conn.close()
        self.tree.delete(*self.tree.get_children())
        for i,r in enumerate(rows):
            tag="alt" if i%2==0 else ""
            cd=f"{r['credit_days']} วัน" if r['credit_days'] else "—"
            self.tree.insert("","end",iid=r["id"],tags=(tag,),
                values=(r["code"],r["name"],r["contact"] or "—",
                        r["phone"] or "—",cd,r["tax_id"] or "—"))
        self.count_lbl.config(text=f"ผู้จำหน่าย {len(rows)} ราย")
    def _sel(self):
        s=self.tree.selection(); return int(s[0]) if s else None
    def _add(self): self._form()
    def _edit(self):
        sid=self._sel()
        if sid: self._form(sid)
    def _gen_code(self):
        conn=get_db()
        rows=conn.execute("SELECT code FROM suppliers WHERE code LIKE 'SUP%'").fetchall()
        conn.close()
        nums=[int(r["code"][3:]) for r in rows if r["code"][3:].isdigit()]
        return f"SUP{(max(nums)+1 if nums else 1):04d}"
    def _form(self,sid=None):
        conn=get_db()
        s=conn.execute("SELECT * FROM suppliers WHERE id=?",(sid,)).fetchone() if sid else None
        conn.close()
        win=tk.Toplevel(self)
        win.title("แก้ไขผู้จำหน่าย" if sid else "เพิ่มผู้จำหน่ายใหม่")
        win.configure(bg=C_BG); win.grab_set(); win.resizable(False,False)
        tk.Frame(win,bg=C_BLUE,height=3).pack(fill=tk.X)
        tk.Label(win,text="แก้ไขผู้จำหน่าย" if sid else "เพิ่มผู้จำหน่ายใหม่",
                 font=F_H1,bg=C_BG,fg=C_TEXT,pady=10).pack(anchor="w",padx=16)
        entries={}
        def row(lbl,key,default="",width=34):
            tk.Label(win,text=lbl,font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(anchor="w",padx=16,pady=(6,2))
            e=field(win,width=width); e.pack(fill=tk.X,padx=16,pady=(0,2),ipady=5)
            val=str(s[key] if s and s[key] is not None else default)
            if val: e.insert(0,val)
            entries[key]=e; return e
        # code row
        tk.Label(win,text="รหัสผู้จำหน่าย",font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(anchor="w",padx=16,pady=(6,2))
        cr=tk.Frame(win,bg=C_BG); cr.pack(fill=tk.X,padx=16,pady=(0,2))
        code_e=field(cr,width=22); code_e.pack(side=tk.LEFT,fill=tk.X,expand=True,ipady=5)
        code_e.insert(0,s["code"] if s else self._gen_code())
        entries["code"]=code_e
        tk.Button(cr,text="⚡ Auto",font=(FM.primary,10,"bold"),bg=C_YELLOW,fg=C_BG,
                  relief=tk.FLAT,cursor="hand2",padx=8,pady=4,bd=0,
                  command=lambda:(code_e.delete(0,tk.END),code_e.insert(0,self._gen_code()))
                  ).pack(side=tk.LEFT,padx=(6,0))
        row("ชื่อบริษัท / ชื่อร้าน","name")
        row("ผู้ติดต่อ","contact")
        # phone + email row
        pr=tk.Frame(win,bg=C_BG); pr.pack(fill=tk.X,padx=16,pady=(0,2))
        for lbl,key in [("โทรศัพท์","phone"),("Email","email")]:
            col=tk.Frame(pr,bg=C_BG); col.pack(side=tk.LEFT,fill=tk.X,expand=True,padx=(0,8))
            tk.Label(col,text=lbl,font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(anchor="w",pady=(6,2))
            e=field(col,width=16); e.pack(fill=tk.X,ipady=5)
            val=str(s[key] if s and s[key] is not None else "")
            if val: e.insert(0,val)
            entries[key]=e
        row("ที่อยู่","address")
        # tax_id + credit_days
        tr=tk.Frame(win,bg=C_BG); tr.pack(fill=tk.X,padx=16,pady=(0,2))
        for lbl,key,def_ in [("เลขผู้เสียภาษี","tax_id",""),("เครดิต (วัน)","credit_days","0")]:
            col=tk.Frame(tr,bg=C_BG); col.pack(side=tk.LEFT,fill=tk.X,expand=True,padx=(0,8))
            tk.Label(col,text=lbl,font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(anchor="w",pady=(6,2))
            e=field(col,width=16); e.pack(fill=tk.X,ipady=5)
            val=str(s[key] if s and s[key] is not None else def_)
            if val: e.insert(0,val)
            entries[key]=e
        row("หมายเหตุ","note")
        def save():
            d={k:e.get().strip() if hasattr(e,"get") else "" for k,e in entries.items()}
            if not d["name"]: messagebox.showerror("ผิดพลาด","กรุณากรอกชื่อบริษัท/ร้าน",parent=win); return
            if not d["code"]: messagebox.showerror("ผิดพลาด","กรุณากรอกรหัสผู้จำหน่าย",parent=win); return
            try: d["credit_days"]=int(d.get("credit_days",0) or 0)
            except: d["credit_days"]=0
            conn2=get_db()
            now=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            try:
                if sid:
                    conn2.execute(
                        "UPDATE suppliers SET code=?,name=?,contact=?,phone=?,email=?,"
                        "address=?,tax_id=?,credit_days=?,note=? WHERE id=?",
                        (d["code"],d["name"],d["contact"],d["phone"],d["email"],
                         d["address"],d["tax_id"],d["credit_days"],d["note"],sid))
                else:
                    conn2.execute(
                        "INSERT INTO suppliers (code,name,contact,phone,email,"
                        "address,tax_id,credit_days,note,active,created_at) VALUES (?,?,?,?,?,?,?,?,?,1,?)",
                        (d["code"],d["name"],d["contact"],d["phone"],d["email"],
                         d["address"],d["tax_id"],d["credit_days"],d["note"],now))
                conn2.commit()
            except Exception as ex:
                messagebox.showerror("ผิดพลาด",str(ex),parent=win); conn2.close(); return
            conn2.close(); win.destroy(); self.load()
        tk.Frame(win,bg=C_BORDER,height=1).pack(fill=tk.X,padx=16,pady=(12,0))
        bf=tk.Frame(win,bg=C_BG,padx=16,pady=12); bf.pack(fill=tk.X)
        accent_btn(bf,"💾  บันทึก",save,pad_x=24,pad_y=10).pack(side=tk.LEFT,padx=(0,8))
        pill_btn(bf,"ยกเลิก",win.destroy,bg=C_SURFACE2,fg=C_MUTED,pad_x=16,pad_y=10).pack(side=tk.LEFT)
        if sid:
            def _del():
                if messagebox.askyesno("ยืนยัน","ลบผู้จำหน่ายนี้?",parent=win):
                    c2=get_db(); c2.execute("UPDATE suppliers SET active=0 WHERE id=?",(sid,)); c2.commit(); c2.close()
                    win.destroy(); self.load()
            pill_btn(bf,"🗑 ลบ",_del,bg="#2d1a00",fg="#ff6b6b",pad_x=14,pad_y=10).pack(side=tk.RIGHT)
        win.bind("<Return>",lambda e:save())
        win.update_idletasks()
        win.minsize(win.winfo_reqwidth(), win.winfo_reqheight())

# ══ TAB: CUSTOMERS ═══════════════════════════════════════
class CustomerTab(tk.Frame):
    PRICE_LEVELS = [
        ("price",   "ราคาปกติ"),
        ("price_a", "ราคา A"),
        ("price_b", "ราคา B"),
        ("price_c", "ราคา C"),
    ]

    def __init__(self,parent):
        super().__init__(parent,bg=C_BG); self._build(); self.load()

    def _build(self):
        top=tk.Frame(self,bg=C_BG,pady=10,padx=12); top.pack(fill=tk.X)
        tk.Label(top,text="👥  ระบบสมาชิก",font=F_H1,bg=C_BG,fg=C_TEXT).pack(side=tk.LEFT)
        bf=tk.Frame(top,bg=C_BG); bf.pack(side=tk.RIGHT)
        accent_btn(bf,"➕  เพิ่มสมาชิก",self._add).pack(side=tk.LEFT,padx=3)
        pill_btn(bf,"✏️  แก้ไข",self._edit,bg=C_SURFACE2,fg=C_YELLOW).pack(side=tk.LEFT,padx=3)
        pill_btn(bf,"📋  ประวัติ",self._hist,bg=C_SURFACE2,fg=C_BLUE).pack(side=tk.LEFT,padx=3)
        ghost_btn(bf,"↺",self.load).pack(side=tk.LEFT,padx=3)
        sf=tk.Frame(self,bg=C_BG,padx=12); sf.pack(fill=tk.X,pady=(0,8))
        self.sv=tk.StringVar(); self.sv.trace_add("write",lambda *a:self.load())
        fe=field(sf,width=30); fe.configure(textvariable=self.sv); fe.pack(side=tk.LEFT,ipady=5)
        tk.Label(sf,text="  🔍",font=(FM.primary,14),bg=C_BG,fg=C_MUTED).pack(side=tk.LEFT)
        cols=("รหัส","ชื่อ","โทรศัพท์","ระดับราคา","แต้มสะสม","ยอดสะสม (฿)","วันที่สมัคร")
        frm,self.tree=make_tree(self,cols,(80,180,110,90,80,110,120))
        frm.pack(fill=tk.BOTH,expand=True,padx=12,pady=(0,12))

    def load(self,*_):
        q=self.sv.get().strip()
        conn=get_db(); rows=conn.execute("SELECT * FROM customers WHERE name LIKE ? OR code LIKE ? OR phone LIKE ? ORDER BY name",(f"%{q}%",f"%{q}%",f"%{q}%")).fetchall(); conn.close()
        _lv_map={"price":"ปกติ","price_a":"ราคา A","price_b":"ราคา B","price_c":"ราคา C"}
        self.tree.delete(*self.tree.get_children())
        for i,r in enumerate(rows):
            lv=_lv_map.get(r["price_level"] if "price_level" in r.keys() else "price","ปกติ")
            self.tree.insert("","end",iid=r["id"],tags=("alt" if i%2==0 else "",),
                values=(r["code"],r["name"],r["phone"] or "",lv,r["points"],f"฿{r['total_spent']:,.2f}",r["created_at"] or ""))

    def _sel(self):
        s=self.tree.selection()
        if not s: messagebox.showwarning("เลือก","กรุณาเลือกสมาชิกก่อน"); return None
        return int(s[0])

    def _form(self,cid=None):
        conn=get_db()
        c=conn.execute("SELECT * FROM customers WHERE id=?",(cid,)).fetchone() if cid else None
        last=conn.execute("SELECT MAX(CAST(SUBSTR(code,2) AS INTEGER)) FROM customers WHERE code LIKE 'M%'").fetchone()[0]; conn.close()
        auto=f"M{(last or 0)+1:04d}"
        win=tk.Toplevel(self); win.title("สมาชิก"); win.configure(bg=C_BG); win.grab_set(); win.resizable(False,False)
        tk.Frame(win,bg=C_ACCENT,height=3).pack(fill=tk.X)
        tk.Label(win,text="แก้ไขสมาชิก" if cid else "เพิ่มสมาชิกใหม่",font=F_H1,bg=C_BG,fg=C_TEXT,pady=12).pack(anchor="w",padx=16)
        fields=[("รหัสสมาชิก","code"),("ชื่อ-นามสกุล","name"),("โทรศัพท์","phone")]
        entries={}
        for lbl,key in fields:
            tk.Label(win,text=lbl,font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(anchor="w",padx=16,pady=(8,2))
            e=field(win,width=32); e.pack(fill=tk.X,padx=16,pady=(0,4),ipady=6)
            if c: e.insert(0,c[key] or "")
            elif key=="code": e.insert(0,auto)
            entries[key]=e
        if cid:
            tk.Label(win,text="แต้มสะสม",font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(anchor="w",padx=16,pady=(8,2))
            pe=field(win,width=32); pe.insert(0,str(c["points"])); pe.pack(fill=tk.X,padx=16,pady=(0,4),ipady=6); entries["points"]=pe
        # ── ระดับราคาขาย ──────────────────────────────────
        tk.Label(win,text="ระดับราคาขาย",font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(anchor="w",padx=16,pady=(10,4))
        pf=tk.Frame(win,bg=C_SURFACE2,padx=14,pady=10); pf.pack(fill=tk.X,padx=16,pady=(0,6))
        price_lv=tk.StringVar(value=(c["price_level"] if c and c["price_level"] else "price"))
        for val,lbl in self.PRICE_LEVELS:
            colors={"price":"#8b949e","price_a":"#4a9eff","price_b":"#3fb950","price_c":"#ffd166"}
            col=colors.get(val,"#8b949e")
            tk.Radiobutton(pf,text=lbl,variable=price_lv,value=val,
                           bg=C_SURFACE2,fg=col,selectcolor=C_BG,
                           activebackground=C_SURFACE2,activeforeground=col,
                           font=(FM.primary,11)).pack(side=tk.LEFT,padx=8)
        def save():
            try:
                d={k:v.get().strip() for k,v in entries.items()}
                if not d["name"]: raise ValueError("กรุณากรอกชื่อ")
            except ValueError as e: messagebox.showerror("ผิดพลาด",str(e)); return
            plevel=price_lv.get()
            conn2=get_db()
            if cid: conn2.execute("UPDATE customers SET code=?,name=?,phone=?,points=?,price_level=? WHERE id=?",(d["code"],d["name"],d["phone"],int(d.get("points",0)),plevel,cid))
            else: conn2.execute("INSERT INTO customers (code,name,phone,price_level,created_at) VALUES (?,?,?,?,?)",(d["code"],d["name"],d["phone"],plevel,date.today().strftime("%Y-%m-%d")))
            conn2.commit(); conn2.close(); win.destroy(); self.load()
        accent_btn(win,"💾  บันทึก",save,pad_x=24,pad_y=10).pack(pady=16)

    def _add(self): self._form()
    def _edit(self):
        cid=self._sel()
        if cid: self._form(cid)
    def _hist(self):
        cid=self._sel()
        if not cid: return
        conn=get_db(); cust=conn.execute("SELECT * FROM customers WHERE id=?",(cid,)).fetchone(); conn.close()
        sales=get_sales_by_customer(cid,limit=50)  # db_sale — Session 8
        win=tk.Toplevel(self); win.title(f"ประวัติ: {cust['name']}"); win.configure(bg=C_BG); win.geometry("560,400".replace(",","x"))
        tk.Frame(win,bg=C_ACCENT,height=3).pack(fill=tk.X)
        tk.Label(win,text=f"👤  {cust['name']}   แต้ม: {cust['points']}   ยอดสะสม: ฿{cust['total_spent']:,.2f}",
                 font=F_H1,bg=C_BG,fg=C_BLUE,pady=10).pack(anchor="w",padx=16)
        try: rate=int(get_setting("point_rate") or 10)
        except: rate=10
        frm,tv=make_tree(win,("เลขบิล","วันเวลา","ยอดสุทธิ","แต้มที่ได้"),(80,160,120,100),height=12)
        frm.pack(fill=tk.BOTH,expand=True,padx=12,pady=8)
        for s in sales: tv.insert("","end",values=(s["id"],s["sale_date"],f"฿{s['net']:,.2f}",int(s["net"]//rate)))


# ══ TAB: PROMOTIONS ══════════════════════════════════════
class PromoTab(tk.Frame):
    def __init__(self,parent):
        super().__init__(parent,bg=C_BG); self._build(); self.load()

    def _build(self):
        top=tk.Frame(self,bg=C_BG,pady=10,padx=12); top.pack(fill=tk.X)
        tk.Label(top,text="💳  โปรโมชั่นและตัดราคา",font=F_H1,bg=C_BG,fg=C_TEXT).pack(side=tk.LEFT)
        bf=tk.Frame(top,bg=C_BG); bf.pack(side=tk.RIGHT)
        accent_btn(bf,"➕  เพิ่มโปรโมชั่น",self._add).pack(side=tk.LEFT,padx=3)
        pill_btn(bf,"✏️  แก้ไข",self._edit,bg=C_SURFACE2,fg=C_YELLOW).pack(side=tk.LEFT,padx=3)
        ghost_btn(bf,"↺",self.load).pack(side=tk.LEFT,padx=3)
        cols=("ID","ชื่อโปรโมชั่น","ประเภท","ส่วนลด","สินค้า","เริ่ม","สิ้นสุด","สถานะ")
        frm,self.tree=make_tree(self,cols,(40,200,80,80,160,100,100,70))
        frm.pack(fill=tk.BOTH,expand=True,padx=12,pady=(0,12))

    def load(self,*_):
        # Session 12: swap → db_promotion.get_all_promotions
        rows=get_all_promotions()
        self.tree.delete(*self.tree.get_children())
        for i,r in enumerate(rows):
            typ="ลด %" if r["type"]=="percent" else "ลด ฿"
            val=f"{r['value']}%" if r["type"]=="percent" else f"฿{r['value']:.0f}"
            self.tree.insert("","end",iid=r["id"],tags=("on" if r["active"] else "off",),
                values=(r["id"],r["name"],typ,val,r["pname"] or "ทุกสินค้า",r["start_date"] or "—",r["end_date"] or "—","เปิด" if r["active"] else "ปิด"))

    def _sel(self):
        s=self.tree.selection()
        if not s: messagebox.showwarning("เลือก","กรุณาเลือกโปรโมชั่น"); return None
        return int(s[0])

    def _form(self,pid=None):
        # Session 12: swap → db_promotion.get_promotion
        p=get_promotion(pid) if pid else None
        prods=get_products_for_sale_list()  # Session 6: swap → db_product
        win=tk.Toplevel(self); win.title("โปรโมชั่น"); win.configure(bg=C_BG); win.grab_set(); win.resizable(False,False)
        tk.Frame(win,bg=C_ACCENT,height=3).pack(fill=tk.X)
        tk.Label(win,text="แก้ไขโปรโมชั่น" if pid else "เพิ่มโปรโมชั่นใหม่",font=F_H1,bg=C_BG,fg=C_TEXT,pady=12).pack(anchor="w",padx=16)
        def le(lbl,key,default=""):
            tk.Label(win,text=lbl,font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(anchor="w",padx=16,pady=(8,2))
            e=field(win,width=36); e.insert(0,str(p[key] if p and p[key] is not None else default))
            e.pack(fill=tk.X,padx=16,pady=(0,4),ipady=6); return e
        ne=le("ชื่อโปรโมชั่น","name"); ve=le("ส่วนลด (% หรือ ฿)","value","0")

        # วันเริ่ม / วันสิ้นสุด — ใช้ ThaiDateEntry (DD/MM/BBBB)
        tk.Label(win,text="วันเริ่ม (DD/MM/BBBB หรือเว้นว่าง)",font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(anchor="w",padx=16,pady=(8,2))
        se2_frame=tk.Frame(win,bg=C_BG); se2_frame.pack(fill=tk.X,padx=16,pady=(0,4))
        se2_tde=ThaiDateEntry(se2_frame,width=14)
        if p and p["start_date"]: se2_tde.set(p["start_date"])
        se2_tde.pack(side=tk.LEFT)
        tk.Button(se2_frame,text="✖ ล้าง",font=F_SMALL,bg=C_SURFACE2,fg=C_MUTED,
                  relief=tk.FLAT,cursor="hand2",bd=0,padx=6,
                  command=lambda:(se2_tde.delete(0,tk.END))).pack(side=tk.LEFT,padx=(6,0))

        tk.Label(win,text="วันสิ้นสุด (DD/MM/BBBB หรือเว้นว่าง)",font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(anchor="w",padx=16,pady=(8,2))
        ee_frame=tk.Frame(win,bg=C_BG); ee_frame.pack(fill=tk.X,padx=16,pady=(0,4))
        ee_tde=ThaiDateEntry(ee_frame,width=14)
        if p and p["end_date"]: ee_tde.set(p["end_date"])
        ee_tde.pack(side=tk.LEFT)
        tk.Button(ee_frame,text="✖ ล้าง",font=F_SMALL,bg=C_SURFACE2,fg=C_MUTED,
                  relief=tk.FLAT,cursor="hand2",bd=0,padx=6,
                  command=lambda:(ee_tde.delete(0,tk.END))).pack(side=tk.LEFT,padx=(6,0))

        # aliases ให้ save() ใช้ได้ผ่าน get_iso()
        class _ISOProxy:
            def __init__(self,tde): self._t=tde
            def get(self): return self._t.get_iso() if self._t.get().strip() else ""
            def strip(self): return self.get()
        se2=_ISOProxy(se2_tde); ee=_ISOProxy(ee_tde)
        tk.Label(win,text="ประเภท",font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(anchor="w",padx=16,pady=(8,2))
        tv2=tk.StringVar(value=p["type"] if p else "percent")
        tf=tk.Frame(win,bg=C_BG); tf.pack(anchor="w",padx=16,pady=(0,8))
        for v,l in [("percent","ลด %"),("fixed","ลด ฿")]:
            tk.Radiobutton(tf,text=l,variable=tv2,value=v,bg=C_BG,fg=C_TEXT,selectcolor=C_SURFACE2,
                           activebackground=C_BG,activeforeground=C_TEXT,font=F_BODY).pack(side=tk.LEFT,padx=8)
        tk.Label(win,text="ใช้กับสินค้า",font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(anchor="w",padx=16,pady=(8,2))
        pv=tk.StringVar()
        pc=ttk.Combobox(win,textvariable=pv,font=F_BODY,state="readonly")
        pnames=["ทุกสินค้า"]+[r["name"] for r in prods]; pids=[0]+[r["id"] for r in prods]
        pc["values"]=pnames
        if p and p["product_id"]:
            try: pc.current(pids.index(p["product_id"]))
            except: pc.current(0)
        else: pc.current(0)
        pc.pack(fill=tk.X,padx=16,pady=(0,8))
        av=tk.BooleanVar(value=bool(p["active"]) if p else True)
        tk.Checkbutton(win,text="เปิดใช้งาน",variable=av,bg=C_BG,fg=C_TEXT,selectcolor=C_SURFACE2,
                       activebackground=C_BG,activeforeground=C_TEXT,font=F_BODY).pack(anchor="w",padx=16)
        def save():
            try:
                nm=ne.get().strip()
                if not nm: raise ValueError("กรุณากรอกชื่อ")
                vv=float(ve.get() or 0); pi=pids[pc.current()]
            except ValueError as e: messagebox.showerror("ผิดพลาด",str(e)); return
            conn2=get_db()
            if pid: conn2.execute("UPDATE promotions SET name=?,type=?,value=?,product_id=?,start_date=?,end_date=?,active=? WHERE id=?",(nm,tv2.get(),vv,pi,se2.get().strip() or None,ee.get().strip() or None,int(av.get()),pid))
            else: conn2.execute("INSERT INTO promotions (name,type,value,product_id,start_date,end_date,active) VALUES (?,?,?,?,?,?,?)",(nm,tv2.get(),vv,pi,se2.get().strip() or None,ee.get().strip() or None,int(av.get())))
            conn2.commit(); conn2.close(); win.destroy(); self.load()
        accent_btn(win,"💾  บันทึก",save,pad_x=24,pad_y=10).pack(pady=16)

    def _add(self): self._form()
    def _edit(self):
        pid=self._sel()
        if pid: self._form(pid)



# ══ TAB: RETURNS ════════════════════════════════════════════════════════
class ReturnTab(tk.Frame):
    def __init__(self,parent):
        super().__init__(parent,bg=C_BG)
        self.sale_id=None
        self.selected_item=None
        self._build(); self.load()

    def _build(self):
        top=tk.Frame(self,bg=C_BG,pady=10,padx=14); top.pack(fill=tk.X)
        tk.Label(top,text="↩  รับคืนสินค้า / คืนเงิน",font=F_H1,bg=C_BG,fg=C_TEXT).pack(side=tk.LEFT)
        pill_btn(top,"↺ รีเฟรช",self.load,bg=C_SURFACE2,fg=C_TEXT).pack(side=tk.RIGHT)

        # Search bill
        sf=tk.Frame(self,bg=C_SURFACE,highlightbackground=C_BORDER,highlightthickness=1)
        sf.pack(fill=tk.X,padx=14,pady=(0,8))
        tk.Frame(sf,bg=C_ACCENT2,height=3).pack(fill=tk.X)
        hf=tk.Frame(sf,bg=C_SURFACE,padx=14,pady=10); hf.pack(fill=tk.X)
        tk.Label(hf,text="ค้นหาบิลที่ต้องการคืน",font=F_H2,bg=C_SURFACE,fg=C_TEXT).pack(anchor="w")
        rf=tk.Frame(hf,bg=C_SURFACE); rf.pack(fill=tk.X,pady=(8,0))
        tk.Label(rf,text="เลขบิล:",font=F_BODY,bg=C_SURFACE,fg=C_MUTED).pack(side=tk.LEFT)
        self.bill_var=tk.StringVar()
        e=field(rf,width=14); e.configure(textvariable=self.bill_var)
        e.pack(side=tk.LEFT,padx=6,ipady=4)
        e.bind("<Return>",self._search_bill)
        accent_btn(rf,"ค้นหา",self._search_bill,pad_x=14,pad_y=5).pack(side=tk.LEFT)
        self.bill_info=tk.Label(hf,text="",font=F_SMALL,bg=C_SURFACE,fg=C_MUTED)
        self.bill_info.pack(anchor="w",pady=(6,0))

        # Split: items list + return form
        cf=tk.Frame(self,bg=C_BG); cf.pack(fill=tk.BOTH,expand=True,padx=14,pady=(0,8))

        lf=tk.Frame(cf,bg=C_SURFACE,highlightbackground=C_BORDER,highlightthickness=1)
        lf.pack(side=tk.LEFT,fill=tk.BOTH,expand=True,padx=(0,6))
        tk.Label(lf,text="รายการสินค้าในบิล",font=F_H2,bg=C_SURFACE,fg=C_TEXT,pady=8).pack(anchor="w",padx=12)
        tk.Frame(lf,bg=C_BORDER,height=1).pack(fill=tk.X)
        cols=("ชื่อสินค้า","ราคา","ซื้อ","คืนแล้ว","คืนได้")
        self.item_tree=ttk.Treeview(lf,columns=cols,show="headings",height=8,style="R.Treeview")
        ws=[210,80,70,80,70]
        for col,w in zip(cols,ws):
            self.item_tree.heading(col,text=col)
            self.item_tree.column(col,width=w,anchor="center" if col!="ชื่อสินค้า" else "w")
        self.item_tree.pack(fill=tk.BOTH,expand=True,padx=4,pady=4)
        self.item_tree.bind("<<TreeviewSelect>>",self._on_select)
        self.item_tree.tag_configure("ok",foreground=C_TEXT)
        self.item_tree.tag_configure("off",foreground=C_MUTED)

        # Return form
        rf2=tk.Frame(cf,bg=C_SURFACE,highlightbackground=C_BORDER,highlightthickness=1,width=260)
        rf2.pack(side=tk.RIGHT,fill=tk.Y); rf2.pack_propagate(False)
        tk.Frame(rf2,bg=C_ACCENT2,height=3).pack(fill=tk.X)
        tk.Label(rf2,text="รายละเอียดการคืน",font=F_H2,bg=C_SURFACE,fg=C_TEXT,pady=8).pack(anchor="w",padx=12)
        self.sel_lbl=tk.Label(rf2,text="เลือกสินค้าจากรายการ",font=F_SMALL,bg=C_SURFACE,fg=C_MUTED,wraplength=220,justify="left")
        self.sel_lbl.pack(anchor="w",padx=12)
        tk.Frame(rf2,bg=C_BORDER,height=1).pack(fill=tk.X,pady=6)

        tk.Label(rf2,text="จำนวนที่คืน:",font=F_BODY,bg=C_SURFACE,fg=C_TEXT).pack(anchor="w",padx=12)
        self.qty_e=field(rf2,width=20,justify="right",font=(FM.primary,17,"bold"),fg=C_YELLOW)
        self.qty_e.insert(0,"1"); self.qty_e.pack(fill=tk.X,padx=12,pady=(2,8),ipady=5)
        self.qty_e.bind("<KeyRelease>",self._calc_refund)

        tk.Label(rf2,text="เหตุผล:",font=F_BODY,bg=C_SURFACE,fg=C_TEXT).pack(anchor="w",padx=12)
        self.reason_var=tk.StringVar(value="สินค้าชำรุด")
        cb=ttk.Combobox(rf2,textvariable=self.reason_var,
                        values=["สินค้าชำรุด","ไม่ตรงตามสั่ง","เปลี่ยนใจ","อื่นๆ"],
                        state="readonly",font=F_BODY)
        cb.pack(fill=tk.X,padx=12,pady=(2,8))

        self.restock_var=tk.BooleanVar(value=True)
        tk.Checkbutton(rf2,text="คืนสินค้าเข้าสต็อก",variable=self.restock_var,
                       bg=C_SURFACE,fg=C_TEXT,selectcolor=C_SURFACE2,
                       activebackground=C_SURFACE,activeforeground=C_TEXT,font=F_BODY
                       ).pack(anchor="w",padx=12)

        tk.Frame(rf2,bg=C_BORDER,height=1).pack(fill=tk.X,pady=8)
        refund_f=tk.Frame(rf2,bg=C_SURFACE,padx=12); refund_f.pack(fill=tk.X)
        tk.Label(refund_f,text="ยอดคืนเงิน:",font=F_BODY,bg=C_SURFACE,fg=C_MUTED).pack(side=tk.LEFT)
        self.refund_var=tk.StringVar(value="฿ 0.00")
        tk.Label(refund_f,textvariable=self.refund_var,font=F_NUM_SM,bg=C_SURFACE,fg=C_ACCENT2).pack(side=tk.RIGHT)

        tk.Button(rf2,text="↩  ยืนยันรับคืน",command=self._confirm_return,
                  font=(FM.primary,14,"bold"),bg=C_ACCENT2,fg=C_WHITE,
                  relief=tk.FLAT,cursor="hand2",pady=10,
                  activebackground=C_ACCENT2,bd=0).pack(fill=tk.X,padx=12,pady=12)

        # History
        hf2=tk.Frame(self,bg=C_SURFACE,highlightbackground=C_BORDER,highlightthickness=1)
        hf2.pack(fill=tk.X,padx=14,pady=(0,10))
        tk.Label(hf2,text="ประวัติการรับคืน",font=F_H2,bg=C_SURFACE,fg=C_TEXT,pady=8).pack(anchor="w",padx=12)
        tk.Frame(hf2,bg=C_BORDER,height=1).pack(fill=tk.X)
        cols2=("วันที่","เลขบิล","สินค้า","จำนวน","คืนเงิน","เหตุผล","สต็อก")
        self.hist_tree=ttk.Treeview(hf2,columns=cols2,show="headings",height=5,style="R.Treeview")
        ws2=[130,70,180,70,90,100,70]
        for col,w in zip(cols2,ws2):
            self.hist_tree.heading(col,text=col)
            self.hist_tree.column(col,width=w,anchor="center" if col!="สินค้า" else "w")
        self.hist_tree.pack(fill=tk.X,padx=4,pady=4)

    def _search_bill(self,event=None):
        bid=self.bill_var.get().strip()
        if not bid: return
        try: bid=int(bid)
        except: messagebox.showwarning("ผิดพลาด","กรุณากรอกเลขบิลให้ถูกต้อง",parent=self.winfo_toplevel()); return
        sale=get_sale_with_customer(bid)   # db_sale — Session 8
        if not sale:
            messagebox.showwarning("ไม่พบ","ไม่พบบิล #%d"%bid,parent=self.winfo_toplevel()); return
        items=get_sale_items(bid)           # db_sale — Session 8
        self.bill_items={it["id"]:dict(it) for it in items}
        conn=get_db()
        cname=sale["cn"] or "ไม่ระบุ"
        self.bill_info.config(
            text="บิล #{}  |  {}  |  ลูกค้า: {}  |  ยอด: ฿{:,.2f}".format(sale["id"],sale["sale_date"],cname,sale["net"]),
            fg=C_BLUE)
        self.item_tree.delete(*self.item_tree.get_children())
        self.sale_id=bid
        for it in items:
            ret=conn.execute("SELECT COALESCE(SUM(qty),0) FROM returns WHERE sale_id=? AND product_id=?",(bid,it["product_id"])).fetchone()[0]
            can=it["qty"]-ret
            tag="ok" if can>0 else "off"
            self.item_tree.insert("","end",iid=it["id"],tags=(tag,),
                values=(it["name"],"฿%.2f"%it["price"],it["qty"],ret,can))
        conn.close()
        self.selected_item=None
        self.sel_lbl.config(text="เลือกสินค้าจากรายการ")
        self.refund_var.set("฿ 0.00")

    def _on_select(self,event=None):
        sel=self.item_tree.selection()
        if not sel: return
        vals=self.item_tree.item(sel[0])["values"]
        name,price,bought,returned,can=vals
        if int(can)<=0:
            self.sel_lbl.config(text="คืนครบแล้ว",fg=C_MUTED)
            self.selected_item=None; return
        price_num=float(str(price).replace("฿","").strip())
        self.selected_item={"iid":int(sel[0]),"name":name,"price":price_num,"max":int(can)}
        info="สินค้า: %s\nราคา: ฿%.2f  |คืนได้สูงสุด %s ชิ้น"%(name,price_num,can)
        self.sel_lbl.config(text=info,fg=C_TEXT)
        self.qty_e.delete(0,tk.END); self.qty_e.insert(0,"1")
        self._calc_refund()

    def _calc_refund(self,event=None):
        if not self.selected_item:
            self.refund_var.set("฿ 0.00"); return
        try: qty=int(self.qty_e.get() or 0)
        except: qty=0
        self.refund_var.set("฿ {:,.2f}".format(qty*self.selected_item["price"]))

    def _confirm_return(self):
        if not self.selected_item:
            messagebox.showwarning("เลือกสินค้า","กรุณาเลือกสินค้าก่อน",parent=self.winfo_toplevel()); return
        try: qty=int(self.qty_e.get() or 0)
        except: messagebox.showerror("ผิดพลาด","กรุณากรอกจำนวน",parent=self.winfo_toplevel()); return
        if qty<=0 or qty>self.selected_item["max"]:
            messagebox.showwarning("จำนวนไม่ถูกต้อง","กรอก 1-%d"%self.selected_item["max"],parent=self.winfo_toplevel()); return
        refund=qty*self.selected_item["price"]
        name=self.selected_item["name"]
        reason=self.reason_var.get()
        restock=int(self.restock_var.get())
        rst_txt="คืนสต็อก" if restock else "ไม่คืนสต็อก"
        msg="รับคืน: {}\nจำนวน: {} ชิ้น\nคืนเงิน: ฿{:,.2f}\nเหตุผล: {}\n{}\n\nยืนยัน?".format(name,qty,refund,reason,rst_txt)
        if not messagebox.askyesno("ยืนยัน",msg,parent=self.winfo_toplevel()): return
        pid=get_sale_item_product_id(self.selected_item["iid"])  # db_sale — Session 8
        if pid is None: return
        conn=get_db()
        now=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        conn.execute("INSERT INTO returns (return_date,sale_id,product_id,name,price,qty,refund,reason,restock) VALUES (?,?,?,?,?,?,?,?,?)",
                     (now,self.sale_id,pid,name,self.selected_item["price"],qty,refund,reason,restock))
        if restock:
            conn.execute("UPDATE products SET stock=stock+? WHERE id=?",(qty,pid))
        conn.commit(); conn.close()
        messagebox.showinfo("รับคืนสำเร็จ",
            "✅ รับคืนเรียบร้อย\n{}  {} ชิ้น\nคืนเงินลูกค้า ฿{:,.2f}".format(name,qty,refund),
            parent=self.winfo_toplevel())
        self._search_bill()
        self.load()

    def load(self,*_):
        conn=get_db()
        try:
            rows=conn.execute("SELECT * FROM returns ORDER BY return_date DESC LIMIT 50").fetchall()
        except: rows=[]
        conn.close()
        self.hist_tree.delete(*self.hist_tree.get_children())
        for i,r in enumerate(rows):
            self.hist_tree.insert("","end",tags=("alt" if i%2==0 else "",),
                values=(r["return_date"],r["sale_id"],r["name"],r["qty"],
            "฿{:,.2f}".format(r["refund"]),r["reason"] or "-","คืน" if r["restock"] else "ไม่คืน"))


# ══ TAB: ประวัติยกเลิก / คืนสินค้า ════════════════════════
class HistoryTab(tk.Frame):
    """รวมประวัติยกเลิกบิล + คืนสินค้า หน้าเดียว มี filter วันที่"""
    def __init__(self, parent):
        super().__init__(parent, bg=C_BG)
        self._build(); self._set_dates("month", 0)

    def _build(self):
        top = tk.Frame(self, bg=C_BG, pady=10, padx=14); top.pack(fill=tk.X)
        tk.Label(top, text="🗂️  ประวัติยกเลิก / คืนสินค้า", font=F_H1, bg=C_BG, fg=C_TEXT).pack(side=tk.LEFT)
        ctrl = tk.Frame(top, bg=C_BG); ctrl.pack(side=tk.RIGHT)
        qf = tk.Frame(ctrl, bg=C_BG); qf.pack(side=tk.LEFT, padx=(0,6))
        for txt, cmd in [("📅 วันนี้", lambda: self._set_dates(0,0)),
                         ("⏳ เมื่อวาน", lambda: self._set_dates(-1,-1)),
                         ("📆 7วัน", lambda: self._set_dates(-6,0)),
                         ("📋 เดือนนี้", lambda: self._set_dates("month",0))]:
            pill_btn(qf, txt, cmd, bg=C_SURFACE2, fg=C_BLUE, pad_x=7, pad_y=3).pack(side=tk.LEFT, padx=2)
        tk.Label(ctrl, text="จาก", font=F_SMALL, bg=C_BG, fg=C_MUTED).pack(side=tk.LEFT, padx=(4,2))
        self.df = ThaiDateEntry(ctrl, width=11); self.df.pack(side=tk.LEFT, padx=(0,6))
        tk.Label(ctrl, text="ถึง", font=F_SMALL, bg=C_BG, fg=C_MUTED).pack(side=tk.LEFT, padx=(6,2))
        self.dt = ThaiDateEntry(ctrl, width=11); self.dt.pack(side=tk.LEFT, padx=(0,4))
        pill_btn(ctrl, "🔍 ดูข้อมูล", self.load, bg=C_ACCENT, fg=C_BG, pad_x=10, pad_y=3).pack(side=tk.LEFT, padx=(6,0))

        fb = tk.Frame(self, bg=C_BG, padx=14); fb.pack(fill=tk.X, pady=(0,4))
        tk.Label(fb, text="ประเภท:", font=F_SMALL, bg=C_BG, fg=C_MUTED).pack(side=tk.LEFT)
        self.type_var = tk.StringVar(value="ทั้งหมด")
        cb = ttk.Combobox(fb, textvariable=self.type_var,
                          values=["ทั้งหมด","🚫 ยกเลิกบิล","↩ คืนสินค้า"],
                          width=16, state="readonly", font=F_BODY)
        cb.pack(side=tk.LEFT, padx=(4,0)); cb.bind("<<ComboboxSelected>>", lambda e: self.load())

        kf = tk.Frame(self, bg=C_BG, padx=12, pady=4); kf.pack(fill=tk.X)
        self.sv = {}
        for key, lbl, icon, col in [("vc","บิลยกเลิก","🚫","#ff6b6b"),
                                    ("vs","ยอดบิลยกเลิก","💸","#ff6b6b"),
                                    ("rc","รายการคืน","↩","#e67e22"),
                                    ("rs","คืนเงินรวม","💵","#e67e22")]:
            card = tk.Frame(kf, bg=C_SURFACE2, padx=14, pady=10)
            card.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=4)
            tk.Label(card, text=f"{icon} {lbl}", font=F_SMALL, bg=C_SURFACE2, fg=C_MUTED).pack(anchor="w")
            v = tk.StringVar(value="—"); self.sv[key] = v
            tk.Label(card, textvariable=v, font=(FM.primary,16,"bold"), bg=C_SURFACE2, fg=col).pack(anchor="w")

        cols = ("เวลา","ประเภท","เลขบิล","รายละเอียด","จำนวน","ยอด (฿)","โดย","เหตุผล")
        frm, self.tree = make_tree(self, cols, (140,110,60,260,55,90,110,150), height=16)
        frm.pack(fill=tk.BOTH, expand=True, padx=12, pady=(2,8))
        self.tree.tag_configure("void", foreground="#ff6b6b")
        self.tree.tag_configure("ret", foreground="#e67e22")

    def _set_dates(self, fr, to):
        from datetime import timedelta
        today = date.today()
        d_from = today.replace(day=1) if fr=="month" else today+timedelta(days=fr)
        d_to = today+timedelta(days=to) if isinstance(to,int) else today
        self.df.set(d_from.strftime("%Y-%m-%d")); self.dt.set(d_to.strftime("%Y-%m-%d"))
        self.load()

    def load(self, *_):
        d0 = self.df.get_iso(); d1 = self.dt.get_iso()
        tfilter = self.type_var.get() if hasattr(self,"type_var") else "ทั้งหมด"
        voids, rets = [], []
        conn = get_db()
        try:
            voids = conn.execute(
                "SELECT sa.sale_id, sa.at, sa.by_name, sa.reason, COALESCE(s.net,0) net "
                "FROM sale_audit sa LEFT JOIN sales s ON s.id=sa.sale_id "
                "WHERE sa.action='void' AND date(sa.at) BETWEEN ? AND ? ORDER BY sa.at DESC",
                (d0, d1)).fetchall()
        except Exception as e:
            print(f"[history void] {e}")
        try:
            rets = conn.execute(
                "SELECT return_date, sale_id, name, qty, refund, reason, restock "
                "FROM returns WHERE date(return_date) BETWEEN ? AND ? ORDER BY return_date DESC",
                (d0, d1)).fetchall()
        except Exception as e:
            print(f"[history ret] {e}")
        conn.close()

        vc = len(voids); vs = sum(float(r["net"] or 0) for r in voids)
        rc = len(rets);  rs = sum(float(r["refund"] or 0) for r in rets)
        self.sv["vc"].set(f"{vc} บิล");    self.sv["vs"].set(f"฿{vs:,.2f}")
        self.sv["rc"].set(f"{rc} รายการ"); self.sv["rs"].set(f"฿{rs:,.2f}")

        rows = []
        if tfilter in ("ทั้งหมด", "🚫 ยกเลิกบิล"):
            for r in voids:
                rows.append((r["at"] or "",
                    (to_thai_dt(r["at"]), "🚫 ยกเลิกบิล", r["sale_id"], "ยกเลิกทั้งบิล",
                     "—", f"฿{float(r['net'] or 0):,.2f}", r["by_name"] or "—", r["reason"] or "—"),
                    "void"))
        if tfilter in ("ทั้งหมด", "↩ คืนสินค้า"):
            for r in rets:
                rows.append((r["return_date"] or "",
                    (to_thai_dt(r["return_date"]), "↩ คืนสินค้า", r["sale_id"], r["name"] or "—",
                     r["qty"], f"฿{float(r['refund'] or 0):,.2f}", "—", r["reason"] or "—"),
                    "ret"))
        rows.sort(key=lambda x: x[0], reverse=True)

        self.tree.delete(*self.tree.get_children())
        for _ts, vals, tag in rows:
            self.tree.insert("", "end", tags=(tag,), values=vals)
        if not rows:
            self.tree.insert("", "end",
                values=("—", "(ไม่มีข้อมูลในช่วงวันที่นี้)", "", "", "", "", "", ""))


# ══ TAB: รายการขายสินค้า (สินค้า×สต็อก — หน้าควบคุมสินค้า) ════
class ProductSalesTab(tk.Frame):
    """รายงานยอดขายระดับสินค้า + สต็อกคงเหลือ + สีเตือน + drill-down"""
    _PAY = {"เงินสด":"cash","QR/พร้อมเพย์":"qr","บัตรเครดิต":"credit","เงินเชื่อ":"debt","ผสม":"mixed"}
    _SORT = {"ขายมากสุด":"qty","กำไรมากสุด":"profit","สต็อกน้อยสุด":"stock",
             "ชื่อสินค้า":"name","รหัสสินค้า":"code"}
    _VIEW = {"ทั้งหมด":"all","ใกล้หมด (≤ ขั้นต่ำ)":"low"}

    def __init__(self, parent):
        super().__init__(parent, bg=C_BG)
        self._rows = {}
        self._build(); self._set_dates(0, 0)

    def _build(self):
        top = tk.Frame(self, bg=C_BG, pady=10, padx=14); top.pack(fill=tk.X)
        tk.Label(top, text="📦  รายการขายสินค้า", font=F_H1, bg=C_BG, fg=C_TEXT).pack(side=tk.LEFT)
        ctrl = tk.Frame(top, bg=C_BG); ctrl.pack(side=tk.RIGHT)
        qf = tk.Frame(ctrl, bg=C_BG); qf.pack(side=tk.LEFT, padx=(0,6))
        for txt, cmd in [("📅 วันนี้", lambda: self._set_dates(0,0)),
                         ("⏳ เมื่อวาน", lambda: self._set_dates(-1,-1)),
                         ("📆 7วัน", lambda: self._set_dates(-6,0)),
                         ("📋 เดือนนี้", lambda: self._set_dates("month",0))]:
            pill_btn(qf, txt, cmd, bg=C_SURFACE2, fg=C_BLUE, pad_x=7, pad_y=3).pack(side=tk.LEFT, padx=2)
        tk.Label(ctrl, text="จาก", font=F_SMALL, bg=C_BG, fg=C_MUTED).pack(side=tk.LEFT, padx=(4,2))
        self.df = ThaiDateEntry(ctrl, width=11); self.df.pack(side=tk.LEFT, padx=(0,6))
        tk.Label(ctrl, text="ถึง", font=F_SMALL, bg=C_BG, fg=C_MUTED).pack(side=tk.LEFT, padx=(6,2))
        self.dt = ThaiDateEntry(ctrl, width=11); self.dt.pack(side=tk.LEFT, padx=(0,4))
        pill_btn(ctrl, "🔍 ดูข้อมูล", self.load, bg=C_ACCENT, fg=C_BG, pad_x=10, pad_y=3).pack(side=tk.LEFT, padx=(6,0))

        # filter row
        fb = tk.Frame(self, bg=C_BG, padx=14); fb.pack(fill=tk.X, pady=(0,2))
        def _mk(lbl, var, vals, w=15):
            tk.Label(fb, text=lbl, font=F_SMALL, bg=C_BG, fg=C_MUTED).pack(side=tk.LEFT, padx=(0,2))
            cbx = ttk.Combobox(fb, textvariable=var, values=vals, width=w, state="readonly", font=F_BODY)
            cbx.pack(side=tk.LEFT, padx=(0,10)); cbx.bind("<<ComboboxSelected>>", lambda e: self.load())
            return cbx
        try: _cats = ["ทั้งหมด"] + CategoryManager.get_all()
        except Exception: _cats = ["ทั้งหมด"]
        self.cat_var  = tk.StringVar(value="ทั้งหมด"); _mk("หมวด:", self.cat_var, _cats, 14)
        self.view_var = tk.StringVar(value="ทั้งหมด"); _mk("มุมมอง:", self.view_var, list(self._VIEW.keys()), 14)
        self.sort_var = tk.StringVar(value="ขายมากสุด"); _mk("เรียง:", self.sort_var, list(self._SORT.keys()), 12)
        self.pay_var  = tk.StringVar(value="ทั้งหมด"); _mk("ชำระ:", self.pay_var, ["ทั้งหมด"]+list(self._PAY.keys()), 12)
        tk.Label(fb, text="ค้นหา:", font=F_SMALL, bg=C_BG, fg=C_MUTED).pack(side=tk.LEFT, padx=(0,2))
        self.search_var = tk.StringVar()
        _se = tk.Entry(fb, textvariable=self.search_var, width=18, bg=C_SURFACE2, fg=C_TEXT,
                       insertbackground=C_TEXT, relief=tk.FLAT, font=F_BODY)
        _se.pack(side=tk.LEFT, ipady=3); _se.bind("<Return>", lambda e: self.load())

        # KPI cards
        kf = tk.Frame(self, bg=C_BG, padx=12, pady=4); kf.pack(fill=tk.X)
        self.sv = {}
        for key, lbl, icon, col in [("n","รายการสินค้า","📦",C_BLUE),
                                    ("q","จำนวนชิ้นรวม","🔢",C_TEXT),
                                    ("s","ยอดขายรวม","💰",C_ACCENT),
                                    ("p","กำไรรวม","📈",C_GREEN)]:
            card = tk.Frame(kf, bg=C_SURFACE2, padx=14, pady=10)
            card.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=4)
            tk.Label(card, text=f"{icon} {lbl}", font=F_SMALL, bg=C_SURFACE2, fg=C_MUTED).pack(anchor="w")
            v = tk.StringVar(value="—"); self.sv[key] = v
            tk.Label(card, textvariable=v, font=(FM.primary,16,"bold"), bg=C_SURFACE2, fg=col).pack(anchor="w")

        # legend
        lg = tk.Frame(self, bg=C_BG, padx=14); lg.pack(fill=tk.X)
        for txt,c in [("● ปกติ",C_GREEN),("● ใกล้หมด",C_YELLOW),("● ต่ำกว่าขั้นต่ำ/หมด","#ff6b6b")]:
            tk.Label(lg, text=txt, font=(FM.primary,9), bg=C_BG, fg=c).pack(side=tk.LEFT, padx=(0,10))
        tk.Label(lg, text="(ดับเบิลคลิกสินค้า = ดูว่าขายในบิลไหนบ้าง)", font=(FM.primary,9),
                 bg=C_BG, fg=C_MUTED).pack(side=tk.RIGHT)

        # table
        cols = ("รหัส","หมวด","ชื่อสินค้า","ขาย","หน่วย","ยอดขาย","ต้นทุน","กำไร","คงเหลือ","ขายล่าสุด")
        frm, self.tree = make_tree(self, cols, (95,85,200,52,52,90,90,90,72,115), height=18)
        frm.pack(fill=tk.BOTH, expand=True, padx=12, pady=(2,8))
        self.tree.tag_configure("st_low",  background="#3a1414")   # ต่ำ/หมด
        self.tree.tag_configure("st_near", background="#33300f")   # ใกล้หมด
        self.tree.bind("<Double-1>", self._drilldown)

    def _set_dates(self, fr, to):
        from datetime import timedelta
        today = date.today()
        d_from = today.replace(day=1) if fr=="month" else today+timedelta(days=fr)
        d_to = today+timedelta(days=to) if isinstance(to,int) else today
        self.df.set(d_from.strftime("%Y-%m-%d")); self.dt.set(d_to.strftime("%Y-%m-%d"))
        self.load()

    def load(self, *_):
        d0 = self.df.get_iso(); d1 = self.dt.get_iso()
        try:
            rows = get_product_sales_report(
                d0, d1,
                customer=None,
                payment=self._PAY.get(self.pay_var.get()),
                category=self.cat_var.get(),
                search=self.search_var.get().strip(),
                sort=self._SORT.get(self.sort_var.get(), "qty"),
                view=self._VIEW.get(self.view_var.get(), "all"))
        except Exception as e:
            print(f"[product sales] {e}"); rows = []

        self._rows = {r["pid"]: r for r in rows}
        n = len(rows)
        q = sum(r["qty"] or 0 for r in rows)
        s = sum(r["sales"] or 0 for r in rows)
        p = sum(r["profit"] or 0 for r in rows)
        self.sv["n"].set(f"{n:,} รายการ"); self.sv["q"].set(f"{q:,.0f} ชิ้น")
        self.sv["s"].set(f"฿{s:,.2f}");    self.sv["p"].set(f"฿{p:,.2f}")

        self.tree.delete(*self.tree.get_children())
        for i, r in enumerate(rows):
            stock = r["stock"] or 0; mn = r["min_stock"] or 0
            if stock <= 0 or stock <= mn:      tag = "st_low"
            elif stock <= mn*1.5:              tag = "st_near"
            else:                              tag = "alt" if i%2==0 else ""
            last = to_thai_dt(r["last_sold"]) if r.get("last_sold") else "—"
            self.tree.insert("","end", iid=str(r["pid"]), tags=(tag,),
                values=(r["code"] or "—", r["category"] or "—", r["name"] or "—",
                        f"{r['qty'] or 0:,.0f}", r["unit"] or "—",
                        f"฿{r['sales'] or 0:,.2f}", f"฿{r['cost'] or 0:,.2f}",
                        f"฿{r['profit'] or 0:,.2f}", f"{stock:,.0f}", last))
        if not rows:
            self.tree.insert("","end", values=("—","(ไม่มีการขายในช่วงวันที่นี้)","","","","","","","",""))

    def _drilldown(self, _=None):
        sel = self.tree.selection()
        if not sel: return
        try: pid = int(sel[0])
        except Exception: return
        r = self._rows.get(pid)
        if not r: return
        d0 = self.df.get_iso(); d1 = self.dt.get_iso()
        try: bills = get_product_sale_detail(d0, d1, pid)
        except Exception as e:
            messagebox.showerror("ผิดพลาด", str(e), parent=self.winfo_toplevel()); return

        win = tk.Toplevel(self); win.title(f"รายละเอียดการขาย: {r['name']}")
        win.configure(bg=C_BG); win.geometry("680x520"); win.grab_set(); win.lift()
        win.bind("<Escape>", lambda e: win.destroy())
        tk.Frame(win, bg=C_ACCENT, height=3).pack(fill=tk.X)
        tk.Label(win, text=f"📦 {r['name']}", font=F_H2, bg=C_BG, fg=C_TEXT).pack(anchor="w", padx=16, pady=(10,2))
        tk.Label(win, text=f"รหัส {r['code'] or '—'}  •  ขายรวม {r['qty'] or 0:,.0f} {r['unit'] or ''}  "
                           f"•  ยอด ฿{r['sales'] or 0:,.2f}  •  คงเหลือ {r['stock'] or 0:,.0f}",
                 font=F_SMALL, bg=C_BG, fg=C_MUTED).pack(anchor="w", padx=16, pady=(0,8))
        cols = ("เลขบิล","เวลา","ลูกค้า","ช่องทาง","จำนวน","ราคา/หน่วย","รวม")
        frm, tv = make_tree(win, cols, (70,150,150,90,55,85,90), height=14)
        frm.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0,8))
        pay_th = {v:k for k,v in self._PAY.items()}
        for i, b in enumerate(bills):
            tv.insert("","end", tags=("alt" if i%2==0 else "",),
                values=(b["sale_id"], to_thai_dt(b["sale_date"]), b["cust"] or "—",
                        pay_th.get(b["payment_method"], b["payment_method"] or "—"),
                        f"{b['qty'] or 0:,.0f}", f"฿{b['price'] or 0:,.2f}",
                        f"฿{b['subtotal'] or 0:,.2f}"))
        if not bills:
            tv.insert("","end", values=("—","(ไม่พบบิล)","","","","",""))
        pill_btn(win, "ปิด (Esc)", win.destroy, bg=C_SURFACE, fg=C_MUTED,
                 pad_x=14, pad_y=8).pack(pady=(0,10))


# ══ TAB: DASHBOARD ═══════════════════════════════════════
class DashboardTab(tk.Frame):
    def __init__(self,parent):
        super().__init__(parent,bg=C_BG)
        self._build(); self.load()

    def _build(self):
        # Top row: date filter + refresh
        top=tk.Frame(self,bg=C_BG,pady=10,padx=14); top.pack(fill=tk.X)
        tk.Label(top,text="Dashboard",font=F_H1,bg=C_BG,fg=C_TEXT).pack(side=tk.LEFT)
        rf=tk.Frame(top,bg=C_BG); rf.pack(side=tk.RIGHT)
        for txt,cmd in [("วันนี้",self._today),("7 วัน",self._week),("30 วัน",self._month),("รีเฟรช",self.load)]:
            pill_btn(rf,txt,cmd,bg=C_SURFACE2,fg=C_TEXT).pack(side=tk.LEFT,padx=2)
        self.period_lbl=tk.Label(top,text="",font=F_SMALL,bg=C_BG,fg=C_MUTED)
        self.period_lbl.pack(side=tk.RIGHT,padx=12)
        self._days=30

        # KPI row
        kpi=tk.Frame(self,bg=C_BG,padx=14); kpi.pack(fill=tk.X,pady=(0,8))
        self.kv=[]
        for lbl,icon,col in [("ยอดขายวันนี้","💰",C_ACCENT),("บิลวันนี้","🧾",C_BLUE),
                               ("กำไรวันนี้","📈",C_GREEN),("สต็อกต่ำ","⚠️",C_YELLOW),
                               ("สต็อกติดลบ","📉","#ff2222")]:
            f=tk.Frame(kpi,bg=C_SURFACE,highlightbackground=C_BORDER,highlightthickness=1,padx=14,pady=10)
            f.pack(side=tk.LEFT,fill=tk.X,expand=True,padx=4)
            tk.Label(f,text=icon+" "+lbl,font=F_SMALL,bg=C_SURFACE,fg=C_MUTED).pack(anchor="w")
            v=tk.StringVar(value="—")
            lbl_w=tk.Label(f,textvariable=v,font=F_NUM_SM,bg=C_SURFACE,fg=col)
            lbl_w.pack(anchor="w",pady=(4,0))
            self.kv.append(v)

        # Charts row
        cf=tk.Frame(self,bg=C_BG,padx=14); cf.pack(fill=tk.BOTH,expand=True,pady=(0,8))

        # Left: bar chart - daily sales
        lf=tk.Frame(cf,bg=C_SURFACE,highlightbackground=C_BORDER,highlightthickness=1)
        lf.pack(side=tk.LEFT,fill=tk.BOTH,expand=True,padx=(0,6))
        tk.Label(lf,text="ยอดขายรายวัน",font=F_H2,bg=C_SURFACE,fg=C_TEXT,pady=8).pack(anchor="w",padx=12)
        tk.Frame(lf,bg=C_ACCENT,height=2).pack(fill=tk.X)
        self.bar_canvas=tk.Canvas(lf,bg=C_SURFACE,highlightthickness=0,height=220)
        self.bar_canvas.pack(fill=tk.BOTH,expand=True,padx=4,pady=8)

        # Right: top products list + donut
        rf2=tk.Frame(cf,bg=C_BG,width=260); rf2.pack(side=tk.RIGHT,fill=tk.Y)
        rf2.pack_propagate(False)

        # Top products
        tf=tk.Frame(rf2,bg=C_SURFACE,highlightbackground=C_BORDER,highlightthickness=1)
        tf.pack(fill=tk.BOTH,expand=True,pady=(0,6))
        tk.Label(tf,text="สินค้าขายดี Top 5",font=F_H2,bg=C_SURFACE,fg=C_TEXT,pady=8).pack(anchor="w",padx=12)
        tk.Frame(tf,bg=C_BLUE,height=2).pack(fill=tk.X)
        self.top_frame=tk.Frame(tf,bg=C_SURFACE); self.top_frame.pack(fill=tk.BOTH,expand=True,padx=10,pady=6)

        # Low stock
        sf2=tk.Frame(rf2,bg=C_SURFACE,highlightbackground=C_BORDER,highlightthickness=1)
        sf2.pack(fill=tk.X)
        tk.Label(sf2,text="⚠️  สต็อกใกล้หมด",font=F_H2,bg=C_SURFACE,fg=C_TEXT,pady=8).pack(anchor="w",padx=12)
        tk.Frame(sf2,bg=C_YELLOW,height=2).pack(fill=tk.X)
        self.stock_frame=tk.Frame(sf2,bg=C_SURFACE); self.stock_frame.pack(fill=tk.BOTH,padx=10,pady=6)

        # Negative stock card
        nf=tk.Frame(rf2,bg=C_SURFACE,highlightbackground="#ff2222",highlightthickness=1)
        nf.pack(fill=tk.X,pady=(6,0))
        neg_hdr=tk.Frame(nf,bg=C_SURFACE); neg_hdr.pack(fill=tk.X)
        tk.Label(neg_hdr,text="📉  สต็อกติดลบ",font=F_H2,bg=C_SURFACE,fg="#ff2222",pady=8).pack(side=tk.LEFT,padx=12)
        self.neg_count_lbl=tk.Label(neg_hdr,text="",font=F_SMALL,bg=C_SURFACE,fg="#ff2222")
        self.neg_count_lbl.pack(side=tk.RIGHT,padx=12)
        tk.Frame(nf,bg="#ff2222",height=2).pack(fill=tk.X)
        self.neg_frame=tk.Frame(nf,bg=C_SURFACE); self.neg_frame.pack(fill=tk.BOTH,padx=10,pady=6)

    def _today(self): self._days=1; self.load()
    def _week(self):  self._days=7; self.load()
    def _month(self): self._days=30; self.load()

    def load(self,*_):
        from datetime import timedelta
        today=date.today()
        d0=(today-timedelta(days=self._days-1)).strftime("%Y-%m-%d")
        d1=today.strftime("%Y-%m-%d")
        lbl_map={1:"วันนี้",7:"7 วันล่าสุด",30:"30 วันล่าสุด"}
        self.period_lbl.config(text=lbl_map.get(self._days,f"{self._days} วัน")+"  "+d0+" ถึง "+d1)

        # KPI today — db_sale Session 8
        kpi=get_today_sales_kpi(d1)
        t_sales =kpi["total_net"]
        t_bills =kpi["total_bills"]
        t_profit=kpi["total_net"]-kpi["total_cost"]
        low_cnt=count_low_stock()
        neg_cnt=count_negative_stock()

        self.kv[0].set(f"฿{t_sales:,.0f}")
        self.kv[1].set(str(t_bills)+" บิล")
        self.kv[2].set(f"฿{t_profit:,.0f}")
        self.kv[3].set(str(low_cnt)+" รายการ")
        self.kv[4].set(str(neg_cnt)+" รายการ" if neg_cnt>0 else "ปกติ")

        # Daily sales for bar chart — db_sale Session 8
        daily=get_daily_sales_range(d0,d1)

        # Top 5 products — db_sale Session 8
        top5=get_top_products(d0,d1,limit=5)

        # Low stock / Negative stock
        lows=get_low_stock_products(limit=6)
        negs=get_negative_stock_products(limit=10)

        self._draw_bars(daily)
        self._draw_top5(top5)
        self._draw_low(lows)
        self._draw_neg(negs)

    def _draw_neg(self,negs):
        """วาดรายการสินค้าสต็อกติดลบ"""
        for w in self.neg_frame.winfo_children(): w.destroy()
        # อัปเดต count label
        cnt=len(negs)
        self.neg_count_lbl.config(
            text=f"{cnt} รายการ" if cnt>0 else "ปกติ ✅",
            fg="#ff2222" if cnt>0 else C_GREEN)
        if not negs:
            tk.Label(self.neg_frame,text="✅  ไม่มีสต็อกติดลบ",
                     font=F_SMALL,bg=C_SURFACE,fg=C_GREEN).pack(pady=4)
            return
        # Header
        hdr=tk.Frame(self.neg_frame,bg=C_SURFACE); hdr.pack(fill=tk.X,pady=(0,2))
        tk.Label(hdr,text="ชื่อสินค้า",font=(FM.primary,10),
                 bg=C_SURFACE,fg=C_MUTED,width=22,anchor="w").pack(side=tk.LEFT)
        tk.Label(hdr,text="คงเหลือ",font=(FM.primary,10),
                 bg=C_SURFACE,fg=C_MUTED,width=8,anchor="e").pack(side=tk.RIGHT)
        # Rows
        for i,r in enumerate(negs):
            row_bg="#1a1520" if i%2==0 else C_SURFACE
            row=tk.Frame(self.neg_frame,bg=row_bg); row.pack(fill=tk.X,pady=1)
            tk.Label(row,text="🔴",font=(FM.primary,9),
                     bg=row_bg,fg="#ff2222").pack(side=tk.LEFT,padx=(0,4))
            tk.Label(row,text=r["name"][:22],font=(FM.primary,11),
                     bg=row_bg,fg="#ffcccc",anchor="w").pack(side=tk.LEFT,fill=tk.X,expand=True)
            tk.Label(row,text=str(r["stock"]),font=(FM.primary,11,"bold"),
                     bg=row_bg,fg="#ff2222",anchor="e",width=6).pack(side=tk.RIGHT,padx=4)

    def _has_min_stock(self,conn):
        try: conn.execute("SELECT min_stock FROM products LIMIT 1"); return True
        except: return False

    def _draw_bars(self,daily):
        c=self.bar_canvas; c.delete("all")
        c.update_idletasks()
        W=c.winfo_width() or 400; H=c.winfo_height() or 220
        if not daily:
            c.create_text(W//2,H//2,text="ยังไม่มีข้อมูลในช่วงนี้",fill=C_MUTED,font=F_BODY); return
        vals=[r["total"] for r in daily]
        mx=max(vals) if vals else 1
        n=len(daily)
        pad_l,pad_r,pad_t,pad_b=40,10,20,40
        bw_total=(W-pad_l-pad_r)
        bw=max(4,bw_total//n-4)
        gap=max(2,(bw_total-bw*n)//(n+1))
        # grid lines
        for gi in range(4):
            y=pad_t+(H-pad_t-pad_b)*gi//3
            gv=mx*(3-gi)//3
            c.create_line(pad_l,y,W-pad_r,y,fill=C_BORDER,dash=(4,4))
            c.create_text(pad_l-4,y,text=f"{gv/1000:.0f}k" if gv>=1000 else str(int(gv)),
                          anchor="e",fill=C_MUTED,font=(FM.primary,10))
        # bars
        colors=[C_ACCENT,"#00b8a9","#00c9b1","#1D9E75","#0F6E56"]
        for i,r in enumerate(daily):
            x=pad_l+gap+i*(bw+gap)
            h_bar=max(2,int((r["total"]/mx)*(H-pad_t-pad_b)))
            y1=H-pad_b-h_bar; y2=H-pad_b
            col=colors[i%len(colors)]
            c.create_rectangle(x,y1,x+bw,y2,fill=col,outline="",width=0)
            # date label
            d=r["d"][5:]  # MM-DD
            c.create_text(x+bw//2,H-pad_b+6,text=d,fill=C_MUTED,font=(FM.primary,8),anchor="n")
            # value on top
            if h_bar>18:
                c.create_text(x+bw//2,y1-3,text=f"{r['total']:,.0f}",
                              fill=C_TEXT,font=(FM.primary,8),anchor="s")

    def _draw_top5(self,top5):
        for w in self.top_frame.winfo_children(): w.destroy()
        if not top5:
            tk.Label(self.top_frame,text="ยังไม่มีข้อมูล",font=F_SMALL,bg=C_SURFACE,fg=C_MUTED).pack()
            return
        max_qty=top5[0]["tq"] if top5 else 1
        colors=[C_ACCENT,C_BLUE,C_GREEN,C_YELLOW,ORANGE]
        for i,r in enumerate(top5):
            row=tk.Frame(self.top_frame,bg=C_SURFACE); row.pack(fill=tk.X,pady=3)
            tk.Label(row,text=str(i+1),font=(FM.primary,11,"bold"),bg=C_SURFACE,
                     fg=colors[i%len(colors)],width=2).pack(side=tk.LEFT)
            info=tk.Frame(row,bg=C_SURFACE); info.pack(side=tk.LEFT,fill=tk.X,expand=True,padx=6)
            tk.Label(info,text=r["name"][:22],font=F_SMALL,bg=C_SURFACE,fg=C_TEXT,anchor="w").pack(fill=tk.X)
            # bar
            bar_f=tk.Frame(info,bg=C_SURFACE2,height=4); bar_f.pack(fill=tk.X,pady=(2,0))
            pct=int(r["tq"]/max_qty*100)
            tk.Frame(bar_f,bg=colors[i%len(colors)],height=4,width=max(2,pct*2)).pack(side=tk.LEFT)
            tk.Label(row,text=str(r["tq"]),font=(FM.primary,11,"bold"),bg=C_SURFACE,fg=colors[i%len(colors)]).pack(side=tk.RIGHT)

    def _draw_low(self,lows):
        for w in self.stock_frame.winfo_children(): w.destroy()
        if not lows:
            tk.Label(self.stock_frame,text="สต็อกปกติทุกรายการ ✓",font=F_SMALL,bg=C_SURFACE,fg=C_GREEN).pack()
            return
        for r in lows:
            row=tk.Frame(self.stock_frame,bg=C_SURFACE); row.pack(fill=tk.X,pady=2)
            col=C_ACCENT2 if r["stock"]==0 else C_YELLOW
            tk.Label(row,text=r["name"][:20],font=F_SMALL,bg=C_SURFACE,fg=C_TEXT,anchor="w").pack(side=tk.LEFT)
            tk.Label(row,text=str(r["stock"]),font=(FM.primary,11,"bold"),bg=C_SURFACE,fg=col).pack(side=tk.RIGHT)


# ══ TAB: REPORT ══════════════════════════════════════════
class ReportTab(tk.Frame):
    def __init__(self,parent):
        super().__init__(parent,bg=C_BG); self._build(); self.load()

    def _build(self):
        top=tk.Frame(self,bg=C_BG,pady=10,padx=12); top.pack(fill=tk.X)
        tk.Label(top,text="📊  รายงานการขาย",font=F_H1,bg=C_BG,fg=C_TEXT).pack(side=tk.LEFT)
        ef=tk.Frame(top,bg=C_BG); ef.pack(side=tk.RIGHT,padx=(0,10))
        pill_btn(ef,"📥 Excel",self._exp_excel,bg="#1D9E75",fg="#ffffff").pack(side=tk.LEFT,padx=2)
        pill_btn(ef,"📄 CSV",self._exp_csv,bg=C_SURFACE2,fg=C_TEXT).pack(side=tk.LEFT,padx=2)
        ctrl=tk.Frame(top,bg=C_BG); ctrl.pack(side=tk.RIGHT)
        # Quick date buttons
        qf=tk.Frame(ctrl,bg=C_BG); qf.pack(side=tk.LEFT,padx=(0,6))
        for txt,cmd in [("📅 วันนี้",lambda:self._set_dates(0,0)),
                        ("⏳ เมื่อวาน",lambda:self._set_dates(-1,-1)),
                        ("📆 7วัน",lambda:self._set_dates(-6,0)),
                        ("📋 เดือนนี้",lambda:self._set_dates("month",0))]:
            pill_btn(qf,txt,cmd,bg=C_SURFACE2,fg=C_BLUE,pad_x=7,pad_y=3).pack(side=tk.LEFT,padx=2)
        tk.Label(ctrl,text="จาก",font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(side=tk.LEFT,padx=(4,2))
        self.df=ThaiDateEntry(ctrl,width=11); self.df.pack(side=tk.LEFT,padx=(0,6))
        tk.Label(ctrl,text="ถึง",font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(side=tk.LEFT,padx=(6,2))
        self.dt=ThaiDateEntry(ctrl,width=11); self.dt.pack(side=tk.LEFT,padx=(0,4))
        pill_btn(ctrl,"🔍 ดูรายงาน",self.load,bg=C_ACCENT,fg=C_BG,pad_x=10,pad_y=3).pack(side=tk.LEFT,padx=(6,0))

        self._set_dates(0,0)  # init today

        # ── Filter bar ─────────────────────────────────────
        fb = tk.Frame(self, bg=C_BG, padx=12); fb.pack(fill=tk.X, pady=(0,4))
        tk.Label(fb, text="ลูกค้า:", font=F_SMALL, bg=C_BG, fg=C_MUTED).pack(side=tk.LEFT)
        self.cust_var = tk.StringVar(value="ทั้งหมด")
        self.cust_cb  = ttk.Combobox(fb, textvariable=self.cust_var,
                                     values=["ทั้งหมด"], font=F_BODY,
                                     width=18, state="readonly")
        self.cust_cb.pack(side=tk.LEFT, padx=(4,12))

        tk.Label(fb, text="ช่องทางชำระ:", font=F_SMALL, bg=C_BG, fg=C_MUTED).pack(side=tk.LEFT)
        self.pay_var = tk.StringVar(value="ทั้งหมด")
        pay_opts = ["ทั้งหมด","เงินสด","QR/พร้อมเพย์","บัตรเครดิต","เงินเชื่อ","ผสม"]
        ttk.Combobox(fb, textvariable=self.pay_var,
                     values=pay_opts, font=F_BODY,
                     width=14, state="readonly").pack(side=tk.LEFT, padx=(4,12))

        tk.Label(fb, text="ค้นหาสินค้า:", font=F_SMALL, bg=C_BG, fg=C_MUTED).pack(side=tk.LEFT)
        self.prod_sv = tk.StringVar()
        field(fb, width=16).configure(textvariable=self.prod_sv)
        field(fb, width=16).pack(side=tk.LEFT, padx=(4,0), ipady=4)

        pill_btn(fb, "✖ ล้าง filter", self._clear_filters,
                 bg=C_SURFACE2, fg=C_MUTED, pad_x=8, pad_y=3).pack(side=tk.RIGHT)

        # stat cards row
        cf=tk.Frame(self,bg=C_BG,padx=12); cf.pack(fill=tk.X,pady=(0,8))
        self.sv_list=[]
        for lbl,icon,color in [("บิล","🧾",C_TEXT),("ยอดขาย","💰",C_ACCENT),("ต้นทุน","📦",C_MUTED),("กำไร","📈",C_GREEN),("ส่วนลด","🏷️",C_YELLOW)]:
            v=tk.StringVar(value="0" if lbl=="บิล" else "฿0")
            stat_card(cf,lbl,v,icon,color); self.sv_list.append(v)

        nb=ttk.Notebook(self,style="R.TNotebook"); nb.pack(fill=tk.BOTH,expand=True,padx=12,pady=(0,4))
        # Tab 1: รายการบิล (เพิ่ม ต้นทุน กำไร อัตรากำไร%)
        bf=tk.Frame(nb,bg=C_BG); nb.add(bf,text="  รายการบิล  ")
        cols=("เลขบิล","วันเวลา","ลูกค้า","ช่องทาง","ยอดสุทธิ","ต้นทุน","กำไร","กำไร%")
        frm,self.bt=make_tree(bf,cols,(55,135,100,80,80,80,75,65))
        frm.pack(fill=tk.BOTH,expand=True)
        self.bt.tag_configure("profit_high",foreground="#3fb950")
        self.bt.tag_configure("profit_low", foreground="#e67e22")
        self.bt.tag_configure("loss",       foreground="#ff4444")
        self.bt.bind("<<TreeviewSelect>>",self._show_items)
        self.bt.bind("<Double-1>", self._open_bill_detail)
        self.bt.bind("<Button-3>", self._show_bill_menu)
        self._build_bill_menu()
        # Tab 2: สินค้าขายดี
        tf=tk.Frame(nb,bg=C_BG); nb.add(tf,text="  สินค้าขายดี  ")
        frm2,self.tt=make_tree(tf,("อันดับ","ชื่อสินค้า","จำนวนขาย","ยอดขาย","ต้นทุน","กำไร","กำไร%"),(40,200,80,90,80,75,65))
        frm2.pack(fill=tk.BOTH,expand=True)
        # Tab 3: สรุปกำไรรายวัน
        pf=tk.Frame(nb,bg=C_BG); nb.add(pf,text="  กำไรรายวัน  ")
        frm3,self.pt=make_tree(pf,("วันที่","บิล","ยอดขาย","ต้นทุน","กำไร","กำไร%"),(110,50,100,100,90,70))
        frm3.pack(fill=tk.BOTH,expand=True)
        self.pt.tag_configure("profit_high",foreground="#3fb950")

        tk.Frame(self,bg=C_BORDER,height=1).pack(fill=tk.X,padx=12)
        tk.Label(self,text="รายการในบิลที่เลือก",font=F_SMALL,bg=C_BG,fg=C_MUTED,pady=6).pack(anchor="w",padx=12)
        frm3,self.it=make_tree(self,("ชื่อสินค้า","ราคา","จำนวน","รวม"),(250,100,80,100),height=4)
        frm3.pack(fill=tk.X,padx=12,pady=(0,10))

    # _pick_date ถูกแทนด้วย show_thai_calendar (shared helper ด้านบน)

    def _to_thai(self,ymd):   return _iso_to_th(ymd)
    def _from_thai(self,thai): return _th_to_iso(thai)

    def _set_dates(self,fr,to):
        from datetime import timedelta
        today=date.today()
        d_from=today.replace(day=1) if fr=="month" else today+timedelta(days=fr)
        d_to=today+timedelta(days=to) if isinstance(to,int) else today
        self.df.set(d_from.strftime("%Y-%m-%d"))
        self.dt.set(d_to.strftime("%Y-%m-%d"))
        try: self.load()
        except: pass

    def _clear_filters(self):
        self.cust_var.set("ทั้งหมด")
        self.pay_var.set("ทั้งหมด")
        self.prod_sv.set("")
        self.load()

    def _refresh_cust_list(self):
        """โหลด dropdown ลูกค้าจาก DB — db_sale Session 9"""
        names = ["ทั้งหมด"] + get_customer_names()
        self.cust_cb["values"] = names

    def load(self,*_):
        d0 = self.df.get_iso()
        d1 = self.dt.get_iso()

        cust_filter = self.cust_var.get() if hasattr(self,"cust_var") else "ทั้งหมด"
        pay_filter  = self.pay_var.get()  if hasattr(self,"pay_var")  else "ทั้งหมด"
        prod_filter = self.prod_sv.get().strip() if hasattr(self,"prod_sv") else ""

        # แปลง pay_filter UI → payment_method ใน DB
        pay_map = {"เงินสด":"cash","QR/พร้อมเพย์":"qr","บัตรเครดิต":"credit",
                   "เงินเชื่อ":"debt","ผสม":"mixed"}
        pay_db = pay_map.get(pay_filter) if pay_filter != "ทั้งหมด" else None

        # โหลด dropdown ลูกค้า (ครั้งแรก) — db_sale Session 9
        if hasattr(self,"cust_cb") and len(self.cust_cb["values"]) <= 1:
            self._refresh_cust_list()

        # ── Bill list + cost map — db_sale Session 9 ─────────
        sales, cost_map = get_report_bills(d0, d1, cust_filter, pay_db)

        # ── filter สินค้าในบิล — db_sale Session 9 ───────────
        if prod_filter:
            sale_ids_with_prod = get_report_bill_ids_by_product(
                d0, d1, prod_filter, cust_filter, pay_db
            )
            sales = [s for s in sales if s["id"] in sale_ids_with_prod]

        # ── Top products — db_sale Session 9 ─────────────────
        top = get_report_top_products(d0, d1, cust_filter, pay_db, limit=20)

        # ── Daily summary — db_sale Session 9 ────────────────
        daily, daily_cost = get_report_daily_summary(d0, d1, cust_filter, pay_db)

        # ── KPI cards (ไม่นับบิลที่ยกเลิก) ────────────────
        _active = [s for s in sales if (s.get("status") or "normal") != "void"]
        income = sum(s["net"] for s in _active)
        cost   = sum(cost_map.get(s["id"], 0) for s in _active)
        profit = income - cost
        disc   = sum((s["discount"] or 0)+(s["promo_disc"] or 0) for s in _active)
        for v, val in zip(self.sv_list,
            [str(len(_active)), f"฿{income:,.2f}", f"฿{cost:,.2f}",
             f"฿{profit:,.2f}", f"฿{disc:,.2f}"]):
            v.set(val)

        # ── Bill tree ────────────────────────────────────
        self.bt.delete(*self.bt.get_children())
        try:
            _cn=get_db()
            _ret_ids={r[0] for r in _cn.execute("SELECT DISTINCT sale_id FROM returns").fetchall()}
            _cn.close()
        except Exception:
            _ret_ids=set()
        for i, s in enumerate(sales):
            c_bill  = cost_map.get(s["id"], 0)
            p_bill  = s["net"] - c_bill
            pct     = (p_bill/s["net"]*100) if s["net"] > 0 else 0
            voided  = (s.get("status") or "normal") == "void"
            tag     = "loss" if voided else ("profit_high" if pct >= 20 else ("profit_low" if pct >= 0 else "loss"))
            pm      = {"cash":"เงินสด","qr":"QR","credit":"บัตร",
                       "debt":"เชื่อ","mixed":"ผสม"}.get(s["payment_method"] or "cash","—")
            bill_no = (f"✖ {s['id']}" if voided else
                       (f"↩ {s['id']}" if s["id"] in _ret_ids else s["id"]))
            self.bt.insert("","end",iid=s["id"],
                tags=(tag,"alt" if i%2==0 else ""),
                values=(bill_no, to_thai_dt(s["sale_date"]), s["cn"] or "—", pm,
                        f"฿{s['net']:,.2f}", f"฿{c_bill:,.2f}",
                        f"฿{p_bill:,.2f}", f"{pct:.1f}%"))

        # ── Top products tree ────────────────────────────
        self.tt.delete(*self.tt.get_children())
        for i, r in enumerate(top, 1):
            tc   = r["tc"] or 0
            tp   = (r["ts"] or 0) - tc
            pct  = (tp/r["ts"]*100) if r["ts"] and r["ts"] > 0 else 0
            self.tt.insert("","end",
                tags=("alt" if i%2==0 else "",),
                values=(i, r["name"], r["tq"],
                        f"฿{r['ts']:,.2f}", f"฿{tc:,.2f}",
                        f"฿{tp:,.2f}", f"{pct:.1f}%"))

        # ── Daily profit tree ────────────────────────────
        self.pt.delete(*self.pt.get_children())
        for i, r in enumerate(daily):
            tc  = daily_cost.get(r["d"], 0)
            net = r["net"] or 0
            tp  = net - tc
            pct = (tp/net*100) if net > 0 else 0
            tag = "profit_high" if pct >= 20 else "alt" if i%2==0 else ""
            d_th = to_thai_date(r["d"])
            self.pt.insert("","end", tags=(tag,),
                values=(d_th, r["bills"],
                        f"฿{net:,.2f}", f"฿{tc:,.2f}",
                        f"฿{tp:,.2f}", f"{pct:.1f}%"))

    def _show_items(self,event=None):
        s=self.bt.selection()
        if not s: return
        items=get_sale_items(int(s[0]))  # db_sale — Session 8
        self.it.delete(*self.it.get_children())
        for it in items: self.it.insert("","end",values=(it["name"],f"฿{it['price']:.2f}",it["qty"],f"฿{it['subtotal']:.2f}"))

    # ════ รายละเอียดบิล (Double-click) ════════════════════
    def _open_bill_detail(self, event=None):
        if event is not None and hasattr(event, "y"):
            iid = self.bt.identify_row(event.y)
            if iid: self.bt.selection_set(iid)
        sel = self.bt.selection()
        if not sel: return
        try: sid = int(sel[0])
        except Exception: return
        self._bill_detail_window(sid)

    def _bill_detail_window(self, sid):
        try:
            self._build_bill_detail(sid)
        except Exception as _e:
            import traceback
            messagebox.showerror("รายละเอียดบิล — เกิดข้อผิดพลาด",
                f"{_e}\n\n{traceback.format_exc()[-1600:]}",
                parent=self.winfo_toplevel())

    def _build_bill_detail(self, sid):
        try:
            sale = get_sale_with_customer(sid)
            items = [dict(it) for it in get_sale_items(sid)]
        except Exception as e:
            messagebox.showerror("ผิดพลาด", f"โหลดบิลไม่ได้: {e}", parent=self.winfo_toplevel()); return
        if not sale:
            messagebox.showwarning("ไม่พบบิล", f"ไม่พบบิล #{sid}", parent=self.winfo_toplevel()); return

        # สถานะจากตาราง returns (ของจริง)
        total_qty = sum(int(it.get("qty",0) or 0) for it in items)
        ret_qty = 0
        try:
            _c = get_db()
            ret_qty = _c.execute("SELECT COALESCE(SUM(qty),0) FROM returns WHERE sale_id=?", (sid,)).fetchone()[0] or 0
            _c.close()
        except Exception: ret_qty = 0
        if ret_qty <= 0:            status_txt, status_fg = "● ปกติ", "#3fb950"
        elif ret_qty >= total_qty:  status_txt, status_fg = "↩ คืนสินค้า (ทั้งบิล)", "#e67e22"
        else:                       status_txt, status_fg = "↩ คืนสินค้า (บางส่วน)", "#e67e22"
        if (sale.get("status") or "normal") == "void":
            status_txt, status_fg = "✖ ยกเลิกแล้ว", "#ff4444"

        # สรุป
        cost_total = sum((it.get("cost",0) or 0)*(it.get("qty",0) or 0) for it in items)
        net = sale["net"] or 0
        profit = net - cost_total
        disc_total = (sale["discount"] or 0) + (sale["promo_disc"] or 0)
        vat_txt = "—"
        try:
            if get_setting("vat_enabled") == "1":
                rate = float(get_setting("vat_percent") or get_setting("vat_rate") or 0)
                if rate > 0:
                    vat_txt = f"฿{net*rate/(100+rate):,.2f} (รวมใน {rate:g}%)"
        except Exception: vat_txt = "—"
        pm = {"cash":"เงินสด","qr":"QR/พร้อมเพย์","credit":"บัตรเครดิต",
              "debt":"เงินเชื่อ","mixed":"ผสม"}.get(sale["payment_method"] or "cash","—")

        win = tk.Toplevel(self); win.title(f"รายละเอียดบิลขาย #{sid}")
        win.configure(bg=C_BG); win.geometry("760x680"); win.grab_set(); win.lift()
        win.bind("<Escape>", lambda e: win.destroy())
        win.bind("<Return>", lambda e: win.destroy())
        tk.Frame(win, bg=C_ACCENT, height=3).pack(fill=tk.X)

        hf = tk.Frame(win, bg=C_BG, padx=16, pady=10); hf.pack(fill=tk.X)
        tk.Label(hf, text=f"🧾 บิลขาย #{sid}", font=F_H1, bg=C_BG, fg=C_TEXT).pack(side=tk.LEFT)
        tk.Label(hf, text=status_txt, font=(FM.primary,12,"bold"), bg=C_BG, fg=status_fg).pack(side=tk.RIGHT)

        info = tk.Frame(win, bg=C_SURFACE2, padx=14, pady=10); info.pack(fill=tk.X, padx=12)
        for i,(k,v) in enumerate([
                ("วันที่เวลา", to_thai_dt(sale["sale_date"])),
                ("ลูกค้า", sale["cn"] or "—"),
                ("ช่องทางชำระ", pm),
                ("ยอดรวม", f"฿{(sale['total'] or 0):,.2f}"),
                ("ส่วนลด", f"฿{disc_total:,.2f}"),
                ("VAT", vat_txt),
                ("กำไร", f"฿{profit:,.2f}"),
                ("ยอดสุทธิ", f"฿{net:,.2f}"),
                ("หมายเหตุ", sale["note"] or "—")]):
            cell = tk.Frame(info, bg=C_SURFACE2); cell.grid(row=i//2, column=i%2, sticky="w", padx=(0,30), pady=2)
            tk.Label(cell, text=k+":", font=F_SMALL, bg=C_SURFACE2, fg=C_MUTED, width=12, anchor="w").pack(side=tk.LEFT)
            tk.Label(cell, text=str(v), font=F_BODY, bg=C_SURFACE2, fg=C_TEXT).pack(side=tk.LEFT)
        if (sale.get("status") or "normal") == "void":
            vinfo = tk.Frame(win, bg="#3a1414", padx=14, pady=6); vinfo.pack(fill=tk.X, padx=12, pady=(4,0))
            tk.Label(vinfo, text=f"✖ ยกเลิกโดย {sale.get('void_by') or '—'}  เมื่อ {sale.get('void_at') or '—'}  |  เหตุผล: {sale.get('void_reason') or '—'}",
                     font=F_SMALL, bg="#3a1414", fg="#ff8888", anchor="w").pack(anchor="w")

        sf = tk.Frame(win, bg=C_BG, padx=12); sf.pack(fill=tk.X, pady=(8,2))
        tk.Label(sf, text="🔍 ค้นหาสินค้าในบิล:", font=F_SMALL, bg=C_BG, fg=C_MUTED).pack(side=tk.LEFT)
        qv = tk.StringVar()
        qe = tk.Entry(sf, textvariable=qv, width=26, bg=C_SURFACE2, fg=C_TEXT,
                      insertbackground=C_TEXT, relief=tk.FLAT, font=F_BODY)
        qe.pack(side=tk.LEFT, padx=(6,0), ipady=3)

        # ── ตารางรายการสินค้า (JOIN products: barcode/หน่วย/โปรโมชั่น) ──
        try:
            frm, it_tree = make_tree(win,
                ("บาร์โค้ด","ชื่อสินค้า","จำนวน","คืนแล้ว","หน่วย","ราคา/หน่วย","ส่วนลด","โปรโมชั่น","รวม"),
                (95,160,48,52,55,75,62,95,80), height=10)
            frm.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0,4))
            it_tree.tag_configure("ret_full", foreground="#e67e22")   # คืนครบ = ส้ม

            # ยอดคืนต่อสินค้า (query เดียว)
            ret_map = {}
            try:
                _crm = get_db()
                for _rr in _crm.execute(
                    "SELECT product_id, COALESCE(SUM(qty),0) q FROM returns WHERE sale_id=? GROUP BY product_id",
                    (sid,)).fetchall():
                    ret_map[_rr["product_id"]] = int(_rr["q"] or 0)
                _crm.close()
            except Exception as _re:
                print(f"[bill detail returns] {_re}")

            # โหลดรายการ + ข้อมูลสินค้า (query เดียว) — fallback เป็นข้อมูลพื้นฐานถ้า JOIN ล้มเหลว
            rows_data = items
            try:
                _cj = get_db()
                _r = _cj.execute(
                    "SELECT si.product_id, si.name, si.price, si.cost, si.qty, si.subtotal, "
                    "p.barcode AS p_barcode, p.unit AS p_unit, "
                    "p.promotion_enabled, p.promotion_type, p.promotion_value, p.promotion_price, "
                    "p.promotion_start, p.promotion_end, p.promotion_buy_qty, p.promotion_free_qty "
                    "FROM sale_items si LEFT JOIN products p ON p.id = si.product_id "
                    "WHERE si.sale_id=?", (sid,)
                ).fetchall()
                _cj.close()
                if _r: rows_data = [dict(x) for x in _r]
            except Exception as _je:
                print(f"[bill detail join] {_je}")

            def _promo_label(it):
                try:
                    return _receipt_promo_label({
                        "promotion_enabled": it.get("promotion_enabled"),
                        "promotion_type":    it.get("promotion_type"),
                        "promotion_value":   it.get("promotion_value"),
                        "promotion_price":   it.get("promotion_price"),
                        "promotion_start":   it.get("promotion_start"),
                        "promotion_end":     it.get("promotion_end"),
                        "promotion_buy_qty": it.get("promotion_buy_qty"),
                        "promotion_free_qty":it.get("promotion_free_qty"),
                    })
                except Exception:
                    return ""

            def _fill(q=""):
                it_tree.delete(*it_tree.get_children())
                ql = q.lower().strip()
                shown = 0
                for i, it in enumerate(rows_data):
                    nm = str(it.get("name","") or "")
                    bc = str(it.get("p_barcode") or "")
                    un = str(it.get("p_unit") or "")
                    if ql and ql not in nm.lower() and ql not in bc.lower(): continue
                    qn = int(it.get("qty",0) or 0)
                    pr = float(it.get("price",0) or 0)
                    sub = float(it.get("subtotal",0) or 0)
                    d = max(0.0, pr*qn - sub)
                    pl = _promo_label(it)
                    rq = int(ret_map.get(it.get("product_id"), 0))
                    rq_txt = str(rq) if rq > 0 else "—"
                    if rq >= qn and rq > 0: row_tags = ("ret_full",)
                    elif i % 2 == 0:        row_tags = ("alt",)
                    else:                   row_tags = ()
                    it_tree.insert("","end", iid=str(i), tags=row_tags,
                        values=(bc or "—", nm, qn, rq_txt, un or "—", f"฿{pr:,.2f}",
                                (f"฿{d:,.2f}" if d > 0.005 else "—"),
                                pl or "—", f"฿{sub:,.2f}"))
                    shown += 1
                if shown == 0:
                    it_tree.insert("","end",
                        values=("—","(ไม่พบรายการสินค้าในบิลนี้)","","","","","","",""))
            _fill()
            qv.trace_add("write", lambda *_: _fill(qv.get()))
        except Exception as _te:
            import traceback; traceback.print_exc()
            tk.Label(win, text=f"⚠️ โหลดรายการสินค้าไม่ได้: {_te}",
                     bg=C_BG, fg="#ff6b6b", font=F_BODY, wraplength=700,
                     justify="left").pack(fill=tk.X, padx=12, pady=8)

        sumb = tk.Frame(win, bg=C_SURFACE2, padx=14, pady=8); sumb.pack(fill=tk.X, padx=12, pady=(2,4))
        for lbl, val, fg in [("รายการ", f"{len(items)}", C_TEXT),
                             ("ชิ้นรวม", f"{total_qty}", C_TEXT),
                             ("ต้นทุนรวม", f"฿{cost_total:,.2f}", C_MUTED),
                             ("กำไรรวม", f"฿{profit:,.2f}", C_GREEN if profit>=0 else "#ff4444"),
                             ("VAT", vat_txt, C_YELLOW),
                             ("ยอดสุทธิ", f"฿{net:,.2f}", C_ACCENT)]:
            col = tk.Frame(sumb, bg=C_SURFACE2); col.pack(side=tk.LEFT, fill=tk.X, expand=True)
            tk.Label(col, text=lbl, font=(FM.primary,10), bg=C_SURFACE2, fg=C_MUTED).pack(anchor="w")
            tk.Label(col, text=val, font=(FM.primary,12,"bold"), bg=C_SURFACE2, fg=fg).pack(anchor="w")

        bf = tk.Frame(win, bg=C_BG, padx=12, pady=8); bf.pack(fill=tk.X)
        accent_btn(bf, "🖨️  พิมพ์ใบเสร็จซ้ำ", lambda: self._reprint(sid), pad_x=16, pad_y=8).pack(side=tk.LEFT)
        pill_btn(bf, "⧉ Copy เลขบิล", lambda: self._copy_bill_no(sid), bg=C_SURFACE2, fg=C_BLUE, pad_x=12, pad_y=8).pack(side=tk.LEFT, padx=(6,0))

        def _return_selected():
            try:
                _rows = rows_data; _tree = it_tree
            except Exception:
                messagebox.showinfo("คืนสินค้า", "ไม่พบรายการสินค้า", parent=win); return
            sel = _tree.selection()
            if not sel:
                messagebox.showinfo("คืนสินค้า", "เลือกสินค้าที่ต้องการคืนก่อน (คลิกที่รายการในตาราง)", parent=win); return
            try: idx = int(sel[0])
            except Exception: return
            if idx < 0 or idx >= len(_rows): return
            it = _rows[idx]
            pid = it.get("product_id"); nm = str(it.get("name","") or "")
            price = float(it.get("price",0) or 0); sold = int(it.get("qty",0) or 0)
            try:
                _cr = get_db()
                already = _cr.execute("SELECT COALESCE(SUM(qty),0) FROM returns WHERE sale_id=? AND product_id=?",
                                      (sid, pid)).fetchone()[0] or 0
                _cr.close()
            except Exception: already = 0
            maxq = sold - int(already)
            if maxq <= 0:
                messagebox.showinfo("คืนครบแล้ว", f"{nm}\nคืนครบจำนวนที่ขายแล้ว ({sold} ชิ้น)", parent=win); return
            self._return_dialog(win, sid, pid, nm, price, maxq,
                                on_done=lambda: (self.load(), win.destroy(), self._bill_detail_window(sid)))

        if (sale.get("status") or "normal") != "void":
            pill_btn(bf, "↩ คืนสินค้า", _return_selected, bg="#2a2410", fg="#f4c542",
                     pad_x=12, pad_y=8).pack(side=tk.LEFT, padx=(6,0))
        if current_staff.get("role") in ("owner","admin"):
            if (sale.get("status") or "normal") == "void":
                tk.Label(bf, text="✖ บิลนี้ถูกยกเลิกแล้ว", font=(FM.primary,11,"bold"),
                         bg=C_BG, fg="#ff6b6b").pack(side=tk.LEFT, padx=(10,0))
            else:
                pill_btn(bf, "🚫 ยกเลิกบิล", lambda: self._void_bill(sid, on_done=win.destroy),
                         bg="#3a1414", fg="#ff6b6b", pad_x=12, pad_y=8).pack(side=tk.LEFT, padx=(6,0))
        pill_btn(bf, "ปิด (Esc)", win.destroy, bg=C_SURFACE, fg=C_MUTED, pad_x=14, pad_y=8).pack(side=tk.RIGHT)
        qe.focus_set()

    def _reprint(self, sid):
        try:
            print_receipt(sid)
        except Exception as e:
            messagebox.showerror("พิมพ์ซ้ำไม่สำเร็จ", str(e), parent=self.winfo_toplevel())

    def _return_dialog(self, parent, sid, pid, name, price, maxq, on_done=None):
        """คืนสินค้ารายตัว — mirror ระบบคืนสินค้าเดิม (returns + restock + movement)"""
        dlg = tk.Toplevel(parent); dlg.title("คืนสินค้า"); dlg.configure(bg=C_BG)
        dlg.geometry("460x380"); dlg.grab_set(); dlg.lift()
        dlg.bind("<Escape>", lambda e: dlg.destroy())
        tk.Frame(dlg, bg="#f4c542", height=3).pack(fill=tk.X)
        tk.Label(dlg, text="↩ คืนสินค้า", font=F_H1, bg=C_BG, fg=C_TEXT).pack(anchor="w", padx=16, pady=(10,2))
        tk.Label(dlg, text=name, font=F_BODY, bg=C_BG, fg=C_TEXT, wraplength=420, justify="left").pack(anchor="w", padx=16)
        tk.Label(dlg, text=f"คืนได้สูงสุด {maxq} ชิ้น  •  ฿{price:,.2f}/หน่วย",
                 font=F_SMALL, bg=C_BG, fg=C_MUTED).pack(anchor="w", padx=16, pady=(2,8))

        body = tk.Frame(dlg, bg=C_BG, padx=16); body.pack(fill=tk.X)
        tk.Label(body, text="จำนวนที่คืน", font=F_SMALL, bg=C_BG, fg=C_MUTED).pack(anchor="w")
        qv = tk.StringVar(value="1")
        qe = tk.Entry(body, textvariable=qv, width=10, bg=C_SURFACE2, fg=C_TEXT,
                      insertbackground=C_TEXT, relief=tk.FLAT, font=F_BODY)
        qe.pack(anchor="w", ipady=4, pady=(0,6))
        tk.Label(body, text="เหตุผลการคืน", font=F_SMALL, bg=C_BG, fg=C_MUTED).pack(anchor="w")
        rv = tk.StringVar(value="ลูกค้าเปลี่ยนใจ")
        ttk.Combobox(body, textvariable=rv,
                     values=["ลูกค้าเปลี่ยนใจ","สินค้าชำรุด","สินค้าหมดอายุ","ขายผิด","อื่นๆ"],
                     width=30, font=F_BODY).pack(anchor="w", pady=(0,6))
        restock_v = tk.BooleanVar(value=True)
        tk.Checkbutton(body, text="คืนเข้าสต็อก (เพิ่มสต็อกกลับ)", variable=restock_v,
                       bg=C_BG, fg=C_TEXT, selectcolor=C_SURFACE2, activebackground=C_BG,
                       activeforeground=C_TEXT, font=F_BODY).pack(anchor="w")
        refund_lbl = tk.Label(body, text="", font=(FM.primary,13,"bold"), bg=C_BG, fg="#f4c542")
        refund_lbl.pack(anchor="w", pady=(8,0))
        def _calc(*_):
            try: q = int(qv.get() or 0)
            except Exception: q = 0
            refund_lbl.config(text=f"คืนเงินลูกค้า: ฿{max(0,q)*price:,.2f}")
        qv.trace_add("write", _calc); _calc()

        def _confirm():
            try: q = int(qv.get() or 0)
            except Exception:
                messagebox.showerror("ผิดพลาด", "กรอกจำนวนเป็นตัวเลข", parent=dlg); return
            if q <= 0 or q > maxq:
                messagebox.showwarning("จำนวนไม่ถูกต้อง", f"กรอก 1–{maxq}", parent=dlg); return
            reason = (rv.get() or "").strip() or "(ไม่ระบุ)"
            restock = 1 if restock_v.get() else 0
            refund = q * price
            if not messagebox.askyesno("ยืนยันคืนสินค้า",
                f"คืน {name}\nจำนวน {q} ชิ้น\nคืนเงิน ฿{refund:,.2f}\nเหตุผล: {reason}\n"
                f"{'เข้าสต็อก' if restock else 'ไม่เข้าสต็อก'}\n\nยืนยัน?",
                parent=dlg): return
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            try:
                conn = get_db()
                conn.execute(
                    "INSERT INTO returns (return_date,sale_id,product_id,name,price,qty,refund,reason,restock) "
                    "VALUES (?,?,?,?,?,?,?,?,?)",
                    (now, sid, pid, name, price, q, refund, reason, restock))
                if restock:
                    conn.execute("UPDATE products SET stock=stock+? WHERE id=?", (q, pid))
                conn.commit(); conn.close()
            except Exception as e:
                messagebox.showerror("คืนไม่สำเร็จ", str(e), parent=dlg); return
            if restock:
                try:
                    record_movement(pid, q, "return", ref_id=sid, ref_type="return",
                                    note=f"คืนสินค้าบิล#{sid}", cost=0)
                except Exception as _e:
                    print(f"[return StockCard] {_e}")
            messagebox.showinfo("คืนสำเร็จ",
                f"✅ รับคืน {name}\n{q} ชิ้น  คืนเงิน ฿{refund:,.2f}", parent=dlg)
            dlg.destroy()
            if callable(on_done):
                try: on_done()
                except Exception: pass

        btn = tk.Frame(dlg, bg=C_BG, padx=16); btn.pack(fill=tk.X, pady=12)
        accent_btn(btn, "↩ ยืนยันคืน", _confirm, pad_x=16, pad_y=8).pack(side=tk.LEFT)
        pill_btn(btn, "ยกเลิก", dlg.destroy, bg=C_SURFACE, fg=C_MUTED, pad_x=14, pad_y=8).pack(side=tk.RIGHT)
        qe.focus_set()

    def _void_bill(self, sid, on_done=None):
        # สิทธิ์: ยกเลิกบิลย้อนหลัง = เฉพาะเจ้าของร้าน/แอดมิน
        if current_staff.get("role") not in ("owner","admin"):
            messagebox.showwarning("ไม่มีสิทธิ์",
                "ยกเลิกบิลย้อนหลังได้เฉพาะเจ้าของร้าน/แอดมิน",
                parent=self.winfo_toplevel()); return
        try:
            sale = get_sale_with_customer(sid)
            items = [dict(x) for x in get_sale_items(sid)]
        except Exception as e:
            messagebox.showerror("ผิดพลาด", str(e), parent=self.winfo_toplevel()); return
        if not sale:
            return
        if (sale.get("status") or "normal") == "void":
            messagebox.showinfo("ยกเลิกแล้ว", f"บิล #{sid} ถูกยกเลิกไปแล้ว", parent=self.winfo_toplevel()); return
        # บิลเงินเชื่อ: หนี้คงค้างจะถูกล้างอัตโนมัติเมื่อ void (ผ่าน db_receivable)
        # ถ้าลูกค้าชำระมาบางส่วนแล้ว ต้องคืนเงินส่วนนั้นเอง
        is_debt = (sale.get("payment_method") or "") in ("debt","credit") or ("เงิน ซื้อ" in (sale.get("note") or ""))
        paid_partial = 0.0
        if is_debt:
            try:
                _c2 = get_db()
                paid_partial = _c2.execute(
                    "SELECT COALESCE(SUM(amount),0) FROM receivable_payments WHERE sale_id=?", (sid,)
                ).fetchone()[0] or 0
                _c2.close()
            except Exception: paid_partial = 0.0
        # เหตุผล
        reason = simpledialog.askstring("ยกเลิกบิล", f"เหตุผลการยกเลิกบิล #{sid}:",
                                        parent=self.winfo_toplevel())
        if reason is None: return
        reason = reason.strip() or "(ไม่ระบุ)"
        net = sale["net"] or 0
        n_items = len(items); n_qty = sum(int(it.get("qty",0) or 0) for it in items)
        extra = ""
        if is_debt:
            extra = "\n\n📋 บิลเงินเชื่อ: หนี้คงค้างจะถูกล้างอัตโนมัติ"
            if paid_partial > 0.009:
                extra += f"\n⚠️ ลูกค้าชำระมาแล้ว ฿{paid_partial:,.2f} — กรุณาคืนเงินส่วนนี้ให้ลูกค้าด้วยตนเอง"
        if not messagebox.askyesno("ยืนยันยกเลิกบิล",
            f"ยกเลิกบิล #{sid}?\n\nคืนสต็อก {n_qty} ชิ้น ({n_items} รายการ)\n"
            f"ยอด ฿{net:,.2f}\nเหตุผล: {reason}{extra}\n\n⚠️ การกระทำนี้ย้อนกลับไม่ได้",
            parent=self.winfo_toplevel()): return
        by = current_staff.get("name","")
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            conn = get_db()
            # 1) คืนสต็อก (mirror ระบบคืนสินค้า: stock += qty)
            for it in items:
                conn.execute("UPDATE products SET stock=stock+? WHERE id=?",
                             (int(it.get("qty",0) or 0), it.get("product_id")))
            # 2) คืนแต้ม + ยอดสะสม ให้ตรงกับ save_sale (points += net//rate, total_spent += net)
            if sale.get("customer_id"):
                try: rate = int(get_setting("point_rate") or 10)
                except: rate = 10
                pts = int(net // rate) if rate > 0 else 0
                conn.execute(
                    "UPDATE customers SET points=MAX(0,points-?), total_spent=MAX(0,total_spent-?) WHERE id=?",
                    (pts, net, sale["customer_id"]))
            # 3) mark voided
            conn.execute("UPDATE sales SET status='void',void_at=?,void_by=?,void_reason=? WHERE id=?",
                         (now, by, reason, sid))
            # 4) audit
            conn.execute("INSERT INTO sale_audit (sale_id,action,by_name,at,reason) VALUES (?,?,?,?,?)",
                         (sid, "void", by, now, reason))
            conn.commit(); conn.close()
        except Exception as e:
            messagebox.showerror("ยกเลิกไม่สำเร็จ", str(e), parent=self.winfo_toplevel()); return
        # 5) Stock Card หลัง commit (กัน DB lock)
        try:
            for it in items:
                record_movement(it.get("product_id"), int(it.get("qty",0) or 0), "void",
                    ref_id=sid, ref_type="void", note=f"ยกเลิกบิล#{sid}", cost=it.get("cost",0) or 0)
        except Exception as e:
            print(f"[void StockCard] {e}")
        messagebox.showinfo("ยกเลิกบิลแล้ว",
            f"✅ ยกเลิกบิล #{sid} เรียบร้อย\nคืนสต็อก {n_qty} ชิ้น" +
            ("\n📋 ล้างยอดลูกหนี้แล้ว" if is_debt else ""),
            parent=self.winfo_toplevel())
        try:
            root = self.winfo_toplevel()
            if hasattr(root, "_ar_tab"): root._ar_tab.refresh_all()
        except Exception: pass
        self.load()
        if callable(on_done):
            try: on_done()
            except Exception: pass

    def _copy_bill_no(self, sid):
        try:
            self.clipboard_clear(); self.clipboard_append(str(sid)); self.update_idletasks()
            messagebox.showinfo("คัดลอกแล้ว", f"คัดลอกเลขบิล #{sid}", parent=self.winfo_toplevel())
        except Exception: pass

    def _build_bill_menu(self):
        m = tk.Menu(self, tearoff=0, bg=C_SURFACE, fg=C_TEXT,
                    activebackground=C_ACCENT, activeforeground=C_BG, bd=0, font=(FM.primary,11))
        m.add_command(label="🔍  ดูรายละเอียดบิล", command=lambda: self._open_bill_detail())
        m.add_command(label="🖨️  พิมพ์ใบเสร็จซ้ำ", command=self._menu_reprint)
        m.add_command(label="⧉  Copy เลขบิล", command=self._menu_copy)
        m.add_separator()
        m.add_command(label="🚫  ยกเลิกบิล (ผ่านรายละเอียด)", command=lambda: self._open_bill_detail())
        self._bill_menu = m

    def _show_bill_menu(self, event):
        iid = self.bt.identify_row(event.y)
        if iid: self.bt.selection_set(iid)
        try: self._bill_menu.tk_popup(event.x_root, event.y_root)
        finally: self._bill_menu.grab_release()

    def _menu_reprint(self):
        sel = self.bt.selection()
        if sel: self._reprint(int(sel[0]))

    def _menu_copy(self):
        sel = self.bt.selection()
        if sel: self._copy_bill_no(int(sel[0]))


    def _report_data(self):
        d0,d1 = self.df.get_iso(), self.dt.get_iso()
        data  = get_report_data(d0,d1)   # db_sale — Session 8
        sales = data["sales"]
        items = data["items"]
        top   = data["top"]
        prods = get_all_products_active()  # db_product
        return sales, items, top, prods, d0, d1

    def _sheets(self):
        sales,items,top,prods,d0,d1 = self._report_data()
        h1=["เลขบิล","วันเวลา","ลูกค้า","ยอดรวม","ส่วนลด","โปรโมชั่น","ยอดสุทธิ","รับเงิน","เงินทอน"]
        r1=[(s["id"],s["sale_date"],s["cn"] or "",s["total"],s["discount"] or 0,s["promo_disc"] or 0,s["net"],s["paid"],s["change_amt"]) for s in sales]
        h2=["วันที่","เลขบิล","ชื่อสินค้า","ราคา","จำนวน","รวม"]
        r2=[(it["sdate"],it["sale_id"],it["name"],it["price"],it["qty"],it["subtotal"]) for it in items]
        h3=["อันดับ","ชื่อสินค้า","จำนวนขาย","ยอดขาย","กำไร"]
        r3=[(i+1,t["name"],t["tq"],round(t["ts"],2),round(t["tp"],2)) for i,t in enumerate(top)]
        h4=["ID","บาร์โค้ด","ชื่อสินค้า","ราคาขาย","ต้นทุน","สต็อก","หน่วย"]
        r4=[(p["id"],p["barcode"] or "",p["name"],p["price"],p["cost"],p["stock"],p["unit"]) for p in prods]
        # Sanitize date for filename (remove / and : which are invalid in Windows paths)
        d0s = d0.replace("/","-").replace(":","-")
        d1s = d1.replace("/","-").replace(":","-")
        name = "report_" + d0s + "_" + d1s
        return name, [("รายการบิล",h1,r1),("รายการในบิล",h2,r2),("สินค้าขายดี",h3,r3),("สินค้า-สต็อก",h4,r4)]

    def _exp_excel(self):
        if not HAS_EXCEL:
            ans = messagebox.askyesno("ต้องติดตั้ง openpyxl","ยังไม่มี openpyxl\nต้องการคำแนะนำการติดตั้งไหม?")
            if ans:
                messagebox.showinfo("วิธีติดตั้ง","เปิด Command Prompt แล้วพิมพ์:\n\npip install openpyxl\n\nแล้วเปิดโปรแกรมใหม่อีกครั้ง")
            return
        from tkinter import filedialog
        from datetime import datetime as _dt
        default_name = "report_" + _dt.now().strftime("%Y%m%d_%H%M%S") + ".xlsx"
        path = filedialog.asksaveasfilename(
            title="บันทึก Excel",
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel","*.xlsx"),("All","*.*")])
        if not path: return
        _, sheets = self._sheets()
        path = _do_export_excel(path, sheets)
        if path:
            if messagebox.askyesno("Excel สำเร็จ", "บันทึกแล้ว:\n" + path + "\n\nเปิดไฟล์เลยไหม?"):
                _open_path(path)

    def _exp_csv(self):
        from tkinter import filedialog
        folder = filedialog.askdirectory(title="เลือกโฟลเดอร์บันทึก")
        if not folder: return
        sales,items,top,prods,d0,d1 = self._report_data()
        paths = []
        h1=["เลขบิล","วันเวลา","ลูกค้า","ยอดรวม","ส่วนลด","โปรโมชั่น","ยอดสุทธิ","รับเงิน","เงินทอน"]
        r1=[(s["id"],s["sale_date"],s["cn"] or "",s["total"],s["discount"] or 0,s["promo_disc"] or 0,s["net"],s["paid"],s["change_amt"]) for s in sales]
        d0s=d0.replace("/","-"); d1s=d1.replace("/","-")
        paths.append(_do_export_csv(os.path.join(folder,"bills_"+d0s+"_"+d1s), h1, r1))
        h4=["ID","บาร์โค้ด","ชื่อสินค้า","ราคาขาย","ต้นทุน","สต็อก","หน่วย"]
        r4=[(p["id"],p["barcode"] or "",p["name"],p["price"],p["cost"],p["stock"],p["unit"]) for p in prods]
        paths.append(_do_export_csv(os.path.join(folder,"products"), h4, r4))
        msg = "บันทึก " + str(len(paths)) + " ไฟล์ในโฟลเดอร์ exports/"
        if messagebox.showinfo("CSV สำเร็จ", msg) is None and sys.platform == "win32":
            _open_path("exports")

# ══ TAB: SETTINGS ════════════════════════════════════════
# ══ TAB: QUICK SALE SETTINGS ═════════════════════════════════
class QSSaleTab(tk.Frame):
    def __init__(self,parent):
        super().__init__(parent,bg=C_BG); self._build()

    def _build(self):
        tk.Frame(self,bg=C_YELLOW,height=4).pack(fill=tk.X)
        hf=tk.Frame(self,bg=C_BG,padx=20,pady=14); hf.pack(fill=tk.X)
        tk.Label(hf,text="⚡  Quick Sale Mode",font=F_H1,bg=C_BG,fg=C_YELLOW).pack(anchor="w")
        tk.Label(hf,text="โหมดขายด่วน — F1 เริ่มขาย • สแกนสินค้า • F5 บันทึก+พิมพ์ทันที",
                 font=F_SMALL,bg=C_BG,fg=C_MUTED).pack(anchor="w")

        # Enable toggle
        sf=tk.Frame(self,bg=C_SURFACE,highlightbackground=C_BORDER,highlightthickness=1)
        sf.pack(fill=tk.X,padx=20,pady=(0,10))
        tf=tk.Frame(sf,bg=C_SURFACE,padx=16,pady=12); tf.pack(fill=tk.X)
        self.qs_var=tk.BooleanVar(value=get_setting("quick_sale_mode")=="1")
        tk.Checkbutton(tf,text="☑  เปิดใช้งาน Quick Sale Mode",
                       variable=self.qs_var,bg=C_SURFACE,fg=C_TEXT,
                       selectcolor=C_SURFACE2,activebackground=C_SURFACE,
                       activeforeground=C_TEXT,font=(FM.primary,13,"bold"),
                       command=self._save).pack(anchor="w")
        tk.Label(tf,text="เมื่อเปิด Quick Sale Mode: F1=เริ่มขาย, F5=บันทึก+พิมพ์ทันที (ไม่เปิดหน้าต่างชำระเงิน)",
                 font=F_SMALL,bg=C_SURFACE,fg=C_MUTED,wraplength=500).pack(anchor="w",pady=(4,0))

        # Payment method
        pf=tk.Frame(self,bg=C_SURFACE,highlightbackground=C_BORDER,highlightthickness=1)
        pf.pack(fill=tk.X,padx=20,pady=(0,10))
        tk.Frame(pf,bg=C_BLUE,height=2).pack(fill=tk.X)
        tk.Label(pf,text="วิธีชำระเงินเริ่มต้น",font=F_H2,bg=C_SURFACE,fg=C_TEXT,pady=8).pack(anchor="w",padx=16)
        self.qs_pay=tk.StringVar(value=get_setting("quick_sale_payment") or "cash")
        prf=tk.Frame(pf,bg=C_SURFACE,padx=16); prf.pack(anchor="w",pady=(0,12))
        for v,lbl,col in [("cash","💵 เงินสด",C_GREEN),
                          ("qr","📱 QR พร้อมเพย์",C_BLUE),
                          ("transfer","🏦 เงินโอน",C_YELLOW)]:
            tk.Radiobutton(prf,text=lbl,variable=self.qs_pay,value=v,
                           bg=C_SURFACE,fg=col,selectcolor=C_SURFACE2,
                           activebackground=C_SURFACE,activeforeground=col,
                           font=F_BODY,command=self._save).pack(anchor="w",pady=2)

        # Options
        of=tk.Frame(self,bg=C_SURFACE,highlightbackground=C_BORDER,highlightthickness=1)
        of.pack(fill=tk.X,padx=20,pady=(0,10))
        tk.Frame(of,bg=C_GREEN,height=2).pack(fill=tk.X)
        tk.Label(of,text="ตัวเลือกเพิ่มเติม",font=F_H2,bg=C_SURFACE,fg=C_TEXT,pady=8).pack(anchor="w",padx=16)
        opts=[
            ("พิมพ์ใบเสร็จอัตโนมัติหลังบันทึก","quick_sale_print",True),
            ("เปิดบิลใหม่อัตโนมัติหลังบันทึก","quick_sale_new_bill",True),
        ]
        self.opt_vars={}
        ofp=tk.Frame(of,bg=C_SURFACE,padx=16); ofp.pack(anchor="w",pady=(0,12))
        for lbl,key,default in opts:
            v=tk.BooleanVar(value=get_setting(key)!="0")
            tk.Checkbutton(ofp,text=lbl,variable=v,bg=C_SURFACE,fg=C_TEXT,
                           selectcolor=C_SURFACE2,activebackground=C_SURFACE,
                           activeforeground=C_TEXT,font=F_BODY,
                           command=self._save).pack(anchor="w",pady=2)
            self.opt_vars[key]=v

        # Status preview
        sp=tk.Frame(self,bg=C_SURFACE2,padx=16,pady=12,highlightbackground=C_BORDER,highlightthickness=1)
        sp.pack(fill=tk.X,padx=20,pady=(0,10))
        tk.Label(sp,text="สถานะปัจจุบัน",font=F_SMALL,bg=C_SURFACE2,fg=C_MUTED).pack(anchor="w")
        self.status_lbl=tk.Label(sp,text="",font=(FM.primary,13,"bold"),bg=C_SURFACE2,fg=C_TEXT)
        self.status_lbl.pack(anchor="w",pady=(4,0))
        self._update_status()

    def _update_status(self):
        if self.qs_var.get():
            pay_th={"cash":"เงินสด","qr":"QR พร้อมเพย์","transfer":"เงินโอน"}.get(self.qs_pay.get(),"เงินสด")
            self.status_lbl.config(text="⚡ Quick Sale Mode: เปิดอยู่ | ชำระด้วย {}".format(pay_th),fg=C_YELLOW)
        else:
            self.status_lbl.config(text="○ โหมดปกติ (Normal Mode)",fg=C_MUTED)

    def _save(self):
        set_setting("quick_sale_mode","1" if self.qs_var.get() else "0")
        set_setting("quick_sale_payment",self.qs_pay.get())
        for key,v in self.opt_vars.items():
            set_setting(key,"1" if v.get() else "0")
        self._update_status()

        # Refresh SaleTab status bar
        try:
            self.winfo_toplevel().tabs[0]._refresh_qs_bar()
        except: pass
class SettingsTab(tk.Frame):
    _ACCENT = "#00d4aa"
    _MENU = [
        ("store",     "ร้านค้า",       "🏪"),
        ("receipt",   "ใบเสร็จ",       "🧾"),
        ("printer",   "เครื่องพิมพ์",  "🖨"),
        ("inventory", "สต็อก",         "📦"),
        ("database",  "ฐานข้อมูล",     "💾"),
        ("vat",       "ภาษี VAT",      "🧮"),
        ("system",    "ระบบ",          "⚙️"),
    ]

    def __init__(self, parent):
        super().__init__(parent, bg=C_BG)
        self.entries     = {}
        self._panels     = {}
        self._menu_btns  = {}
        self.lsv = tk.BooleanVar(value=get_setting("low_stock_alert")=="1")
        self.neg_policy  = tk.StringVar(value=get_setting("neg_stock_policy") or "warn")
        self.shift_req   = tk.BooleanVar(value=get_setting("shift_required") == "1")
        # ── Printer/Receipt UX (Phase A) ──
        self.paper_size    = tk.StringVar(value=get_setting("paper_size") or "80mm")
        self.preview_print = tk.BooleanVar(value=get_setting("preview_before_print") != "0")
        self.auto_print    = tk.BooleanVar(value=get_setting("auto_print_receipt") != "0")
        self.print_copy    = tk.BooleanVar(value=get_setting("print_copy") == "1")
        self.escpos_direct = tk.BooleanVar(value=get_setting("escpos_direct") != "0")
        self.show_vat      = tk.BooleanVar(value=get_setting("show_vat") != "0")
        self.show_staff    = tk.BooleanVar(value=get_setting("show_staff") == "1")
        self.show_points   = tk.BooleanVar(value=get_setting("show_points") != "0")
        self.db_mode_var = tk.StringVar(
            value=DB_CONFIG.get("mode","sqlite_offline") if HAS_SERVICES else "sqlite_offline")
        self._build()

    def _build(self):
        tk.Frame(self, bg=self._ACCENT, height=3).pack(fill=tk.X)
        body = tk.Frame(self, bg=C_BG); body.pack(fill=tk.BOTH, expand=True)

        # Sidebar
        sb = tk.Frame(body, bg=C_SURFACE2, width=210)
        sb.pack(side=tk.LEFT, fill=tk.Y); sb.pack_propagate(False)
        tk.Label(sb, text="  ⚙️  ตั้งค่า", font=(FM.primary,13,"bold"),
                 bg=C_SURFACE2, fg=C_TEXT, pady=16, anchor="w").pack(fill=tk.X)
        tk.Frame(sb, bg=C_BORDER, height=1).pack(fill=tk.X)
        for key, label, icon in self._MENU:
            self._make_menu_btn(sb, key, label, icon)
        tk.Label(sb, text="RakComSoft 2026", font=(FM.primary,9),
                 bg=C_SURFACE2, fg=C_MUTED).pack(side=tk.BOTTOM, pady=8)

        # Panel area
        self._panel_area = tk.Frame(body, bg=C_BG)
        self._panel_area.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self._panels["store"]     = self._build_store()
        self._panels["receipt"]   = self._build_receipt()
        self._panels["printer"]   = self._build_printer()
        self._panels["inventory"] = self._build_inventory()
        self._panels["database"]  = self._build_database()
        self._panels["vat"]       = self._build_vat()
        self._panels["system"]    = self._build_system()
        self._show_panel("store")

    def _make_menu_btn(self, parent, key, label, icon):
        f = tk.Frame(parent, bg=C_SURFACE2, cursor="hand2"); f.pack(fill=tk.X)
        icon_lbl = tk.Label(f, text=icon, font=(FM.primary, 13),
                            bg=C_SURFACE2, fg=C_MUTED, anchor="center", width=2)
        icon_lbl.pack(side=tk.LEFT, padx=(12, 2), pady=10)
        text_lbl = tk.Label(f, text=label, font=(FM.primary, 12),
                            bg=C_SURFACE2, fg=C_MUTED, anchor="w")
        text_lbl.pack(side=tk.LEFT, fill=tk.X, expand=True, pady=10)
        for w in (f, icon_lbl, text_lbl):
            w.bind("<Button-1>", lambda e, k=key: self._show_panel(k))
        self._menu_btns[key] = (f, icon_lbl, text_lbl)

    def _show_panel(self, key):
        for k, (f, icon_lbl, text_lbl) in self._menu_btns.items():
            if k == key:
                f.config(bg=C_BG, highlightbackground=self._ACCENT, highlightthickness=3)
                icon_lbl.config(bg=C_BG, fg=self._ACCENT)
                text_lbl.config(bg=C_BG, fg=self._ACCENT)
            else:
                f.config(bg=C_SURFACE2, highlightthickness=0)
                icon_lbl.config(bg=C_SURFACE2, fg=C_MUTED)
                text_lbl.config(bg=C_SURFACE2, fg=C_MUTED)
        for k, p in self._panels.items():
            p.pack_forget()
        self._panels[key].pack(fill=tk.BOTH, expand=True)

    def _scrollable(self, parent):
        outer = tk.Frame(parent, bg=C_BG); outer.pack(fill=tk.BOTH, expand=True)
        canvas = tk.Canvas(outer, bg=C_BG, highlightthickness=0)
        vsb = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview,
                            style="R.Vertical.TScrollbar")
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        inner = tk.Frame(canvas, bg=C_BG)
        wid = canvas.create_window((0,0), window=inner, anchor="nw")
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(wid, width=e.width))
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)),"units"))
        return outer, inner

    def _panel_hdr(self, parent, title, subtitle):
        tk.Frame(parent, bg=self._ACCENT, height=2).pack(fill=tk.X)
        hf = tk.Frame(parent, bg=C_BG, padx=24, pady=14); hf.pack(fill=tk.X)
        tk.Label(hf, text=title, font=F_H1, bg=C_BG, fg=C_TEXT).pack(anchor="w")
        tk.Label(hf, text=subtitle, font=F_SMALL, bg=C_BG, fg=C_MUTED).pack(anchor="w")

    def _card(self, parent, title=None):
        cf = tk.Frame(parent, bg=C_SURFACE2,
                      highlightbackground=C_BORDER, highlightthickness=1)
        cf.pack(fill=tk.X, padx=20, pady=(0,12))
        if title:
            tk.Label(cf, text=title, font=(FM.primary,11,"bold"),
                     bg=C_SURFACE2, fg=C_TEXT, padx=16, pady=10).pack(anchor="w")
            tk.Frame(cf, bg=C_BORDER, height=1).pack(fill=tk.X)
        inner = tk.Frame(cf, bg=C_SURFACE2, padx=16, pady=12); inner.pack(fill=tk.X)
        return inner

    def _frow(self, parent, label, key, width=36):
        tk.Label(parent, text=label, font=F_SMALL,
                 bg=C_SURFACE2, fg=C_MUTED).pack(anchor="w", pady=(6,1))
        e = field(parent, width=width); e.insert(0, get_setting(key))
        e.pack(anchor="w", ipady=5); self.entries[key] = e

    def _save_row(self, parent, extra_btns=()):
        tk.Frame(parent, bg=C_BG, height=1).pack(pady=4)
        bf = tk.Frame(parent, bg=C_BG); bf.pack(anchor="e", padx=20, pady=(0,16))
        for lbl, cmd, bg, fg in extra_btns:
            pill_btn(bf, lbl, cmd, bg=bg, fg=fg, pad_x=14, pad_y=7).pack(side=tk.LEFT, padx=(0,6))
        accent_btn(bf, "💾  บันทึก", self._save, pad_x=20, pad_y=8).pack(side=tk.LEFT)

    # ════ PANELS ════════════════════════════════════════
    def _build_store(self):
        p = tk.Frame(self._panel_area, bg=C_BG)
        self._panel_hdr(p, "🏪  ร้านค้า", "ข้อมูลพื้นฐานที่แสดงบนใบเสร็จและรายงาน")
        _, inner = self._scrollable(p)
        c = self._card(inner, "ข้อมูลร้านค้า")
        for lbl, key in [("ชื่อร้านค้า","store_name"),("ที่อยู่","store_address"),("โทรศัพท์","store_phone")]:
            self._frow(c, lbl, key)
        c2 = self._card(inner, "การเงิน")
        self._frow(c2, "พร้อมเพย์  (เบอร์ 10 หลัก หรือเลขนิติบุคคล 13 หลัก)", "promptpay_id")
        self._frow(c2, "อัตราแต้มสะสม (฿ ต่อ 1 แต้ม)", "point_rate")
        hf = tk.Frame(c2, bg="#1a2a1a", padx=10, pady=8); hf.pack(fill=tk.X, pady=(8,0))
        tk.Label(hf, text="💡  บุคคลธรรมดา → เบอร์ 10 หลัก เช่น 0812345678", font=F_SMALL, bg="#1a2a1a", fg="#6db86d").pack(anchor="w")
        tk.Label(hf, text="    บริษัท/หจก. → เลขนิติบุคคล 13 หลัก", font=F_SMALL, bg="#1a2a1a", fg="#6db86d").pack(anchor="w")
        self._save_row(inner)
        return p

    def _build_receipt(self):
        p = tk.Frame(self._panel_area, bg=C_BG)
        self._panel_hdr(p, "🧾  ใบเสร็จ", "รูปแบบและการแสดงผลของเอกสารใบเสร็จ")
        _, inner = self._scrollable(p)

        # ── ข้อความ ──
        c = self._card(inner, "ข้อความบนใบเสร็จ")
        self._frow(c, "ข้อความหัวใบเสร็จ (เพิ่มเติม — เว้นว่างได้)", "receipt_header")
        self._frow(c, "ข้อความท้ายใบเสร็จ", "receipt_footer")

        # ── การแสดงผล (wired) ──
        c2 = self._card(inner, "แสดงข้อมูลบนใบเสร็จ")
        for txt, var in [("แสดงข้อมูลภาษี VAT", self.show_vat),
                         ("แสดงชื่อพนักงาน", self.show_staff),
                         ("แสดงแต้มสมาชิก", self.show_points)]:
            tk.Checkbutton(c2, text=txt, variable=var, bg=C_SURFACE2, fg=C_TEXT,
                           selectcolor=C_BG, activebackground=C_SURFACE2,
                           activeforeground=C_TEXT, font=F_BODY).pack(anchor="w", pady=2)
        tk.Label(c2, text="💡 ชื่อพนักงาน = ผู้ที่ล็อกอินขณะพิมพ์",
                 font=(FM.primary,10), bg=C_SURFACE2, fg=C_MUTED).pack(anchor="w", pady=(4,0))

        # ── คุณภาพตัวอักษร / ฟอนต์ (สำหรับเครื่องพิมพ์ความร้อน) ──
        cq = self._card(inner, "คุณภาพตัวอักษร / ฟอนต์")
        _fontmap = {"auto":"อัตโนมัติ","leelawadee":"Leelawadee UI","noto":"Noto Sans Thai",
                    "sarabun":"TH Sarabun","tahoma":"Tahoma","kanit":"Kanit"}
        rowf = tk.Frame(cq, bg=C_SURFACE2); rowf.pack(fill=tk.X, pady=(0,4))
        tk.Label(rowf, text="ฟอนต์", font=F_SMALL, bg=C_SURFACE2, fg=C_MUTED).pack(side=tk.LEFT, padx=(0,8))
        self._rfont_cb = ttk.Combobox(rowf, values=["อัตโนมัติ","Leelawadee UI","Noto Sans Thai","TH Sarabun","Tahoma","Kanit"],
                                       width=16, state="readonly", font=F_BODY)
        self._rfont_cb.set(_fontmap.get((get_setting("receipt_font") or "auto").strip(), "อัตโนมัติ"))
        self._rfont_cb.pack(side=tk.LEFT)
        rowf2 = tk.Frame(cq, bg=C_SURFACE2); rowf2.pack(fill=tk.X, pady=(4,4))
        tk.Label(rowf2, text="ขนาดตัวอักษร", font=F_SMALL, bg=C_SURFACE2, fg=C_MUTED).pack(side=tk.LEFT, padx=(0,8))
        _szm = {"small":"เล็ก","medium":"กลาง","large":"ใหญ่"}
        self._rsize_cb = ttk.Combobox(rowf2, values=["เล็ก","กลาง","ใหญ่"], width=8, state="readonly", font=F_BODY)
        self._rsize_cb.set(_szm.get((get_setting("receipt_font_size") or "medium").strip(), "กลาง"))
        self._rsize_cb.pack(side=tk.LEFT, padx=(0,16))
        tk.Label(rowf2, text="ระยะบรรทัด", font=F_SMALL, bg=C_SURFACE2, fg=C_MUTED).pack(side=tk.LEFT, padx=(0,8))
        _spm = {"compact":"กระชับ","normal":"ปกติ","relaxed":"โปร่ง"}
        self._rspace_cb = ttk.Combobox(rowf2, values=["กระชับ","ปกติ","โปร่ง"], width=8, state="readonly", font=F_BODY)
        self._rspace_cb.set(_spm.get((get_setting("receipt_line_spacing") or "normal").strip(), "ปกติ"))
        self._rspace_cb.pack(side=tk.LEFT)
        self.r_sharpen = tk.BooleanVar(value=(get_setting("receipt_sharpen") or "1") != "0")
        tk.Checkbutton(cq, text="เพิ่มความคม (ลดเบลอ/ฟุ้ง — แนะนำเปิด)", variable=self.r_sharpen,
                       bg=C_SURFACE2, fg=C_TEXT, selectcolor=C_BG, activebackground=C_SURFACE2,
                       activeforeground=C_TEXT, font=F_BODY).pack(anchor="w", pady=(6,0))
        tk.Label(cq, text="💡 Leelawadee UI / Noto = คมบนความร้อน • เปลี่ยนแล้วกดบันทึก แล้วทดสอบพิมพ์",
                 font=(FM.primary,10), bg=C_SURFACE2, fg=C_MUTED).pack(anchor="w", pady=(4,0))

        # ── ระยะขอบ + โลโก้ ──
        c3 = self._card(inner, "ระยะขอบ / โลโก้")
        mf = tk.Frame(c3, bg=C_SURFACE2); mf.pack(fill=tk.X)
        for lbl, key in [("ขอบซ้าย", "left_margin"), ("ขอบขวา", "right_margin")]:
            col = tk.Frame(mf, bg=C_SURFACE2); col.pack(side=tk.LEFT, padx=(0,16))
            tk.Label(col, text=lbl, font=F_SMALL, bg=C_SURFACE2, fg=C_MUTED).pack(anchor="w")
            e = field(col, width=8); e.insert(0, get_setting(key))
            e.pack(anchor="w", ipady=4); self.entries[key] = e
        # ── โลโก้ใบเสร็จ (ใช้งานได้จริง) ──
        self.logo_on = tk.BooleanVar(value=(get_setting("logo_enabled") or "1") != "0")
        tk.Checkbutton(c3, text="เปิดใช้งานโลโก้บนใบเสร็จ", variable=self.logo_on,
                       bg=C_SURFACE2, fg=C_TEXT, selectcolor=C_BG, activebackground=C_SURFACE2,
                       activeforeground=C_TEXT, font=F_BODY).pack(anchor="w", pady=(10,2))
        tk.Label(c3, text="โลโก้ใบเสร็จ  (PNG / JPG / JPEG / WEBP)", font=F_SMALL,
                 bg=C_SURFACE2, fg=C_MUTED).pack(anchor="w", pady=(2,2))
        lrow = tk.Frame(c3, bg=C_SURFACE2); lrow.pack(fill=tk.X)
        e_logo = field(lrow, width=38)
        try: e_logo.insert(0, get_setting("logo_path") or "")
        except Exception: pass
        e_logo.configure(state="readonly")
        e_logo.pack(side=tk.LEFT, ipady=4, fill=tk.X, expand=True)
        self.entries["logo_path"] = e_logo
        # ปุ่ม
        brow = tk.Frame(c3, bg=C_SURFACE2); brow.pack(fill=tk.X, pady=(6,2))
        pill_btn(brow, "📁 เลือกไฟล์", self._pick_logo, bg=C_SURFACE, fg=C_BLUE, pad_x=12, pad_y=6).pack(side=tk.LEFT, padx=(0,6))
        pill_btn(brow, "🧹 ล้างโลโก้", self._clear_logo, bg=C_SURFACE, fg=C_MUTED, pad_x=12, pad_y=6).pack(side=tk.LEFT, padx=(0,6))
        pill_btn(brow, "🖨️ ทดสอบพิมพ์", self._test_logo_print, bg=C_SURFACE, fg=C_GREEN, pad_x=12, pad_y=6).pack(side=tk.LEFT)
        # ขนาดโลโก้
        szrow = tk.Frame(c3, bg=C_SURFACE2); szrow.pack(fill=tk.X, pady=(8,2))
        tk.Label(szrow, text="ขนาดโลโก้:", font=F_SMALL, bg=C_SURFACE2, fg=C_MUTED).pack(side=tk.LEFT, padx=(0,8))
        _szmap = {"small":"เล็ก","medium":"กลาง","large":"ใหญ่"}
        _cur = (get_setting("receipt_logo_size") or "medium").strip()
        self._logo_sz_cb = ttk.Combobox(szrow, values=["เล็ก","กลาง","ใหญ่"],
                                         width=10, state="readonly", font=F_BODY)
        self._logo_sz_cb.set(_szmap.get(_cur, "กลาง"))
        self._logo_sz_cb.pack(side=tk.LEFT)
        self._logo_sz_cb.bind("<<ComboboxSelected>>", lambda e: self._refresh_logo_preview())
        # Preview (กรอบพื้นเข้ม จัดกลาง จำกัดขนาด)
        tk.Label(c3, text="ตัวอย่าง (ตามที่จะพิมพ์จริง):", font=F_SMALL,
                 bg=C_SURFACE2, fg=C_MUTED).pack(anchor="w", pady=(8,2))
        self._logo_canvas = tk.Canvas(c3, width=280, height=140, bg="#0d1117",
                                      highlightthickness=1, highlightbackground="#2b3340")
        self._logo_canvas.pack(anchor="w", pady=(0,4))
        tk.Label(c3, text="💡 เลือกภาพอะไรก็ได้ ระบบย่อ/แปลงขาวดำให้อัตโนมัติ — ไม่ต้องแต่งรูปเอง",
                 font=(FM.primary,10), bg=C_SURFACE2, fg="#6db86d").pack(anchor="w", pady=(2,0))
        self._refresh_logo_preview()

        self._save_row(inner)
        return p

    def _build_printer(self):
        p = tk.Frame(self._panel_area, bg=C_BG)
        self._panel_hdr(p, "🖨️  เครื่องพิมพ์", "ตั้งค่าอุปกรณ์เครื่องพิมพ์และกระดาษ")
        _, inner = self._scrollable(p)

        # ── การตั้งค่าพื้นฐาน ──
        c = self._card(inner, "การตั้งค่าพื้นฐาน")
        tk.Label(c, text="เครื่องพิมพ์ (จากที่ติดตั้งใน Windows)", font=F_SMALL,
                 bg=C_SURFACE2, fg=C_MUTED).pack(anchor="w", pady=(0,2))
        rowp2 = tk.Frame(c, bg=C_SURFACE2); rowp2.pack(fill=tk.X, pady=(0,4))
        _printers = self._list_windows_printers()
        cb = ttk.Combobox(rowp2, values=_printers, width=34, font=F_BODY, state="normal")
        cb.set(get_setting("printer_name"))
        cb.pack(side=tk.LEFT, ipady=2); self.entries["printer_name"] = cb
        pill_btn(rowp2, "🔄", lambda: cb.configure(values=self._list_windows_printers()),
                 bg=C_SURFACE, fg=C_BLUE, pad_x=10, pad_y=6).pack(side=tk.LEFT, padx=(8,0))
        if not _printers:
            tk.Label(c, text="💡 ไม่พบรายชื่อเครื่องพิมพ์ — พิมพ์ชื่อเองได้ หรือใช้ VID/PID ใน ▼ ตั้งค่าขั้นสูง",
                     font=(FM.primary,10), bg=C_SURFACE2, fg=C_MUTED).pack(anchor="w", pady=(0,4))

        # ขนาดกระดาษ
        tk.Label(c, text="ขนาดกระดาษ", font=F_SMALL, bg=C_SURFACE2, fg=C_MUTED).pack(anchor="w", pady=(6,2))
        psf = tk.Frame(c, bg=C_SURFACE2); psf.pack(fill=tk.X)
        for val, lbl in [("58mm","58mm"),("80mm","80mm"),("9x5.5",'9" x 5.5" ต่อเนื่อง'),("A4","A4")]:
            tk.Radiobutton(psf, text=lbl, variable=self.paper_size, value=val,
                           bg=C_SURFACE2, fg=C_TEXT, selectcolor=C_BG,
                           activebackground=C_SURFACE2, activeforeground=C_TEXT,
                           font=F_BODY).pack(side=tk.LEFT, padx=(0,14))

        # ตัวเลือก
        of = tk.Frame(c, bg=C_SURFACE2); of.pack(fill=tk.X, pady=(8,0))
        for txt, var in [("พิมพ์ใบเสร็จอัตโนมัติหลังขาย", self.auto_print),
                         ("แสดงตัวอย่างก่อนพิมพ์", self.preview_print),
                         ("พิมพ์สำเนา (สำเนาลูกค้า / ร้าน / บัญชี)", self.print_copy)]:
            tk.Checkbutton(of, text=txt, variable=var, bg=C_SURFACE2, fg=C_TEXT,
                           selectcolor=C_BG, activebackground=C_SURFACE2,
                           activeforeground=C_TEXT, font=F_BODY).pack(anchor="w", pady=1)
        tk.Label(c, text="🔧 ขนาดกระดาษ: บันทึกค่าไว้ จะมีผลเมื่อเปิดใช้งานในเวอร์ชันถัดไป",
                 font=(FM.primary,10), bg=C_SURFACE2, fg="#e67e22").pack(anchor="w", pady=(8,0))
        tk.Label(c, text="ℹ️ ตัดกระดาษ/ลิ้นชักเงิน ตั้งค่าได้ที่ Windows → Printer Properties",
                 font=(FM.primary,10), bg=C_SURFACE2, fg=C_MUTED).pack(anchor="w", pady=(2,0))

        # ── Advanced (collapsible) ──
        adv_bar = tk.Frame(inner, bg=C_BG); adv_bar.pack(fill=tk.X, padx=20, pady=(0,2))
        adv_state = {"open": False}
        adv_btn = tk.Label(adv_bar, text="▼  ตั้งค่าขั้นสูง (Advanced)",
                           font=(FM.primary,11,"bold"), bg=C_BG, fg=C_MUTED, cursor="hand2")
        adv_btn.pack(anchor="w")

        adv = tk.Frame(inner, bg=C_BG)   # ยังไม่ pack — ซ่อนไว้
        ca = self._card(adv, "การเชื่อมต่อ (เฉพาะผู้ดูแล)")
        for lbl, key in [("Printer VID  (เช่น 04b8)","printer_vid"),
                         ("Printer PID  (เช่น 0202)","printer_pid"),
                         ("COM Port  (เช่น COM3)","com_port"),
                         ("Baudrate  (เช่น 9600)","baudrate"),
                         ("Encoding  (เช่น tis-620 / cp874)","encoding")]:
            self._frow(ca, lbl, key, width=20)
        tk.Checkbutton(ca, text="ESC/POS direct mode (USB RAW)", variable=self.escpos_direct,
                       bg=C_SURFACE2, fg=C_TEXT, selectcolor=C_BG,
                       activebackground=C_SURFACE2, font=F_BODY).pack(anchor="w", pady=(6,0))
        pill_btn(ca, "🔍  ค้นหา VID/PID อัตโนมัติ (USB)", self._detect_printers,
                 bg=C_SURFACE, fg=C_BLUE, pad_x=12, pad_y=6).pack(anchor="w", pady=(6,0))
        hf = tk.Frame(ca, bg=C_SURFACE, padx=10, pady=8); hf.pack(fill=tk.X, pady=(8,0))
        tk.Label(hf, text="วิธีหา VID/PID: Device Manager → Properties → Hardware IDs → VID_XXXX&PID_XXXX",
                 font=F_SMALL, bg=C_SURFACE, fg=C_MUTED, wraplength=520, justify="left").pack(anchor="w")

        def _toggle_adv():
            if adv_state["open"]:
                adv.pack_forget(); adv_btn.config(text="▼  ตั้งค่าขั้นสูง (Advanced)")
                adv_state["open"] = False
            else:
                adv.pack(fill=tk.X, after=adv_bar); adv_btn.config(text="▲  ตั้งค่าขั้นสูง (Advanced)")
                adv_state["open"] = True
        adv_btn.bind("<Button-1>", lambda e: _toggle_adv())

        self._save_row(inner, extra_btns=[("🖨️  ทดสอบพิมพ์", self._test_print, C_SURFACE2, C_TEXT)])
        return p

    def _list_windows_printers(self):
        """รายชื่อเครื่องพิมพ์ที่ติดตั้งใน Windows — หลายวิธี fallback (ไม่บังคับต้องมี pywin32)"""
        names = []
        def _add(n):
            n = (n or "").strip()
            if n and n not in names:
                names.append(n)

        # 1) pywin32 (ถ้ามี — ครบสุด รวม network)
        try:
            import win32print
            flags = win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS
            for p in win32print.EnumPrinters(flags):
                _add(p[2])
            if names:
                return names
        except Exception:
            pass

        # 2) winreg — มากับ Python ทุกเครื่อง Windows ไม่ต้องติดตั้งอะไร
        try:
            import winreg
            sources = [
                (winreg.HKEY_CURRENT_USER,
                 r"Software\Microsoft\Windows NT\CurrentVersion\Devices", "values"),
                (winreg.HKEY_LOCAL_MACHINE,
                 r"SYSTEM\CurrentControlSet\Control\Print\Printers", "subkeys"),
            ]
            for root, path, kind in sources:
                try:
                    k = winreg.OpenKey(root, path)
                except Exception:
                    continue
                try:
                    i = 0
                    while True:
                        try:
                            n = winreg.EnumValue(k, i)[0] if kind == "values" \
                                else winreg.EnumKey(k, i)
                        except OSError:
                            break
                        _add(n); i += 1
                finally:
                    winreg.CloseKey(k)
            if names:
                return names
        except Exception:
            pass

        # 3) PowerShell Get-Printer (สำรองสุดท้าย)
        try:
            import subprocess
            out = subprocess.run(
                ["powershell", "-NoProfile", "-Command",
                 "Get-Printer | Select-Object -ExpandProperty Name"],
                capture_output=True, text=True, timeout=8,
                creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0))
            if out.returncode == 0:
                for ln in out.stdout.splitlines():
                    _add(ln)
        except Exception:
            pass

        return names

    def _detect_printers(self):
        """ค้นหาอุปกรณ์ USB (best-effort) → เลือกเพื่อเติม VID/PID ในตั้งค่าขั้นสูง"""
        try:
            import usb.core, usb.util
        except Exception:
            messagebox.showinfo("ค้นหาอัตโนมัติ",
                "ต้องติดตั้ง pyusb ก่อน  (pip install pyusb)\n\n"
                "หรือกรอก VID/PID เองใน ▼ ตั้งค่าขั้นสูง",
                parent=self.winfo_toplevel()); return
        found = []
        try:
            for d in usb.core.find(find_all=True):
                vid = f"{d.idVendor:04x}"; pid = f"{d.idProduct:04x}"
                name = ""
                try: name = usb.util.get_string(d, d.iProduct) or ""
                except Exception: pass
                found.append((vid, pid, name))
        except Exception as ex:
            messagebox.showwarning("ค้นหาอัตโนมัติ",
                f"สแกน USB ไม่สำเร็จ:\n{ex}\n\nลองกรอก VID/PID เองในตั้งค่าขั้นสูง",
                parent=self.winfo_toplevel()); return
        if not found:
            messagebox.showinfo("ค้นหาอัตโนมัติ", "ไม่พบอุปกรณ์ USB",
                                parent=self.winfo_toplevel()); return

        win = tk.Toplevel(self); win.title("เลือกเครื่องพิมพ์"); win.configure(bg=C_BG)
        win.grab_set(); win.resizable(False, False); win.lift()
        tk.Label(win, text="พบอุปกรณ์ USB — เลือกเครื่องพิมพ์",
                 font=F_H2, bg=C_BG, fg=C_TEXT, pady=8).pack(padx=16)
        lb = tk.Listbox(win, width=46, height=min(12, len(found)),
                        bg=C_SURFACE, fg=C_TEXT, font=(FM.primary,11),
                        selectbackground=C_ACCENT, relief=tk.FLAT)
        for vid, pid, name in found:
            lb.insert(tk.END, f"{vid}:{pid}   {name}".rstrip())
        lb.pack(padx=16, pady=8, fill=tk.X)

        def choose():
            sel = lb.curselection()
            if not sel: return
            vid, pid, name = found[sel[0]]
            for key, val in [("printer_vid", vid), ("printer_pid", pid),
                             ("printer_name", name or f"{vid}:{pid}")]:
                e = self.entries.get(key)
                if e is not None:
                    e.delete(0, tk.END); e.insert(0, val)
            win.destroy()
            messagebox.showinfo("เลือกแล้ว",
                f"ตั้งค่า VID={vid} PID={pid}\n\nอย่าลืมกด 💾 บันทึก",
                parent=self.winfo_toplevel())

        bf = tk.Frame(win, bg=C_BG, pady=8); bf.pack(fill=tk.X, padx=16)
        accent_btn(bf, "✔  เลือก", choose).pack(side=tk.LEFT)
        pill_btn(bf, "ยกเลิก", win.destroy, bg=C_SURFACE, fg=C_MUTED).pack(side=tk.RIGHT)

    def _build_inventory(self):
        p = tk.Frame(self._panel_area, bg=C_BG)
        self._panel_hdr(p, "📦  สต็อก", "นโยบายการควบคุมจำนวนสินค้าคงเหลือ")
        _, inner = self._scrollable(p)
        c = self._card(inner, "นโยบายสต็อกติดลบ")
        for val, lbl, desc, col in [
            ("allow",  "✅  อนุญาตให้ติดลบ", "ขายได้ทันที ไม่ตรวจสอบ  เหมาะ: ร้านขายส่ง/อะไหล่","#3fb950"),
            ("warn",   "⚠️  แจ้งเตือนแต่ยังขายได้  (ค่าเริ่มต้น)", "แสดงคำเตือนและถาม  เหมาะ: ร้านค้าทั่วไป","#e67e22"),
            ("strict", "🚫  ห้ามสต็อกติดลบ", "บล็อกการขายถ้าสต็อกไม่พอ  เหมาะ: ควบคุมเข้มงวด","#ff6b6b"),
        ]:
            rf = tk.Frame(c, bg=C_SURFACE2); rf.pack(fill=tk.X, pady=3)
            tk.Radiobutton(rf, text=lbl, variable=self.neg_policy, value=val,
                           bg=C_SURFACE2, fg=col, selectcolor=C_BG,
                           activebackground=C_SURFACE2, activeforeground=col,
                           font=(FM.primary,11,"bold")).pack(anchor="w")
            tk.Label(rf, text=f"     {desc}", font=(FM.primary,10),
                     bg=C_SURFACE2, fg=C_MUTED).pack(anchor="w")
        c2 = self._card(inner, "การแจ้งเตือน")
        tk.Checkbutton(c2, text="แจ้งเตือนสต็อกต่ำเมื่อเปิดโปรแกรม",
                       variable=self.lsv, bg=C_SURFACE2, fg=C_TEXT,
                       selectcolor=C_BG, activebackground=C_SURFACE2, font=F_BODY).pack(anchor="w")
        c3 = self._card(inner, "ระบบกะ (Shift)")
        tk.Checkbutton(c3, text="บังคับเปิดกะก่อนขายสินค้า",
                       variable=self.shift_req, bg=C_SURFACE2, fg=C_TEXT,
                       selectcolor=C_BG, activebackground=C_SURFACE2,
                       font=(FM.primary,11,"bold")).pack(anchor="w")
        tk.Label(c3, text="     เมื่อเปิด: ต้องกด “เปิดกะ” ก่อนจึงจะขายได้  (ปิด = ขายได้ตามปกติ)",
                 font=(FM.primary,10), bg=C_SURFACE2, fg=C_MUTED).pack(anchor="w")
        self._save_row(inner)
        return p

    def _build_database(self):
        p = tk.Frame(self._panel_area, bg=C_BG)
        self._panel_hdr(p, "💾  ฐานข้อมูล", "เลือกโหมดการเก็บข้อมูล — รองรับ Cloud ในอนาคต")
        _, inner = self._scrollable(p)
        c = self._card(inner, "Database Mode")
        for val, lbl, desc, col, ena in [
            ("sqlite_offline","💾  SQLite Offline  (ปัจจุบัน)","ข้อมูลบนเครื่อง — เร็ว เสถียร","#3fb950",True),
            ("sqlite_network","🔗  SQLite Network Share","แชร์ .db ผ่าน LAN — หลายเครื่องในวงเดียวกัน","#58a6ff",True),
            ("cloud","☁️  Cloud Edition","API + PostgreSQL — Multi-branch, Mobile","#ffd166",False),
        ]:
            rf = tk.Frame(c, bg=C_SURFACE2); rf.pack(fill=tk.X, pady=3)
            rh = tk.Frame(rf, bg=C_SURFACE2); rh.pack(fill=tk.X)
            tk.Radiobutton(rh, text=lbl, variable=self.db_mode_var, value=val,
                           bg=C_SURFACE2, fg=col, selectcolor=C_BG,
                           activebackground=C_SURFACE2, activeforeground=col,
                           font=(FM.primary,11,"bold"),
                           state="normal" if ena else "disabled").pack(side=tk.LEFT)
            if not ena:
                tk.Label(rh, text="เร็วๆ นี้", font=(FM.primary,9),
                         bg="#332200", fg="#ffd166", padx=6, pady=2).pack(side=tk.LEFT, padx=6)
            tk.Label(rf, text=f"     {desc}", font=(FM.primary,10),
                     bg=C_SURFACE2, fg=C_MUTED).pack(anchor="w")
        if not HAS_SERVICES:
            wf = tk.Frame(c, bg="#2d1a00", padx=10, pady=8); wf.pack(fill=tk.X, pady=(8,0))
            tk.Label(wf, text="⚠️  ไม่พบ db_services.py — วางไว้ข้างๆ rakcomsoft.py",
                     font=F_SMALL, bg="#2d1a00", fg="#e67e22").pack(anchor="w")
        c2 = self._card(inner, "Roadmap")
        for txt in ["PostgreSQL","Cloud Sync","Multi Branch","REST API"]:
            tk.Label(c2, text=f"○  {txt}  (อนาคต)", font=F_SMALL, bg=C_SURFACE2, fg=C_MUTED).pack(anchor="w", pady=1)
        self._save_row(inner)
        return p

    def _build_vat(self):
        p=tk.Frame(self._panel_area,bg=C_BG)
        self._panel_hdr(p,"🧮  ภาษี VAT","ตั้งค่าภาษีมูลค่าเพิ่ม — รองรับ VAT ใน/นอก")
        _,inner=self._scrollable(p)
        c1=self._card(inner,"การใช้งาน VAT")
        vat_en=tk.BooleanVar(value=get_setting("vat_enabled")=="1")
        def _toggle():
            val = "1" if vat_en.get() else "0"
            set_setting("vat_enabled", val)
            print(f"[VAT] vat_enabled = {val}")
            try: _upd_prev()
            except: pass
        tk.Checkbutton(c1,text="ใช้งาน VAT (ภาษีมูลค่าเพิ่ม)",variable=vat_en,command=_toggle,
                       bg=C_SURFACE2,fg=C_TEXT,selectcolor=C_BG,activebackground=C_SURFACE2,
                       font=(FM.primary,12,"bold")).pack(anchor="w",pady=(0,4))
        tk.Label(c1,text="เปิดเมื่อร้านจด VAT กับสรรพากร",font=F_SMALL,bg=C_SURFACE2,fg=C_MUTED).pack(anchor="w")
        c2=self._card(inner,"อัตราภาษี")
        r2=tk.Frame(c2,bg=C_SURFACE2); r2.pack(anchor="w")
        tk.Label(r2,text="VAT %:",font=F_BODY,bg=C_SURFACE2,fg=C_TEXT).pack(side=tk.LEFT)
        vat_pct=tk.StringVar(value=get_setting("vat_percent") or "7")
        pe=field(r2,width=6,justify="right",font=(FM.primary,14,"bold")); pe.configure(textvariable=vat_pct); pe.pack(side=tk.LEFT,padx=(8,4),ipady=4)
        tk.Label(r2,text="%",font=F_BODY,bg=C_SURFACE2,fg=C_MUTED).pack(side=tk.LEFT)
        def _save_pct(*_):
            try: set_setting("vat_percent",str(float(vat_pct.get() or 0))); _upd_prev()
            except: pass
        pe.bind("<FocusOut>",_save_pct); pe.bind("<Return>",_save_pct)
        c3=self._card(inner,"รูปแบบ VAT")
        vat_type=tk.StringVar(value=get_setting("vat_type") or "excluded")
        def _save_type(): set_setting("vat_type",vat_type.get()); _upd_prev()
        for val,lbl,desc in [
            ("excluded","VAT นอก (Exclusive)","ราคา+VAT=ยอดรวม  ฿100+฿7=฿107"),
            ("included","VAT ใน (Inclusive)","ราคารวม VAT  ฿107→VAT฿7→ก่อน VAT฿100"),
        ]:
            rf=tk.Frame(c3,bg=C_SURFACE2); rf.pack(anchor="w",pady=3)
            tk.Radiobutton(rf,text=lbl,variable=vat_type,value=val,command=_save_type,
                           bg=C_SURFACE2,fg=C_TEXT,selectcolor=C_BG,activebackground=C_SURFACE2,
                           font=(FM.primary,11,"bold")).pack(side=tk.LEFT)
            tk.Label(rf,text=f"  {desc}",font=F_SMALL,bg=C_SURFACE2,fg=C_MUTED).pack(side=tk.LEFT)
        c4=self._card(inner,"ข้อมูลผู้ประกอบการ (ใบกำกับภาษี)")
        for lbl,key in [("เลขผู้เสียภาษี 13 หลัก","store_tax_id"),("สำนักงาน/สาขา","store_branch")]:
            tk.Label(c4,text=lbl,font=F_SMALL,bg=C_SURFACE2,fg=C_MUTED).pack(anchor="w",pady=(6,2))
            e=field(c4,width=28); e.pack(fill=tk.X,ipady=5)
            v=get_setting(key) or ""
            if v: e.insert(0,v)
            def _sv(ev,k=key,en=e): set_setting(k,en.get().strip())
            e.bind("<FocusOut>",_sv); e.bind("<Return>",_sv)
        c5=self._card(inner,"ตัวอย่าง Preview")
        prev_lbl=tk.Label(c5,text="",font=("Courier New",11),bg=C_SURFACE2,fg=C_TEXT,justify="left",pady=6); prev_lbl.pack(anchor="w",padx=8)
        def _upd_prev(*_):
            try:
                pct=float(vat_pct.get() or 0); vt=vat_type.get(); en=vat_en.get()
                if not en or pct==0: prev_lbl.config(text="  VAT ปิดอยู่",fg=C_MUTED); return
                if vt=="excluded":
                    prev_lbl.config(text=f"  VAT นอก:\n  ยอดสินค้า  ฿100.00\n  VAT {pct:.0f}%     ฿{100*pct/100:6.2f}\n  รวม        ฿{100+100*pct/100:.2f}",fg=C_TEXT)
                else:
                    bf=round(10000/(100+pct),2); va=round(100-bf,2)
                    prev_lbl.config(text=f"  VAT ใน:\n  ราคาขาย   ฿100.00\n  VAT {pct:.0f}% ใน ฿{va:6.2f}\n  ก่อน VAT  ฿{bf:.2f}",fg=C_TEXT)
            except: pass
        _upd_prev()
        return p

    def _build_system(self):
        p = tk.Frame(self._panel_area, bg=C_BG)
        self._panel_hdr(p, "⚙️  ระบบ", "ข้อมูลเวอร์ชันและการตั้งค่าทั่วไป")
        _, inner = self._scrollable(p)
        c = self._card(inner, "เกี่ยวกับโปรแกรม")
        for lbl, val in [("โปรแกรม","RakComSoft 2026"),("เวอร์ชัน",f"v{APP_VERSION}"),
                         ("ฐานข้อมูล","SQLite 3"),("UI Framework","Python Tkinter")]:
            row = tk.Frame(c, bg=C_SURFACE2); row.pack(fill=tk.X, pady=2)
            tk.Label(row, text=lbl, font=F_SMALL, bg=C_SURFACE2, fg=C_MUTED, width=14, anchor="w").pack(side=tk.LEFT)
            tk.Label(row, text=val, font=F_BODY, bg=C_SURFACE2, fg=C_TEXT).pack(side=tk.LEFT)
        c2 = self._card(inner, "License")
        tk.Button(c2, text="🔑  License / เกี่ยวกับโปรแกรม",
                  font=F_BODY, bg=C_SURFACE2, fg=C_TEXT,
                  relief=tk.FLAT, cursor="hand2", padx=12, pady=6,
                  command=self._show_license).pack(anchor="w")

        # ── การอัปเดตโปรแกรม ──
        c3 = self._card(inner, "การอัปเดตโปรแกรม")
        r1 = tk.Frame(c3, bg=C_SURFACE2); r1.pack(fill=tk.X, pady=3)
        tk.Label(r1, text="ช่องทางอัปเดต", font=F_SMALL, bg=C_SURFACE2, fg=C_MUTED,
                 width=16, anchor="w").pack(side=tk.LEFT)
        self._upd_ch = tk.StringVar(value=get_setting("update_channel") or "stable")
        ttk.Combobox(r1, textvariable=self._upd_ch, values=["stable","beta"],
                     width=12, state="readonly", font=F_BODY).pack(side=tk.LEFT)
        tk.Label(c3, text="stable = เสถียร แนะนำสำหรับร้านค้า  •  beta = รุ่นทดลอง",
                 font=(FM.primary,9), bg=C_SURFACE2, fg=C_MUTED).pack(anchor="w", pady=(0,4))
        self._upd_auto = tk.BooleanVar(value=(get_setting("update_autocheck") or "1") != "0")
        tk.Checkbutton(c3, text="ตรวจสอบอัปเดตอัตโนมัติเมื่อเปิดโปรแกรม",
                       variable=self._upd_auto, bg=C_SURFACE2, fg=C_TEXT,
                       selectcolor=C_SURFACE, activebackground=C_SURFACE2,
                       activeforeground=C_TEXT, font=F_BODY).pack(anchor="w")
        tk.Button(c3, text="🔄  ตรวจสอบอัปเดตเดี๋ยวนี้",
                  font=F_BODY, bg=C_ACCENT, fg=C_BG, relief=tk.FLAT, cursor="hand2",
                  padx=12, pady=6,
                  command=lambda: self.winfo_toplevel()._show_update_page()).pack(anchor="w", pady=(6,0))
        return p

    # ════ ACTIONS ════════════════════════════════════════
    def _save(self):
        for key, e in self.entries.items(): set_setting(key, e.get().strip())
        # การอัปเดต (channel + auto-check)
        if hasattr(self, "_upd_ch"):
            set_setting("update_channel", self._upd_ch.get().strip() or "stable")
        if hasattr(self, "_upd_auto"):
            set_setting("update_autocheck", "1" if self._upd_auto.get() else "0")
        # ขนาดโลโก้ (combobox เก็บค่าไทย -> แปลงเป็น key อังกฤษ)
        if hasattr(self, "_logo_sz_cb"):
            _rev = {"เล็ก":"small","กลาง":"medium","ใหญ่":"large"}
            set_setting("receipt_logo_size", _rev.get(self._logo_sz_cb.get().strip(), "medium"))
        # คุณภาพตัวอักษร / ฟอนต์
        if hasattr(self, "_rfont_cb"):
            _fr = {"อัตโนมัติ":"auto","Leelawadee UI":"leelawadee","Noto Sans Thai":"noto",
                   "TH Sarabun":"sarabun","Tahoma":"tahoma","Kanit":"kanit"}
            set_setting("receipt_font", _fr.get(self._rfont_cb.get().strip(), "auto"))
        if hasattr(self, "_rsize_cb"):
            _sr = {"เล็ก":"small","กลาง":"medium","ใหญ่":"large"}
            set_setting("receipt_font_size", _sr.get(self._rsize_cb.get().strip(), "medium"))
        if hasattr(self, "_rspace_cb"):
            _pr = {"กระชับ":"compact","ปกติ":"normal","โปร่ง":"relaxed"}
            set_setting("receipt_line_spacing", _pr.get(self._rspace_cb.get().strip(), "normal"))
        if hasattr(self, "r_sharpen"):
            set_setting("receipt_sharpen", "1" if self.r_sharpen.get() else "0")
        if hasattr(self, "logo_on"):
            set_setting("logo_enabled", "1" if self.logo_on.get() else "0")
        set_setting("low_stock_alert", "1" if self.lsv.get() else "0")
        set_setting("neg_stock_policy", self.neg_policy.get())
        if hasattr(self, "shift_req"):
            set_setting("shift_required", "1" if self.shift_req.get() else "0")
        # ── Printer/Receipt UX (Phase A) ──
        if hasattr(self, "paper_size"):
            set_setting("paper_size", self.paper_size.get())
            for var, key in [(self.preview_print,"preview_before_print"),
                             (self.auto_print,"auto_print_receipt"),
                             (self.print_copy,"print_copy"),
                             (self.escpos_direct,"escpos_direct"),
                             (self.show_vat,"show_vat"),(self.show_staff,"show_staff"),
                             (self.show_points,"show_points")]:
                set_setting(key, "1" if var.get() else "0")
        if HAS_SERVICES and hasattr(self, "db_mode_var"):
            new_mode = self.db_mode_var.get()
            if new_mode != DB_CONFIG.get("mode"):
                set_db_mode(new_mode); Services.reset()
        messagebox.showinfo("✅  บันทึกแล้ว", "บันทึกการตั้งค่าเรียบร้อย")

    def _pick_logo(self):
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            title="เลือกไฟล์โลโก้",
            filetypes=[("รูปภาพ","*.png *.jpg *.jpeg *.webp"),
                       ("PNG","*.png"),("JPEG","*.jpg *.jpeg"),("WEBP","*.webp"),("ทั้งหมด","*.*")],
            parent=self.winfo_toplevel())
        if not path: return
        e = self.entries.get("logo_path")
        if e is not None:
            e.configure(state="normal"); e.delete(0, tk.END); e.insert(0, path); e.configure(state="readonly")
        self._refresh_logo_preview()

    def _clear_logo(self):
        e = self.entries.get("logo_path")
        if e is not None:
            e.configure(state="normal"); e.delete(0, tk.END); e.configure(state="readonly")
        self._refresh_logo_preview()

    def _refresh_logo_preview(self):
        cv = getattr(self, "_logo_canvas", None)
        if cv is None: return
        cv.delete("all")
        W = int(cv["width"]); H = int(cv["height"])
        path = ""
        e = self.entries.get("logo_path")
        if e is not None:
            try: path = e.get().strip()
            except Exception: path = ""
        if not path:
            cv.create_text(W//2, H//2, text="— ไม่มีโลโก้ —", fill="#6b7280", font=(FM.primary,11))
            return
        try:
            import receipt_logo
            img = receipt_logo.load_receipt_logo(path)
            if img is None: raise ValueError("load failed")
            gray = receipt_logo._flatten_white(img)
            _rev = {"เล็ก":"small","กลาง":"medium","ใหญ่":"large"}
            sz = _rev.get(self._logo_sz_cb.get().strip(), "medium") if hasattr(self,"_logo_sz_cb") else "medium"
            prevw = {"small":120,"medium":180,"large":240}.get(sz, 180)
            prevw = min(prevw, W-16)
            small = receipt_logo.resize_receipt_logo(gray, prevw)
            bw = small.convert("1").convert("L")           # โชว์ผลขาวดำจริง
            if bw.height > H-12:                            # จำกัดสูงไม่ให้ล้น
                r = (H-12)/float(bw.height)
                from PIL import Image as _IM
                bw = bw.resize((max(1,int(bw.width*r)), H-12),
                               getattr(_IM,"LANCZOS",1))
            self._logo_tkimg = ITk.PhotoImage(bw)           # เก็บ ref กัน GC
            cv.create_image(W//2, H//2, image=self._logo_tkimg)
        except Exception:
            cv.create_text(W//2, H//2, text="⚠️ โหลดภาพไม่ได้", fill="#e67e22", font=(FM.primary,11))

    def _test_logo_print(self):
        # บันทึกค่าโลโก้ปัจจุบันก่อน (ให้ ทดสอบพิมพ์ ใช้ภาพที่เพิ่งเลือก)
        e = self.entries.get("logo_path")
        if e is not None:
            try: set_setting("logo_path", e.get().strip())
            except Exception: pass
        if hasattr(self, "_logo_sz_cb"):
            _rev = {"เล็ก":"small","กลาง":"medium","ใหญ่":"large"}
            set_setting("receipt_logo_size", _rev.get(self._logo_sz_cb.get().strip(), "medium"))
        self._test_print()

    def _test_print(self):
        # ชื่อเครื่องพิมพ์ Windows ที่เลือกใน dropdown
        pname = ""
        e = self.entries.get("printer_name")
        if e is not None:
            try: pname = e.get().strip()
            except Exception: pname = ""

        test_text = (
            "================================\n"
            "        RakComSoft 2026\n"
            "      ทดสอบพิมพ์ / TEST PRINT\n"
            f"   {pname or '-'}\n"
            f"   {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n"
            "================================\n\n\n"
        )

        # 1) เครื่องพิมพ์ Windows (วิธีหลักตอนนี้) — ไม่ต้องลง python-escpos
        if pname:
            import receipt_raster
            ok, err = _win_print_raw(pname, receipt_raster.lines_to_escpos_raster(test_text.splitlines()))
            if ok:
                messagebox.showinfo("✅ ทดสอบพิมพ์",
                    f"ส่งงานพิมพ์ไปที่ \"{pname}\" แล้ว", parent=self.winfo_toplevel())
            else:
                messagebox.showerror("ทดสอบพิมพ์ไม่สำเร็จ",
                    f"พิมพ์ไปที่ \"{pname}\" ไม่ได้\n{err}\n\n"
                    "ตรวจสอบว่าเครื่องพิมพ์เปิดอยู่และตั้งเป็นพร้อมใช้งานใน Windows",
                    parent=self.winfo_toplevel())
            return

        # 2) Fallback: ESC/POS USB (VID/PID) — ของเดิม เผื่อใช้ thermal โดยตรง
        if not HAS_ESCPOS:
            messagebox.showinfo("ทดสอบพิมพ์",
                "เลือกเครื่องพิมพ์จาก dropdown ด้านบนก่อน\n\n"
                "หรือถ้าต้องการต่อ thermal โดยตรงผ่าน USB:\n"
                "  • pip install python-escpos\n"
                "  • กรอก VID/PID ใน ▼ ตั้งค่าขั้นสูง",
                parent=self.winfo_toplevel()); return
        vid = self.entries["printer_vid"].get().strip() if "printer_vid" in self.entries else ""
        pid = self.entries["printer_pid"].get().strip() if "printer_pid" in self.entries else ""
        if not vid or not pid:
            messagebox.showwarning("ขาดข้อมูล",
                "เลือกเครื่องพิมพ์จาก dropdown ด้านบน\nหรือกรอก VID/PID ใน ▼ ตั้งค่าขั้นสูง แล้วกดบันทึก",
                parent=self.winfo_toplevel()); return
        try:
            p=Usb(int(vid,16),int(pid,16),timeout=0,in_ep=0x82,out_ep=0x01)
            p.set(align="center",bold=True,double_height=True); p.text("Rakcomsoft\n")
            p.set(align="center",bold=False,double_height=False)
            p.text("Thermal Test OK\n"+datetime.now().strftime("%d/%m/%Y %H:%M:%S")+"\n")
            p.text("="*42+"\n\n\n"); p.cut(); p.close()
            messagebox.showinfo("✅","Test print สำเร็จ", parent=self.winfo_toplevel())
        except Exception as e:
            messagebox.showerror("ผิดพลาด",f"เชื่อมต่อไม่ได้:\n{e}\n\nตรวจสอบสาย USB และ VID/PID",
                                 parent=self.winfo_toplevel())

    def _show_license(self):
        try:
            from license_system import check_license
            info = check_license()
            messagebox.showinfo("License",
                f"Status : {info.get('status','—')}\nEdition: {info.get('edition','—')}\n"
                f"Expire : {info.get('expire_date','—')}\nHW ID  : {info.get('hardware_id','—')}")
        except Exception:
            messagebox.showinfo("เกี่ยวกับ","RakComSoft 2026 v2.1\nrakcomshop.com")


# ══ TAB: EOD ══════════════════════════════════════════════════════════════
class EODTab(tk.Frame):
    """สรุปเงินประจำวัน — End of Day Report"""

    def __init__(self, parent):
        super().__init__(parent, bg=C_BG)
        self._build()

    def _build(self):
        # ── Header ──────────────────────────────────────────
        hdr = tk.Frame(self, bg=C_BG, padx=14, pady=8)
        hdr.pack(fill=tk.X)
        tk.Label(hdr, text="💵  สรุปเงินประจำวัน",
                 font=F_H1, bg=C_BG, fg=C_TEXT).pack(side=tk.LEFT)

        # date range
        df = tk.Frame(hdr, bg=C_BG); df.pack(side=tk.RIGHT)
        tk.Label(df, text="วันที่:", font=F_BODY,
                 bg=C_BG, fg=C_MUTED).pack(side=tk.LEFT, padx=(0,4))
        self._eod_d0 = ThaiDateEntry(df, width=11)
        self._eod_d0.set(date.today().strftime("%Y-%m-%d"))
        self._eod_d0.pack(side=tk.LEFT, padx=2)
        tk.Label(df, text="ถึง", font=F_SMALL,
                 bg=C_BG, fg=C_MUTED).pack(side=tk.LEFT, padx=2)
        self._eod_d1 = ThaiDateEntry(df, width=11)
        self._eod_d1.set(date.today().strftime("%Y-%m-%d"))
        self._eod_d1.pack(side=tk.LEFT, padx=2)

        tk.Button(df, text="🔍 ค้นหา", command=self.load,
                  font=F_BODY, bg=C_ACCENT, fg=C_BG,
                  relief=tk.FLAT, cursor="hand2",
                  padx=12, pady=4, bd=0,
                  activebackground=C_ACCENT).pack(side=tk.LEFT, padx=(8,0))
        tk.Button(df, text="🖨️ พิมพ์", command=self._print,
                  font=F_BODY, bg=C_SURFACE2, fg=C_TEXT,
                  relief=tk.FLAT, cursor="hand2",
                  padx=12, pady=4, bd=0,
                  activebackground=C_SURFACE2).pack(side=tk.LEFT, padx=(4,0))

        tk.Frame(self, bg=C_BORDER, height=1).pack(fill=tk.X)

        # ── Body: 2 คอลัมน์ ──────────────────────────────────
        body = tk.Frame(self, bg=C_BG); body.pack(fill=tk.BOTH, expand=True, padx=14, pady=10)
        left  = tk.Frame(body, bg=C_BG); left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0,8))
        right = tk.Frame(body, bg=C_BG); right.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        # ── LEFT: ยอดขายแยกประเภทชำระ ───────────────────────
        def section(parent, title, accent):
            f = tk.Frame(parent, bg=C_SURFACE,
                         highlightbackground=C_BORDER, highlightthickness=1)
            f.pack(fill=tk.X, pady=(0,10))
            tk.Frame(f, bg=accent, height=3).pack(fill=tk.X)
            tk.Label(f, text=title, font=F_H2,
                     bg=C_SURFACE, fg=C_TEXT, padx=14, pady=8).pack(anchor="w")
            tk.Frame(f, bg=C_BORDER, height=1).pack(fill=tk.X)
            inner = tk.Frame(f, bg=C_SURFACE, padx=14, pady=8)
            inner.pack(fill=tk.X)
            return inner

        # ยอดขายแยกตามประเภทชำระ
        pay_inner = section(left, "💰  ยอดขายแยกตามประเภทชำระ", "#4a9eff")
        self._pay_rows = {}
        for method, label, color in [
            ("cash",    "เงินสด",      "#3fb950"),
            ("qr",      "QR/พร้อมเพย์","#00d4aa"),
            ("credit",  "บัตรเครดิต",  "#4a9eff"),
            ("debt",    "เงินเชื่อ",   "#e67e22"),
            ("mixed",   "ชำระผสม",     "#9b59b6"),
        ]:
            row = tk.Frame(pay_inner, bg=C_SURFACE); row.pack(fill=tk.X, pady=2)
            tk.Label(row, text=label, font=F_BODY,
                     bg=C_SURFACE, fg=C_MUTED, width=14, anchor="w").pack(side=tk.LEFT)
            lv = tk.Label(row, text="฿0.00", font=F_BODY,
                          bg=C_SURFACE, fg=color, anchor="e")
            lv.pack(side=tk.RIGHT)
            self._pay_rows[method] = lv
        tk.Frame(pay_inner, bg=C_BORDER, height=1).pack(fill=tk.X, pady=(6,4))
        row = tk.Frame(pay_inner, bg=C_SURFACE); row.pack(fill=tk.X)
        tk.Label(row, text="รวมทั้งสิ้น", font=(FM.primary,13,"bold"),
                 bg=C_SURFACE, fg=C_TEXT, width=14, anchor="w").pack(side=tk.LEFT)
        self._total_lbl = tk.Label(row, text="฿0.00",
                                   font=(FM.primary,16,"bold"),
                                   bg=C_SURFACE, fg=C_ACCENT, anchor="e")
        self._total_lbl.pack(side=tk.RIGHT)

        # จำนวนบิล
        bill_inner = section(left, "🧾  จำนวนบิล", "#ffd166")
        self._bill_rows = {}
        for method, label in [
            ("cash","เงินสด"), ("qr","QR"), ("credit","บัตร"),
            ("debt","เชื่อ"), ("mixed","ผสม")
        ]:
            row = tk.Frame(bill_inner, bg=C_SURFACE); row.pack(fill=tk.X, pady=2)
            tk.Label(row, text=label, font=F_BODY,
                     bg=C_SURFACE, fg=C_MUTED, width=14, anchor="w").pack(side=tk.LEFT)
            lv = tk.Label(row, text="0 บิล", font=F_BODY,
                          bg=C_SURFACE, fg=C_TEXT, anchor="e")
            lv.pack(side=tk.RIGHT)
            self._bill_rows[method] = lv
        tk.Frame(bill_inner, bg=C_BORDER, height=1).pack(fill=tk.X, pady=(6,4))
        row = tk.Frame(bill_inner, bg=C_SURFACE); row.pack(fill=tk.X)
        tk.Label(row, text="รวมบิลทั้งหมด", font=(FM.primary,13,"bold"),
                 bg=C_SURFACE, fg=C_TEXT, width=14, anchor="w").pack(side=tk.LEFT)
        self._bill_total_lbl = tk.Label(row, text="0 บิล",
                                        font=(FM.primary,14,"bold"),
                                        bg=C_SURFACE, fg="#ffd166", anchor="e")
        self._bill_total_lbl.pack(side=tk.RIGHT)

        # ── RIGHT: สรุปยอดขาย + ต้นทุน + กำไร ──────────────
        sum_inner = section(right, "📊  สรุปยอดขาย", "#3fb950")
        self._sum_rows = {}
        for key, label, color in [
            ("gross",   "ก่อนหักส่วนลด",   C_TEXT),
            ("disc",    "ส่วนลดรวม",       "#e67e22"),
            ("promo",   "ส่วนลดโปรโมชั่น", "#e67e22"),
            ("net",     "ยอดรวมสุทธิ",     C_ACCENT),
            ("sep",     "─────────────",   C_BORDER),
            ("cost",    "ต้นทุนรวม",       "#ff6b6b"),
            ("profit",  "กำไรรวม",         "#3fb950"),
            ("margin",  "อัตรากำไร (%)",   "#3fb950"),
        ]:
            if key == "sep":
                tk.Frame(sum_inner, bg=C_BORDER, height=1).pack(fill=tk.X, pady=6)
                continue
            row = tk.Frame(sum_inner, bg=C_SURFACE); row.pack(fill=tk.X, pady=2)
            is_bold = key in ("net","profit")
            fnt = (FM.primary, 14, "bold") if is_bold else F_BODY
            tk.Label(row, text=label, font=fnt,
                     bg=C_SURFACE, fg=C_MUTED if not is_bold else C_TEXT,
                     width=18, anchor="w").pack(side=tk.LEFT)
            lv = tk.Label(row, text="฿0.00" if key!="margin" else "0.00%",
                          font=fnt, bg=C_SURFACE, fg=color, anchor="e")
            lv.pack(side=tk.RIGHT)
            self._sum_rows[key] = lv

        # ลูกหนี้
        debt_inner = section(right, "📋  ลูกหนี้คงค้าง", "#e67e22")
        self._debt_rows = {}
        for key, label, color in [
            ("new_debt",   "เกิดใหม่วันนี้",   "#e67e22"),
            ("total_debt", "ค้างชำระทั้งหมด",  "#ff6b6b"),
            ("debtors",    "จำนวนลูกหนี้",     C_MUTED),
        ]:
            row = tk.Frame(debt_inner, bg=C_SURFACE); row.pack(fill=tk.X, pady=2)
            tk.Label(row, text=label, font=F_BODY,
                     bg=C_SURFACE, fg=C_MUTED, width=18, anchor="w").pack(side=tk.LEFT)
            lv = tk.Label(row, text="฿0.00" if "debt" in key else "0 ราย",
                          font=F_BODY, bg=C_SURFACE, fg=color, anchor="e")
            lv.pack(side=tk.RIGHT)
            self._debt_rows[key] = lv

        self.load()

    # ── LOAD DATA ────────────────────────────────────────────
    def load(self, *_):
        d0 = self._eod_d0.get_iso()
        d1 = self._eod_d1.get_iso()
        if not d0 or not d1: return

        # ── ยอดขายแยก payment_method — db_sale Session 8 ─
        pay_map  = {"cash":0,"qr":0,"credit":0,"debt":0,"mixed":0}
        bill_map = {"cash":0,"qr":0,"credit":0,"debt":0,"mixed":0}
        for r in get_eod_payment_summary(d0,d1):
            m = r["payment_method"] or "cash"
            if m not in pay_map: m = "cash"
            pay_map[m]  += r["s"] or 0
            bill_map[m] += r["c"] or 0

        total_net   = sum(pay_map.values())
        total_bills = sum(bill_map.values())

        for m, lv in self._pay_rows.items():
            lv.config(text=f"฿{pay_map[m]:,.2f}")
        self._total_lbl.config(text=f"฿{total_net:,.2f}")
        for m, lv in self._bill_rows.items():
            lv.config(text=f"{bill_map[m]:,} บิล")
        self._bill_total_lbl.config(text=f"{total_bills:,} บิล")

        # ── สรุปยอด — db_sale Session 8 ──────────────────
        r      = get_eod_summary(d0,d1)
        gross  = r["gross"]; disc = r["disc"]
        promo  = r["promo"]; net  = r["net"]
        cost   = get_eod_cost(d0,d1)
        profit = net - cost
        margin = (profit / net * 100) if net > 0 else 0

        self._sum_rows["gross"].config(text=f"฿{gross:,.2f}")
        self._sum_rows["disc"].config(text=f"-฿{disc:,.2f}")
        self._sum_rows["promo"].config(text=f"-฿{promo:,.2f}")
        self._sum_rows["net"].config(text=f"฿{net:,.2f}")
        self._sum_rows["cost"].config(text=f"฿{cost:,.2f}")
        self._sum_rows["profit"].config(
            text=f"฿{profit:,.2f}",
            fg="#3fb950" if profit>=0 else "#ff6b6b")
        self._sum_rows["margin"].config(
            text=f"{margin:.2f}%",
            fg="#3fb950" if profit>=0 else "#ff6b6b")

        # ── ลูกหนี้ — ★ single source of truth (db_receivable) ──
        debt = get_total_debt_summary(date_from=d0, date_to=d1)
        self._debt_rows["new_debt"].config(text=f"฿{debt['new_debt']:,.2f}")
        self._debt_rows["total_debt"].config(text=f"฿{debt['total_debt']:,.2f}")
        self._debt_rows["debtors"].config(text=f"{debt['debtors']:,} ราย")

    # ── PRINT ─────────────────────────────────────────────────
    def _print(self):
        d0 = self._eod_d0.get_iso(); d1 = self._eod_d1.get_iso()
        rows   = get_eod_payment_summary(d0,d1)   # db_sale — Session 8
        r      = get_eod_summary(d0,d1)            # db_sale — Session 8
        cost   = get_eod_cost(d0,d1)               # db_sale — Session 8
        net    = r["net"]
        profit = net - cost

        W = 42
        store = get_setting("store_name") or "ร้านค้า"
        lines = [
            "="*W, store.center(W),
            "สรุปเงินประจำวัน".center(W), "="*W,
            f"วันที่: {d0}" + (f" ถึง {d1}" if d0!=d1 else ""),
            "-"*W, "ยอดขายแยกประเภทชำระ", "-"*W,
        ]
        pay_lbl = {"cash":"เงินสด","qr":"QR/พร้อมเพย์",
                   "credit":"บัตรเครดิต","debt":"เงินเชื่อ","mixed":"ผสม"}
        total_n = 0
        for row in rows:
            m = row["payment_method"] or "cash"
            lbl = pay_lbl.get(m, m)
            lines.append(f"{lbl:<20} ฿{row['s']:>10,.2f}  ({row['c']} บิล)")
            total_n += row["c"]
        lines += [
            "-"*W,
            f"{'รวมทั้งสิ้น':<20} ฿{net:>10,.2f}  ({total_n} บิล)",
            "="*W, "สรุปกำไร", "-"*W,
            f"{'ก่อนหักส่วนลด':<20} ฿{r['gross']:>10,.2f}",
            f"{'ส่วนลด':<20}-฿{r['disc']:>10,.2f}",
            f"{'ยอดสุทธิ':<20} ฿{net:>10,.2f}",
            f"{'ต้นทุน':<20} ฿{cost:>10,.2f}",
            f"{'กำไร':<20} ฿{profit:>10,.2f}",
            "="*W,
        ]
        text = "\n".join(lines)
        import tempfile, os
        tmp = tempfile.NamedTemporaryFile(suffix=".txt",
                                          mode="w", encoding="utf-8",
                                          delete=False)
        tmp.write(text); tmp.close()
        win = tk.Toplevel(self.winfo_toplevel())
        win.title("สรุปเงินประจำวัน"); win.configure(bg=C_BG)
        win.geometry("520x600")
        t = tk.Text(win, font=(FM.primary,11), bg="#1a1a2e",
                    fg="#e8e8e8", relief=tk.FLAT, padx=14, pady=12)
        t.pack(fill=tk.BOTH, expand=True)
        t.insert("1.0", text); t.config(state="disabled")
        bf = tk.Frame(win, bg=C_BG, pady=8); bf.pack(fill=tk.X)
        if sys.platform == "win32":
            tk.Button(bf, text="🖨️  พิมพ์",
                      command=lambda: os.startfile(tmp.name, "print"),
                      font=F_BODY, bg=C_ACCENT, fg=C_BG,
                      relief=tk.FLAT, padx=14, pady=6, bd=0,
                      cursor="hand2").pack(side=tk.LEFT, padx=14)
        tk.Button(bf, text="📋  คัดลอก",
                  command=lambda:(win.clipboard_clear(),
                                  win.clipboard_append(text)),
                  font=F_BODY, bg=C_SURFACE2, fg=C_TEXT,
                  relief=tk.FLAT, padx=14, pady=6, bd=0,
                  cursor="hand2").pack(side=tk.LEFT, padx=4)
        tk.Button(bf, text="ปิด", command=win.destroy,
                  font=F_BODY, bg=C_SURFACE, fg=C_MUTED,
                  relief=tk.FLAT, padx=14, pady=6, bd=0,
                  cursor="hand2").pack(side=tk.RIGHT, padx=14)

# ══ MAIN ════════════════════════════════════════════════

def _show_import_preview(parent, products):
    """แสดง Preview สินค้าก่อน Import"""
    win = tk.Toplevel(parent)
    win.title("Preview — Import สินค้า")
    win.configure(bg=C_BG)
    win.geometry("960x680")
    win.lift()

    tk.Frame(win,bg=C_ACCENT,height=4).pack(fill=tk.X)
    hf = tk.Frame(win,bg=C_BG,padx=16,pady=10); hf.pack(fill=tk.X)
    tk.Label(hf,text="📥  Preview Import สินค้า",font=F_H1,bg=C_BG,fg=C_TEXT).pack(side=tk.LEFT)
    tk.Label(hf,text="พบ {} รายการ".format(len(products)),
             font=F_H2,bg=C_BG,fg=C_ACCENT).pack(side=tk.RIGHT)

    # Options
    of = tk.Frame(win,bg=C_SURFACE2,padx=16,pady=8); of.pack(fill=tk.X)
    skip_var  = tk.BooleanVar(value=True)
    update_var= tk.BooleanVar(value=False)
    tk.Checkbutton(of,text="ข้ามสินค้าที่มีบาร์โค้ดซ้ำในระบบแล้ว",
                   variable=skip_var,bg=C_SURFACE2,fg=C_TEXT,
                   selectcolor=C_SURFACE,activebackground=C_SURFACE2,font=F_BODY).pack(side=tk.LEFT,padx=(0,16))
    tk.Checkbutton(of,text="อัพเดตสินค้าที่มีอยู่แล้ว (แทนที่ข้อมูลเดิม)",
                   variable=update_var,bg=C_SURFACE2,fg=C_TEXT,
                   selectcolor=C_SURFACE,activebackground=C_SURFACE2,font=F_BODY).pack(side=tk.LEFT)

    # Table
    tf = tk.Frame(win,bg=C_BG); tf.pack(fill=tk.BOTH,expand=True,padx=12,pady=(6,0))
    cols=("บาร์โค้ด","ชื่อสินค้า","หมวดหมู่","หน่วย","ราคาขาย","ต้นทุน","สต็อก")
    tv = ttk.Treeview(tf,columns=cols,show="headings",height=14,style="R.Treeview")
    ws2=[110,260,90,70,80,80,60]
    for col,w in zip(cols,ws2):
        tv.heading(col,text=col)
        tv.column(col,width=w,anchor="center" if col not in("ชื่อสินค้า","หมวดหมู่") else "w")
    sb2=ttk.Scrollbar(tf,orient="vertical",command=tv.yview,style="R.Vertical.TScrollbar")
    tv.configure(yscrollcommand=sb2.set)
    tv.pack(side=tk.LEFT,fill=tk.BOTH,expand=True)
    sb2.pack(side=tk.RIGHT,fill=tk.Y)

    preview_limit=200
    for p in products[:preview_limit]:
        tv.insert("","end",values=(
            p['barcode'] or '—',
            p['name'],
            p['category'],
            p['unit'],
            "฿{:.2f}".format(p['price']),
            "฿{:.2f}".format(p['cost']),
            p['stock'],
        ))
    if len(products)>preview_limit:
        tv.insert("","end",values=("...","แสดง {} จาก {:,} รายการ (Import ครบทั้งหมด)".format(preview_limit,len(products)),"","","","",""))

    # Bottom
    bf = tk.Frame(win,bg=C_BG,padx=12,pady=10); bf.pack(fill=tk.X)
    status_lbl = tk.Label(bf,text="",font=F_SMALL,bg=C_BG,fg=C_MUTED)
    status_lbl.pack(side=tk.LEFT)

    # Progress bar
    prog_f=tk.Frame(win,bg=C_BG,padx=12); prog_f.pack(fill=tk.X,pady=(0,4))
    prog_lbl=tk.Label(prog_f,text="",font=F_SMALL,bg=C_BG,fg=C_MUTED); prog_lbl.pack(anchor="w")
    prog_bar=ttk.Progressbar(prog_f,orient="horizontal",mode="determinate",length=400)
    prog_bar.pack(fill=tk.X,pady=(2,0))

    import_btn_ref=[None]

    def do_import():
        import threading, time, sqlite3
        if import_btn_ref[0]: import_btn_ref[0].config(state="disabled")

        def run():
            t0=time.time()
            # Open NEW connection in this thread (SQLite thread safety)
            conn=sqlite3.connect(DB_FILE)
            conn.row_factory=sqlite3.Row
            total=len(products)

            # Pre-load existing barcodes
            existing_map={}
            if skip_var.get() or update_var.get():
                rows=conn.execute("SELECT id,barcode FROM products WHERE barcode!=''").fetchall()
                existing_map={r["barcode"]:r["id"] for r in rows}

            add_rows=[]; upd_rows=[]; skipped=0
            for i,p in enumerate(products):
                bc=p["barcode"]
                eid=existing_map.get(bc) if bc else None
                if eid:
                    if update_var.get():
                        upd_rows.append((p["name"],p["category"],p["unit"],p["price"],p["cost"],
                                         p["price_a"],p["price_b"],p["price_c"],p["stock"],p["min_stock"],eid))
                    elif skip_var.get(): skipped+=1
                    else:
                        add_rows.append((bc,p["name"],p["category"],p["unit"],p["price"],p["cost"],
                                         p["price_a"],p["price_b"],p["price_c"],p["stock"],p["min_stock"]))
                else:
                    add_rows.append((bc,p["name"],p["category"],p["unit"],p["price"],p["cost"],
                                     p["price_a"],p["price_b"],p["price_c"],p["stock"],p["min_stock"]))
                if i%500==0:
                    pct=int((i+1)/total*80)
                    win.after(0,lambda v=pct,ii=i:(
                        prog_bar.config(value=v),
                        prog_lbl.config(text="เตรียมข้อมูล {}/{} รายการ".format(ii+1,total))))

            win.after(0,lambda:(prog_bar.config(value=85),prog_lbl.config(text="กำลังบันทึกลงฐานข้อมูล...")))

            # Single transaction bulk insert
            conn.execute("BEGIN")
            if add_rows:
                conn.executemany("INSERT INTO products (barcode,name,category,unit,price,cost,price_a,price_b,price_c,stock,min_stock) VALUES (?,?,?,?,?,?,?,?,?,?,?)",add_rows)
            if upd_rows:
                conn.executemany("UPDATE products SET name=?,category=?,unit=?,price=?,cost=?,price_a=?,price_b=?,price_c=?,stock=?,min_stock=? WHERE id=?",upd_rows)
            conn.execute("COMMIT")
            conn.close()

            elapsed=time.time()-t0
            msg="✅ Import สำเร็จ\n\nเพิ่มใหม่: {:,}\nอัพเดต: {:,}\nข้าม: {:,}\n\nเวลา: {:.1f} วินาที".format(len(add_rows),len(upd_rows),skipped,elapsed)
            def done():
                prog_bar.config(value=100)
                prog_lbl.config(text="Import เสร็จสิ้น ใช้เวลา {:.1f} วินาที".format(elapsed))
                messagebox.showinfo("Import สำเร็จ",msg,parent=win)
                win.destroy()
                try: parent.tabs[1].load()
                except: pass
            win.after(0,done)

        threading.Thread(target=run,daemon=True).start()

    b=accent_btn(bf,"📥  Import {} รายการ".format(len(products)),do_import,pad_x=20,pad_y=8); b.pack(side=tk.RIGHT); import_btn_ref[0]=b
    pill_btn(bf,"ยกเลิก",win.destroy,bg=C_SURFACE,fg=C_MUTED,pad_x=16,pad_y=8).pack(side=tk.RIGHT,padx=(0,6))


class ShiftReportTab(tk.Frame):
    """รายงานกะ — รายการกะ + รายละเอียด + พิมพ์สรุป (Feature Phase 4)"""

    def __init__(self, parent):
        super().__init__(parent, bg=C_BG)
        self._rows = {}
        self._build()
        self.load()

    def _build(self):
        hdr = tk.Frame(self, bg=C_BG, padx=14, pady=8); hdr.pack(fill=tk.X)
        tk.Label(hdr, text="🕐  รายงานกะ", font=F_H1, bg=C_BG, fg=C_TEXT).pack(side=tk.LEFT)
        tk.Button(hdr, text="🖨️ พิมพ์สรุปกะที่เลือก", command=self._print_selected,
                  font=F_BODY, bg=C_ACCENT, fg=C_BG, relief=tk.FLAT, cursor="hand2",
                  padx=12, pady=4, bd=0).pack(side=tk.RIGHT)
        tk.Button(hdr, text="🔄 รีเฟรช", command=self.load,
                  font=F_BODY, bg=C_SURFACE2, fg=C_TEXT, relief=tk.FLAT, cursor="hand2",
                  padx=12, pady=4, bd=0).pack(side=tk.RIGHT, padx=(0,6))
        tk.Frame(self, bg=C_BORDER, height=1).pack(fill=tk.X)

        cols  = ("id","open","close","staff","opening","sales","bills",
                 "expected","counted","variance","status")
        heads = {"id":"#","open":"เปิด","close":"ปิด","staff":"พนักงาน","opening":"ตั้งต้น",
                 "sales":"ยอดขาย","bills":"บิล","expected":"ควรมี","counted":"นับได้",
                 "variance":"ส่วนต่าง","status":"สถานะ"}
        widths = {"id":40,"open":135,"close":135,"staff":110,"opening":90,"sales":100,
                  "bills":55,"expected":100,"counted":100,"variance":90,"status":70}
        wrap = tk.Frame(self, bg=C_BG); wrap.pack(fill=tk.BOTH, expand=True, padx=14, pady=10)
        self.tree = ttk.Treeview(wrap, columns=cols, show="headings", height=18)
        for c in cols:
            self.tree.heading(c, text=heads[c])
            self.tree.column(c, width=widths[c], anchor="center")
        self.tree.column("staff", anchor="w")
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb = ttk.Scrollbar(wrap, orient="vertical", command=self.tree.yview)
        sb.pack(side=tk.RIGHT, fill=tk.Y); self.tree.configure(yscrollcommand=sb.set)
        self.tree.tag_configure("short", foreground="#ff6b6b")
        self.tree.tag_configure("over",  foreground="#ffd166")
        self.tree.bind("<Double-1>", lambda e: self._print_selected())

    def load(self, *_):
        import shift_service
        self.tree.delete(*self.tree.get_children())
        self._rows = {}
        for sh in shift_service.list_shifts():
            summ = shift_service.get_shift_summary(sh["id"])
            self._rows[str(sh["id"])] = summ
            if summ["status"] == "open":
                counted, var, tag = "", "", ()
            else:
                v = summ.get("cash_variance") or 0
                counted = f"{float(summ.get('closing_cash_counted') or 0):,.2f}"
                var = f"{v:+,.2f}"
                tag = ("short",) if v < 0 else (("over",) if v > 0 else ())
            self.tree.insert("", "end", iid=str(sh["id"]), tags=tag, values=(
                sh["id"],
                (sh.get("open_time") or "")[:16],
                (sh.get("close_time") or "")[:16] if sh.get("close_time") else "—",
                sh.get("staff_name") or "-",
                f"{float(sh.get('opening_cash') or 0):,.2f}",
                f"{float(summ.get('total_sales') or 0):,.2f}",
                summ.get("total_bills") or 0,
                f"{float(summ.get('expected_cash') or 0):,.2f}",
                counted, var,
                "เปิดอยู่" if summ["status"] == "open" else "ปิดแล้ว",
            ))

    def _print_selected(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("เลือกกะ", "กรุณาเลือกกะที่ต้องการพิมพ์สรุปก่อน",
                                parent=self.winfo_toplevel()); return
        summ = self._rows.get(sel[0])
        if summ:
            self._show_summary(summ)

    def _show_summary(self, s):
        W = 38
        def line(lbl, amt): return f"{lbl:<20}{amt:>18,.2f}"
        L = ["=" * W, "สรุปกะ".center(W), "=" * W,
             f"กะ #: {s['id']}",
             f"พนักงาน: {s.get('staff_name') or '-'}",
             f"เปิด: {s.get('open_time') or '-'}",
             f"ปิด:  {s.get('close_time') or '-'}",
             "-" * W,
             line("เงินตั้งต้น", float(s.get('opening_cash') or 0)),
             line("ยอดขายในกะ", float(s.get('total_sales') or 0)),
             f"{'จำนวนบิล':<20}{(s.get('total_bills') or 0):>18}",
             line("ขายเงินสด", float(s.get('cash_sales') or 0)),
             line("เงินที่ควรมี", float(s.get('expected_cash') or 0))]
        if s["status"] == "closed":
            L += [line("นับได้", float(s.get('closing_cash_counted') or 0)),
                  line("ส่วนต่าง", float(s.get('cash_variance') or 0))]
        else:
            L += ["(กะยังเปิดอยู่ — ยังไม่ปิด)".center(W)]
        L += ["=" * W]
        text = "\n".join(L)

        win = tk.Toplevel(self); win.title(f"สรุปกะ #{s['id']}")
        win.configure(bg=C_BG); win.geometry("420x470"); win.lift()
        t = tk.Text(win, font=(FM.primary, 11), bg="#1a1a2e", fg="#e8e8e8",
                    relief=tk.FLAT, padx=14, pady=12)
        t.pack(fill=tk.BOTH, expand=True, padx=10, pady=8)
        t.insert("1.0", text); t.config(state=tk.DISABLED)
        bf = tk.Frame(win, bg=C_BG, pady=8); bf.pack(fill=tk.X, padx=10)
        pill_btn(bf, "📋  คัดลอก",
                 lambda: (win.clipboard_clear(), win.clipboard_append(text)),
                 bg=C_SURFACE2, fg=C_TEXT).pack(side=tk.LEFT)
        pill_btn(bf, "ปิด", win.destroy, bg=C_SURFACE, fg=C_MUTED).pack(side=tk.RIGHT)


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        global _responsive
        _responsive = ResponsiveManager(self)
        _responsive.maximize()
        # License check
        if HAS_BACKUP:
            auto_backup_if_needed(on_done=self._on_backup_done)
        if HAS_LICENSE:
            self._license_status = check_license()
            self.after(500, self._check_license_on_start)
        else:
            self._license_status = {"status":"trial","edition":"trial","days_left":7,"message":"","hardware_id":"N/A"}
        self.title("RakComSoft 2026")
        apply_fonts(self)
        self.geometry("1260x760"); self.minsize(1000,640); self.configure(bg=C_BG)
        _style()

        # ── App Icon (Title Bar + Taskbar + Alt+Tab) ──────
        try:
            _ico = os.path.join(os.path.dirname(__file__), "rakcomsoft.ico")
            if os.path.exists(_ico): self.iconbitmap(_ico)
        except Exception: pass

        # ── Topbar ──
        bar=tk.Frame(self,bg=C_SURFACE,height=54)
        bar.pack(fill=tk.X); bar.pack_propagate(False)
        # left accent stripe
        tk.Frame(bar,bg=C_ACCENT,width=4).pack(side=tk.LEFT,fill=tk.Y)
        # logo
        # Logo image + brand text
        _logo_f=tk.Frame(bar,bg=C_SURFACE); _logo_f.pack(side=tk.LEFT,padx=(8,4))
        try:
            _png=os.path.join(os.path.dirname(__file__),"rakcomsoft_40.png")
            if os.path.exists(_png):
                self._logo_ph=tk.PhotoImage(file=_png)
                tk.Label(_logo_f,image=self._logo_ph,bg=C_SURFACE,
                         bd=0,padx=0,pady=0).pack(side=tk.LEFT,padx=(0,8))
        except Exception as _le:
            print(f"[Logo] {_le}")
        _txt_f=tk.Frame(_logo_f,bg=C_SURFACE); _txt_f.pack(side=tk.LEFT)
        tk.Label(_txt_f,text="RakComSoft",font=F_LOGO,bg=C_SURFACE,fg=C_ACCENT,pady=0).pack(anchor="w")
        tk.Frame(bar,bg=C_BORDER,width=1).pack(side=tk.LEFT,fill=tk.Y,padx=8)
        store=get_setting("store_name") or ""
        tk.Label(bar,text=store,font=(FM.primary,13),bg=C_SURFACE,fg=C_MUTED).pack(side=tk.LEFT)
        # clock
        # ── Right frame คงที่ — ป้องกัน layout shift จาก clock ──
        right_f = tk.Frame(bar, bg=C_SURFACE, width=240)
        right_f.pack(side=tk.RIGHT, fill=tk.Y)
        right_f.pack_propagate(False)

        self.date_lbl = tk.Label(right_f, text="",
                                  font=(FM.primary, 11),
                                  bg=C_SURFACE, fg=C_MUTED,
                                  anchor="center")
        self.date_lbl.pack(side=tk.LEFT, fill=tk.Y, padx=(6,2))

        # clock container — fixed width ป้องกัน layout shift
        clk_wrap = tk.Frame(right_f, bg=C_SURFACE, width=90)
        clk_wrap.pack(side=tk.LEFT, fill=tk.Y, padx=(2,6))
        clk_wrap.pack_propagate(False)
        self.clk = tk.Label(clk_wrap, text="",
                             font=(FM.primary, 11),
                             bg=C_SURFACE, fg=C_MUTED,
                             anchor="center")
        self.clk.pack(fill=tk.BOTH, expand=True)
        self._tick()
        # Top bar buttons — ใช้ Label แทน Button เพื่อไม่ให้กระพริบ/เคลื่อนไหว
        def _nav_lbl(parent, text, fg, cmd):
            lb = tk.Label(parent, text=text, font=(FM.primary,12),
                         bg=C_SURFACE, fg=fg, cursor="hand2", padx=12)
            lb.pack(side=tk.RIGHT, fill=tk.Y)
            lb.bind("<Button-1>", lambda e: cmd())
            lb.bind("<Enter>",    lambda e: lb.config(fg=C_ACCENT if fg!=C_YELLOW else C_YELLOW))
            lb.bind("<Leave>",    lambda e: lb.config(fg=fg))
            return lb

        tk.Frame(bar,bg=C_BORDER,width=1).pack(side=tk.RIGHT,fill=tk.Y,padx=4)
        self._topbar_tools_lbl = _nav_lbl(bar,"เครื่องมือ ▾",C_TEXT,lambda:self._tools_menu(bar))

        tk.Frame(bar,bg=C_BORDER,width=1).pack(side=tk.RIGHT,fill=tk.Y,padx=4)
        self._topbar_more_lbl = _nav_lbl(bar,"➕  เพิ่มเติม ▾",C_YELLOW,lambda:None)
        self._topbar_more_lbl.bind("<Button-1>",lambda e:self._more_menu(self._topbar_more_lbl))

        tk.Frame(bar,bg=C_BORDER,width=1).pack(side=tk.RIGHT,fill=tk.Y,padx=4)
        self._topbar_docs_lbl = _nav_lbl(bar,"📄  เอกสาร ▾",C_BLUE,lambda:None)
        self._topbar_docs_lbl.bind("<Button-1>",lambda e:self._docs_menu(self._topbar_docs_lbl))

        tk.Frame(bar,bg=C_BORDER,width=1).pack(side=tk.RIGHT,fill=tk.Y,padx=4)
        self._topbar_barcode_lbl = _nav_lbl(bar,"🏷️  พิมพ์บาร์โค้ด",C_ACCENT,self._open_barcode_printer)

        # ── Shift open/close (Phase 3) — สร้างเองเพราะสี/ข้อความเปลี่ยนตามสถานะ ──
        tk.Frame(bar,bg=C_BORDER,width=1).pack(side=tk.RIGHT,fill=tk.Y,padx=4)
        self._shift_lbl_fg = C_BLUE
        self._topbar_shift_lbl = tk.Label(bar, text="🕐  เปิดกะ", font=(FM.primary,12),
                                          bg=C_SURFACE, fg=C_BLUE, cursor="hand2", padx=12)
        self._topbar_shift_lbl.pack(side=tk.RIGHT, fill=tk.Y)
        self._topbar_shift_lbl.bind("<Button-1>", lambda e: self._shift_action())
        self._topbar_shift_lbl.bind("<Enter>", lambda e: self._topbar_shift_lbl.config(fg=C_ACCENT))
        self._topbar_shift_lbl.bind("<Leave>", lambda e: self._topbar_shift_lbl.config(fg=self._shift_lbl_fg))
        self._refresh_shift_label()

        # ── Notebook ──
        self.nb=ttk.Notebook(self,style="R.TNotebook"); nb=self.nb; nb.pack(fill=tk.BOTH,expand=True)
        # 8 Main Tabs (ตั้งค่า ย้ายไปเพิ่มหลัง ลูกหนี้ — ดูด้านล่าง)
        tabs=[(SaleTab,"  🛒  ขายสินค้า"),(ProductTab,"  📦  สินค้า/สต็อก"),
              (CustomerTab,"  👥  สมาชิก"),(ReportTab,"  📊  รายงาน"),
              (ProductSalesTab,"  📋  รายการขายสินค้า"),(StockCardTab,"  📑  Stock Card"),
              (DashboardTab,"  📈  Dashboard"),(EODTab,"  💵  สรุปเงิน"),
              (StaffTab,"  👤  พนักงาน")]
        self.tabs=[]
        for cls,lbl in tabs:
            t=cls(nb); nb.add(t,text=lbl); self.tabs.append(t)

        # ── (เครดิตลูกค้า/AR ย้ายไปเมนู "เอกสาร ▾" — ไม่อยู่บนแท็บหลักแล้ว) ──
        init_receivable_schema(DB_FILE)
        stock_service.ensure_cost_columns(DB_FILE)   # เฟส1: avg_cost/last_cost
        stock_service.ensure_purchase_table(DB_FILE)  # เฟส2: purchase_history
        CategoryManager.ensure(); UnitManager.ensure()  # หมวดหมู่/หน่วยนับ (ตารางแยก)
        promotion_engine.ensure_promo_columns(DB_FILE)   # เฟสA: โปรโมชั่นผูกสินค้า
        unit_service.ensure_table(DB_FILE)                # เฟส4: หน่วยขายหลายขนาด

        # ── ตั้งค่า Tab (ขวาสุด — หลังลูกหนี้) ──
        _settings_tab = SettingsTab(nb)
        nb.add(_settings_tab, text="  ⚙️  ตั้งค่า")
        self.tabs.append(_settings_tab)

        nb.bind("<<NotebookTabChanged>>",self._on_tab)

        self.after(100,lambda:show_login(self,self._after_login))

    def _after_login(self):
        check_low_stock(self)
        self._apply_tab_permissions()
        # เช็กอัปเดตเงียบๆ หลัง login (ถ้ายังไม่ตั้งค่า repo จะไม่ทำอะไร)
        self.after(3000, self._startup_update_check)

    def _apply_tab_permissions(self):
        """
        Enforce tab and top-bar permissions for the current logged-in staff.
        Tab index map:
          0=ขายสินค้า  1=สินค้า/สต็อก  2=สมาชิก  3=รายงาน  4=รายการขายสินค้า
          5=Stock Card  6=Dashboard  7=สรุปเงิน(EOD)  8=พนักงาน  9=ตั้งค่า
          (เครดิตลูกค้า/คืนสินค้า/เอกสารย้อนหลัง → เมนู "เอกสาร ▾")
        """
        role = current_staff.get("role", "cashier")
        is_owner  = role in ("owner", "admin")
        is_mgr    = role == "manager"
        is_cashier = role == "cashier"

        # ── Notebook tab gating ─────────────────────────────
        # 3 = Reports, 4 = Dashboard, 5 = EOD, 6 = Staff,
        # 7 = ลูกหนี้ (AR, manager+), 8 = ตั้งค่า (Settings)
        tab_perms = {
            3: "can_reports",        # รายงาน
            4: "can_reports",        # รายการขายสินค้า
            5: "can_stock_adjust",   # Stock Card
            6: "can_dashboard",      # Dashboard
            7: "can_reports",        # EOD สรุปเงิน
            8: "can_manage_employees", # พนักงาน
            9: "can_settings",       # ตั้งค่า
        }
        nb = self.nb
        for idx in range(nb.index("end")):
            perm = tab_perms.get(idx)
            if perm is None:
                # Tabs 0,1,2 — always accessible
                nb.tab(idx, state="normal")
            else:
                allowed = is_owner or has_permission(perm)
                nb.tab(idx, state="normal" if allowed else "disabled")

        # ── Top-bar button gating ────────────────────────────
        # Store references on first call so we can update them
        try:
            if hasattr(self, "_topbar_barcode_lbl"):
                barcode_ok = is_owner or has_permission("can_barcode_print")
                self._topbar_barcode_lbl.config(
                    fg=C_ACCENT if barcode_ok else C_MUTED,
                    cursor="hand2" if barcode_ok else "arrow",
                )
                if barcode_ok:
                    self._topbar_barcode_lbl.bind("<Button-1>", lambda e: self._open_barcode_printer())
                else:
                    self._topbar_barcode_lbl.bind("<Button-1>",
                        lambda e: require_permission("can_barcode_print", self))

            if hasattr(self, "_topbar_tools_lbl"):
                tools_ok = is_owner or has_permission("can_system_tools")
                self._topbar_tools_lbl.config(
                    fg=C_TEXT if tools_ok else C_MUTED,
                    cursor="hand2" if tools_ok else "arrow",
                )
                if tools_ok:
                    self._topbar_tools_lbl.bind("<Button-1>", lambda e: self._tools_menu(self._topbar_tools_lbl))
                else:
                    self._topbar_tools_lbl.bind("<Button-1>",
                        lambda e: require_permission("can_system_tools", self))

            if hasattr(self, "_topbar_more_lbl"):
                more_ok = is_owner or not is_cashier
                self._topbar_more_lbl.config(
                    fg=C_YELLOW if more_ok else C_MUTED,
                    cursor="hand2" if more_ok else "arrow",
                )
                if more_ok:
                    self._topbar_more_lbl.bind("<Button-1>", lambda e: self._more_menu(self._topbar_more_lbl))
                else:
                    self._topbar_more_lbl.bind("<Button-1>",
                        lambda e: messagebox.showwarning(
                            "ไม่มีสิทธิ์","คุณไม่มีสิทธิ์ใช้งานส่วนนี้\n\nกรุณาติดต่อเจ้าของร้าน",
                            parent=self))
        except Exception as _e:
            print(f"[Permissions] _apply_tab_permissions topbar error: {_e}")

    def _on_tab(self,e):
        i=e.widget.index("current")
        # ── Permission guard — block bypass via keyboard/direct call ──
        tab_perms = {
            3: "can_reports",
            4: "can_reports",
            5: "can_stock_adjust",
            6: "can_dashboard",
            7: "can_reports",
            8: "can_manage_employees",
            9: "can_settings",
        }
        perm = tab_perms.get(i)
        if perm and not has_permission(perm):
            # Force back to tab 0 (sale screen)
            last = getattr(self, "_last_tab", 0)
            safe = last if last not in tab_perms else 0
            self.nb.select(safe)
            require_permission(perm, self)
            return
        self._last_tab=i
        try:
            if i!=1 and hasattr(self.tabs[1],"reset_filters"):
                self.tabs[1].reset_filters()
        except: pass
        reload_map={1:"load",2:"load",3:"load",4:"load",5:"load",6:"load"}
        if i==0:
            try: self.tabs[0]._refresh_qs_bar()
            except: pass
            # Refresh debt label ถ้าลูกค้ายังถูกเลือกอยู่
            try:
                sale_tab = self.tabs[0]
                if sale_tab.customer:
                    sale_tab._update_debt_label()
            except: pass
        if i==7:
            # EODTab (index 7 หลังแทรก Stock Card) — reset วันที่เป็นวันนี้แล้ว load
            try:
                eod=self.tabs[7]
                today=date.today().strftime("%Y-%m-%d")
                eod.d0.set(today); eod.d1.set(today)
                eod.load()
            except: pass
        if i in reload_map: getattr(self.tabs[i],reload_map[i])()

    def _more_menu_nb(self, nb):
        """เปิด popup เมื่อคลิก tab เพิ่มเติม ▾"""
        # หาตำแหน่ง tab เพิ่มเติม
        try:
            tab_bbox = nb.bbox(7)
            x = nb.winfo_rootx() + tab_bbox[0]
            y = nb.winfo_rooty() + tab_bbox[1] + tab_bbox[3]
        except:
            x = nb.winfo_rootx()
            y = nb.winfo_rooty() + 30
        m=tk.Menu(self,tearoff=0,bg=C_SURFACE2,fg=C_TEXT,
                  activebackground=C_ACCENT,activeforeground=C_BG,
                  font=(FM.primary,12),relief=tk.FLAT,bd=0)
        m.add_command(label="🏭  ผู้จำหน่าย",
                      command=lambda:self._guarded_extra("ผู้จำหน่าย",SupplierTab,"can_manage_suppliers"))
        m.add_command(label="💳  โปรโมชั่น",
                      command=lambda:self._guarded_extra("โปรโมชั่น",PromoTab,"can_manage_promotions"))
        m.add_separator()
        m.add_command(label="⚡  Quick Sale",
                      command=lambda:self._open_extra("Quick Sale",QSSaleTab))
        m.add_command(label="🕐  รายงานกะ",
                      command=lambda:self._open_extra("รายงานกะ",ShiftReportTab))
        m.add_separator()
        for lbl in ["📥  Import/Export","☁️  Cloud Sync","🔌  API",
                    "⚖️  เครื่องชั่ง","🛠️  Tools"]:
            m.add_command(label=lbl,state="disabled",foreground="#555")
        try: m.tk_popup(x, y)
        finally: m.grab_release()

    def _more_menu(self, bar):
        """เมนู เพิ่มเติม"""
        m=tk.Menu(self,tearoff=0,bg=C_SURFACE2,fg=C_TEXT,
                  activebackground=C_ACCENT,activeforeground=C_BG,
                  font=(FM.primary,12),relief=tk.FLAT,bd=0)
        m.add_command(label="🏭  ผู้จำหน่าย",
                      command=lambda:self._guarded_extra("ผู้จำหน่าย",SupplierTab,"can_manage_suppliers"))
        m.add_command(label="💳  โปรโมชั่น",
                      command=lambda:self._guarded_extra("โปรโมชั่น",PromoTab,"can_manage_promotions"))
        m.add_separator()
        m.add_command(label="⚡  Quick Sale",
                      command=lambda:self._open_extra("Quick Sale",QSSaleTab))
        m.add_command(label="🕐  รายงานกะ",
                      command=lambda:self._open_extra("รายงานกะ",ShiftReportTab))
        m.add_separator()
        for lbl in ["📥  Import/Export","☁️  Cloud Sync","🔌  API",
                    "⚖️  เครื่องชั่ง","🛠️  Tools"]:
            m.add_command(label=lbl,state="disabled",foreground="#555")
        try:
            x = bar.winfo_rootx()
            y = bar.winfo_rooty() + bar.winfo_height()
            m.tk_popup(x, y)
        finally: m.grab_release()

    def _docs_menu(self, bar):
        """เมนู เอกสาร — งานเครดิต/เอกสาร/ธุรกิจ (รองรับขยายอนาคต)"""
        m=tk.Menu(self,tearoff=0,bg=C_SURFACE2,fg=C_TEXT,
                  activebackground=C_ACCENT,activeforeground=C_BG,
                  font=(FM.primary,12),relief=tk.FLAT,bd=0)
        m.add_command(label="💳  เครดิตลูกค้า", command=self._open_credit)
        m.add_command(label="↩  คืนสินค้า",
                      command=lambda:self._open_extra("คืนสินค้า",ReturnTab))
        m.add_command(label="🗂️  เอกสารย้อนหลัง (ยกเลิก/คืน)",
                      command=lambda:self._open_extra("เอกสารย้อนหลัง",HistoryTab))
        m.add_separator()
        for lbl in ["🧾  ใบเสนอราคา","📑  ใบสั่งซื้อ","📨  วางบิล","💵  รับชำระหนี้"]:
            m.add_command(label=lbl+"   (เร็วๆ นี้)", state="disabled", foreground="#555")
        try:
            x=bar.winfo_rootx(); y=bar.winfo_rooty()+bar.winfo_height()
            m.tk_popup(x,y)
        finally: m.grab_release()

    def _open_credit(self):
        """เปิดหน้าเครดิตลูกค้า (AR) ใน Toplevel — เก็บ instance ไว้เพื่อ refresh หลังขาย"""
        if not require_permission("can_reports", self):
            return
        win=tk.Toplevel(self); win.title("เครดิตลูกค้า")
        win.configure(bg=C_BG); win.geometry("1100x680"); win.lift()
        try:
            _ico=os.path.join(os.path.dirname(__file__),"rakcomsoft.ico")
            if os.path.exists(_ico): win.iconbitmap(_ico)
        except: pass
        self._ar_tab = ReceivableTab(win, db_path=DB_FILE)
        self._ar_tab.pack(fill=tk.BOTH, expand=True)
        try: self._ar_tab.refresh_all()
        except: pass
        def _close():
            try: del self._ar_tab
            except: pass
            win.destroy()
        win.protocol("WM_DELETE_WINDOW", _close)
        win.bind("<Escape>", lambda e: _close())

    def _guarded_extra(self, title, cls, perm_key):
        """Open extra window only if permission is granted."""
        if not require_permission(perm_key, self):
            return
        self._open_extra(title, cls)

    def _open_extra(self, title, cls):
        """เปิด tab เสริมใน Toplevel window"""
        win=tk.Toplevel(self); win.title(title)
        win.configure(bg=C_BG); win.geometry("1100x680"); win.lift()
        try:
            _ico=os.path.join(os.path.dirname(__file__),"rakcomsoft.ico")
            if os.path.exists(_ico): win.iconbitmap(_ico)
        except: pass
        t=cls(win); t.pack(fill=tk.BOTH,expand=True)
        try: t.load()
        except: pass

    def _tools_menu(self,parent):
        if not require_permission("can_system_tools", self):
            return
        m=tk.Menu(self,tearoff=0,bg=C_SURFACE,fg=C_TEXT,
                  activebackground=C_ACCENT,activeforeground=C_BG,
                  font=F_BODY,relief=tk.FLAT,bd=0)
        m.add_command(label="📥  Import Excel (RakComSoft เดิม)",command=self._import_excel)
        m.add_separator()
        m.add_command(label="💾  สำรองข้อมูล (CSV)",command=self._export_backup)
        m.add_separator()
        m.add_command(label="💾  สำรองข้อมูล & กู้คืน",command=self._show_backup_page)
        m.add_separator()
        m.add_command(label="🔑  License / เกี่ยวกับโปรแกรม",command=self._show_license_page)
        m.add_separator()
        m.add_command(label="🔄  ตรวจสอบอัปเดตโปรแกรม",command=self._show_update_page)
        x=parent.winfo_rootx()+parent.winfo_width()-120
        y=parent.winfo_rooty()+parent.winfo_height()
        m.tk_popup(x,y)

    def _export_backup(self):
        from tkinter import filedialog
        import csv as csv_mod
        path=filedialog.asksaveasfilename(defaultextension=".csv",
            filetypes=[("CSV","*.csv")],title="บันทึกข้อมูลสำรอง")
        if not path: return
        conn=get_db()
        rows=conn.execute("SELECT * FROM products WHERE active=1").fetchall(); conn.close()
        with open(path,"w",newline="",encoding="utf-8-sig") as f:
            w=csv_mod.writer(f)
            w.writerow(["barcode","name","category","unit","price","cost","price_a","price_b","price_c","stock","min_stock"])
            for r in rows: w.writerow([r["barcode"],r["name"],r["category"],r["unit"],r["price"],r["cost"],r.get("price_a",0),r.get("price_b",0),r.get("price_c",0),r["stock"],r["min_stock"]])
        messagebox.showinfo("✅ สำเร็จ","สำรองข้อมูลเรียบร้อย\n{}".format(path))

    def _import_excel(self):
        import threading
        from tkinter import filedialog
        path=filedialog.askopenfilename(
            filetypes=[("Excel",("*.xls","*.xlsx","*.xlsm")),("All","*.*")],
            title="เลือกไฟล์ Excel จาก RakComSoft เดิม")
        if not path: return

        # Show loading dialog while parsing
        loading=tk.Toplevel(self)
        loading.title("กำลังอ่านไฟล์")
        loading.configure(bg=C_BG)
        loading.resizable(False,False)
        loading.grab_set()
        loading.geometry("340x120")
        loading.update_idletasks()
        x=(loading.winfo_screenwidth()-340)//2
        y=(loading.winfo_screenheight()-120)//2
        loading.geometry("340x120+{}+{}".format(x,y))
        tk.Frame(loading,bg=C_ACCENT,height=3).pack(fill=tk.X)
        tk.Label(loading,text="กำลังอ่านไฟล์ Excel...",font=F_H2,bg=C_BG,fg=C_TEXT,pady=10).pack()
        pb=ttk.Progressbar(loading,mode="indeterminate",length=280)
        pb.pack(pady=8); pb.start(10)
        tk.Label(loading,text="กรุณารอสักครู่...",font=F_SMALL,bg=C_BG,fg=C_MUTED).pack()
        loading.update()

        result={"products":None,"error":None}

        def parse_thread():
            try:
                result["products"]=_parse_rakcom_excel(path)
            except Exception as e:
                result["error"]=str(e)

        def check_done():
            if t.is_alive():
                self.after(100,check_done)
                return
            pb.stop()
            loading.destroy()
            if result["error"]:
                messagebox.showerror("อ่านไฟล์ไม่ได้",
                    "เกิดข้อผิดพลาด:\n{}\n\nตรวจสอบว่าติดตั้ง xlrd และ openpyxl แล้ว".format(result["error"]))
                return
            products=result["products"]
            if not products:
                messagebox.showwarning("ไม่พบข้อมูล","ไม่พบข้อมูลสินค้าในไฟล์"); return
            _show_import_preview(self, products)

        t=threading.Thread(target=parse_thread,daemon=True)
        t.start()
        self.after(100,check_done)

    def _check_license_on_start(self):
        """แสดงสถานะ License ตอนเปิดโปรแกรม"""
        if not HAS_LICENSE: return
        s = self._license_status
        if s["status"] == "expired":
            from tkinter import messagebox
            messagebox.showwarning(
                "หมดระยะเวลาทดลองใช้งาน",
                "หมดระยะเวลาทดลองใช้งานแล้ว\n\n"
                "กรุณาติดต่อ RakComSoft\n"
                "LINE: rakcomshop\n"
                "โทร: 097-335-7599")
        elif s["status"] == "trial" and s["days_left"] <= 3:
            from tkinter import messagebox
            messagebox.showinfo(
                "แจ้งเตือน Trial",
                f"RakComSoft 2026 Trial\nเหลืออีก {s['days_left']} วัน\n\n"
                "ติดต่อ LINE: rakcomshop\n"
                "โทร: 097-335-7599")

    def _show_license_page(self):
        """หน้า License"""
        s = self._license_status if HAS_LICENSE else check_license()
        hw_id = s.get("hardware_id", "N/A")
        win = tk.Toplevel(self)
        win.title("License — RakComSoft 2026")
        win.configure(bg="#0d1117")
        win.resizable(False, False)
        win.grab_set(); win.lift()
        win.geometry("460x580")
        BG="#0d1117"; SF="#161b22"; ACC="#00d4aa"; TX="#e6edf3"; MT="#8b949e"
        YL="#ffd166"; GR="#3fb950"; RD="#ff6b6b"; BD="#30363d"
        tk.Frame(win,bg=ACC,height=4).pack(fill=tk.X)
        tk.Label(win,text="◆  RakComSoft 2026",font=(FM.primary,18,"bold"),
                 bg=BG,fg=ACC,pady=14).pack()
        tk.Label(win,text="Version: 1.0.0",font=(FM.primary,10),bg=BG,fg=MT).pack()
        from license_system import EDITION_NAMES
        ed_lbl=EDITION_NAMES.get(s.get("edition","trial"),"Trial")
        tk.Label(win,text=f"Edition: {ed_lbl}",font=(FM.primary,11,"bold"),
                 bg=BG,fg=YL,pady=4).pack()
        # Status badge
        st=s.get("status","trial")
        if st=="licensed": badge_text,badge_color="✓  ลงทะเบียนแล้ว",GR
        elif st=="trial":  badge_text,badge_color=f"⚠  Trial — เหลือ {s['days_left']} วัน",YL
        else:              badge_text,badge_color="✗  หมดอายุ",RD
        bf=tk.Frame(win,bg=SF,padx=20,pady=10,highlightbackground=BD,highlightthickness=1)
        bf.pack(fill=tk.X,padx=20,pady=(10,0))
        tk.Label(bf,text=badge_text,font=(FM.primary,12,"bold"),bg=SF,fg=badge_color).pack()
        # Hardware ID
        hf=tk.Frame(win,bg=SF,padx=20,pady=12,highlightbackground=BD,highlightthickness=1)
        hf.pack(fill=tk.X,padx=20,pady=(8,0))
        tk.Label(hf,text="Hardware ID",font=(FM.primary,9),bg=SF,fg=MT).pack(anchor="w")
        tk.Label(hf,text=hw_id,font=("Consolas",14,"bold"),bg=SF,fg=ACC).pack(anchor="w")
        tk.Button(hf,text="📋 คัดลอกรหัสเครื่อง",
                  command=lambda:(win.clipboard_clear(),win.clipboard_append(hw_id)),
                  font=(FM.primary,9),bg=ACC,fg=BG,relief=tk.FLAT,
                  cursor="hand2",padx=8,pady=4).pack(anchor="w",pady=(6,0))
        # License Key input
        lf=tk.Frame(win,bg=SF,padx=20,pady=12,highlightbackground=BD,highlightthickness=1)
        lf.pack(fill=tk.X,padx=20,pady=(8,0))
        tk.Label(lf,text="License Key",font=(FM.primary,9),bg=SF,fg=MT).pack(anchor="w")
        key_row=tk.Frame(lf,bg="#161b22"); key_row.pack(fill=tk.X,pady=(4,0))
        key_e=tk.Entry(key_row,font=("Consolas",10),bg="#0d1117",fg=TX,
                       insertbackground=TX,relief=tk.FLAT)
        key_e.pack(side=tk.LEFT,fill=tk.X,expand=True,ipady=6)
        def paste_key():
            try:
                txt=win.clipboard_get().strip()
                key_e.delete(0,tk.END)
                key_e.insert(0,txt)
            except: pass
        tk.Button(key_row,text="📋 วาง",command=paste_key,
                  font=(FM.primary,10),bg=ACC,fg=BG,relief=tk.FLAT,
                  cursor="hand2",padx=8,pady=3).pack(side=tk.LEFT,padx=(6,0))
        def activate():
            key=key_e.get().strip()
            if not key: messagebox.showwarning("ข้อผิดพลาด","กรุณากรอก License Key",parent=win); return
            ok,msg=TrialManager.activate_license(key)
            if ok:
                messagebox.showinfo("สำเร็จ",msg,parent=win)
                self._license_status=check_license(); win.destroy()
            else:
                messagebox.showerror("ผิดพลาด",msg,parent=win)
        tk.Button(lf,text="🔑  ลงทะเบียนโปรแกรม",command=activate,
                  font=(FM.primary,11,"bold"),bg=ACC,fg=BG,
                  relief=tk.FLAT,cursor="hand2",pady=8).pack(fill=tk.X,pady=(8,0))
        # Contact
        cf=tk.Frame(win,bg=BG,pady=12); cf.pack()
        tk.Label(cf,text="LINE: rakcomshop  |  โทร: 097-335-7599",
                 font=(FM.primary,9),bg=BG,fg=MT).pack()

    def _open_barcode_printer(self, items=None):
        """เปิดหน้าพิมพ์บาร์โค้ด"""
        if not require_permission("can_barcode_print", self):
            return
        if not HAS_BARCODE:
            messagebox.showerror("ต้องติดตั้ง Library",
                "กรุณาติดตั้ง library ก่อน:\n\n"
                "pip install reportlab pillow")
            return
        BarcodePrinterWindow(self, initial_items=items)

    def _on_backup_done(self, path):
        """callback หลัง auto backup เสร็จ"""
        try:
            if path:
                self.after(0, lambda: self.title(f"Rakcomsoft — Backup สำเร็จ"))
                self.after(3000, lambda: self.title("Rakcomsoft"))
        except: pass

    # ══ UPDATE SYSTEM ════════════════════════════════════════
    def _show_update_page(self):
        """หน้าอัปเดตโปรแกรม — เวอร์ชัน/Release Note/ดาวน์โหลด+ติดตั้ง"""
        import threading
        if _updater is None:
            messagebox.showerror("ระบบอัปเดต",
                "ไม่พบไฟล์ updater.py ในโฟลเดอร์โปรแกรม\nกรุณาวาง updater.py + RakUpdater.pyw ไว้ที่ D:\\pos",
                parent=self); return

        win = tk.Toplevel(self); win.title("อัปเดตโปรแกรม")
        win.configure(bg=C_BG); win.geometry("560x560"); win.grab_set(); win.lift()
        win.bind("<Escape>", lambda e: win.destroy())
        tk.Frame(win, bg=C_ACCENT, height=3).pack(fill=tk.X)
        tk.Label(win, text="🔄  อัปเดตโปรแกรม RakComSoft 2026", font=F_H1,
                 bg=C_BG, fg=C_TEXT).pack(anchor="w", padx=18, pady=(12,4))

        info = tk.Frame(win, bg=C_SURFACE2, padx=16, pady=12); info.pack(fill=tk.X, padx=16)
        cur = tk.StringVar(value=f"v{APP_VERSION}")
        latest = tk.StringVar(value="— (กดตรวจสอบ)")
        pub = tk.StringVar(value="—")
        for lbl, var in [("เวอร์ชันปัจจุบัน:", cur), ("เวอร์ชันล่าสุด:", latest), ("เผยแพร่เมื่อ:", pub)]:
            r = tk.Frame(info, bg=C_SURFACE2); r.pack(fill=tk.X, pady=1)
            tk.Label(r, text=lbl, font=F_SMALL, bg=C_SURFACE2, fg=C_MUTED, width=16, anchor="w").pack(side=tk.LEFT)
            tk.Label(r, textvariable=var, font=(FM.primary,12,"bold"), bg=C_SURFACE2, fg=C_TEXT).pack(side=tk.LEFT)

        chf = tk.Frame(win, bg=C_BG, padx=16); chf.pack(fill=tk.X, pady=(8,2))
        tk.Label(chf, text="ช่องทาง:", font=F_SMALL, bg=C_BG, fg=C_MUTED).pack(side=tk.LEFT)
        ch_var = tk.StringVar(value=get_setting("update_channel") or "stable")
        ch_cb = ttk.Combobox(chf, textvariable=ch_var, values=["stable","beta"],
                             width=10, state="readonly", font=F_BODY); ch_cb.pack(side=tk.LEFT, padx=6)

        tk.Label(win, text="รายละเอียดการอัปเดต (Release Note):", font=F_SMALL,
                 bg=C_BG, fg=C_MUTED).pack(anchor="w", padx=18, pady=(8,2))
        nf = tk.Frame(win, bg=C_BG, padx=16); nf.pack(fill=tk.BOTH, expand=True)
        notes = tk.Text(nf, height=8, bg=C_SURFACE2, fg=C_TEXT, font=F_BODY,
                        wrap="word", relief=tk.FLAT, padx=10, pady=8)
        notes.pack(fill=tk.BOTH, expand=True); notes.insert("1.0", "กด \"ตรวจสอบอัปเดต\" เพื่อดูรายละเอียด"); notes.config(state="disabled")

        pb = ttk.Progressbar(win, mode="determinate", length=520, maximum=100)
        pb.pack(padx=18, pady=(8,2))
        status = tk.Label(win, text="", font=F_SMALL, bg=C_BG, fg=C_MUTED); status.pack(anchor="w", padx=18)

        bf = tk.Frame(win, bg=C_BG, padx=16, pady=10); bf.pack(fill=tk.X)
        state = {"url": "", "latest": APP_VERSION, "busy": False}

        def _set_notes(txt):
            notes.config(state="normal"); notes.delete("1.0","end")
            notes.insert("1.0", txt or "(ไม่มีรายละเอียด)"); notes.config(state="disabled")

        def _check():
            if state["busy"]: return
            state["busy"]=True; status.config(text="กำลังตรวจสอบ...", fg=C_BLUE); pb.config(value=0)
            ch = ch_var.get()
            def work():
                r = _updater.check_for_update(APP_VERSION, channel=ch)
                def done():
                    state["busy"]=False
                    if not r["ok"] and r["error"]:
                        status.config(text="❌ "+r["error"], fg="#ff6b6b")
                        latest.set("ตรวจสอบไม่ได้"); return
                    latest.set("v"+r["latest"]); pub.set(r["published"] or "—")
                    _set_notes(r["notes"]); state["url"]=r["url"]; state["latest"]=r["latest"]
                    if r["available"] and r["url"]:
                        status.config(text="🎉 พบเวอร์ชันใหม่! กด \"ดาวน์โหลด & อัปเดต\"", fg=C_GREEN)
                        up_btn.config(state="normal")
                    elif r["available"] and not r["url"]:
                        status.config(text="พบเวอร์ชันใหม่ แต่ไม่มีไฟล์ update.zip ใน release", fg=C_YELLOW)
                    else:
                        status.config(text="✅ ใช้เวอร์ชันล่าสุดอยู่แล้ว", fg=C_GREEN)
                win.after(0, done)
            threading.Thread(target=work, daemon=True).start()

        def _update():
            if state["busy"] or not state["url"]: return
            if not messagebox.askyesno("ยืนยันอัปเดต",
                f"อัปเดตเป็นเวอร์ชัน {state['latest']}?\n\n"
                "โปรแกรมจะสำรองไฟล์เดิม แล้วปิดเพื่อติดตั้ง จากนั้นเปิดใหม่อัตโนมัติ\n"
                "(ฐานข้อมูลจะไม่ถูกแตะต้อง)", parent=win): return
            state["busy"]=True; up_btn.config(state="disabled"); chk_btn.config(state="disabled")
            ad = os.path.dirname(os.path.abspath(__file__))
            zip_dest = os.path.join(ad, "update.zip")
            def prog(done, total):
                pct = int(done*100/total) if total else 0
                win.after(0, lambda: (pb.config(value=pct),
                          status.config(text=f"กำลังดาวน์โหลด... {pct}%  ({done//1024:,} KB)", fg=C_BLUE)))
            def work():
                try:
                    _updater.download_update(state["url"], zip_dest, progress_cb=prog)
                    _updater.write_job(zip_dest, state["latest"])
                    _updater.log(f"DOWNLOAD OK v{state['latest']} → เริ่ม RakUpdater", ad)
                    def launch():
                        status.config(text="✅ ดาวน์โหลดเสร็จ — กำลังเริ่มตัวติดตั้ง...", fg=C_GREEN)
                        pywin = os.path.join(os.path.dirname(sys.executable), "pythonw.exe")
                        rk = os.path.join(ad, "RakUpdater.pyw")
                        try:
                            import subprocess
                            subprocess.Popen([pywin if os.path.exists(pywin) else sys.executable, rk], cwd=ad)
                            self.after(800, lambda: (self.quit(), self.destroy()))
                        except Exception as e:
                            messagebox.showerror("เริ่มตัวติดตั้งไม่ได้", str(e), parent=win)
                            state["busy"]=False; chk_btn.config(state="normal")
                    win.after(0, launch)
                except Exception as e:
                    win.after(0, lambda: (status.config(text="❌ "+str(e), fg="#ff6b6b"),
                              up_btn.config(state="normal"), chk_btn.config(state="normal")))
                    state["busy"]=False
            threading.Thread(target=work, daemon=True).start()

        chk_btn = accent_btn(bf, "🔍  ตรวจสอบอัปเดต", _check, pad_x=16, pad_y=8); chk_btn.pack(side=tk.LEFT)
        up_btn = pill_btn(bf, "⬇  ดาวน์โหลด & อัปเดต", _update, bg=C_GREEN, fg=C_BG, pad_x=14, pad_y=8)
        up_btn.pack(side=tk.LEFT, padx=(8,0)); up_btn.config(state="disabled")
        pill_btn(bf, "ปิด", win.destroy, bg=C_SURFACE, fg=C_MUTED, pad_x=14, pad_y=8).pack(side=tk.RIGHT)
        win.after(300, _check)   # เช็กอัตโนมัติเมื่อเปิดหน้า

    def _startup_update_check(self):
        """เช็กอัปเดตเงียบๆ ตอนเปิดโปรแกรม — ถ้ามีใหม่แสดง popup"""
        import threading
        if _updater is None: return
        if (get_setting("update_autocheck") or "1") == "0": return   # ปิดเช็กอัตโนมัติ
        ch = get_setting("update_channel") or "stable"
        def work():
            try:
                r = _updater.check_for_update(APP_VERSION, channel=ch)
            except Exception:
                return
            if r.get("available") and r.get("url"):
                def popup():
                    note = (r.get("notes") or "").strip()
                    if len(note) > 280: note = note[:280] + "..."
                    if messagebox.askyesno("พบอัปเดตใหม่",
                        f"พบเวอร์ชันใหม่ v{r['latest']} (ปัจจุบัน v{APP_VERSION})\n\n"
                        f"{note}\n\nต้องการอัปเดตตอนนี้หรือไม่?"):
                        self._show_update_page()
                self.after(0, popup)
        threading.Thread(target=work, daemon=True).start()

    def _show_backup_page(self):
        """หน้าจอ Backup & Recovery"""
        win=tk.Toplevel(self); win.title("สำรองข้อมูล & กู้คืน")
        win.configure(bg=C_BG); win.geometry("720x600"); win.lift()
        BG=C_BG; SF=C_SURFACE; ACC=C_ACCENT; TX=C_TEXT; MT=C_MUTED
        YL=C_YELLOW; GR=C_GREEN; RD=C_ACCENT2; BD=C_BORDER

        tk.Frame(win,bg=ACC,height=4).pack(fill=tk.X)
        hf=tk.Frame(win,bg=BG,padx=16,pady=10); hf.pack(fill=tk.X)
        tk.Label(hf,text="💾  สำรองข้อมูล & กู้คืน",font=F_H1,bg=BG,fg=TX).pack(side=tk.LEFT)

        nb=ttk.Notebook(win); nb.pack(fill=tk.BOTH,expand=True,padx=12,pady=(0,8))

        # ── Tab 1: สถานะ & Backup ──
        t1=tk.Frame(nb,bg=BG); nb.add(t1,text="  💾 สำรองข้อมูล  ")
        if HAS_BACKUP:
            s=get_backup_status()
            # Status card
            sc=tk.Frame(t1,bg=SF,padx=16,pady=12,highlightbackground=BD,highlightthickness=1)
            sc.pack(fill=tk.X,padx=12,pady=(10,6))
            tk.Label(sc,text="สถานะ Backup ล่าสุด",font=F_H2,bg=SF,fg=TX).pack(anchor="w")
            status_color=GR if s["last_success"] else RD
            status_icon="✅" if s["last_success"] else "❌"
            tk.Label(sc,text=f"{status_icon} {s['last_backup']}  |  {s['last_size']}  |  {s['backup_count']} ไฟล์",
                     font=F_BODY,bg=SF,fg=status_color).pack(anchor="w",pady=(4,0))
            if s["error_msg"]:
                tk.Label(sc,text=f"⚠️ {s['error_msg']}",font=F_SMALL,bg=SF,fg=YL).pack(anchor="w")
            # Multi-location status
            try:
                for icon,lbl,typ in get_multi_status_summary():
                    col={"ok":GR,"error":RD,"skip":MT}.get(typ,MT)
                    tk.Label(sc,text=f"{icon} {lbl}",font=F_SMALL,bg=SF,fg=col).pack(anchor="w")
            except: pass
            # GDrive status
            gc=tk.Frame(t1,bg=SF,padx=16,pady=12,highlightbackground=BD,highlightthickness=1)
            gc.pack(fill=tk.X,padx=12,pady=(0,6))
            tk.Label(gc,text="Google Drive",font=F_H2,bg=SF,fg=TX).pack(anchor="w")
            gdrive_color=GR if s["gdrive_ok"] else MT
            tk.Label(gc,text=("☁️ " if s["gdrive_ok"] else "⭕ ")+s["gdrive_account"],
                     font=F_BODY,bg=SF,fg=gdrive_color).pack(anchor="w",pady=(4,0))
            gbf=tk.Frame(gc,bg=SF); gbf.pack(anchor="w",pady=(6,0))
            if not s["gdrive_ok"]:
                def do_connect():
                    ok,msg=connect_gdrive()
                    messagebox.showinfo("Google Drive",
                        f"เชื่อมต่อสำเร็จ: {msg}" if ok else f"ไม่สำเร็จ: {msg}",parent=win)
                    win.destroy(); self._show_backup_page()
                tk.Button(gbf,text="🔗 เชื่อมต่อ Google Drive",command=do_connect,
                          font=F_BODY,bg=ACC,fg=BG,relief=tk.FLAT,cursor="hand2",
                          padx=10,pady=4).pack(side=tk.LEFT)
            else:
                def do_disconnect():
                    if messagebox.askyesno("ยืนยัน","ยกเลิกการเชื่อมต่อ Google Drive?",parent=win):
                        disconnect_gdrive(); win.destroy(); self._show_backup_page()
                tk.Button(gbf,text="🔗 ยกเลิกการเชื่อมต่อ",command=do_disconnect,
                          font=F_BODY,bg=SF,fg=RD,relief=tk.FLAT,cursor="hand2",
                          padx=10,pady=4).pack(side=tk.LEFT)
            # Manual backup buttons
            bf=tk.Frame(t1,bg=BG,padx=12); bf.pack(fill=tk.X,pady=4)
            def do_local_backup():
                import threading
                btn_local.config(state=tk.DISABLED,text="⏳ กำลัง Backup หลายตำแหน่ง...")
                def run():
                    r=create_multi_location_backup("manual")
                    lines=[]
                    lines.append("✅ Local Backup สำเร็จ" if r["local"]["ok"] else "❌ Local Backup ล้มเหลว")
                    if r["drive_d"].get("found"):
                        lines.append("✅ Drive D สำเร็จ" if r["drive_d"]["ok"] else f"❌ Drive D: {r['drive_d']['msg']}")
                    else:
                        lines.append("⭕ Drive D — ไม่พบ")
                    for ext in r.get("external",[]):
                        lines.append(f"✅ USB {ext['drive']} สำเร็จ" if ext["ok"] else f"❌ USB {ext['drive']}: {ext['msg']}")
                    if not r.get("external"):
                        lines.append("⭕ USB Drive — ไม่พบ")
                    lines.append("✅ Google Drive สำเร็จ" if r["gdrive"]["ok"] else f"⭕ Google Drive — {r['gdrive']['msg']}")
                    msg="\n".join(lines)
                    win.after(0,lambda:(
                        messagebox.showinfo("Backup สำเร็จ",msg,parent=win),
                        btn_local.config(state=tk.NORMAL,text="💾 Backup เฉพาะเครื่อง (หลายตำแหน่ง)")))
                threading.Thread(target=run,daemon=True).start()
            def do_gdrive_backup():
                import threading
                btn_gdrive.config(state=tk.DISABLED,text="กำลัง Upload...")
                def run():
                    path=create_local_backup("manual")
                    if path:
                        ok=upload_to_gdrive(path,"Manual")
                        msg="Upload สำเร็จ!" if ok else "Upload ไม่สำเร็จ ดู error ใน backup/backup.log"
                    else: msg="Backup ไม่สำเร็จ"
                    win.after(0,lambda:(
                        messagebox.showinfo("Google Drive",msg,parent=win),
                        btn_gdrive.config(state=tk.NORMAL,text="☁️ Backup ไป Google Drive")))
                threading.Thread(target=run,daemon=True).start()
            btn_local=tk.Button(bf,text="💾 Backup เฉพาะเครื่อง (หลายตำแหน่ง)",command=do_local_backup,
                                font=F_H2,bg=ACC,fg=BG,relief=tk.FLAT,cursor="hand2",pady=10)
            btn_local.pack(fill=tk.X,pady=(0,4))
            btn_gdrive=tk.Button(bf,text="☁️ Backup ไป Google Drive",command=do_gdrive_backup,
                                 font=F_H2,bg="#1565c0",fg=TX,relief=tk.FLAT,cursor="hand2",pady=10)
            btn_gdrive.pack(fill=tk.X)
        else:
            tk.Label(t1,text="⚠️ ไม่พบ backup_system.py\nกรุณาวางไฟล์ไว้ที่ C:\\POS\\",
                     font=F_H2,bg=BG,fg=YL,justify=tk.CENTER).pack(expand=True)

        # ── Tab 2: กู้คืนข้อมูล ──
        t2=tk.Frame(nb,bg=BG); nb.add(t2,text="  🔄 กู้คืนข้อมูล  ")
        tk.Label(t2,text="⚠️ การกู้คืนจะแทนที่ข้อมูลปัจจุบัน กรุณาระวัง",
                 font=F_SMALL,bg=BG,fg=YL,pady=6).pack(anchor="w",padx=12)
        if HAS_BACKUP:
            backups=list_local_backups()
            tf=tk.Frame(t2,bg=BG); tf.pack(fill=tk.BOTH,expand=True,padx=12)
            cols=("ชื่อไฟล์","วันที่","ขนาด","ประเภท")
            tv=ttk.Treeview(tf,columns=cols,show="headings",height=10,style="R.Treeview")
            for col,w in zip(cols,[260,120,80,70]):
                tv.heading(col,text=col); tv.column(col,width=w,anchor="w" if col=="ชื่อไฟล์" else "center")
            sb=ttk.Scrollbar(tf,orient="vertical",command=tv.yview,style="R.Vertical.TScrollbar")
            tv.configure(yscrollcommand=sb.set)
            tv.pack(side=tk.LEFT,fill=tk.BOTH,expand=True); sb.pack(side=tk.RIGHT,fill=tk.Y)
            for b in backups:
                tv.insert("","end",iid=b["path"],values=(b["name"],b["date"],b["size"],b["type"]))
            def do_restore():
                sel=tv.selection()
                if not sel: messagebox.showwarning("เลือก","กรุณาเลือก Backup ก่อน",parent=win); return
                path=sel[0]
                if not messagebox.askyesno("ยืนยันกู้คืน",
                    f"กู้คืนจาก:\n{os.path.basename(path)}\n\nข้อมูลปัจจุบันจะถูกแทนที่\nต้องการดำเนินการต่อ?",
                    parent=win): return
                ok,msg=restore_from_zip(path)
                messagebox.showinfo("กู้คืน",msg,parent=win)
                if ok: win.destroy()
            def browse_restore():
                from tkinter import filedialog
                path=filedialog.askopenfilename(title="เลือกไฟล์ ZIP",
                    filetypes=[("ZIP","*.zip"),("All","*.*")],parent=win)
                if path:
                    if not messagebox.askyesno("ยืนยัน",f"กู้คืนจาก:\n{path}",parent=win): return
                    ok,msg=restore_from_zip(path)
                    messagebox.showinfo("กู้คืน",msg,parent=win)
                    if ok: win.destroy()
            rbf=tk.Frame(t2,bg=BG,padx=12,pady=8); rbf.pack(fill=tk.X)
            tk.Button(rbf,text="🔄 กู้คืนจาก Backup ที่เลือก",command=do_restore,
                      font=F_H2,bg="#c0392b",fg=TX,relief=tk.FLAT,cursor="hand2",pady=8
                      ).pack(side=tk.LEFT,fill=tk.X,expand=True,padx=(0,4))
            tk.Button(rbf,text="📂 เลือกไฟล์ ZIP...",command=browse_restore,
                      font=F_BODY,bg=SF,fg=MT,relief=tk.FLAT,cursor="hand2",pady=8
                      ).pack(side=tk.LEFT)

    def _tick(self):
        now = datetime.now()
        # แยก date/time — update label แยกกัน ไม่ดัน layout
        d_th = f"{now.day:02d}/{now.month:02d}/{now.year+543}"
        self.date_lbl.config(text=f"📅 {d_th}")
        self.clk.config(text=f"🕐 {now.strftime('%H:%M:%S')}")
        self.after(1000, self._tick)

    # ══ SHIFT (Open/Close) — Feature Phase 3 ════════════════
    def _refresh_shift_label(self):
        """อัปเดตปุ่มกะบนแถบบน ตามสถานะกะปัจจุบัน"""
        try:
            import shift_service
            if not hasattr(self, "_topbar_shift_lbl"):
                return
            if shift_service.get_open_shift():
                self._topbar_shift_lbl.config(text="🔒  ปิดกะ", fg=C_YELLOW)
                self._shift_lbl_fg = C_YELLOW
            else:
                self._topbar_shift_lbl.config(text="🕐  เปิดกะ", fg=C_BLUE)
                self._shift_lbl_fg = C_BLUE
        except Exception as e:
            print(f"[Shift] refresh label: {e}")

    def _shift_action(self):
        """คลิกปุ่มกะ — เปิดหรือปิดตามสถานะ"""
        import shift_service
        if shift_service.get_open_shift():
            self._close_shift_dialog()
        else:
            self._open_shift_dialog()

    def _open_shift_dialog(self):
        import shift_service
        win = tk.Toplevel(self); win.title("เปิดกะ"); win.configure(bg=C_BG)
        win.grab_set(); win.resizable(False, False); win.lift()
        tk.Frame(win, bg=C_BLUE, height=4).pack(fill=tk.X)
        tk.Label(win, text="🕐  เปิดกะใหม่", font=F_H1, bg=C_BG, fg=C_TEXT, pady=10).pack()
        tk.Label(win, text=f"พนักงาน: {current_staff.get('name','-')}",
                 font=F_BODY, bg=C_BG, fg=C_MUTED).pack()
        bf = tk.Frame(win, bg=C_BG, padx=24, pady=12); bf.pack(fill=tk.X)
        tk.Label(bf, text="เงินสดตั้งต้นในลิ้นชัก (฿)",
                 font=F_BODY, bg=C_BG, fg=C_TEXT).pack(anchor="w")
        e = field(bf, font=(FM.primary, 18, "bold"), justify="right")
        e.insert(0, "0"); e.pack(fill=tk.X, ipady=6, pady=(4, 0))
        e.focus_set(); e.select_range(0, tk.END)

        def do_open():
            try: oc = float(e.get() or 0)
            except: oc = 0.0
            try:
                sid = shift_service.open_shift(
                    current_staff.get("id", 0), current_staff.get("name", ""), oc)
            except Exception as ex:
                messagebox.showwarning("เปิดกะไม่สำเร็จ", str(ex), parent=win); return
            win.destroy(); self._refresh_shift_label()
            messagebox.showinfo("✅ เปิดกะแล้ว",
                f"เปิดกะ #{sid}\nเงินตั้งต้น ฿{oc:,.2f}", parent=self)

        win.bind("<Return>", lambda ev: do_open())
        btnf = tk.Frame(win, bg=C_BG, padx=24); btnf.pack(fill=tk.X, pady=(0, 14))
        accent_btn(btnf, "✔  เปิดกะ", do_open).pack(side=tk.LEFT)
        pill_btn(btnf, "ยกเลิก", win.destroy, bg=C_SURFACE, fg=C_MUTED).pack(side=tk.RIGHT)

    def _close_shift_dialog(self):
        import shift_service
        sh = shift_service.get_open_shift()
        if not sh:
            self._refresh_shift_label(); return
        summ = shift_service.get_shift_summary(sh["id"])

        win = tk.Toplevel(self); win.title("ปิดกะ"); win.configure(bg=C_BG)
        win.grab_set(); win.resizable(False, False); win.lift()
        tk.Frame(win, bg=C_YELLOW, height=4).pack(fill=tk.X)
        tk.Label(win, text="🔒  ปิดกะ", font=F_H1, bg=C_BG, fg=C_TEXT, pady=10).pack()

        info = tk.Frame(win, bg=C_SURFACE, padx=20, pady=12)
        info.pack(fill=tk.X, padx=16)
        def row(lbl, val, color=C_TEXT):
            r = tk.Frame(info, bg=C_SURFACE); r.pack(fill=tk.X, pady=2)
            tk.Label(r, text=lbl, font=F_BODY, bg=C_SURFACE, fg=C_MUTED).pack(side=tk.LEFT)
            tk.Label(r, text=val, font=F_BODY, bg=C_SURFACE, fg=color).pack(side=tk.RIGHT)
        row("พนักงาน", sh.get("staff_name") or "-")
        row("เวลาเปิด", sh.get("open_time") or "-")
        row("เงินตั้งต้น", f"฿{float(sh['opening_cash']):,.2f}")
        row("ยอดขายในกะ", f"฿{summ['total_sales']:,.2f}  ({summ['total_bills']} บิล)")
        row("ขายเงินสด", f"฿{summ['cash_sales']:,.2f}", C_GREEN)
        row("เงินสดที่ควรมี", f"฿{summ['expected_cash']:,.2f}", C_ACCENT)

        cf = tk.Frame(win, bg=C_BG, padx=20, pady=10); cf.pack(fill=tk.X)
        tk.Label(cf, text="นับเงินสดจริงในลิ้นชัก (฿)",
                 font=F_BODY, bg=C_BG, fg=C_TEXT).pack(anchor="w")
        e = field(cf, font=(FM.primary, 18, "bold"), justify="right")
        e.insert(0, f"{summ['expected_cash']:.2f}"); e.pack(fill=tk.X, ipady=6, pady=(4, 0))
        var_lbl = tk.Label(win, text="", font=F_H2, bg=C_BG); var_lbl.pack(pady=(2, 4))

        def upd_var(*_):
            try: counted = float(e.get() or 0)
            except: counted = 0.0
            diff = round(counted - summ["expected_cash"], 2)
            if diff == 0:   var_lbl.config(text="ตรงพอดี ✓", fg=C_GREEN)
            elif diff > 0:  var_lbl.config(text=f"เกิน ฿{diff:,.2f}", fg=C_YELLOW)
            else:           var_lbl.config(text=f"ขาด ฿{abs(diff):,.2f}", fg=C_ACCENT2)
        e.bind("<KeyRelease>", upd_var); upd_var()
        e.focus_set(); e.select_range(0, tk.END)

        def do_close():
            try: counted = float(e.get() or 0)
            except: counted = 0.0
            try:
                res = shift_service.close_shift(sh["id"], counted)
            except Exception as ex:
                messagebox.showwarning("ปิดกะไม่สำเร็จ", str(ex), parent=win); return
            win.destroy(); self._refresh_shift_label()
            sign = "เกิน" if res["variance"] > 0 else ("ขาด" if res["variance"] < 0 else "ตรงพอดี")
            messagebox.showinfo("✅ ปิดกะแล้ว",
                f"กะ #{res['shift_id']}\n"
                f"ยอดขาย ฿{res['total_sales']:,.2f} ({res['total_bills']} บิล)\n"
                f"ควรมี ฿{res['expected_cash']:,.2f} | นับได้ ฿{res['counted']:,.2f}\n"
                f"ส่วนต่าง: {sign} ฿{abs(res['variance']):,.2f}", parent=self)

        win.bind("<Return>", lambda ev: do_close())
        btnf = tk.Frame(win, bg=C_BG, padx=20); btnf.pack(fill=tk.X, pady=(0, 14))
        accent_btn(btnf, "✔  ยืนยันปิดกะ", do_close).pack(side=tk.LEFT)
        pill_btn(btnf, "ยกเลิก", win.destroy, bg=C_SURFACE, fg=C_MUTED).pack(side=tk.RIGHT)

if __name__=="__main__":
    init_db(); App().mainloop()
